#!/usr/bin/env python3
"""
DeskEli - Sistema Completo de Gestión de Incidencias
Con: LDAP, JWT, SLA, Exportación, Temas, Audit Trail, Búsqueda Global
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, has_request_context
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, emit, join_room, leave_room
from flask_wtf.csrf import CSRFProtect
from sqlalchemy import func
from datetime import datetime, timedelta
import jwt
import os
import json
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

# GUARDAR CLAUDE API KEY COMO VARIABLE GLOBAL
CLAUDE_API_KEY = os.getenv('ANTHROPIC_API_KEY', '').strip()
print(f"[APP] CLAUDE_API_KEY cargada: {bool(CLAUDE_API_KEY)}")

from io import BytesIO, StringIO
import csv
import openpyxl
from openpyxl.styles import PatternFill
import uuid
import requests
from threading import Thread
import socket
from collections import defaultdict
import time
import gzip
import shutil
from pathlib import Path
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import bleach
import re
import hashlib
import secrets

# ═════════════════════════════════════════════════════════════════════════════
# RATE LIMITING RNF-03-07 - 120 req/min por IP
# ═════════════════════════════════════════════════════════════════════════════

request_counts = defaultdict(list)
RATE_LIMIT = 120  # requests
RATE_WINDOW = 60  # seconds
MAX_IPS_TRACKED = 10000  # Prevenir memory leak

# Bloqueo de cuenta por intentos fallidos (configurable vía .env)
MAX_FAILED_LOGIN_ATTEMPTS = int(os.getenv('MAX_FAILED_LOGIN_ATTEMPTS', '5'))
LOCKOUT_DURATION_MINUTES = int(os.getenv('LOCKOUT_DURATION_MINUTES', '15'))

def rate_limit_check():
    """Middleware de rate limiting (120 req/min por IP) - SECURITY FIX 8: mejorar cleanup"""
    global request_counts
    ip = request.remote_addr
    now = time.time()

    # SECURITY FIX 8: Limpiar requests antiguos más eficientemente
    request_counts[ip] = [t for t in request_counts[ip] if now - t < RATE_WINDOW]

    # Verificar límite
    if len(request_counts[ip]) >= RATE_LIMIT:
        return jsonify({'success': False, 'error': 'Rate limit exceeded: 120 req/min'}), 429

    # Registrar request
    request_counts[ip].append(now)

    # SECURITY FIX 8: Limpiar IPs antiguas si dict crece demasiado (memory leak prevention)
    # Limpiar IPs que no tengan requests en la última ventana
    if len(request_counts) > MAX_IPS_TRACKED:
        # Remover IPs con requests expirados (sin requests activos)
        expired_ips = [ip_key for ip_key, times in request_counts.items()
                      if not times or max(times) < now - RATE_WINDOW]
        for expired_ip in expired_ips[:len(expired_ips) // 2]:  # Remover mitad de expirados
            if expired_ip in request_counts:
                del request_counts[expired_ip]

    return None

# ═════════════════════════════════════════════════════════════════════════════
# CONFIGURACION
# ═════════════════════════════════════════════════════════════════════════════

from crypto_utils import (
    init_crypto, encrypt_secret, decrypt_secret,
    encrypt_bytes, decrypt_bytes, has_key as crypto_has_key,
)
from password_policy import validate_password
from ldap_auth import LdapConfig, authenticate_user as ldap_authenticate_user, test_ldap_connection, LDAP_AVAILABLE
from file_compression import compress_upload
from microsoft_auth import (
    build_auth_url as ms_build_auth_url,
    exchange_code_for_token as ms_exchange_code,
    get_user_info as ms_get_user_info,
    extract_user_data_from_token as ms_extract_token_data,
    MSAL_AVAILABLE,
)

app = Flask(__name__)
# Usar BD en directorio raíz, no en instance/
db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ticketdesk_v2.db')

# Leer DATABASE_URL del .env o usar SQLite por defecto
db_url = os.getenv('DATABASE_URL', f'sqlite:///{db_path}')
# Convertir postgres:// a postgresql://
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url

# SECRET_KEY: en producción es obligatoria; en dev se usa un valor temporal con warning
_FLASK_ENV = os.getenv('FLASK_ENV', 'development').lower()
_IS_PRODUCTION = _FLASK_ENV == 'production'
_secret_key = os.getenv('SECRET_KEY')
if not _secret_key:
    if _IS_PRODUCTION:
        raise RuntimeError(
            'SECRET_KEY no está definida en el entorno y FLASK_ENV=production. '
            'Generala con: python -c "import secrets; print(secrets.token_urlsafe(32))" '
            'y agregala al archivo .env'
        )
    _secret_key = 'dev_key_do_not_use_in_production_change_this'
    print('[WARN] SECRET_KEY no definida — usando valor de desarrollo. NO usar en producción.')
app.config['SECRET_KEY'] = _secret_key

# Inicializar cifrado de secretos en BD (Fernet con DB_ENCRYPTION_KEY)
init_crypto(_IS_PRODUCTION)

app.config['SESSION_COOKIE_HTTPONLY'] = True
# Cookies solo por HTTPS en producción (HTTP local en desarrollo)
app.config['SESSION_COOKIE_SECURE'] = _IS_PRODUCTION
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['PREFERRED_URL_SCHEME'] = 'https' if _IS_PRODUCTION else 'http'

# Carpetas de uploads (adjuntos)
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'subtasks')
TICKET_UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'tickets')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TICKET_UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TICKET_UPLOAD_FOLDER'] = TICKET_UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 25 * 1024 * 1024  # 25 MB por archivo

ALLOWED_EXTENSIONS = {
    'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'txt', 'csv',
    'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp',
    'zip', 'rar', '7z',
    'log', 'msg', 'eml'
}

def _allowed_attachment(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db = SQLAlchemy(app)

# SECURITY FIX 5: Reducir CORS a dominios específicos (no "*")
# H-05: en producción se descartan orígenes localhost / 127.0.0.1 aunque estén en .env
_raw_origins = [o.strip() for o in os.getenv('ALLOWED_ORIGINS', 'http://localhost:5050').split(',') if o.strip()]
if _IS_PRODUCTION:
    allowed_origins = [o for o in _raw_origins
                       if 'localhost' not in o.lower() and '127.0.0.1' not in o]
    if not allowed_origins:
        raise RuntimeError(
            'ALLOWED_ORIGINS solo contiene localhost/127.0.0.1 en FLASK_ENV=production. '
            'Definí al menos un origen real (ej. http://10.161.55.5:5050) en .env'
        )
    _dropped = [o for o in _raw_origins if o not in allowed_origins]
    if _dropped:
        print(f'[INFO] CORS producción: orígenes localhost descartados: {_dropped}')
else:
    allowed_origins = _raw_origins
CORS(app, resources={r"/api/*": {"origins": allowed_origins}}, supports_credentials=True)

# SECURITY FIX 4: CSRF protection con Flask-WTF
# Estrategia: opt-in. CSRFProtect se inicializa para generar tokens, pero el chequeo
# automático en cada request queda apagado (WTF_CSRF_CHECK_DEFAULT=False). Los forms HTML
# de auth (login, cambio de contraseña) validan el token manualmente.
# Los /api/* quedan defendidos por SameSite=Lax + CORS estricto + rate limit.
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_CHECK_DEFAULT'] = False  # validación manual en endpoints sensibles
app.config['WTF_CSRF_TIME_LIMIT'] = None      # token vivo durante la sesión
app.config['WTF_CSRF_SSL_STRICT'] = _IS_PRODUCTION
csrf = CSRFProtect(app)

socketio = SocketIO(app, cors_allowed_origins=allowed_origins, async_mode='threading')

# Rate limiting middleware
@app.before_request
def apply_rate_limit():
    """Aplicar rate limit a endpoints de API (RNF-03-07)"""
    if request.path.startswith('/api/'):
        result = rate_limit_check()
        if result:
            return result

def get_public_base_url():
    """Devuelve la URL pública base del sistema, siempre correcta incluso desde
    threads background (schedulers) donde no hay request context.

    Prioridad:
    1. Env var PUBLIC_URL (recomendado en producción, ej: https://deskeli.eliotproyectos.tech)
    2. request.host_url si hay contexto Flask activo
    3. Primera URL válida (no-localhost) de ALLOWED_ORIGINS
    4. Fallback: http://localhost:5050
    """
    # 1. Variable de entorno explícita
    public_url = os.getenv('PUBLIC_URL', '').strip().rstrip('/')
    if public_url:
        return public_url

    # 2. Request context activo
    try:
        if has_request_context():
            return request.host_url.rstrip('/')
    except Exception:
        pass

    # 3. Buscar en ALLOWED_ORIGINS la primera URL no-localhost
    origins = os.getenv('ALLOWED_ORIGINS', '').split(',')
    for origin in origins:
        origin = origin.strip().rstrip('/')
        if origin and 'localhost' not in origin.lower() and '127.0.0.1' not in origin:
            return origin

    # 4. Fallback
    return 'http://localhost:5050'


# SECURITY FIX 2: @after_request decorator con security headers
@app.after_request
def apply_security_headers(response):
    """Aplicar security headers (HSTS, CSP, X-Frame-Options, etc.)"""
    # HSTS: Force HTTPS por 1 año
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'

    # CSP: Content Security Policy - prevenir XSS
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline'; img-src 'self' data:; font-src 'self'"

    # X-Frame-Options: prevenir clickjacking
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'

    # X-Content-Type-Options: prevenir MIME sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'

    # X-XSS-Protection: habilitar XSS filter
    response.headers['X-XSS-Protection'] = '1; mode=block'

    # Referrer-Policy: limitar referrer info
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'

    # Permissions-Policy: deshabilitar features no necesarias
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'

    # CACHE para assets estáticos: 30 días (CDN-friendly)
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=2592000, immutable'
    elif request.path.startswith('/api/'):
        # Las APIs no se cachean (datos dinámicos)
        response.headers['Cache-Control'] = 'no-store'

    return response

# Colores por empresa
COMPANY_COLORS = {
    'eliot':     {'name': 'Manufacturas Eliot', 'primary': '#2563eb', 'secondary': '#1e40af', 'icon': '🏭', 'logo': '/static/img/eliot.jpg'},
    'pash':      {'name': 'Pash',                 'primary': '#7c3aed', 'secondary': '#6d28d9', 'icon': '💻', 'logo': '/static/img/pash.png'},
    'primatela': {'name': 'Primatela',            'primary': '#059669', 'secondary': '#047857', 'icon': '🌿', 'logo': '/static/img/primatela.jpg'}
}
COMPANY_LOGOS = {code: cfg['logo'] for code, cfg in COMPANY_COLORS.items()}

THEMES = {
    'blue':     {'primary': '#2563eb', 'name': 'Azul Profesional'},
    'indigo':   {'primary': '#4f46e5', 'name': 'Índigo'},
    'purple':   {'primary': '#7c3aed', 'name': 'Púrpura'},
    'violet':   {'primary': '#8b5cf6', 'name': 'Violeta'},
    'fuchsia':  {'primary': '#d946ef', 'name': 'Fucsia'},
    'pink':     {'primary': '#ec4899', 'name': 'Rosa'},
    'rose':     {'primary': '#f43f5e', 'name': 'Rosado'},
    'red':      {'primary': '#dc2626', 'name': 'Rojo'},
    'orange':   {'primary': '#ea580c', 'name': 'Naranja'},
    'amber':    {'primary': '#d97706', 'name': 'Ámbar'},
    'yellow':   {'primary': '#ca8a04', 'name': 'Amarillo'},
    'lime':     {'primary': '#84cc16', 'name': 'Lima'},
    'green':    {'primary': '#059669', 'name': 'Verde'},
    'emerald':  {'primary': '#10b981', 'name': 'Esmeralda'},
    'teal':     {'primary': '#0d9488', 'name': 'Teal'},
    'cyan':     {'primary': '#0891b2', 'name': 'Cian'},
    'sky':      {'primary': '#0284c7', 'name': 'Cielo'},
    'navy':     {'primary': '#1e3a8a', 'name': 'Azul Marino'},
    'slate':    {'primary': '#475569', 'name': 'Pizarra'},
    'gray':     {'primary': '#4b5563', 'name': 'Gris'},
    'stone':    {'primary': '#78716c', 'name': 'Piedra'},
    'brown':    {'primary': '#92400e', 'name': 'Café'},
    'midnight': {'primary': '#111827', 'name': 'Medianoche'},
    'coral':    {'primary': '#fb7185', 'name': 'Coral'},
    'orange_black': {
        'primary': '#ea580c',
        'name': 'Naranja → Negro (degradado)',
        'gradient': 'linear-gradient(135deg, #ea580c 0%, #1f2937 100%)'
    },
    'purple_black': {
        'primary': '#7c3aed',
        'name': 'Púrpura → Negro (degradado)',
        'gradient': 'linear-gradient(135deg, #7c3aed 0%, #1f2937 100%)'
    }
}

# ═════════════════════════════════════════════════════════════════════════════
# MODELOS DE DATOS
# ═════════════════════════════════════════════════════════════════════════════

class Company(db.Model):
    __tablename__ = 'companies'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)  # eliot, pash, primatela
    name = db.Column(db.String(100), nullable=False)
    icon = db.Column(db.String(50), default='🏢')
    primary_color = db.Column(db.String(7), default='#2563eb')
    secondary_color = db.Column(db.String(7), default='#1e40af')
    ldap_server = db.Column(db.String(255))
    ldap_base_dn = db.Column(db.String(255))
    ldap_bind_user = db.Column(db.String(255))
    ldap_bind_password = db.Column(db.String(255))
    # Microsoft Entra ID (Azure AD) — para autenticación por OAuth 2.0
    microsoft_tenant_id = db.Column(db.String(100))         # GUID del tenant
    microsoft_client_id = db.Column(db.String(100))         # Application (client) ID
    microsoft_client_secret = db.Column(db.String(500))     # Client secret cifrado
    microsoft_enabled = db.Column(db.Boolean, default=False)  # ¿La empresa usa Microsoft?
    # Configuración SMTP por empresa (si está, sobrescribe la global)
    smtp_host = db.Column(db.String(255))
    smtp_port = db.Column(db.Integer)
    smtp_user = db.Column(db.String(255))
    smtp_password = db.Column(db.String(255))
    smtp_from = db.Column(db.String(255))
    smtp_security = db.Column(db.String(10))  # 'tls' o 'ssl'
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # admin, technician, employee (base role para permisos)
    role_label = db.Column(db.String(80))  # Rol personalizado PRIMARIO (legacy / display por defecto)
    extra_role_labels = db.Column(db.Text)  # Lista adicional de roles personalizados (JSON array de etiquetas)
    company = db.Column(db.String(20), nullable=False, index=True)  # eliot, pash, primatela
    password_hash = db.Column(db.String(255))
    is_active = db.Column(db.Boolean, default=True)
    last_login = db.Column(db.DateTime)
    force_logout_at = db.Column(db.DateTime)  # Si > session.login_at → sesión expulsada por admin
    must_change_password = db.Column(db.Boolean, default=False)  # Forzar cambio en próximo login
    failed_login_attempts = db.Column(db.Integer, default=0)     # Contador de intentos fallidos consecutivos
    locked_until = db.Column(db.DateTime)                        # Si > now() la cuenta está bloqueada
    microsoft_object_id = db.Column(db.String(100), unique=True, index=True)  # OID de Microsoft Entra (identificador único)
    # Datos de contacto del usuario (se pre-rellenan al crear un ticket)
    area = db.Column(db.String(120))      # Área/departamento
    location = db.Column(db.String(120))  # Sede/oficina/piso
    phone = db.Column(db.String(40))      # Teléfono / extensión
    # Si el usuario es un ESPEJO de un tecnico de Eliot en pash/primatela,
    # aqui va el id del usuario origen en eliot. NULL = usuario local normal.
    mirrored_from_id = db.Column(db.Integer, index=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    __table_args__ = (db.UniqueConstraint('username', 'company', name='_user_company_uc'),)

class Ticket(db.Model):
    __tablename__ = 'tickets'
    id = db.Column(db.Integer, primary_key=True)
    ticket_number = db.Column(db.String(50), unique=True, nullable=False)  # TKT-ELIOT-00001 | AUTO-YYYYMMDDHHMMSS-xxxx
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    category = db.Column(db.String(100), default='General')
    status = db.Column(db.String(20), default='open')  # open, in_progress, resolved
    priority = db.Column(db.String(20), default='medium')  # low, medium, high, critical
    company = db.Column(db.String(20), nullable=False)
    creator_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    assignee_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    sla_minutes = db.Column(db.Integer)
    sla_deadline = db.Column(db.DateTime)
    sla_alerts_sent = db.Column(db.String(20), default='')  # CSV de thresholds enviados, ej: "30,60,100"
    rating = db.Column(db.Integer)  # 1-5 stars (CSAT)
    rating_comment = db.Column(db.Text)  # Comentario textual del usuario al calificar
    rating_nps = db.Column(db.Integer)  # NPS 0-10 (¿recomendarías el servicio?)
    rating_at = db.Column(db.DateTime)  # cuándo se calificó
    reminder_sent_at = db.Column(db.DateTime)  # 48h después de resolver, si no calificó
    time_worked_seconds = db.Column(db.Integer, default=0)
    version = db.Column(db.Integer, default=1)  # Optimistic locking
    # Datos de contacto del solicitante (snapshot al momento de crear el ticket)
    user_area = db.Column(db.String(120))      # Área/departamento donde trabaja
    user_location = db.Column(db.String(120))  # Sede/edificio/piso
    user_phone = db.Column(db.String(40))      # Teléfono de contacto / extensión
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    resolved_at = db.Column(db.DateTime)
    creator = db.relationship('User', foreign_keys=[creator_id], backref='created_tickets')
    assignee = db.relationship('User', foreign_keys=[assignee_id], backref='assigned_tickets')
    messages = db.relationship('Message', backref='ticket', cascade='all, delete-orphan')

    @property
    def sla_remaining(self):
        if not self.sla_deadline:
            return None
        remaining = self.sla_deadline - datetime.now()
        return max(0, int(remaining.total_seconds() / 60))

class Message(db.Model):
    __tablename__ = 'messages'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False)
    subtask_id = db.Column(db.Integer, db.ForeignKey('subtasks.id'), nullable=True, index=True)  # Si != null, es un comentario de subtarea
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    text = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User', backref='messages')

class TokenBlacklist(db.Model):
    __tablename__ = 'token_blacklist'
    id = db.Column(db.Integer, primary_key=True)
    jti = db.Column(db.String(255), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)


class ApiKey(db.Model):
    """API Key para integraciones externas (proveedores que crean tickets vía API)."""
    __tablename__ = 'api_keys'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)          # Ej: "Proveedor Aranda", "Sistema RRHH"
    token_prefix = db.Column(db.String(12), nullable=False)   # Primeros 8 chars visibles (para reconocer sin descifrar)
    token_hash = db.Column(db.String(128), unique=True, nullable=False)  # SHA-256 del token completo
    company = db.Column(db.String(20), nullable=False, index=True)       # Empresa asociada
    scopes = db.Column(db.String(200), default='tickets:create,tickets:read')  # CSV de scopes
    is_active = db.Column(db.Boolean, default=True)
    created_by = db.Column(db.Integer)                        # user_id del admin que la creó
    created_at = db.Column(db.DateTime, default=datetime.now)
    last_used_at = db.Column(db.DateTime)
    last_used_ip = db.Column(db.String(45))
    expires_at = db.Column(db.DateTime)                       # Opcional; NULL = no expira
    usage_count = db.Column(db.Integer, default=0)


class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    action = db.Column(db.String(100), nullable=False)
    user_id = db.Column(db.Integer)
    entity_type = db.Column(db.String(50))  # ticket, user, config
    entity_id = db.Column(db.Integer)
    description = db.Column(db.Text)
    ip_addr = db.Column(db.String(45))
    created_at = db.Column(db.DateTime, default=datetime.now)

class Config(db.Model):
    __tablename__ = 'configs'
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class Template(db.Model):
    __tablename__ = 'templates'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    title_template = db.Column(db.String(200), nullable=False)
    description_template = db.Column(db.Text)
    category = db.Column(db.String(100))
    priority = db.Column(db.String(20), default='medium')
    company = db.Column(db.String(20), nullable=False)
    is_system = db.Column(db.Boolean, default=False)
    # form_fields: JSON array de campos personalizados que se renderizan como formulario
    # Ej: [{"name":"equipo","label":"Equipo afectado","type":"text","required":true,"placeholder":"Ej: SAP PRD"}]
    # Tipos soportados: text, textarea, select, date
    form_fields = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)

class Server(db.Model):
    __tablename__ = 'servers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    ip_address = db.Column(db.String(45), nullable=False)
    port = db.Column(db.Integer, default=443)
    description = db.Column(db.String(255))
    is_critical = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    company = db.Column(db.String(20), nullable=False)
    last_ping = db.Column(db.DateTime)
    last_status = db.Column(db.String(20), default='unknown')
    last_latency_ms = db.Column(db.Float, default=0)
    consecutive_failures = db.Column(db.Integer, default=0)
    alarm_active = db.Column(db.Boolean, default=False)
    is_online = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

class UserSession(db.Model):
    __tablename__ = 'user_sessions'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    session_token = db.Column(db.String(255), unique=True)
    ip_addr = db.Column(db.String(45))
    login_time = db.Column(db.DateTime, default=datetime.now)
    last_activity = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User', backref='sessions')

class Tag(db.Model):
    __tablename__ = 'tags'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    color = db.Column(db.String(20), default='#2563eb')
    icon = db.Column(db.String(10), default='🏷️')
    company = db.Column(db.String(20), nullable=False, index=True)
    description = db.Column(db.String(200), default='')
    usage_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)


class BotKnowledge(db.Model):
    __tablename__ = 'bot_knowledge'
    id = db.Column(db.Integer, primary_key=True)
    keywords = db.Column(db.String(500), nullable=False)  # Palabras clave separadas por comas
    question = db.Column(db.String(200), nullable=False)
    answer = db.Column(db.Text, nullable=False)
    category = db.Column(db.String(100))
    priority = db.Column(db.String(20), default='medium')  # Prioridad del ticket si se crea
    created_at = db.Column(db.DateTime, default=datetime.now)

class TechnicianProfile(db.Model):
    """Perfiles de habilidades de técnicos para asignación automática."""
    __tablename__ = 'technician_profiles'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, unique=True)
    company = db.Column(db.String(20), nullable=False, index=True)
    skills = db.Column(db.String(500), default='')  # CSV: Red,Hardware,Email
    skill_levels = db.Column(db.Text, default='{}')  # JSON: {"Red": 80, "Hardware": 60}
    max_tickets = db.Column(db.Integer, default=5)  # Máximo tickets activos
    is_available = db.Column(db.Boolean, default=True)
    avg_resolution_minutes = db.Column(db.Integer, default=0)
    tickets_resolved_total = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    user = db.relationship('User', backref='profile')

    def get_skills_list(self):
        return [s.strip() for s in self.skills.split(',') if s.strip()]

    def get_skill_level(self, skill_name):
        try:
            levels = json.loads(self.skill_levels or '{}')
            return int(levels.get(skill_name, 0))
        except:
            return 0

class AgentAction(db.Model):
    """Log de decisiones tomadas por los 4 agentes del orchestrator."""
    __tablename__ = 'agent_actions'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False, index=True)
    company = db.Column(db.String(20), nullable=False, index=True)
    agent_name = db.Column(db.String(50), nullable=False)  # classifier, assignor, responder, escalator
    action_type = db.Column(db.String(50), nullable=False)  # classify, assign, respond, escalate
    input_data = db.Column(db.Text)  # JSON con inputs
    output_data = db.Column(db.Text)  # JSON con resultados
    confidence = db.Column(db.Integer, default=0)  # 0-100
    used_llm = db.Column(db.Boolean, default=False)
    duration_ms = db.Column(db.Integer, default=0)
    success = db.Column(db.Boolean, default=True)
    error_msg = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now, index=True)
    ticket = db.relationship('Ticket', backref='agent_actions')

class Subrole(db.Model):
    """Catálogo de subroles/especializaciones técnicas (Infraestructura, SAP MM, etc.)."""
    __tablename__ = 'subroles'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    icon = db.Column(db.String(10), default='🔧')
    company = db.Column(db.String(20))  # NULL = global a todas las empresas
    is_system = db.Column(db.Boolean, default=False)  # los pre-seed no se pueden borrar
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    __table_args__ = (db.UniqueConstraint('name', 'company', name='_subrole_company_uc'),)


class UserSubrole(db.Model):
    """Asignación M:N entre usuarios y subroles."""
    __tablename__ = 'user_subroles'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    subrole_id = db.Column(db.Integer, db.ForeignKey('subroles.id'), nullable=False, index=True)
    assigned_at = db.Column(db.DateTime, default=datetime.now)
    __table_args__ = (db.UniqueConstraint('user_id', 'subrole_id', name='_user_subrole_uc'),)

    user = db.relationship('User', backref=db.backref('subrole_assignments', cascade='all, delete-orphan'))
    subrole = db.relationship('Subrole')


class MailboxConfig(db.Model):
    """Buzón IMAP del que el sistema lee correos para crear tickets automáticamente."""
    __tablename__ = 'mailbox_configs'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # ej. "Soporte General"
    company = db.Column(db.String(20), nullable=False, index=True)
    # Conexión IMAP
    imap_host = db.Column(db.String(200), nullable=False)
    imap_port = db.Column(db.Integer, default=993)
    imap_user = db.Column(db.String(200), nullable=False)
    imap_password = db.Column(db.String(500))  # texto plano (mejorable a cifrado)
    use_ssl = db.Column(db.Boolean, default=True)
    folder = db.Column(db.String(100), default='INBOX')
    # Autenticación: 'password' (legacy) o 'oauth2' (M365)
    auth_type = db.Column(db.String(20), default='password')
    oauth_tenant_id = db.Column(db.String(100))     # Azure AD Tenant ID
    oauth_client_id = db.Column(db.String(100))     # App registration Client ID
    oauth_client_secret = db.Column(db.String(500)) # App registration Secret
    # Reglas de creación de ticket
    default_priority = db.Column(db.String(20), default='medium')
    default_category = db.Column(db.String(100), default='Email')
    poll_interval_minutes = db.Column(db.Integer, default=5)
    # Estado
    is_active = db.Column(db.Boolean, default=True)
    last_check_at = db.Column(db.DateTime)
    last_status = db.Column(db.String(20))  # ok / error
    last_error = db.Column(db.Text)
    last_uid = db.Column(db.Integer, default=0)  # UID más alto procesado
    tickets_created = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)


class MailboxEmail(db.Model):
    """Registro de correos procesados para evitar duplicados."""
    __tablename__ = 'mailbox_emails'
    id = db.Column(db.Integer, primary_key=True)
    mailbox_id = db.Column(db.Integer, db.ForeignKey('mailbox_configs.id'), nullable=False, index=True)
    message_id = db.Column(db.String(500), index=True)  # Message-ID del email
    subject = db.Column(db.String(500))
    sender = db.Column(db.String(200))
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'))
    processed_at = db.Column(db.DateTime, default=datetime.now)
    __table_args__ = (db.UniqueConstraint('mailbox_id', 'message_id', name='_mailbox_message_uc'),)


class Webhook(db.Model):
    __tablename__ = 'webhooks'
    id = db.Column(db.Integer, primary_key=True)
    company = db.Column(db.String(20), nullable=False)
    url = db.Column(db.String(500), nullable=False)
    events = db.Column(db.String(200))  # JSON: ticket_created, ticket_resolved, sla_warning
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

class ReportRecipient(db.Model):
    """Destinatarios de reportes automáticos. Configurable desde admin UI."""
    __tablename__ = 'report_recipients'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(200), nullable=False)
    company = db.Column(db.String(20), nullable=False, index=True)  # eliot|pash|primatela
    title = db.Column(db.String(120))  # cargo: "Gerente TI", "Jefe Soporte"
    team_user_ids = db.Column(db.Text)  # JSON array de user.id que conforman el equipo. NULL/'' = toda la empresa
    # Especialistas CC: reciben el mismo reporte por email en copia (email de cada user).
    # NO afecta el filtro de datos — es solo para propagar el envio.
    cc_user_ids = db.Column(db.Text)     # JSON array de user.id a copiar en el envio
    send_quincenal = db.Column(db.Boolean, default=True)
    send_monthly = db.Column(db.Boolean, default=True)
    send_annual = db.Column(db.Boolean, default=True)
    # Alerta lunes: casos del grupo con SLA vencido o >6 dias abiertos
    send_monday_stuck = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    last_sent_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def get_team_ids(self):
        if not self.team_user_ids:
            return []
        try:
            return [int(x) for x in json.loads(self.team_user_ids) if str(x).isdigit() or isinstance(x, int)]
        except Exception:
            return []

    def set_team_ids(self, ids):
        ids = [int(i) for i in (ids or []) if str(i).isdigit() or isinstance(i, int)]
        self.team_user_ids = json.dumps(ids) if ids else None

    def get_cc_ids(self):
        if not self.cc_user_ids:
            return []
        try:
            return [int(x) for x in json.loads(self.cc_user_ids) if str(x).isdigit() or isinstance(x, int)]
        except Exception:
            return []

    def set_cc_ids(self, ids):
        ids = [int(i) for i in (ids or []) if str(i).isdigit() or isinstance(i, int)]
        self.cc_user_ids = json.dumps(ids) if ids else None


class Subtask(db.Model):
    __tablename__ = 'subtasks'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False, index=True)
    subtask_number = db.Column(db.String(40), unique=True)  # TKT-ELIOT-00063-S01
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text)
    category = db.Column(db.String(100), default='General')
    status = db.Column(db.String(20), default='open')  # open, in_progress, resolved
    priority = db.Column(db.String(20), default='medium')  # low, medium, high, critical
    sla_minutes = db.Column(db.Integer)
    sla_deadline = db.Column(db.DateTime)
    time_worked_seconds = db.Column(db.Integer, default=0)
    assignee_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_by_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    order_idx = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    resolved_at = db.Column(db.DateTime)
    completed_at = db.Column(db.DateTime)

    assignee = db.relationship('User', foreign_keys=[assignee_id])
    created_by = db.relationship('User', foreign_keys=[created_by_id])
    ticket = db.relationship('Ticket', backref=db.backref('subtasks', cascade='all, delete-orphan', order_by='Subtask.order_idx'))

    @property
    def sla_remaining(self):
        if not self.sla_deadline:
            return None
        remaining = self.sla_deadline - datetime.now()
        return max(0, int(remaining.total_seconds() / 60))

    @property
    def sla_expired(self):
        if not self.sla_deadline:
            return False
        return self.sla_deadline < datetime.now() and self.status != 'resolved'


class SubtaskAttachment(db.Model):
    __tablename__ = 'subtask_attachments'
    id = db.Column(db.Integer, primary_key=True)
    subtask_id = db.Column(db.Integer, db.ForeignKey('subtasks.id'), nullable=False, index=True)
    original_name = db.Column(db.String(255), nullable=False)
    stored_name = db.Column(db.String(255), nullable=False)  # nombre único en disco
    mime_type = db.Column(db.String(120))
    size_bytes = db.Column(db.Integer)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    uploaded_at = db.Column(db.DateTime, default=datetime.now)

    uploaded_by = db.relationship('User', foreign_keys=[uploaded_by_id])
    subtask = db.relationship('Subtask', backref=db.backref('attachments', cascade='all, delete-orphan'))


class Guion(db.Model):
    """Guión (script) preconfigurado que la API externa puede invocar.
    Cuando el proveedor manda 'guion_id' o 'guion_code', el sistema toma las
    subtareas y responsables ya definidos, sin necesidad de que el proveedor
    conozca emails ni estructuras."""
    __tablename__ = 'guiones'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False, index=True)   # 'auditoria-q1', 'onboarding-user'
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    company = db.Column(db.String(20), nullable=False, index=True)              # empresa dueña del guion
    default_priority = db.Column(db.String(20), default='medium')
    default_category = db.Column(db.String(100), default='General')
    is_active = db.Column(db.Boolean, default=True)
    created_by_id = db.Column(db.Integer)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


class GuionSubtask(db.Model):
    """Cada paso/control preconfigurado dentro de un guión."""
    __tablename__ = 'guion_subtasks'
    id = db.Column(db.Integer, primary_key=True)
    guion_id = db.Column(db.Integer, db.ForeignKey('guiones.id'), nullable=False, index=True)
    order_idx = db.Column(db.Integer, default=0)
    title = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text)
    category = db.Column(db.String(100))
    priority = db.Column(db.String(20), default='medium')                       # low/medium/high/critical
    assignee_id = db.Column(db.Integer, db.ForeignKey('users.id'))              # técnico asignado por default
    created_at = db.Column(db.DateTime, default=datetime.now)

    guion = db.relationship('Guion', backref=db.backref('subtasks', cascade='all, delete-orphan', order_by='GuionSubtask.order_idx'))
    assignee = db.relationship('User', foreign_keys=[assignee_id])


class UserGuion(db.Model):
    """M:N: qué especialistas están asignados a qué guiones.
    Cuando la API externa invoca un guion y una subtarea no tiene assignee_id fijo,
    se distribuye entre los técnicos asignados aquí (balanceo por carga).
    Además la UI de Gestión de Usuarios muestra/edita esta relación por técnico."""
    __tablename__ = 'user_guiones'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    guion_id = db.Column(db.Integer, db.ForeignKey('guiones.id'), nullable=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    __table_args__ = (db.UniqueConstraint('user_id', 'guion_id', name='uq_user_guion'),)

    user = db.relationship('User', foreign_keys=[user_id], backref=db.backref('guion_assignments', cascade='all, delete-orphan'))
    guion = db.relationship('Guion', foreign_keys=[guion_id], backref=db.backref('user_assignments', cascade='all, delete-orphan'))


class TicketAttachment(db.Model):
    __tablename__ = 'ticket_attachments'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False, index=True)
    original_name = db.Column(db.String(255), nullable=False)
    stored_name = db.Column(db.String(255), nullable=False)
    mime_type = db.Column(db.String(120))
    size_bytes = db.Column(db.Integer)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    uploaded_at = db.Column(db.DateTime, default=datetime.now)

    uploaded_by = db.relationship('User', foreign_keys=[uploaded_by_id])
    ticket = db.relationship('Ticket', backref=db.backref('attachments', cascade='all, delete-orphan'))


class KnowledgeArticle(db.Model):
    """Artículo de la base de conocimiento. Puede ser por empresa o global."""
    __tablename__ = 'knowledge_articles'
    id = db.Column(db.Integer, primary_key=True)
    company = db.Column(db.String(20), index=True)  # NULL = visible para todas las empresas
    title = db.Column(db.String(200), nullable=False)
    slug = db.Column(db.String(220), unique=True, nullable=False, index=True)
    body = db.Column(db.Text, nullable=False)  # markdown
    excerpt = db.Column(db.String(300))  # resumen corto para listados
    category = db.Column(db.String(80), index=True)
    tags = db.Column(db.Text)  # CSV o JSON array
    author_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    is_published = db.Column(db.Boolean, default=True, index=True)
    is_public = db.Column(db.Boolean, default=False)  # accesible sin login
    views = db.Column(db.Integer, default=0)
    helpful_count = db.Column(db.Integer, default=0)
    not_helpful_count = db.Column(db.Integer, default=0)
    version = db.Column(db.Integer, default=1)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    author = db.relationship('User', foreign_keys=[author_id])


class KnowledgeArticleFeedback(db.Model):
    __tablename__ = 'knowledge_article_feedback'
    id = db.Column(db.Integer, primary_key=True)
    article_id = db.Column(db.Integer, db.ForeignKey('knowledge_articles.id'), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))  # nullable si es anonimo
    is_helpful = db.Column(db.Boolean, nullable=False)
    comment = db.Column(db.Text)
    ip_addr = db.Column(db.String(45))
    created_at = db.Column(db.DateTime, default=datetime.now)

    article = db.relationship('KnowledgeArticle', backref=db.backref('feedback_entries', cascade='all, delete-orphan'))


class ApprovalWorkflow(db.Model):
    """Plantilla reutilizable de flujo de aprobación.

    Un workflow define QUÉ tickets requieren aprobación (por categoría/prioridad)
    y QUIÉN debe aprobar (lista ordenada de aprobadores por rol/usuario).
    """
    __tablename__ = 'approval_workflows'
    id = db.Column(db.Integer, primary_key=True)
    company = db.Column(db.String(20), nullable=False, index=True)
    name = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text)
    # Condiciones que activan el workflow (todas se cumplen con AND)
    trigger_category = db.Column(db.String(80))  # ej "Accesos"; None = cualquier categoría
    trigger_priority = db.Column(db.String(20))  # low|medium|high|critical; None = cualquiera
    trigger_template_name = db.Column(db.String(200))  # match por plantilla usada
    # Lista de aprobadores en orden. JSON array de {"order": 1, "user_id": 5, "role_label": "Jefe"}
    approvers_json = db.Column(db.Text, nullable=False, default='[]')
    is_active = db.Column(db.Boolean, default=True, index=True)
    created_by_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)


class TimeEntry(db.Model):
    """Entrada de tiempo trabajado por un usuario en un ticket.

    Cada entry representa una sesión de trabajo:
      - Cronómetro: started_at con click; ended_at cuando el user hace stop
      - Manual: started_at + ended_at ingresados a mano (con duration derivado)

    Un mismo user puede tener múltiples entries en el mismo ticket (varias sesiones).
    Solo puede haber UNA entry activa (ended_at NULL) por (user, ticket).
    """
    __tablename__ = 'time_entries'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    company = db.Column(db.String(20), index=True)  # denormalizado para queries de reportes
    started_at = db.Column(db.DateTime, nullable=False)
    ended_at = db.Column(db.DateTime)  # NULL si está corriendo el cronómetro
    duration_seconds = db.Column(db.Integer, default=0)  # calculado al detener o ingresado manual
    notes = db.Column(db.Text)
    is_manual = db.Column(db.Boolean, default=False)  # True = ingreso manual, False = cronómetro
    created_at = db.Column(db.DateTime, default=datetime.now, index=True)

    ticket = db.relationship('Ticket', backref=db.backref('time_entries', cascade='all, delete-orphan'))
    user = db.relationship('User', foreign_keys=[user_id])


class Approval(db.Model):
    """Instancia concreta de una aprobación pendiente sobre un ticket.

    Cuando un workflow se dispara, se crean N Approval records (uno por paso).
    Solo el que tiene `order` mínimo con status `pending` es el activo.
    """
    __tablename__ = 'approvals'
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False, index=True)
    workflow_id = db.Column(db.Integer, db.ForeignKey('approval_workflows.id'))
    approver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    approver_role_label = db.Column(db.String(120))  # snapshot del rol/etiqueta
    order = db.Column(db.Integer, nullable=False, default=1)
    status = db.Column(db.String(20), default='pending', index=True)  # pending, approved, rejected, skipped
    comment = db.Column(db.Text)
    token = db.Column(db.String(80), unique=True, index=True)  # link seguro por email
    decision_at = db.Column(db.DateTime)
    notified_at = db.Column(db.DateTime)  # cuándo se envió el email
    created_at = db.Column(db.DateTime, default=datetime.now)

    ticket = db.relationship('Ticket', backref=db.backref('approvals', cascade='all, delete-orphan'))
    approver = db.relationship('User', foreign_keys=[approver_id])
    workflow = db.relationship('ApprovalWorkflow')

# ═════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE UTILIDAD
# ═════════════════════════════════════════════════════════════════════════════

def get_next_ticket_number(company):
    """Genera próximo número de ticket: TKT-ELIOT-00042.
    Encuentra el MÁXIMO número de la serie TKT-{COMPANY}-NNNNN (no por id, por el sufijo).
    Anti-colisión: si por alguna razón el candidato ya existe, avanza hasta encontrar uno libre.
    """
    prefix = f"TKT-{company.upper()}-"
    # Traer todos los tickets TKT- de esta empresa para calcular el máximo real del sufijo numérico
    rows = Ticket.query.filter(
        Ticket.company == company,
        Ticket.ticket_number.like(prefix + '%')
    ).with_entities(Ticket.ticket_number).all()

    max_num = 0
    for (tn,) in rows:
        try:
            suffix = tn.split('-')[-1]
            n = int(suffix)
            if n > max_num:
                max_num = n
        except (ValueError, IndexError, AttributeError):
            continue

    num = max_num + 1

    # Anti-colisión defensiva (concurrencia / duplicados): avanzar hasta encontrar libre
    for _ in range(100):
        candidate = f"{prefix}{num:05d}"
        if not Ticket.query.filter_by(ticket_number=candidate).first():
            return candidate
        num += 1

    # Fallback extremo: agregar timestamp para garantizar unicidad
    return f"{prefix}{num:05d}-{int(datetime.now().timestamp())}"

def generate_jwt(user_id, company):
    """Genera JWT con JTI para revocación"""
    jti = str(uuid.uuid4())
    payload = {
        'user_id': user_id,
        'company': company,
        'jti': jti,
        'iat': datetime.utcnow(),
        'exp': datetime.utcnow() + timedelta(hours=8)
    }
    return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256'), jti

def verify_jwt(token):
    """Verifica JWT y que no esté en blacklist (SECURITY FIX 6: forzar verificación de exp)"""
    try:
        # SECURITY FIX 6: Forzar verificación de exp con options
        payload = jwt.decode(
            token,
            app.config['SECRET_KEY'],
            algorithms=['HS256'],
            options={"verify_exp": True}  # Forzar verificación de expiración
        )

        # Verificar si está en blacklist
        blacklist = TokenBlacklist.query.filter_by(jti=payload['jti']).first()
        if blacklist:
            return None

        return payload
    except jwt.ExpiredSignatureError:
        print(f'[Auth] JWT token expired')
        return None
    except (jwt.InvalidTokenError, jwt.DecodeError, KeyError) as e:
        print(f'[Auth] JWT verification failed: {e}')
        return None

def sanitize_input(text, max_length=None, allowed_chars=None):
    """SECURITY FIX 12: Sanitizar inputs - validar length y caracteres especiales"""
    if not text:
        return text

    # Validar longitud
    if max_length and len(text) > max_length:
        text = text[:max_length]

    # Validar caracteres permitidos (opcional)
    if allowed_chars:
        text = ''.join(c for c in text if c in allowed_chars)

    # Eliminar caracteres de control
    text = ''.join(char for char in text if ord(char) >= 32 or char in '\t\n\r')

    return text

# ═════════════════════════════════════════════════════════════════════════════
# CACHE EN MEMORIA (TTL simple, sin dependencias externas)
# Si en producción se prefiere Redis, basta cambiar las 3 funciones por su versión Redis.
# ═════════════════════════════════════════════════════════════════════════════

import threading as _threading
_cache_store = {}      # key → (value, expires_at)
_cache_lock = _threading.RLock()
_cache_stats = {'hits': 0, 'misses': 0, 'sets': 0}

def cache_get(key):
    """Devuelve el valor cacheado o None si expiró/no existe."""
    with _cache_lock:
        entry = _cache_store.get(key)
        if not entry:
            _cache_stats['misses'] += 1
            return None
        value, expires_at = entry
        if expires_at and datetime.now() > expires_at:
            _cache_store.pop(key, None)
            _cache_stats['misses'] += 1
            return None
        _cache_stats['hits'] += 1
        return value

def cache_set(key, value, ttl_seconds=300):
    """Guarda value con un TTL (default 5 min)."""
    with _cache_lock:
        expires_at = datetime.now() + timedelta(seconds=ttl_seconds) if ttl_seconds else None
        _cache_store[key] = (value, expires_at)
        _cache_stats['sets'] += 1

def cache_delete(key_or_prefix, prefix_match=False):
    """Borra una key específica o todas las que empiecen con un prefix."""
    with _cache_lock:
        if prefix_match:
            keys = [k for k in _cache_store.keys() if k.startswith(key_or_prefix)]
            for k in keys:
                _cache_store.pop(k, None)
            return len(keys)
        else:
            return 1 if _cache_store.pop(key_or_prefix, None) else 0


def sanitize_html(html_content: str) -> str:
    """
    SECURITY FIX 3: Sanitizar HTML con bleach (XSS prevention).

    Permite solo tags HTML seguros y elimina scripts/contenido malicioso.
    """
    allowed_tags = ['p', 'br', 'strong', 'em', 'u', 'a', 'ul', 'ol', 'li', 'blockquote', 'code', 'pre']
    allowed_attributes = {'a': ['href', 'title']}
    return bleach.clean(html_content, tags=allowed_tags, attributes=allowed_attributes, strip=True)


# ═════════════════════════════════════════════════════════════════════════════
# RENDERIZADO DE MARKDOWN LIGERO PARA DESCRIPCIONES DE TICKETS
# ═════════════════════════════════════════════════════════════════════════════

def render_ticket_text(raw_text):
    """Convierte texto plano con markdown ligero a HTML seguro para mostrar.
    Soporta:
    - **negrita**
    - *itálica*
    - --- (separador horizontal)
    - Saltos de línea (preservados via <br>)
    - URLs auto-link
    """
    import re
    from markupsafe import Markup, escape
    if not raw_text:
        return Markup('')
    text = str(raw_text)
    # 1) Escapar HTML para prevenir XSS
    text = escape(text)
    # 2) Convertir separadores --- a <hr> (deben estar en su propia línea)
    text = re.sub(r'(?m)^\s*---+\s*$', '<hr class="ticket-hr">', str(text))
    # 3) Negritas **texto**
    text = re.sub(r'\*\*([^\*\n]+?)\*\*', r'<strong>\1</strong>', text)
    # 4) Itálicas *texto* (después de negritas para no romperlas)
    text = re.sub(r'(?<!\*)\*([^\*\n]+?)\*(?!\*)', r'<em>\1</em>', text)
    # 5) URLs → links clickables
    text = re.sub(
        r'(https?://[^\s<>"\']+)',
        r'<a href="\1" target="_blank" rel="noopener" style="color:#2563eb;text-decoration:underline;">\1</a>',
        text
    )
    # 6) Saltos de línea → <br>
    text = text.replace('\n', '<br>\n')
    return Markup(text)


# Registrar como filtro Jinja
@app.template_filter('ticket_md')
def jinja_ticket_md(s):
    return render_ticket_text(s)

# ═════════════════════════════════════════════════════════════════════════════
# FIX 7: FUNCIONES DE VALIDACIÓN EXTRAÍDAS (Eliminar repetición)
# ═════════════════════════════════════════════════════════════════════════════

def is_admin(session_data: dict) -> bool:
    """Valida si usuario en sesión es admin."""
    return session_data.get('user_id') is not None and session_data.get('role') == 'admin'

def is_technician(session_data: dict) -> bool:
    """Valida si usuario en sesión es técnico."""
    return session_data.get('user_id') is not None and session_data.get('role') == 'technician'

def is_authenticated(session_data: dict) -> bool:
    """Valida si usuario tiene sesión activa."""
    return session_data.get('user_id') is not None

def check_company_access(session_data: dict, target_company: str) -> bool:
    """
    Valida que usuario pueda acceder a la empresa target (segregación multi-tenant).

    Previene que usuarios de Pash vean datos de Eliot, etc.
    """
    if not is_authenticated(session_data):
        return False
    return session_data.get('company') == target_company

# ═════════════════════════════════════════════════════════════════════════════
# FIX 8: HELPERS DE BASE DE DATOS (Consolidar queries repetidas)
# ═════════════════════════════════════════════════════════════════════════════

def get_tickets_by_company(company: str, filters: dict = None) -> list:
    """
    Obtiene tickets de una empresa con filtros opcionales.

    Args:
        company: Código de empresa (eliot, pash, primatela)
        filters: Dict con keys (status, priority, assignee_id, category)

    Returns:
        Lista de tickets
    """
    query = Ticket.query.filter_by(company=company)

    if filters:
        if filters.get('status'):
            query = query.filter_by(status=filters['status'])
        if filters.get('priority'):
            query = query.filter_by(priority=filters['priority'])
        if filters.get('assignee_id'):
            query = query.filter_by(assignee_id=filters['assignee_id'])
        if filters.get('category'):
            query = query.filter_by(category=filters['category'])

    return query.all()

def get_sla_config(priority: str) -> int:
    """
    Obtiene minutos de SLA para una prioridad dada.

    Args:
        priority: Prioridad del ticket (low, medium, high, critical)

    Returns:
        Minutos de SLA (defaults si no configurado)
    """
    defaults = {'low': 480, 'medium': 240, 'high': 120, 'critical': 60}
    config = Config.query.filter_by(key=f'sla_{priority}').first()
    return int(config.value) if config else defaults.get(priority, 240)

def validate_company_access(ticket_or_entity):
    """SECURITY FIX 9: Validar que usuario tiene acceso a la empresa de la entidad"""
    if not hasattr(ticket_or_entity, 'company'):
        return True  # Entidad sin company, permitir

    if 'company' not in session:
        return False  # Sin sesión

    # SECURITY FIX 9: Verificar que company_id coincide
    return ticket_or_entity.company == session['company']

def log_audit(action: str, user_id: int = None, entity_type: str = None, entity_id: int = None, description: str = None) -> None:
    """
    Registra acción en audit trail con protección contra log injection.

    Args:
        action: Acción realizada (login, create_ticket, etc.)
        user_id: ID del usuario que realiza la acción
        entity_type: Tipo de entidad afectada (ticket, user, config)
        entity_id: ID de la entidad
        description: Descripción de la acción
    """
    try:
        ip_addr = request.remote_addr if request else '0.0.0.0'
    except (RuntimeError, AttributeError):
        # RuntimeError: no application context, AttributeError: request not available
        ip_addr = '0.0.0.0'

    # SECURITY FIX 11: Sanitizar description para no guardar credenciales
    if description:
        # Remover passwords, tokens, claves sensibles
        description = re.sub(r'password["\']?\s*[:=]\s*[^\s]+', 'password=***', description, flags=re.IGNORECASE)
        description = re.sub(r'token["\']?\s*[:=]\s*[^\s]+', 'token=***', description, flags=re.IGNORECASE)
        description = re.sub(r'api[_-]?key["\']?\s*[:=]\s*[^\s]+', 'api_key=***', description, flags=re.IGNORECASE)

    log = AuditLog(
        action=action,
        user_id=user_id,
        entity_type=entity_type,
        entity_id=entity_id,
        description=description,
        ip_addr=ip_addr
    )
    db.session.add(log)
    db.session.commit()

# ═════════════════════════════════════════════════════════════════════════════
# AUTENTICACION
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user.role == 'employee':
            return redirect(url_for('employee_dashboard'))
        elif user.role == 'technician':
            return redirect(url_for('technician_dashboard'))
        elif user.role == 'admin':
            return redirect(url_for('admin_dashboard'))

    return redirect(url_for('login'))

@app.route('/bot')
def bot_chat():
    """Interfaz del bot de soporte inteligente"""
    return render_template('bot.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Mostrar formulario de login
    if request.method == 'GET':
        return render_template('login_v2.html', companies=COMPANY_COLORS)

    if request.method == 'POST':
        # CSRF: validar token del form
        from flask_wtf.csrf import validate_csrf, CSRFError
        from wtforms import ValidationError
        try:
            validate_csrf(request.form.get('csrf_token'))
        except (CSRFError, ValidationError):
            log_audit('login_csrf_failed', None, 'auth', None,
                      f'CSRF inválido en login desde {request.remote_addr}')
            return render_template('login_v2.html', companies=COMPANY_COLORS,
                                 error='Token de seguridad inválido. Recargá la página e intentá de nuevo.')

        company = request.form.get('company', '').strip()
        # Aceptamos tanto 'email' (nuevo nombre del campo) como 'username' (compatibilidad)
        login_id = (request.form.get('email') or request.form.get('username') or '').strip()
        password = request.form.get('password', '')

        # SECURITY FIX 12: Validar input length y sanitizar
        if not company or not login_id or not password:
            return render_template('login_v2.html', companies=COMPANY_COLORS,
                                 error='Correo, contraseña y empresa son requeridos')

        if len(login_id) > 200 or len(company) > 20 or len(password) > 255:
            return render_template('login_v2.html', companies=COMPANY_COLORS,
                                 error='Entrada inválida')

        # Validar que company existe
        company_obj = Company.query.filter_by(code=company).first()
        if not company_obj:
            return render_template('login_v2.html', companies=COMPANY_COLORS,
                                 error='Empresa no válida')

        # Buscar por EMAIL primero (case-insensitive); si no encuentra, fallback a username (compatibilidad)
        login_lower = login_id.lower()
        user = User.query.filter(
            db.func.lower(User.email) == login_lower,
            User.company == company
        ).first()
        if not user:
            user = User.query.filter_by(username=login_id, company=company).first()

        if user:
            # Bloquear login si la cuenta está inactiva
            if not user.is_active:
                log_audit('login_failed', None, 'auth', None, f'Login bloqueado (inactivo): {login_id}@{company}')
                return render_template('login_v2.html', companies=COMPANY_COLORS,
                                       error='Cuenta inactiva. Contacta al administrador.')

            # Bloqueo por intentos fallidos: si locked_until está en el futuro, rechazar
            now_dt = datetime.now()
            if user.locked_until and user.locked_until > now_dt:
                remaining = int((user.locked_until - now_dt).total_seconds() / 60) + 1
                log_audit('login_blocked_locked', None, 'auth', user.id,
                          f'Login rechazado: cuenta bloqueada hasta {user.locked_until.isoformat(timespec="seconds")}')
                return render_template('login_v2.html', companies=COMPANY_COLORS,
                                     error=f'Cuenta bloqueada por demasiados intentos fallidos. Intenta de nuevo en {remaining} minuto(s) o contacta al administrador.')

            # ─── Autenticación: primero LDAP (si está configurado), luego hash local ───
            authenticated = False
            ldap_status = 'not_configured'

            if LDAP_AVAILABLE and (company_obj.ldap_server or '').strip():
                ldap_cfg = LdapConfig.from_company(company_obj, decrypt_secret)
                if ldap_cfg:
                    ldap_status, ldap_full_name, ldap_msg = ldap_authenticate_user(ldap_cfg, user.username, password)
                    if ldap_status == 'ok':
                        authenticated = True
                        log_audit('login_ldap_ok', user.id, 'auth', user.id,
                                  f'Usuario {user.username}@{company} autenticado por LDAP')
                        # Actualizar nombre si LDAP devolvió uno y el local está vacío
                        if ldap_full_name and not user.name:
                            user.name = ldap_full_name
                    elif ldap_status == 'bad_credentials':
                        # LDAP rechazó la contraseña explícitamente → NO caer a hash local
                        user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
                        error_msg = 'Contraseña incorrecta'
                        if user.failed_login_attempts >= MAX_FAILED_LOGIN_ATTEMPTS:
                            user.locked_until = now_dt + timedelta(minutes=LOCKOUT_DURATION_MINUTES)
                            user.failed_login_attempts = 0
                            log_audit('account_locked', None, 'auth', user.id,
                                      f'Cuenta {user.username}@{user.company} bloqueada (LDAP) por {LOCKOUT_DURATION_MINUTES} min desde {request.remote_addr}')
                            error_msg = f'Demasiados intentos fallidos. Cuenta bloqueada por {LOCKOUT_DURATION_MINUTES} minutos.'
                        else:
                            remaining = MAX_FAILED_LOGIN_ATTEMPTS - user.failed_login_attempts
                            log_audit('login_ldap_failed', None, 'auth', user.id,
                                      f'LDAP rechazó credenciales de {login_id}@{company} ({user.failed_login_attempts}/{MAX_FAILED_LOGIN_ATTEMPTS}): {ldap_msg}')
                            error_msg = f'Contraseña incorrecta. Te quedan {remaining} intento(s).'
                        db.session.commit()
                        return render_template('login_v2.html', companies=COMPANY_COLORS, error=error_msg)
                    else:
                        # LDAP no responde → caer a hash local con warning
                        log_audit('login_ldap_unreachable', None, 'auth', user.id,
                                  f'LDAP no disponible para {company}: {ldap_msg}. Fallback a hash local.')

            # Si no se autenticó por LDAP, intentar hash local
            if not authenticated:
                if user.password_hash:
                    password_hash = hashlib.pbkdf2_hmac('sha256', password.encode(), user.username.encode(), 100000)
                    if password_hash.hex() != user.password_hash:
                        # Incrementar contador de intentos fallidos
                        user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
                        error_msg = 'Contraseña incorrecta'
                        if user.failed_login_attempts >= MAX_FAILED_LOGIN_ATTEMPTS:
                            user.locked_until = now_dt + timedelta(minutes=LOCKOUT_DURATION_MINUTES)
                            user.failed_login_attempts = 0
                            log_audit('account_locked', None, 'auth', user.id,
                                      f'Cuenta {user.username}@{user.company} bloqueada por {LOCKOUT_DURATION_MINUTES} min tras {MAX_FAILED_LOGIN_ATTEMPTS} intentos fallidos desde {request.remote_addr}')
                            error_msg = f'Demasiados intentos fallidos. Cuenta bloqueada por {LOCKOUT_DURATION_MINUTES} minutos.'
                        else:
                            remaining = MAX_FAILED_LOGIN_ATTEMPTS - user.failed_login_attempts
                            log_audit('login_failed', None, 'auth', user.id,
                                      f'Login fallido para {login_id}@{company} ({user.failed_login_attempts}/{MAX_FAILED_LOGIN_ATTEMPTS})')
                            error_msg = f'Contraseña incorrecta. Te quedan {remaining} intento(s) antes del bloqueo temporal.'
                        db.session.commit()
                        return render_template('login_v2.html', companies=COMPANY_COLORS, error=error_msg)
                else:
                    # Usuario sin hash configurado y LDAP no autenticó → rechazar
                    log_audit('login_failed', None, 'auth', None, f'Login rechazado (sin password_hash, LDAP={ldap_status}): {login_id}@{company}')
                    return render_template('login_v2.html', companies=COMPANY_COLORS,
                                         error='Usuario sin contraseña configurada. Contacta al administrador.')

            # Login exitoso: resetear contador y desbloquear
            user.failed_login_attempts = 0
            user.locked_until = None

            session['user_id'] = user.id
            session['username'] = user.username
            session['name'] = user.name
            session['role'] = user.role
            session['company'] = user.company
            session['login_at'] = datetime.now().isoformat(timespec='seconds')
            session.permanent = True

            user.last_login = datetime.now()
            db.session.commit()

            log_audit('login', user.id, 'user', user.id, f'Usuario {user.email or user.username} inició sesión desde {request.remote_addr}')

            # Si el usuario debe cambiar contraseña (primer login o reset), redirigir a la pantalla forzada
            if user.must_change_password:
                return redirect(url_for('force_change_password'))

            if user.role == 'employee':
                return redirect(url_for('employee_dashboard'))
            elif user.role == 'technician':
                return redirect(url_for('technician_dashboard'))
            elif user.role == 'admin':
                return redirect(url_for('admin_dashboard'))

        log_audit('login_failed', None, 'auth', None, f'Login fallido - no encontrado: {login_id}@{company}')
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                             error='Correo o usuario no encontrado en esta empresa')

    return render_template('login_v2.html', companies=COMPANY_COLORS)


# ============ FORCED LOGOUT / KICK ============
# Caché en memoria para evitar query a BD en cada request.
# Estructura: {user_id: force_logout_iso_timestamp_str}
_force_logout_cache = {}
_force_logout_cache_loaded_at = None


def _refresh_force_logout_cache():
    """Recarga el caché desde BD (los usuarios con force_logout_at no nulo)."""
    global _force_logout_cache, _force_logout_cache_loaded_at
    try:
        users = User.query.filter(User.force_logout_at.isnot(None)).all()
        _force_logout_cache = {u.id: u.force_logout_at for u in users}
        _force_logout_cache_loaded_at = datetime.now()
    except Exception as e:
        print(f'[force-logout] Error refrescando cache: {e}')


@app.before_request
def _check_force_logout():
    """Si el admin expulsó al usuario, invalidar su sesión."""
    if 'user_id' not in session:
        return None
    # Solo verificar para rutas autenticadas (no /login, /static, /logout)
    path = request.path or ''
    if path.startswith('/static/') or path in ('/login', '/logout'):
        return None

    uid = session.get('user_id')
    if not uid:
        return None

    # Refrescar caché cada 30 segundos
    global _force_logout_cache_loaded_at
    if (_force_logout_cache_loaded_at is None
        or (datetime.now() - _force_logout_cache_loaded_at).total_seconds() > 30):
        _refresh_force_logout_cache()

    forced_at = _force_logout_cache.get(uid)
    if not forced_at:
        return None

    # Comparar session.login_at con force_logout_at
    login_at_str = session.get('login_at')
    if not login_at_str:
        # Sesión vieja sin login_at → forzar logout
        session.clear()
        if request.path.startswith('/api/'):
            return jsonify({'success': False, 'error': 'Sesión expulsada por administrador', 'forced_logout': True}), 401
        return redirect(url_for('login'))

    try:
        login_at = datetime.fromisoformat(login_at_str)
    except ValueError:
        session.clear()
        return redirect(url_for('login'))

    if login_at < forced_at:
        # La sesión inició ANTES de la expulsión → invalidar
        log_audit('forced_logout', uid, 'user', uid, f'Sesión invalidada por kick admin')
        session.clear()
        if request.path.startswith('/api/'):
            return jsonify({'success': False, 'error': 'Has sido expulsado del sistema por un administrador', 'forced_logout': True}), 401
        return redirect(url_for('login') + '?forced_logout=1')

    return None


@app.route('/api/company-theme')
def api_company_theme():
    """Retorna CSS con colores de empresa del usuario (RF-04-02)"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    company = Company.query.filter_by(code=session['company']).first()
    if not company:
        return jsonify({'success': False}), 404

    # Retornar JSON con colores para aplicar dinámicamente
    return jsonify({
        'success': True,
        'company_code': company.code,
        'company_name': company.name,
        'icon': company.icon,
        'logo_url': COMPANY_LOGOS.get(company.code, ''),
        'primary_color': company.primary_color,
        'secondary_color': company.secondary_color
    })

@app.route('/theme.css')
def theme_css():
    """CSS dinámico con colores de empresa (RF-04-02)"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    company = Company.query.filter_by(code=session['company']).first()
    if not company:
        return jsonify({'error': 'Company not found'}), 404

    # Generar CSS dinámico con colores de empresa
    css = f"""
    :root {{
        --primary-color: {company.primary_color};
        --secondary-color: {company.secondary_color};
    }}

    .header {{
        background: linear-gradient(135deg, {company.primary_color} 0%, {company.secondary_color} 100%);
    }}

    .btn-primary, .tab-btn.active {{
        background: {company.primary_color};
    }}

    .btn-primary:hover {{
        background: {company.secondary_color};
    }}

    .metric-card {{
        border-left-color: {company.primary_color};
    }}

    .metric-number {{
        color: {company.primary_color};
    }}

    .kanban-badge, .stat-value, .card-number {{
        background: {company.primary_color};
    }}

    .tab-btn.active {{
        color: {company.primary_color};
        border-bottom-color: {company.primary_color};
    }}

    a {{
        color: {company.primary_color};
    }}
    """
    return css, 200, {'Content-Type': 'text/css; charset=utf-8'}

@app.route('/cambiar-contrasena', methods=['GET', 'POST'])
def force_change_password():
    """Pantalla obligatoria de cambio de contraseña cuando User.must_change_password=True."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return redirect(url_for('login'))

    # Si por alguna razón el usuario llega acá pero ya no necesita cambio, mandalo al dashboard
    if not user.must_change_password:
        return redirect(_dashboard_url_for(user.role))

    if request.method == 'GET':
        return render_template('force_change_password.html', user_name=user.name)

    # CSRF: validar token del form
    from flask_wtf.csrf import validate_csrf, CSRFError
    from wtforms import ValidationError
    try:
        validate_csrf(request.form.get('csrf_token'))
    except (CSRFError, ValidationError):
        log_audit('force_change_csrf_failed', user.id, 'user', user.id,
                  f'CSRF inválido en cambio forzado desde {request.remote_addr}')
        return render_template('force_change_password.html', user_name=user.name,
                               error='Token de seguridad inválido. Recargá e intentá de nuevo.')

    current = request.form.get('current_password', '')
    new_pw = request.form.get('new_password', '')
    confirm = request.form.get('confirm_password', '')

    # Validaciones
    if not current or not new_pw or not confirm:
        return render_template('force_change_password.html', user_name=user.name,
                               error='Completa todos los campos.')
    if new_pw != confirm:
        return render_template('force_change_password.html', user_name=user.name,
                               error='La confirmación no coincide con la nueva contraseña.')
    if new_pw == current:
        return render_template('force_change_password.html', user_name=user.name,
                               error='La nueva contraseña debe ser diferente a la actual.')
    ok, err = validate_password(new_pw, username=user.username)
    if not ok:
        return render_template('force_change_password.html', user_name=user.name, error=err)

    # Verificar contraseña actual
    current_hash = hashlib.pbkdf2_hmac('sha256', current.encode(), user.username.encode(), 100000).hex()
    if current_hash != user.password_hash:
        log_audit('force_change_password_failed', user.id, 'user', user.id,
                  f'Contraseña actual incorrecta al intentar cambio forzado')
        return render_template('force_change_password.html', user_name=user.name,
                               error='La contraseña actual es incorrecta.')

    # Guardar nueva contraseña y bajar la bandera
    user.password_hash = hashlib.pbkdf2_hmac('sha256', new_pw.encode(), user.username.encode(), 100000).hex()
    user.must_change_password = False
    db.session.commit()

    # Si el usuario es un tecnico origen en Eliot (o un espejo que apunta a uno),
    # sincronizar el nuevo password_hash a todos los espejos hermanos.
    _sync_password_to_mirrors(user)

    log_audit('force_change_password_ok', user.id, 'user', user.id,
              f'Usuario {user.username}@{user.company} cambió su contraseña en primer login')

    return redirect(_dashboard_url_for(user.role))


def _sync_password_to_mirrors(user):
    """Si el usuario tiene espejos (o es un espejo), propagar password_hash y
    must_change_password al origen y a los demas espejos para mantener login unificado."""
    try:
        # Encontrar el "origen" (el usuario en Eliot)
        if user.mirrored_from_id:
            source = User.query.get(user.mirrored_from_id)
        elif user.company == MIRROR_SOURCE_COMPANY and user.role == 'technician':
            source = user
        else:
            return
        if not source:
            return
        # Sincronizar el origen (si el que cambio fue un espejo)
        if source.id != user.id:
            source.password_hash = user.password_hash
            source.must_change_password = user.must_change_password
        # Y todos los espejos
        for m in User.query.filter_by(mirrored_from_id=source.id).all():
            if m.id != user.id:
                m.password_hash = user.password_hash
                m.must_change_password = user.must_change_password
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'[mirror] Error sincronizando password a espejos: {e}')


def _dashboard_url_for(role):
    if role == 'employee':
        return url_for('employee_dashboard')
    if role == 'technician':
        return url_for('technician_dashboard')
    if role == 'admin':
        return url_for('admin_dashboard')
    return url_for('login')


@app.before_request
def _enforce_password_change():
    """Si el usuario tiene must_change_password=True, redirigir/bloquear todo
    hasta que cambie la contraseña.

    Excepción: usuarios que entraron por SSO Microsoft (session.login_provider='microsoft')
    no tienen contraseña local que cambiar, así que se omite."""
    if 'user_id' not in session:
        return None
    # Usuarios logueados por Microsoft NO usan contraseña local → skip
    if session.get('login_provider') == 'microsoft':
        return None
    path = request.path or ''
    # Rutas siempre permitidas mientras debe cambiar contraseña
    if (path.startswith('/static/')
            or path in ('/cambiar-contrasena', '/logout', '/favicon.ico')
            or path.startswith('/api/auth/logout')):
        return None

    user = User.query.get(session['user_id'])
    if not user or not user.must_change_password:
        return None

    # Petición de API → JSON 403; HTML → redirect
    if path.startswith('/api/'):
        return jsonify({
            'success': False,
            'error': 'Debes cambiar tu contraseña antes de continuar.',
            'must_change_password': True,
            'redirect': url_for('force_change_password')
        }), 403
    return redirect(url_for('force_change_password'))


@app.route('/logout')
def logout():
    user_id = session.get('user_id')
    if user_id:
        log_audit('logout', user_id, 'user', user_id, f'Usuario cerró sesión')
    session.clear()
    return redirect(url_for('login'))


# ═════════════════════════════════════════════════════════════════════════════
# MICROSOFT ENTRA ID (Azure AD) — LOGIN VIA OAUTH 2.0
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/auth/microsoft/login', methods=['GET', 'POST'])
def auth_microsoft_login():
    """Inicia el flujo OAuth 2.0 con Microsoft. Recibe ?company=CODIGO."""
    company_code = (request.args.get('company') or '').strip()
    if not company_code:
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error='Seleccioná la empresa antes de continuar con Microsoft.')

    company = Company.query.filter_by(code=company_code).first()
    if not company:
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error='Empresa no válida.')

    if not company.microsoft_enabled or not company.microsoft_tenant_id or not company.microsoft_client_id:
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error=f'{company.name} no tiene configurado Microsoft Entra ID. Contactá al administrador.')

    if not MSAL_AVAILABLE:
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error='La librería MSAL no está disponible en el servidor.')

    # Descifrar client secret
    client_secret = decrypt_secret(company.microsoft_client_secret) or ''
    if not client_secret:
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error=f'{company.name}: falta el Client Secret de Microsoft.')

    # Redirect URI — debe coincidir exactamente con lo configurado en Azure App Registration
    redirect_uri = url_for('auth_microsoft_callback', _external=True, _scheme='https' if _IS_PRODUCTION else 'http')

    try:
        auth_url, state = ms_build_auth_url(
            tenant_id=company.microsoft_tenant_id,
            client_id=company.microsoft_client_id,
            client_secret=client_secret,
            redirect_uri=redirect_uri,
        )
    except Exception as e:
        log_audit('ms_auth_error', None, 'auth', None, f'Error construyendo URL Microsoft: {e}')
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error=f'Error iniciando login con Microsoft: {e}')

    # Guardar state + company en session para validar en callback
    session['ms_oauth_state'] = state
    session['ms_oauth_company'] = company_code
    return redirect(auth_url)


@app.route('/auth/microsoft/callback', methods=['GET', 'POST'])
def auth_microsoft_callback():
    """Callback de Microsoft con el authorization code."""
    # Microsoft puede mandar los params en query string (GET) o form (POST con response_mode=form_post)
    def _param(key):
        return request.args.get(key) or request.form.get(key)

    error = _param('error')
    if error:
        error_desc = _param('error_description') or ''
        log_audit('ms_auth_denied', None, 'auth', None, f'Microsoft rechazó login: {error} - {error_desc}')
        session.pop('ms_oauth_state', None)
        session.pop('ms_oauth_company', None)
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error=f'Microsoft rechazó el login: {error_desc or error}')

    code = _param('code')
    state = _param('state')
    if not code or not state:
        # Log detallado para diagnóstico
        args_keys = list(request.args.keys())
        referrer = request.headers.get('Referer', '')
        log_audit('ms_auth_invalid_callback', None, 'auth', None,
                  f'Callback sin code/state. Args: {args_keys}. Referer: {referrer[:200]}. From: {request.remote_addr}')
        # Si no llegaste desde Microsoft, mostrar mensaje explicativo
        if 'login.microsoftonline.com' not in referrer and 'login.live.com' not in referrer:
            return render_template('login_v2.html', companies=COMPANY_COLORS,
                                   error='Esta URL solo funciona cuando Microsoft te redirige. '
                                         'Volvé a /login y click "Iniciar sesión con Microsoft".')
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error=f'Callback inválido desde Microsoft (params: {args_keys}). '
                                     f'Verificá que el Redirect URI en Azure sea exactamente '
                                     f'https://deskeli.eliotproyectos.tech/auth/microsoft/callback')

    # Validar state contra CSRF
    expected_state = session.get('ms_oauth_state')
    if not expected_state or state != expected_state:
        log_audit('ms_auth_state_mismatch', None, 'auth', None,
                  f'State no coincide (posible CSRF) desde {request.remote_addr}')
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error='Sesión OAuth inválida. Intentá de nuevo.')

    company_code = session.get('ms_oauth_company')
    session.pop('ms_oauth_state', None)
    session.pop('ms_oauth_company', None)

    company = Company.query.filter_by(code=company_code).first() if company_code else None
    if not company:
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error='Empresa perdida durante el flow. Intentá de nuevo.')

    client_secret = decrypt_secret(company.microsoft_client_secret) or ''
    redirect_uri = url_for('auth_microsoft_callback', _external=True, _scheme='https' if _IS_PRODUCTION else 'http')

    try:
        token_result = ms_exchange_code(
            tenant_id=company.microsoft_tenant_id,
            client_id=company.microsoft_client_id,
            client_secret=client_secret,
            code=code,
            redirect_uri=redirect_uri,
        )
    except Exception as e:
        log_audit('ms_auth_token_error', None, 'auth', None, f'Error intercambiando code: {e}')
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error=f'Error obteniendo token de Microsoft: {e}')

    if 'error' in token_result:
        err_desc = token_result.get('error_description', token_result['error'])
        log_audit('ms_auth_token_error', None, 'auth', None, f'Token error: {err_desc}')
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error=f'Microsoft rechazó las credenciales de la app: {err_desc}')

    # Extraer info del usuario del id_token
    user_data = ms_extract_token_data(token_result)
    if not user_data or not user_data.get('id'):
        # Fallback: llamar Graph
        access_token = token_result.get('access_token')
        graph_data = ms_get_user_info(access_token) if access_token else None
        if graph_data:
            user_data = {
                'id': graph_data.get('id'),
                'displayName': graph_data.get('displayName') or '',
                'mail': graph_data.get('mail') or graph_data.get('userPrincipalName') or '',
                'userPrincipalName': graph_data.get('userPrincipalName') or '',
                'givenName': graph_data.get('givenName') or '',
                'surname': graph_data.get('surname') or '',
            }

    if not user_data or not user_data.get('id') or not user_data.get('mail'):
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error='No se pudo obtener info del usuario desde Microsoft.')

    # Buscar usuario en DeskEli: primero por microsoft_object_id, luego por email
    ms_oid = user_data['id']
    email = (user_data.get('mail') or user_data.get('userPrincipalName') or '').lower()
    full_name = user_data.get('displayName') or f"{user_data.get('givenName','')} {user_data.get('surname','')}".strip()

    user = User.query.filter_by(microsoft_object_id=ms_oid).first()
    if not user:
        user = User.query.filter(
            db.func.lower(User.email) == email,
            User.company == company.code
        ).first()
        if user:
            # Enlazar la cuenta local con el OID de Microsoft
            user.microsoft_object_id = ms_oid

    # Auto-provisioning: si el usuario del tenant no existe en DeskEli, crearlo
    if not user:
        base_username = email.split('@')[0] if email else f'ms_{ms_oid[:8]}'
        # Evitar colisiones de username
        username = base_username
        suffix = 1
        while User.query.filter_by(username=username, company=company.code).first():
            suffix += 1
            username = f'{base_username}{suffix}'
        user = User(
            username=username[:80],
            name=full_name[:120] or username,
            email=email[:120],
            role='employee',  # Por default; el admin puede promover después
            company=company.code,
            microsoft_object_id=ms_oid,
            is_active=True,
            must_change_password=False,  # No aplica: entra por Microsoft
        )
        db.session.add(user)
        try:
            db.session.commit()
            log_audit('ms_user_provisioned', None, 'user', user.id,
                      f'Usuario auto-provisionado desde Microsoft: {email} ({company.code}) → username {username}')
        except Exception as e:
            db.session.rollback()
            return render_template('login_v2.html', companies=COMPANY_COLORS,
                                   error=f'Error creando usuario local: {e}')

    if not user.is_active:
        return render_template('login_v2.html', companies=COMPANY_COLORS,
                               error='Cuenta desactivada. Contactá al administrador.')

    # Actualizar datos del usuario si cambiaron en Microsoft
    if full_name and user.name != full_name:
        user.name = full_name[:120]
    user.last_login = datetime.now()
    user.failed_login_attempts = 0
    user.locked_until = None
    # El usuario entra por SSO Microsoft: no necesita cambiar contraseña local
    # (el must_change_password aplica solo cuando entra por email+password local)
    user.must_change_password = False
    db.session.commit()

    log_audit('ms_login', user.id, 'user', user.id,
              f'Login Microsoft OK: {email} desde {request.remote_addr}')

    # Crear sesión Flask
    session['user_id'] = user.id
    session['username'] = user.username
    session['name'] = user.name
    session['role'] = user.role
    session['company'] = user.company
    session['login_at'] = datetime.now().isoformat(timespec='seconds')
    session['login_provider'] = 'microsoft'
    session.permanent = True

    # Redirigir según rol
    if user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    if user.role == 'technician':
        return redirect(url_for('technician_dashboard'))
    return redirect(url_for('employee_dashboard'))

# ═════════════════════════════════════════════════════════════════════════════
# GESTIÓN DE EMPRESAS (MULTI-TENANT) - RF-04
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/companies', methods=['GET'])
def api_list_companies():
    """Listar empresas disponibles"""
    companies = Company.query.filter_by(is_active=True).all()
    result = [{
        'code': c.code,
        'name': c.name,
        'icon': c.icon,
        'primary_color': c.primary_color,
        'secondary_color': c.secondary_color
    } for c in companies]
    return jsonify({'success': True, 'companies': result})

@app.route('/api/company/<company_code>', methods=['GET'])
def api_get_company(company_code):
    """Obtener detalles de empresa"""
    company = Company.query.filter_by(code=company_code).first()
    if not company:
        return jsonify({'success': False}), 404

    return jsonify({
        'success': True,
        'company': {
            'code': company.code,
            'name': company.name,
            'icon': company.icon,
            'primary_color': company.primary_color,
            'secondary_color': company.secondary_color
        }
    })

# ── Endpoints legacy reemplazados por la versión más completa abajo (línea ~4720).
# Se conservan deshabilitados para que no haya colisión de rutas.

@app.route('/api/admin/companies/<int:company_id>/users', methods=['GET'])
def api_admin_company_users(company_id):
    """Admin: Listar usuarios de empresa"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = Company.query.get_or_404(company_id)
    users = User.query.filter_by(company=company.code).all()

    result = [{
        'id': u.id,
        'username': u.username,
        'name': u.name,
        'email': u.email,
        'role': u.role,
        'is_active': u.is_active,
        'last_login': u.last_login.isoformat() if u.last_login else None
    } for u in users]

    return jsonify({'success': True, 'users': result})

# ═════════════════════════════════════════════════════════════════════════════
# PORTALES
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/employee/dashboard')
def employee_dashboard():
    if 'user_id' not in session or session['role'] != 'employee':
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    tickets = Ticket.query.filter_by(creator_id=user.id).all()

    stats = {
        'open': len([t for t in tickets if t.status == 'open']),
        'in_progress': len([t for t in tickets if t.status == 'in_progress']),
        'resolved': len([t for t in tickets if t.status == 'resolved']),
        'avg_rating': sum([t.rating for t in tickets if t.rating]) / len([t for t in tickets if t.rating]) if [t for t in tickets if t.rating] else 0
    }

    theme_name = get_company_theme(session.get('company'))
    theme_color = THEMES.get(theme_name, {}).get('primary', '#2563eb')

    return render_template('employee/dashboard.html',
                         tickets=tickets,
                         stats=stats,
                         company=COMPANY_COLORS[session['company']],
                         theme_color=theme_color,
                         current_theme=theme_name)

@app.route('/employee/create', methods=['GET', 'POST'])
def employee_create():
    if 'user_id' not in session or session['role'] != 'employee':
        return redirect(url_for('login'))

    if request.method == 'POST':
        user = User.query.get(session['user_id'])

        # Validar input
        title = request.form.get('title', '').strip()
        description = request.form.get('description', '').strip()
        category = request.form.get('category', 'General').strip()
        priority = request.form.get('priority', 'medium').strip()
        priority_reason = (request.form.get('priority_reason') or '').strip()
        user_area = (request.form.get('user_area') or '').strip()
        user_location = (request.form.get('user_location') or '').strip()
        user_phone = (request.form.get('user_phone') or '').strip()

        # Si el empleado elevó la prioridad a high o critical, debe venir un motivo
        if priority in ('high', 'critical'):
            if priority_reason and len(priority_reason) >= 10:
                # Anexar el motivo a la descripción con formato destacado
                priority_label = {'high': 'ALTA', 'critical': 'CRÍTICA'}[priority]
                description = (
                    f"{description}\n\n"
                    f"---\n"
                    f"🎯 **MOTIVO DE PRIORIDAD {priority_label}** (indicado por el solicitante):\n"
                    f"{priority_reason}"
                )
            else:
                # Sin motivo → bajar automáticamente a media (evitar abuso de prioridades altas)
                priority = 'medium'

        # Validación contextual: campos personales obligatorios
        def _render_error(msg):
            return render_template('employee/create.html', error=msg, user=user)

        if not title or not description:
            return _render_error('Título y descripción son requeridos')
        if len(title) > 200 or len(title) < 5:
            return _render_error('Título debe tener 5-200 caracteres')
        if len(description) > 5000 or len(description) < 10:
            return _render_error('Descripción debe tener 10-5000 caracteres')
        if not user_area or len(user_area) < 2:
            return _render_error('El campo "Área donde trabaja" es obligatorio')
        if not user_location or len(user_location) < 2:
            return _render_error('El campo "Ubicación en la empresa" es obligatorio')
        if not user_phone or len(user_phone) < 4:
            return _render_error('El campo "Número de contacto" es obligatorio (mín. 4 dígitos)')
        if len(user_area) > 120 or len(user_location) > 120 or len(user_phone) > 40:
            return _render_error('Algún campo de contacto excede el máximo permitido')

        # Actualizar el perfil del usuario con los datos más recientes (pre-llenado futuro)
        try:
            user.area = user_area
            user.location = user_location
            user.phone = user_phone
        except Exception:
            pass

        if priority not in ['low', 'medium', 'high', 'critical']:
            priority = 'medium'

        # Obtener SLA de configuración
        sla_config = Config.query.filter_by(key=f"sla_{priority}").first()
        sla_minutes = int(sla_config.value) if sla_config else 120

        ticket = Ticket(
            ticket_number=get_next_ticket_number(user.company),
            title=title,
            description=description,
            category=category,
            priority=priority,
            creator_id=user.id,
            company=user.company,
            sla_minutes=sla_minutes,
            sla_deadline=compute_sla_deadline(datetime.now(), sla_minutes, user.company),
            user_area=user_area,
            user_location=user_location,
            user_phone=user_phone,
        )

        db.session.add(ticket)
        db.session.commit()

        # Aprobaciones multi-nivel: si algún workflow matchea, entra en cola de aprobación
        try:
            template_name_used = request.form.get('template_name') or None
            match = find_matching_workflow(ticket, template_name=template_name_used)
            if match:
                workflow, steps = match
                create_approvals_for_ticket(ticket, workflow, steps)
                db.session.commit()
                log_audit('ticket_pending_approval', user.id, 'ticket', ticket.id,
                          f"Ticket {ticket.ticket_number} en espera de aprobación (workflow: {workflow.name})")
        except Exception as e:
            print(f"[approvals] Error al procesar workflow para ticket {ticket.id}: {e}")

        # Auditoría específica cuando el empleado elevó la prioridad
        if priority in ('high', 'critical') and priority_reason:
            log_audit(
                'ticket_priority_elevated', user.id, 'ticket', ticket.id,
                f'Empleado {user.email or user.username} eligió prioridad {priority.upper()} '
                f'con motivo: "{priority_reason[:200]}"'
            )

        # Procesar archivos adjuntos (imágenes pegadas o archivos seleccionados)
        attachments_saved = 0
        attachments_errors = []
        attachments_bytes_saved = 0  # Bytes ahorrados por compresión
        try:
            from werkzeug.utils import secure_filename
            files = request.files.getlist('attachments') if 'attachments' in request.files else []
            for f in files:
                if not f or not f.filename:
                    continue
                if not _allowed_attachment(f.filename):
                    attachments_errors.append(f"{f.filename}: tipo no permitido")
                    continue
                try:
                    # Comprimir automáticamente si es imagen (transparente para el usuario)
                    out_bytes, out_filename, out_mime, stats = compress_upload(f)
                    if stats['compressed']:
                        attachments_bytes_saved += (stats['original_size'] - stats['final_size'])

                    safe = secure_filename(out_filename) or 'archivo'
                    ext = safe.rsplit('.', 1)[1].lower() if '.' in safe else ''
                    stored = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex
                    path = os.path.join(app.config['TICKET_UPLOAD_FOLDER'], stored)
                    with open(path, 'wb') as fh:
                        fh.write(out_bytes)
                    db.session.add(TicketAttachment(
                        ticket_id=ticket.id,
                        original_name=(out_filename or f.filename)[:255],
                        stored_name=stored,
                        mime_type=(out_mime or f.mimetype or '')[:120],
                        size_bytes=stats['final_size'],
                        uploaded_by_id=user.id
                    ))
                    attachments_saved += 1
                except Exception as e:
                    attachments_errors.append(f"{f.filename}: {e}")
            if attachments_saved > 0:
                db.session.commit()
                if attachments_bytes_saved > 0:
                    log_audit('attachments_compressed', user.id, 'ticket', ticket.id,
                              f'Ticket {ticket.ticket_number}: {attachments_saved} adjunto(s), '
                              f'{attachments_bytes_saved} bytes ({attachments_bytes_saved/1024:.1f} KB) ahorrados por compresión')
        except Exception as e:
            print(f'[employee_create] Error procesando adjuntos: {e}')

        # Hook del Agent Orchestrator — procesar ticket automáticamente.
        # Si el orchestrator no está disponible o falla, usar assign_ticket_auto como FALLBACK.
        orch = app.config.get('orchestrator')
        orch_ran = False
        if orch is not None:
            try:
                orch.process_new_ticket(ticket)
                orch_ran = True
            except Exception as e:
                print(f'[Orchestrator Hook] Falló: {e}. Intentando fallback assign_ticket_auto...')
                orch_ran = False

        # Si después del orchestrator (o sin él) el ticket sigue sin asignar, intentar fallback
        try:
            db.session.refresh(ticket)
            if not ticket.assignee_id:
                print(f'[fallback-assign] Ticket {ticket.ticket_number} sin asignar (orch_ran={orch_ran}), usando assign_ticket_auto')
                assign_ticket_auto(ticket)
                if ticket.assignee_id:
                    if ticket.status == 'open':
                        ticket.status = 'in_progress'
                    db.session.commit()
                    # Notificar por email al técnico asignado por fallback
                    try:
                        new_tech = User.query.get(ticket.assignee_id)
                        if new_tech:
                            notify_ticket_assigned(
                                ticket=ticket,
                                new_assignee=new_tech,
                                assigned_by_name='Asignacion automatica (fallback por carga)',
                                reason='Orchestrator no disponible, usado balanceo por carga'
                            )
                    except Exception as e_email:
                        print(f'[fallback-assign] email error: {e_email}')
        except Exception as e_fb:
            print(f'[fallback-assign] Error general: {e_fb}')

        log_audit('create_ticket', user.id, 'ticket', ticket.id,
                  f"Ticket {ticket.ticket_number} creado · {attachments_saved} adjunto(s)")

        # Si el cliente pide JSON (AJAX), devolverlo en lugar del redirect
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.is_json or request.form.get('_ajax'):
            return jsonify({
                'success': True,
                'ticket_id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'redirect_url': url_for('employee_ticket', ticket_id=ticket.id),
                'attachments_saved': attachments_saved,
                'attachments_errors': attachments_errors
            })

        return redirect(url_for('employee_ticket', ticket_id=ticket.id))

    # GET: pasar el usuario para pre-llenar campos de contacto si ya los configuró antes
    user = User.query.get(session['user_id'])
    return render_template('employee/create.html', user=user)


@app.route('/technician/create', methods=['GET', 'POST'])
def technician_create():
    """Creación de tickets desde el Portal de Técnicos.
    El técnico puede crear un ticket y opcionalmente auto-asignárselo
    o registrarlo en nombre de otro usuario (creado_para)."""
    if 'user_id' not in session or session['role'] not in ('technician', 'admin'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        tech = User.query.get(session['user_id'])

        title = (request.form.get('title') or '').strip()
        description = (request.form.get('description') or '').strip()
        category = (request.form.get('category') or 'General').strip()
        priority = (request.form.get('priority') or 'medium').strip()
        auto_assign = request.form.get('auto_assign_me') in ('1', 'true', 'on')
        # Crear en nombre de otro usuario (opcional)
        behalf_id_raw = (request.form.get('behalf_of_user_id') or '').strip()

        if not title or not description:
            return render_template('technician/create.html', error='Título y descripción son requeridos')
        if len(title) > 200 or len(title) < 5:
            return render_template('technician/create.html', error='Título debe tener 5-200 caracteres')
        if len(description) > 5000 or len(description) < 10:
            return render_template('technician/create.html', error='Descripción debe tener 10-5000 caracteres')
        if priority not in ('low', 'medium', 'high', 'critical'):
            priority = 'medium'

        # Validar "creado para": debe ser usuario activo de la misma empresa
        creator_id = tech.id
        behalf_user = None
        if behalf_id_raw and behalf_id_raw.isdigit():
            behalf_user = User.query.get(int(behalf_id_raw))
            if behalf_user and behalf_user.company == tech.company and behalf_user.is_active:
                creator_id = behalf_user.id
            else:
                behalf_user = None  # inválido, lo ignoramos

        sla_config = Config.query.filter_by(key=f'sla_{priority}').first()
        sla_minutes = int(sla_config.value) if sla_config else 120

        # Si el técnico marca auto-asignar Y no creó en nombre de otro, lo asigna a sí mismo
        assignee_id = tech.id if auto_assign else None

        # Si el técnico creó "en nombre de", aclarar en la descripción
        final_description = description
        if behalf_user:
            final_description = (
                f"_Ticket registrado por el técnico **{tech.name}** en nombre de **{behalf_user.name}** "
                f"({behalf_user.email}). El usuario reportó el problema directamente al equipo de TI._\n\n"
                + description
            )

        ticket = Ticket(
            ticket_number=get_next_ticket_number(tech.company),
            title=title,
            description=final_description,
            category=category,
            priority=priority,
            creator_id=creator_id,
            assignee_id=assignee_id,
            status='in_progress' if assignee_id else 'open',
            company=tech.company,
            sla_minutes=sla_minutes,
            sla_deadline=compute_sla_deadline(datetime.now(), sla_minutes, tech.company)
        )
        db.session.add(ticket)
        db.session.commit()

        # Adjuntos (mismo procesamiento que employee_create)
        attachments_saved = 0
        attachments_bytes_saved = 0
        try:
            from werkzeug.utils import secure_filename
            files = request.files.getlist('attachments') if 'attachments' in request.files else []
            for f in files:
                if not f or not f.filename:
                    continue
                if not _allowed_attachment(f.filename):
                    continue
                try:
                    # Compresión automática (imágenes: resize + JPEG quality 85)
                    out_bytes, out_filename, out_mime, stats = compress_upload(f)
                    if stats['compressed']:
                        attachments_bytes_saved += (stats['original_size'] - stats['final_size'])

                    safe = secure_filename(out_filename) or 'archivo'
                    ext = safe.rsplit('.', 1)[1].lower() if '.' in safe else ''
                    stored = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex
                    path = os.path.join(app.config['TICKET_UPLOAD_FOLDER'], stored)
                    with open(path, 'wb') as fh:
                        fh.write(out_bytes)
                    db.session.add(TicketAttachment(
                        ticket_id=ticket.id,
                        original_name=(out_filename or f.filename)[:255],
                        stored_name=stored,
                        mime_type=(out_mime or f.mimetype or '')[:120],
                        size_bytes=stats['final_size'],
                        uploaded_by_id=tech.id
                    ))
                    attachments_saved += 1
                except Exception as e:
                    print(f'[technician_create] adjunto error: {e}')
            if attachments_saved > 0:
                db.session.commit()
                if attachments_bytes_saved > 0:
                    log_audit('attachments_compressed', tech.id, 'ticket', ticket.id,
                              f'Ticket {ticket.ticket_number}: {attachments_saved} adjunto(s), '
                              f'{attachments_bytes_saved} bytes ({attachments_bytes_saved/1024:.1f} KB) ahorrados por compresión')
        except Exception as e:
            print(f'[technician_create] Error procesando adjuntos: {e}')

        # Hook orchestrator (solo si NO se auto-asignó manualmente, para no sobrescribir)
        try:
            if not assignee_id:
                orch = app.config.get('orchestrator')
                if orch is not None:
                    orch.process_new_ticket(ticket)
        except Exception as e:
            print(f'[Orchestrator Hook] {e}')

        creator_note = f' en nombre de {behalf_user.username}' if behalf_user else ''
        assign_note = ' (auto-asignado al técnico)' if assignee_id else ''
        log_audit('create_ticket', tech.id, 'ticket', ticket.id,
                  f'Ticket {ticket.ticket_number} creado por técnico {tech.username}{creator_note}{assign_note} · {attachments_saved} adjunto(s)')

        # AJAX → JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.form.get('_ajax'):
            return jsonify({
                'success': True,
                'ticket_id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'redirect_url': url_for('technician_ticket', ticket_id=ticket.id),
                'attachments_saved': attachments_saved,
            })

        return redirect(url_for('technician_ticket', ticket_id=ticket.id))

    return render_template('technician/create.html')


@app.route('/api/technician/company-users', methods=['GET'])
def api_technician_company_users():
    """Devuelve la lista de empleados de la empresa (para el dropdown 'crear en nombre de')."""
    if 'user_id' not in session or session['role'] not in ('technician', 'admin'):
        return jsonify({'success': False}), 401
    company = session['company']
    users = User.query.filter_by(company=company, is_active=True).order_by(User.role.desc(), User.name).all()
    return jsonify({
        'success': True,
        'users': [{
            'id': u.id,
            'name': u.name,
            'email': u.email,
            'role': u.role,
            'username': u.username,
        } for u in users if u.id != session['user_id']]
    })


@app.route('/employee/ticket/<int:ticket_id>')
def employee_ticket(ticket_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    ticket = Ticket.query.get_or_404(ticket_id)
    # El creador puede verlo siempre. Admins/técnicos con scope también.
    role = session.get('role')
    is_creator = ticket.creator_id == session['user_id']
    has_admin_access = role in ('admin', 'technician') and ticket.company in admin_companies_scope()
    if not (is_creator or has_admin_access):
        return redirect(url_for('employee_dashboard'))
    # Si admin abre este URL, redirigir a su vista propia
    if role in ('admin', 'technician') and not is_creator:
        return redirect(url_for('admin_ticket_detail', ticket_id=ticket_id) if role == 'admin' else url_for('technician_ticket', ticket_id=ticket_id))

    # Obtener mensajes con usuario cargado
    messages = Message.query.filter_by(ticket_id=ticket_id)\
        .order_by(Message.created_at.asc()).all()

    assignment_info = get_ticket_assignment_info(ticket)
    return render_template('employee/ticket_detail.html', ticket=ticket, messages=messages, assignment_info=assignment_info)

@app.route('/admin/ticket/<int:ticket_id>')
def admin_ticket_detail(ticket_id):
    """Alias para admin - redirige a la vista de detalle"""
    return technician_ticket(ticket_id)


@app.route('/technician/subtask/<int:subtask_id>')
@app.route('/admin/subtask/<int:subtask_id>')
def technician_subtask_detail(subtask_id):
    """Vista de detalle de una subtarea individual (para técnicos y admins).
    Permite trabajar en la subtarea con un layout similar al del ticket padre,
    con un botón visible para volver al ticket principal."""
    if 'user_id' not in session or session['role'] not in ('technician', 'admin'):
        return redirect(url_for('login'))

    subtask = Subtask.query.get_or_404(subtask_id)
    parent_ticket = Ticket.query.get(subtask.ticket_id)
    if not parent_ticket:
        return redirect(url_for('technician_dashboard'))

    # Validar acceso: misma empresa (o admin master via scope)
    if parent_ticket.company not in admin_companies_scope():
        return redirect(url_for('technician_dashboard'))

    # Resolver assignee y creador
    assignee = User.query.get(subtask.assignee_id) if subtask.assignee_id else None
    creator = User.query.get(subtask.created_by_id) if subtask.created_by_id else None

    # Contar attachments de la subtarea
    attachments_count = SubtaskAttachment.query.filter_by(subtask_id=subtask.id).count()

    return render_template(
        'technician/subtask.html',
        subtask=subtask,
        parent_ticket=parent_ticket,
        assignee=assignee,
        creator=creator,
        attachments_count=attachments_count,
        company=COMPANY_COLORS.get(parent_ticket.company, {}),
    )


@app.route('/technician/ticket/<int:ticket_id>')
def technician_ticket(ticket_id):
    """Vista de detalle de ticket para técnicos Y admins"""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return redirect(url_for('login'))

    ticket = Ticket.query.get_or_404(ticket_id)
    user = User.query.get(session['user_id'])

    # Validar acceso: misma empresa, admin master, o ticket asignado a
    # alguna identidad espejo del usuario (vista consolidada cross-company)
    if not can_user_access_ticket(user, ticket):
        return redirect(url_for('technician_dashboard'))

    # Obtener mensajes con usuario cargado (para msg.user.name)
    messages = Message.query.filter_by(ticket_id=ticket_id)\
        .order_by(Message.created_at.asc()).all()

    # Bot suggestion dummy (no depende de Claude API real)
    bot_suggestion = {
        'issue_type': ticket.category or 'General',
        'confidence': 75,
        'solution': 'Revisar configuración del equipo y reiniciar el servicio afectado.',
        'kb_articles': ['KB-001: Solución de problemas generales']
    }

    # Lista de técnicos para reasignación
    technicians = User.query.filter_by(
        company=session['company'], role='technician', is_active=True
    ).all()

    log_audit('view_ticket', session['user_id'], 'ticket', ticket_id,
              f'Técnico vio ticket {ticket.ticket_number}')

    assignment_info = get_ticket_assignment_info(ticket)

    return render_template('technician/ticket_detail.html',
                           ticket=ticket,
                           messages=messages,
                           bot_suggestion=bot_suggestion,
                           technicians=technicians,
                           user=user,
                           assignment_info=assignment_info)

@app.route('/technician/dashboard')
def technician_dashboard():
    if 'user_id' not in session or session['role'] != 'technician':
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    # Excluir tickets internos del sistema (DMs, chats grupales, etc.)
    INTERNAL_PREFIXES = ('DM-', 'CHAT-')
    tickets = [
        t for t in Ticket.query.filter_by(company=user.company).all()
        if not (t.ticket_number or '').startswith(INTERNAL_PREFIXES)
    ]
    # "De mis grupos" = tickets asignados a tecnicos que comparten al menos 1
    # subrol conmigo (incluyendome). Sin asignar NO se muestran aca.
    group_user_ids = get_my_group_user_ids(user)
    team_queue_objs = sorted(
        [t for t in tickets if t.assignee_id in group_user_ids],
        key=lambda x: x.sla_deadline or datetime.now()
    )

    # "Asignados a mi" — incluye tickets asignados a cualquiera de mis identidades
    # (self + espejos + origen). Esto permite ver tickets de otras empresas cuando
    # el mismo especialista esta replicado (Eliot espejo en Pash/Primatela).
    identity_ids = get_user_identity_ids(user)
    my_queue_objs = [t for t in tickets if t.assignee_id in identity_ids]
    # Y traer tickets de OTRAS empresas asignados a mis espejos/origen
    if len(identity_ids) > 1:
        cross_tickets = [
            t for t in Ticket.query.filter(
                Ticket.assignee_id.in_(identity_ids),
                Ticket.company != user.company
            ).all()
            if not (t.ticket_number or '').startswith(INTERNAL_PREFIXES)
        ]
        my_queue_objs.extend(cross_tickets)
    my_queue_objs.sort(key=lambda x: x.sla_deadline or datetime.now())

    # Enriquecer con info de asignación (quién asignó y cuándo) — evitar N+1 con cache
    _assign_cache = {}
    def _enrich(t):
        if t.id not in _assign_cache:
            _assign_cache[t.id] = get_ticket_assignment_info(t)
        info = _assign_cache[t.id]
        return {
            'id': t.id,
            'ticket_number': t.ticket_number,
            'title': t.title,
            'description': t.description,
            'category': t.category,
            'priority': t.priority,
            'status': t.status,
            'company': t.company,
            'company_label': COMPANY_COLORS.get(t.company, {}).get('name', t.company.title()) if t.company else '',
            'company_icon': COMPANY_COLORS.get(t.company, {}).get('icon', '🏢') if t.company else '🏢',
            'sla_remaining': t.sla_remaining,
            'sla_deadline': t.sla_deadline,
            'created_at': t.created_at,
            'updated_at': t.updated_at,
            'assignee': t.assignee,
            'assigned_by': info.get('by'),
            'assigned_when': info.get('when'),
            'assigned_source': info.get('source'),
        }
    team_queue = [_enrich(t) for t in team_queue_objs]
    my_queue = [_enrich(t) for t in my_queue_objs]
    queue = team_queue  # compatibilidad

    # Calcular stats correctas para el template (incluye tickets cross-company
    # de espejos, para que los KPIs reflejen el trabajo real del especialista)
    my_tickets = my_queue_objs
    now = datetime.now()

    # "A tiempo" considera dos escenarios:
    #   - Resuelto: ¿se resolvió ANTES de que el SLA venciera?
    #   - Activo: ¿el SLA aún está en el futuro?
    def _is_on_time(t):
        if not t.sla_deadline:
            return True  # Sin SLA definido se considera a tiempo
        if t.resolved_at:
            return t.resolved_at <= t.sla_deadline
        return t.sla_deadline > now

    on_time = len([t for t in my_tickets if _is_on_time(t)])
    total_my = len(my_tickets) or 1

    # Tiempo promedio de resolución basado en created_at → resolved_at
    # (más confiable que time_worked_seconds que suele estar en 0)
    resolved_my = [t for t in my_tickets if t.resolved_at and t.created_at]
    if resolved_my:
        durations_hours = [
            max(0, (t.resolved_at - t.created_at).total_seconds() / 3600)
            for t in resolved_my
        ]
        avg_res_hours = sum(durations_hours) / len(durations_hours)
    else:
        avg_res_hours = 0

    unassigned = len([t for t in tickets if not t.assignee_id])

    stats = {
        'sla_health': f'{round(on_time / total_my * 100)}%',
        'on_time': on_time,
        'avg_resolution': f'{round(avg_res_hours, 1)}h',
        'unassigned': unassigned,
        # Mantener anteriores por compatibilidad
        'total_assigned': len(my_tickets),
        'open': len([t for t in tickets if t.status == 'open']),
        'in_progress': len([t for t in tickets if t.status == 'in_progress']),
        'time_worked_hours': sum([(t.time_worked_seconds or 0) for t in my_tickets]) / 3600 if my_tickets else 0
    }

    theme_name = get_company_theme(user.company)
    theme_color = THEMES.get(theme_name, {}).get('primary', '#2563eb')

    return render_template('technician/dashboard.html',
                           queue=queue,
                           my_queue=my_queue,
                           team_queue=team_queue,
                           stats=stats,
                           user=user,
                           company=COMPANY_COLORS.get(session.get('company')),
                           theme_color=theme_color,
                           current_theme=theme_name)

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    scope = admin_companies_scope(user.company, user.role)
    all_tickets = Ticket.query.filter(Ticket.company.in_(scope)).all()

    _sla_ok = len([t for t in all_tickets if not t.sla_deadline or t.sla_deadline > datetime.now()])
    _worked = [t for t in all_tickets if t.time_worked_seconds]
    _rated = [t for t in all_tickets if t.rating]

    metrics = {
        'total_tickets': len(all_tickets),
        'sla_on_time': round((_sla_ok / len(all_tickets) * 100), 1) if all_tickets else 0,
        'avg_resolution': round(sum([t.time_worked_seconds for t in _worked]) / len(_worked) / 3600, 2) if _worked else 0,
        'customer_rating': round(sum([t.rating for t in _rated]) / len(_rated), 1) if _rated else 0
    }

    kanban = {
        'new': [t for t in all_tickets if t.status == 'open'],
        'open': [t for t in all_tickets if t.status == 'open'],
        'in_progress': [t for t in all_tickets if t.status == 'in_progress'],
        'resolved': [t for t in all_tickets if t.status == 'resolved'],
        'closed': [t for t in all_tickets if t.status == 'closed']
    }

    # Obtener configuración actual
    sla_config = {}
    for priority in ['low', 'medium', 'high', 'critical']:
        config = Config.query.filter_by(key=f'sla_{priority}').first()
        sla_config[priority] = int(config.value) if config else {'low': 480, 'medium': 240, 'high': 120, 'critical': 60}[priority]

    theme_name = get_company_theme(user.company)

    # Calcular counts por prioridad
    priority_counts = {
        'critical': len([t for t in all_tickets if t.priority == 'critical']),
        'high': len([t for t in all_tickets if t.priority == 'high']),
        'medium': len([t for t in all_tickets if t.priority == 'medium']),
        'low': len([t for t in all_tickets if t.priority == 'low'])
    }

    company_info = COMPANY_COLORS.get(user.company, {})

    # Calcular métricas adicionales para el dashboard
    total_tickets = len(all_tickets)
    resolved_today = len([t for t in all_tickets
                         if t.status == 'resolved' and t.resolved_at
                         and t.resolved_at.date() == datetime.now().date()])
    in_progress_count = len([t for t in all_tickets if t.status == 'in_progress'])
    avg_seconds = sum([t.time_worked_seconds for t in all_tickets if t.time_worked_seconds]) / max(1, len([t for t in all_tickets if t.time_worked_seconds]))
    avg_time = f"{avg_seconds / 3600:.1f}h" if avg_seconds else '0h'

    # Stats reales por empresa (solo de las visibles en el scope)
    company_stats = []
    for c_code in scope:
        c_tickets = [t for t in all_tickets if t.company == c_code]
        c_sla_ok = len([t for t in c_tickets if not t.sla_deadline or t.sla_deadline > datetime.now() or t.status == 'resolved'])
        c_users = User.query.filter_by(company=c_code).count()
        c_techs = User.query.filter_by(company=c_code, role='technician').count()
        c_meta = COMPANY_COLORS.get(c_code, {})
        company_stats.append({
            'code': c_code,
            'name_short': c_meta.get('name', c_code.title()),
            'name_full': Company.query.filter_by(code=c_code).first().name if Company.query.filter_by(code=c_code).first() else c_meta.get('name', c_code.title()),
            'icon': c_meta.get('icon', '🏢'),
            'primary_color': c_meta.get('primary', '#2563eb'),
            'tickets_count': len(c_tickets),
            'sla_pct': round((c_sla_ok / len(c_tickets) * 100), 0) if c_tickets else 100,
            'users_count': c_users,
            'techs_count': c_techs,
        })

    return render_template('admin/dashboard.html',
                         metrics=metrics,
                         kanban=kanban,
                         priority_counts=priority_counts,
                         sla_config=sla_config,
                         themes=THEMES,
                         current_theme=theme_name,
                         company_info=company_info,
                         now=datetime.now(),
                         ad_config={},
                         total_tickets=total_tickets,
                         resolved_today=resolved_today,
                         in_progress_count=in_progress_count,
                         avg_time=avg_time,
                         scope_companies=scope,
                         is_master=is_master_admin(user.company, user.role),
                         company_stats=company_stats)

@app.route('/admin/themes')
def admin_themes():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    return render_template('admin/themes.html')

@app.route('/admin/tickets')
def admin_tickets():
    """Lista paginada de todos los tickets de la empresa del admin"""
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])

    # Filtros opcionales vía query string
    status_filter = request.args.get('status', '')
    priority_filter = request.args.get('priority', '')
    page = int(request.args.get('page', 1))
    per_page = 20

    scope = admin_companies_scope(user.company, user.role)
    query = Ticket.query.filter(Ticket.company.in_(scope))
    if status_filter and status_filter != 'all':
        query = query.filter_by(status=status_filter)
    if priority_filter and priority_filter != 'all':
        query = query.filter_by(priority=priority_filter)

    query = query.order_by(Ticket.created_at.desc())
    total = query.count()
    tickets = query.offset((page - 1) * per_page).limit(per_page).all()

    technicians = User.query.filter_by(
        company=user.company, role='technician'
    ).all()

    return render_template('admin/tickets.html',
                           tickets=tickets,
                           technicians=technicians,
                           total=total,
                           page=page,
                           per_page=per_page,
                           status_filter=status_filter,
                           priority_filter=priority_filter,
                           is_master=is_master_admin(),
                           scope_companies=scope,
                           now=datetime.now())

@app.route('/admin/config')
def admin_config():
    """Página de configuración unificada (SLA y tema)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    sla_config = {}
    for priority in ['low', 'medium', 'high', 'critical']:
        config = Config.query.filter_by(key=f'sla_{priority}').first()
        defaults = {'low': 480, 'medium': 240, 'high': 120, 'critical': 60}
        sla_config[priority] = int(config.value) if config else defaults[priority]

    theme_name = get_company_theme(session.get('company'))

    # Contadores para badges del sidebar
    company = session.get('company', 'eliot')
    tickets_count = Ticket.query.filter(
        Ticket.company == company,
        Ticket.status.in_(['open', 'in_progress'])
    ).count()
    escalations_count = Ticket.query.filter(
        Ticket.company == company,
        Ticket.priority == 'critical',
        Ticket.status.in_(['open', 'in_progress'])
    ).count()

    # Nombre de empresa
    company_obj = Company.query.filter_by(code=company).first()
    company_name = company_obj.name if company_obj else 'Mi Empresa S.A.'

    return render_template('admin/config_new.html',
                           sla_config=sla_config,
                           themes=THEMES,
                           current_theme=theme_name,
                           tickets_count=tickets_count,
                           escalations_count=escalations_count,
                           company_name=company_name,
                           is_master=is_master_admin())


@app.route('/admin/config-old')
def admin_config_old():
    """Página de configuración antigua (acceso a CRUDs completos)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    sla_config = {}
    for priority in ['low', 'medium', 'high', 'critical']:
        config = Config.query.filter_by(key=f'sla_{priority}').first()
        defaults = {'low': 480, 'medium': 240, 'high': 120, 'critical': 60}
        sla_config[priority] = int(config.value) if config else defaults[priority]

    theme_name = get_company_theme(session.get('company'))

    # Un admin de pash/primatela puede sincronizar tecnicos desde eliot (master)
    can_sync_from_master = (
        session.get('company') != MIRROR_SOURCE_COMPANY
        and session.get('company') in MIRROR_TARGET_COMPANIES
    )

    return render_template('admin/config.html',
                           sla_config=sla_config,
                           themes=THEMES,
                           current_theme=theme_name,
                           is_master=is_master_admin(),
                           can_sync_from_master=can_sync_from_master)

@app.route('/admin/orchestrator')
def admin_orchestrator():
    """Dashboard del Agent Orchestrator - control de los 4 agentes"""
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    data = {}
    orch = app.config.get('orchestrator')
    if orch is not None:
        data = orch.get_dashboard_data(user.company)

    # Master admin puede filtrar por empresa; los demas solo ven la suya.
    is_master = is_master_admin()
    available_companies = []
    if is_master:
        available_companies = [
            {'code': c.code, 'name': c.name}
            for c in Company.query.filter_by(is_active=True).order_by(Company.name).all()
        ]

    return render_template('admin/orchestrator.html',
                           orchestrator_data=data,
                           company_info=COMPANY_COLORS.get(user.company, {}),
                           is_master=is_master,
                           available_companies=available_companies,
                           current_company=user.company)


# ─── Dashboard CSAT ───────────────────────────────────────────────────────────
@app.route('/admin/csat')
def admin_csat():
    """Dashboard de satisfacción del cliente (CSAT + NPS)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    is_master = is_master_admin()
    available_companies = []
    if is_master:
        available_companies = [
            {'code': c.code, 'name': c.name}
            for c in Company.query.filter_by(is_active=True).order_by(Company.name).all()
        ]

    return render_template('admin/csat.html',
                           company_info=COMPANY_COLORS.get(user.company, {}),
                           is_master=is_master,
                           available_companies=available_companies,
                           current_company=user.company)


@app.route('/api/admin/csat/summary')
def api_admin_csat_summary():
    """Métricas de CSAT/NPS para dashboard admin.

    Query params:
      - company: 'eliot'|'pash'|'primatela'|'all' (default: user's company)
      - days: rango en días (default 90)
    """
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401

    company_filter = request.args.get('company', session.get('company'))
    try:
        days = int(request.args.get('days', 90))
    except (ValueError, TypeError):
        days = 90
    days = max(1, min(days, 730))

    since = datetime.now() - timedelta(days=days)
    scope = admin_companies_scope()

    q = Ticket.query.filter(
        Ticket.rating.isnot(None),
        Ticket.rating_at >= since,
        Ticket.company.in_(scope)
    )
    if company_filter and company_filter != 'all':
        if company_filter not in scope:
            return jsonify({'success': False, 'error': 'Sin acceso a esa empresa'}), 403
        q = q.filter(Ticket.company == company_filter)

    rated = q.all()
    total = len(rated)

    if total == 0:
        return jsonify({
            'success': True,
            'total': 0,
            'avg_rating': None,
            'nps': None,
            'distribution': {str(i): 0 for i in range(1, 6)},
            'promoters': 0, 'passives': 0, 'detractors': 0,
            'trend': [],
            'by_technician': [],
            'worst_recent': [],
            'best_recent': [],
            'comments_recent': []
        })

    # Promedio CSAT
    ratings = [t.rating for t in rated]
    avg_rating = round(sum(ratings) / len(ratings), 2)

    # Distribución de estrellas
    distribution = {str(i): 0 for i in range(1, 6)}
    for r in ratings:
        distribution[str(r)] = distribution.get(str(r), 0) + 1

    # NPS
    nps_scores = [t.rating_nps for t in rated if t.rating_nps is not None]
    promoters = sum(1 for n in nps_scores if n >= 9)
    passives = sum(1 for n in nps_scores if 7 <= n <= 8)
    detractors = sum(1 for n in nps_scores if n <= 6)
    nps_total = len(nps_scores)
    nps_score = round(((promoters - detractors) / nps_total) * 100, 1) if nps_total else None

    # Tendencia semanal (últimas 12 semanas)
    trend = []
    from collections import defaultdict
    weekly = defaultdict(list)
    for t in rated:
        # ISO week key: YYYY-Www
        week_key = t.rating_at.strftime('%Y-W%V') if t.rating_at else 'unknown'
        weekly[week_key].append(t.rating)
    for wk in sorted(weekly.keys())[-12:]:
        vals = weekly[wk]
        trend.append({
            'week': wk,
            'avg': round(sum(vals) / len(vals), 2),
            'count': len(vals)
        })

    # Por técnico (top 20 con más calificaciones)
    tech_stats = defaultdict(lambda: {'ratings': [], 'nps': []})
    for t in rated:
        if t.assignee_id:
            tech_stats[t.assignee_id]['ratings'].append(t.rating)
            if t.rating_nps is not None:
                tech_stats[t.assignee_id]['nps'].append(t.rating_nps)

    by_technician = []
    for uid, data in tech_stats.items():
        u = User.query.get(uid)
        if not u:
            continue
        r_list = data['ratings']
        n_list = data['nps']
        n_prom = sum(1 for n in n_list if n >= 9)
        n_det = sum(1 for n in n_list if n <= 6)
        by_technician.append({
            'id': uid,
            'name': u.name,
            'company': u.company,
            'count': len(r_list),
            'avg_rating': round(sum(r_list) / len(r_list), 2),
            'nps': round(((n_prom - n_det) / len(n_list)) * 100, 1) if n_list else None
        })
    by_technician.sort(key=lambda x: (-x['count'], -x['avg_rating']))

    # Peores calificaciones recientes (últimas 15)
    worst = sorted([t for t in rated if t.rating <= 3], key=lambda t: t.rating_at or t.updated_at, reverse=True)[:15]
    worst_recent = [{
        'id': t.id,
        'number': t.ticket_number,
        'title': t.title[:80],
        'rating': t.rating,
        'nps': t.rating_nps,
        'comment': (t.rating_comment or '')[:200],
        'assignee': t.assignee.name if t.assignee else '—',
        'company': t.company,
        'rated_at': t.rating_at.strftime('%Y-%m-%d %H:%M') if t.rating_at else ''
    } for t in worst]

    # Mejores calificaciones recientes (últimas 10 con 5 estrellas)
    best = sorted([t for t in rated if t.rating == 5], key=lambda t: t.rating_at or t.updated_at, reverse=True)[:10]
    best_recent = [{
        'id': t.id,
        'number': t.ticket_number,
        'title': t.title[:80],
        'rating': t.rating,
        'comment': (t.rating_comment or '')[:200],
        'assignee': t.assignee.name if t.assignee else '—',
        'company': t.company,
        'rated_at': t.rating_at.strftime('%Y-%m-%d %H:%M') if t.rating_at else ''
    } for t in best]

    # Últimos comentarios (con cualquier rating)
    with_comments = [t for t in rated if t.rating_comment]
    with_comments.sort(key=lambda t: t.rating_at or t.updated_at, reverse=True)
    comments_recent = [{
        'id': t.id,
        'number': t.ticket_number,
        'rating': t.rating,
        'nps': t.rating_nps,
        'comment': t.rating_comment,
        'assignee': t.assignee.name if t.assignee else '—',
        'company': t.company,
        'rated_at': t.rating_at.strftime('%Y-%m-%d %H:%M') if t.rating_at else ''
    } for t in with_comments[:25]]

    # Response rate: cuántos tickets resueltos fueron calificados
    resolved_q = Ticket.query.filter(
        Ticket.status == 'resolved',
        Ticket.resolved_at >= since,
        Ticket.company.in_(scope)
    )
    if company_filter and company_filter != 'all':
        resolved_q = resolved_q.filter(Ticket.company == company_filter)
    resolved_count = resolved_q.count()
    response_rate = round((total / resolved_count) * 100, 1) if resolved_count else 0

    return jsonify({
        'success': True,
        'total': total,
        'resolved_total': resolved_count,
        'response_rate': response_rate,
        'avg_rating': avg_rating,
        'nps': nps_score,
        'nps_total': nps_total,
        'promoters': promoters,
        'passives': passives,
        'detractors': detractors,
        'distribution': distribution,
        'trend': trend,
        'by_technician': by_technician[:20],
        'worst_recent': worst_recent,
        'best_recent': best_recent,
        'comments_recent': comments_recent,
        'days': days,
        'company_filter': company_filter or 'all'
    })


# ─── Base de Conocimiento (KB) ────────────────────────────────────────────────
import unicodedata


def _kb_slugify(text: str, max_len: int = 200) -> str:
    """Genera un slug URL-safe a partir de un texto."""
    if not text:
        return 'articulo'
    txt = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    txt = txt.lower()
    txt = re.sub(r'[^a-z0-9]+', '-', txt).strip('-')
    return (txt or 'articulo')[:max_len]


def _kb_unique_slug(base_slug: str, exclude_id: int = None) -> str:
    """Devuelve slug único agregando -2, -3, ... si ya existe."""
    slug = base_slug
    n = 2
    while True:
        q = KnowledgeArticle.query.filter_by(slug=slug)
        if exclude_id is not None:
            q = q.filter(KnowledgeArticle.id != exclude_id)
        if not q.first():
            return slug
        slug = f"{base_slug}-{n}"
        n += 1


def _kb_article_scope():
    """Empresas cuyo KB puede ver el usuario actual (incluye globales)."""
    role = session.get('role')
    user_company = session.get('company')
    if role == 'admin' and is_master_admin():
        return [c.code for c in Company.query.filter_by(is_active=True).all()]
    return [user_company] if user_company else []


def _kb_serialize(a, include_body=False):
    return {
        'id': a.id,
        'title': a.title,
        'slug': a.slug,
        'excerpt': a.excerpt or '',
        'category': a.category or '',
        'tags': [t.strip() for t in (a.tags or '').split(',') if t.strip()],
        'company': a.company or 'global',
        'is_published': a.is_published,
        'is_public': a.is_public,
        'views': a.views or 0,
        'helpful_count': a.helpful_count or 0,
        'not_helpful_count': a.not_helpful_count or 0,
        'version': a.version or 1,
        'author': a.author.name if a.author else None,
        'created_at': a.created_at.isoformat() if a.created_at else None,
        'updated_at': a.updated_at.isoformat() if a.updated_at else None,
        'body': a.body if include_body else None,
    }


# ─── Admin: CRUD de artículos ───────────────────────────────
@app.route('/admin/kb')
def admin_kb():
    """Gestor de la base de conocimiento (admin)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    is_master = is_master_admin()
    available_companies = []
    if is_master:
        available_companies = [
            {'code': c.code, 'name': c.name}
            for c in Company.query.filter_by(is_active=True).order_by(Company.name).all()
        ]
    return render_template('admin/kb.html',
                           company_info=COMPANY_COLORS.get(user.company, {}),
                           is_master=is_master,
                           available_companies=available_companies,
                           current_company=user.company)


@app.route('/api/admin/kb/articles', methods=['GET'])
def api_admin_kb_list():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401

    scope = _kb_article_scope()
    company_filter = request.args.get('company')
    q = KnowledgeArticle.query
    if company_filter and company_filter not in ('all', ''):
        if company_filter == 'global':
            q = q.filter(KnowledgeArticle.company.is_(None))
        else:
            if company_filter not in scope:
                return jsonify({'success': False, 'error': 'Sin acceso'}), 403
            q = q.filter(KnowledgeArticle.company == company_filter)
    else:
        q = q.filter(db.or_(KnowledgeArticle.company.in_(scope), KnowledgeArticle.company.is_(None)))

    search = (request.args.get('q') or '').strip()
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(
            KnowledgeArticle.title.ilike(like),
            KnowledgeArticle.body.ilike(like),
            KnowledgeArticle.tags.ilike(like),
            KnowledgeArticle.category.ilike(like)
        ))

    articles = q.order_by(KnowledgeArticle.updated_at.desc()).limit(500).all()
    return jsonify({
        'success': True,
        'count': len(articles),
        'articles': [_kb_serialize(a) for a in articles]
    })


@app.route('/api/admin/kb/articles/<int:article_id>', methods=['GET'])
def api_admin_kb_get(article_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    a = KnowledgeArticle.query.get(article_id)
    if not a:
        return jsonify({'success': False, 'error': 'Artículo no encontrado'}), 404
    scope = _kb_article_scope()
    if a.company and a.company not in scope:
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    return jsonify({'success': True, 'article': _kb_serialize(a, include_body=True)})


@app.route('/api/admin/kb/articles', methods=['POST'])
def api_admin_kb_create():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    data = request.json or {}
    title = (data.get('title') or '').strip()
    body = (data.get('body') or '').strip()
    if not title or not body:
        return jsonify({'success': False, 'error': 'Título y cuerpo son obligatorios'}), 400
    company = data.get('company') or None  # None = global
    if company == 'global':
        company = None
    if company and company not in _kb_article_scope():
        return jsonify({'success': False, 'error': 'Sin acceso a esa empresa'}), 403
    # Solo master admin puede crear artículos globales
    if company is None and not is_master_admin():
        return jsonify({'success': False, 'error': 'Solo el admin master puede crear artículos globales'}), 403

    slug = _kb_unique_slug(_kb_slugify(title))
    a = KnowledgeArticle(
        title=title[:200],
        slug=slug,
        body=body,
        excerpt=(data.get('excerpt') or body[:280]).strip()[:300],
        category=(data.get('category') or '').strip()[:80] or None,
        tags=','.join([t.strip() for t in (data.get('tags') or []) if t.strip()]) if isinstance(data.get('tags'), list) else (data.get('tags') or '')[:500],
        company=company,
        author_id=session['user_id'],
        is_published=bool(data.get('is_published', True)),
        is_public=bool(data.get('is_public', False)),
        version=1
    )
    db.session.add(a)
    db.session.commit()
    log_audit('kb_article_created', session['user_id'], 'kb_article', a.id, f"Artículo KB creado: {title}")
    return jsonify({'success': True, 'article': _kb_serialize(a, include_body=True)})


@app.route('/api/admin/kb/articles/<int:article_id>', methods=['PUT'])
def api_admin_kb_update(article_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    a = KnowledgeArticle.query.get(article_id)
    if not a:
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    scope = _kb_article_scope()
    if a.company and a.company not in scope:
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    data = request.json or {}

    new_title = (data.get('title') or a.title).strip()
    if new_title != a.title:
        a.title = new_title[:200]
        a.slug = _kb_unique_slug(_kb_slugify(new_title), exclude_id=a.id)
    if 'body' in data:
        a.body = (data.get('body') or '').strip()
    if 'excerpt' in data:
        a.excerpt = (data.get('excerpt') or '')[:300]
    if 'category' in data:
        a.category = (data.get('category') or '').strip()[:80] or None
    if 'tags' in data:
        if isinstance(data['tags'], list):
            a.tags = ','.join([t.strip() for t in data['tags'] if t.strip()])
        else:
            a.tags = (data.get('tags') or '')[:500]
    if 'is_published' in data:
        a.is_published = bool(data['is_published'])
    if 'is_public' in data:
        a.is_public = bool(data['is_public'])
    if 'company' in data:
        new_co = data['company'] or None
        if new_co == 'global':
            new_co = None
        if new_co and new_co not in scope:
            return jsonify({'success': False, 'error': 'Sin acceso a esa empresa'}), 403
        if new_co is None and not is_master_admin():
            return jsonify({'success': False, 'error': 'Solo master admin puede globalizar'}), 403
        a.company = new_co

    a.version = (a.version or 1) + 1
    a.updated_at = datetime.now()
    db.session.commit()
    log_audit('kb_article_updated', session['user_id'], 'kb_article', a.id, f"Artículo KB actualizado: {a.title}")
    return jsonify({'success': True, 'article': _kb_serialize(a, include_body=True)})


@app.route('/api/admin/kb/articles/<int:article_id>', methods=['DELETE'])
def api_admin_kb_delete(article_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    a = KnowledgeArticle.query.get(article_id)
    if not a:
        return jsonify({'success': False}), 404
    scope = _kb_article_scope()
    if a.company and a.company not in scope:
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    title = a.title
    db.session.delete(a)
    db.session.commit()
    log_audit('kb_article_deleted', session['user_id'], 'kb_article', article_id, f"Artículo KB eliminado: {title}")
    return jsonify({'success': True})


# ─── Portal público (empleados/técnicos): leer KB ─────────
@app.route('/kb')
def kb_public_index():
    """Portal de base de conocimiento (visible a cualquier usuario autenticado)."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    return render_template('kb/index.html',
                           company_info=COMPANY_COLORS.get(user.company, {}),
                           user=user)


@app.route('/kb/<slug>')
def kb_public_article(slug):
    """Vista de un artículo del KB."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    a = KnowledgeArticle.query.filter_by(slug=slug, is_published=True).first()
    if not a:
        return render_template('kb/not_found.html', user=user), 404
    # Segregación: si el artículo tiene empresa, solo esa empresa lo ve
    if a.company and a.company != user.company and not is_master_admin(user.company, user.role):
        return render_template('kb/not_found.html', user=user), 404
    # Incrementar vistas
    a.views = (a.views or 0) + 1
    db.session.commit()
    return render_template('kb/article.html',
                           article=a,
                           company_info=COMPANY_COLORS.get(user.company, {}),
                           user=user)


@app.route('/api/kb/search')
def api_kb_search():
    """Búsqueda pública de artículos (usuario autenticado)."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    user = User.query.get(session['user_id'])
    search = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()

    q = KnowledgeArticle.query.filter(
        KnowledgeArticle.is_published == True,
        db.or_(
            KnowledgeArticle.company == user.company,
            KnowledgeArticle.company.is_(None)
        )
    )
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(
            KnowledgeArticle.title.ilike(like),
            KnowledgeArticle.body.ilike(like),
            KnowledgeArticle.tags.ilike(like)
        ))
    if category:
        q = q.filter(KnowledgeArticle.category == category)

    articles = q.order_by(KnowledgeArticle.views.desc(), KnowledgeArticle.updated_at.desc()).limit(50).all()

    # Categorías disponibles (para chips en UI)
    cats_q = db.session.query(KnowledgeArticle.category).filter(
        KnowledgeArticle.is_published == True,
        db.or_(
            KnowledgeArticle.company == user.company,
            KnowledgeArticle.company.is_(None)
        ),
        KnowledgeArticle.category.isnot(None)
    ).distinct().all()
    categories = sorted(set(c[0] for c in cats_q if c[0]))

    return jsonify({
        'success': True,
        'count': len(articles),
        'articles': [_kb_serialize(a) for a in articles],
        'categories': categories
    })


# ─── Aprobaciones multi-nivel ──────────────────────────────────────────────────
import secrets


def _approvals_serialize_workflow(w, include_approvers=True):
    approvers = []
    try:
        approvers = json.loads(w.approvers_json or '[]')
    except Exception:
        approvers = []
    d = {
        'id': w.id,
        'company': w.company,
        'name': w.name,
        'description': w.description or '',
        'trigger_category': w.trigger_category or '',
        'trigger_priority': w.trigger_priority or '',
        'trigger_template_name': w.trigger_template_name or '',
        'is_active': w.is_active,
        'created_at': w.created_at.isoformat() if w.created_at else None,
        'approvers_count': len(approvers),
    }
    if include_approvers:
        # Enriquecer con datos del usuario
        enriched = []
        for step in approvers:
            uid = step.get('user_id')
            u = User.query.get(uid) if uid else None
            enriched.append({
                'order': step.get('order', len(enriched) + 1),
                'user_id': uid,
                'user_name': u.name if u else '(usuario eliminado)',
                'user_email': u.email if u else '',
                'role_label': step.get('role_label', '')
            })
        d['approvers'] = enriched
    return d


def _approvals_serialize_approval(a, include_token=False):
    return {
        'id': a.id,
        'ticket_id': a.ticket_id,
        'ticket_number': a.ticket.ticket_number if a.ticket else '',
        'ticket_title': a.ticket.title if a.ticket else '',
        'workflow_id': a.workflow_id,
        'workflow_name': a.workflow.name if a.workflow else '',
        'approver_id': a.approver_id,
        'approver_name': a.approver.name if a.approver else '',
        'approver_role_label': a.approver_role_label or '',
        'order': a.order,
        'status': a.status,
        'comment': a.comment or '',
        'decision_at': a.decision_at.isoformat() if a.decision_at else None,
        'created_at': a.created_at.isoformat() if a.created_at else None,
        'token': a.token if include_token else None
    }


def find_matching_workflow(ticket, template_name=None):
    """Devuelve el primer workflow activo cuyas condiciones matcheen el ticket.
    None si ninguno aplica.
    """
    q = ApprovalWorkflow.query.filter_by(company=ticket.company, is_active=True)
    workflows = q.order_by(ApprovalWorkflow.id.desc()).all()
    for w in workflows:
        # AND semantics: cada trigger no-vacío debe matchear
        if w.trigger_category and (ticket.category or '').lower() != w.trigger_category.lower():
            continue
        if w.trigger_priority and ticket.priority != w.trigger_priority:
            continue
        if w.trigger_template_name and template_name and template_name.lower() != w.trigger_template_name.lower():
            continue
        # Si tiene trigger_template_name pero no viene template en el ticket, no aplica
        if w.trigger_template_name and not template_name:
            continue
        # Validar que tenga al menos 1 aprobador
        try:
            steps = json.loads(w.approvers_json or '[]')
        except Exception:
            steps = []
        if not steps:
            continue
        return w, steps
    return None


def create_approvals_for_ticket(ticket, workflow, steps):
    """Crea los N Approval records y marca el ticket como pending_approval.
    Envía email al primer aprobador."""
    for i, step in enumerate(steps):
        uid = step.get('user_id')
        if not uid:
            continue
        approval = Approval(
            ticket_id=ticket.id,
            workflow_id=workflow.id,
            approver_id=uid,
            approver_role_label=step.get('role_label', ''),
            order=step.get('order', i + 1),
            status='pending',
            token=secrets.token_urlsafe(32)
        )
        db.session.add(approval)
    ticket.status = 'pending_approval'
    ticket.updated_at = datetime.now()
    db.session.flush()

    # Notificar al primer aprobador
    first = Approval.query.filter_by(ticket_id=ticket.id).order_by(Approval.order.asc()).first()
    if first:
        _send_approval_email(first)


def _send_approval_email(approval):
    """Envía email al aprobador con link al ticket. Silencioso si SMTP falla."""
    try:
        approver = approval.approver
        ticket = approval.ticket
        if not approver or not approver.email or not ticket:
            return
        subject = f"[DeskEli] Aprobación requerida · Ticket {ticket.ticket_number}"
        body = f"""
        <html><body style="font-family:Segoe UI,sans-serif;color:#1f2937;">
        <div style="max-width:600px;margin:20px auto;padding:24px;background:#f9fafb;border-radius:10px;">
            <h2 style="color:#7c3aed;margin:0 0 10px;">🔐 Se requiere tu aprobación</h2>
            <p>Hola <strong>{approver.name}</strong>,</p>
            <p>El ticket <strong>{ticket.ticket_number}</strong> ({ticket.title}) requiere tu revisión como {approval.approver_role_label or 'aprobador'}.</p>
            <div style="background:white;padding:16px;border-radius:8px;margin:14px 0;">
                <p><strong>Solicitante:</strong> {ticket.creator.name if ticket.creator else '—'}</p>
                <p><strong>Prioridad:</strong> {ticket.priority}</p>
                <p><strong>Categoría:</strong> {ticket.category or '—'}</p>
                <p><strong>Descripción:</strong></p>
                <div style="background:#f3f4f6;padding:10px;border-radius:4px;font-size:13px;">
                    {(ticket.description or '')[:500]}{'...' if len(ticket.description or '') > 500 else ''}
                </div>
            </div>
            <p>Revisá el detalle y aprobá o rechazá desde el enlace:</p>
            <a href="{request.host_url.rstrip('/')}/approvals/decide/{approval.token}"
               style="display:inline-block;padding:12px 24px;background:#7c3aed;color:white;text-decoration:none;border-radius:6px;font-weight:700;">
                Ver ticket y decidir
            </a>
            <p style="font-size:12px;color:#9ca3af;margin-top:24px;">
                Enviado automáticamente por DeskEli — no responder a este correo.
            </p>
        </div>
        </body></html>
        """
        send_email(
            to_email=approver.email,
            subject=subject,
            body=body,
            company=approval.ticket.company
        )
        approval.notified_at = datetime.now()
        db.session.commit()
    except Exception as e:
        print(f"[approvals] Error enviando email: {e}")


def _finalize_approval_chain(ticket):
    """Después de una decisión, chequea si toda la cadena terminó y actualiza el ticket."""
    all_approvals = Approval.query.filter_by(ticket_id=ticket.id).order_by(Approval.order.asc()).all()
    if not all_approvals:
        return

    # ¿Algún rechazo?
    rejected = [a for a in all_approvals if a.status == 'rejected']
    if rejected:
        ticket.status = 'rejected'
        ticket.updated_at = datetime.now()
        return

    # ¿Todos aprobados?
    pending = [a for a in all_approvals if a.status == 'pending']
    if not pending:
        # Toda la cadena aprobó → el ticket va a la cola normal
        ticket.status = 'open'
        ticket.updated_at = datetime.now()
        return

    # Hay pendientes → notificar al siguiente en la cola
    next_pending = min(pending, key=lambda a: a.order)
    if not next_pending.notified_at:
        _send_approval_email(next_pending)


# Admin: gestor de workflows
@app.route('/admin/approvals')
def admin_approvals():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    is_master = is_master_admin()
    available_companies = []
    if is_master:
        available_companies = [
            {'code': c.code, 'name': c.name}
            for c in Company.query.filter_by(is_active=True).order_by(Company.name).all()
        ]
    return render_template('admin/approvals.html',
                           company_info=COMPANY_COLORS.get(user.company, {}),
                           is_master=is_master,
                           available_companies=available_companies,
                           current_company=user.company)


@app.route('/api/admin/approval-workflows', methods=['GET'])
def api_admin_workflows_list():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    scope = admin_companies_scope()
    company_filter = request.args.get('company')
    q = ApprovalWorkflow.query.filter(ApprovalWorkflow.company.in_(scope))
    if company_filter and company_filter != 'all':
        if company_filter not in scope:
            return jsonify({'success': False, 'error': 'Sin acceso'}), 403
        q = q.filter(ApprovalWorkflow.company == company_filter)
    workflows = q.order_by(ApprovalWorkflow.updated_at.desc()).all()
    return jsonify({'success': True, 'workflows': [_approvals_serialize_workflow(w) for w in workflows]})


@app.route('/api/admin/approval-workflows', methods=['POST'])
def api_admin_workflow_create():
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    data = request.json or {}
    name = (data.get('name') or '').strip()
    company = data.get('company') or session.get('company')
    if not name:
        return jsonify({'success': False, 'error': 'Nombre requerido'}), 400
    if company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso a esa empresa'}), 403

    approvers = data.get('approvers') or []
    if not isinstance(approvers, list) or not approvers:
        return jsonify({'success': False, 'error': 'Al menos un aprobador es requerido'}), 400

    # Validar aprobadores
    clean_approvers = []
    for i, step in enumerate(approvers):
        uid = step.get('user_id')
        if not uid:
            return jsonify({'success': False, 'error': f'Aprobador #{i+1} sin user_id'}), 400
        u = User.query.get(uid)
        if not u or u.company != company:
            return jsonify({'success': False, 'error': f'Aprobador #{i+1} inválido o de otra empresa'}), 400
        clean_approvers.append({
            'order': i + 1,
            'user_id': uid,
            'role_label': (step.get('role_label') or '').strip()[:120]
        })

    w = ApprovalWorkflow(
        company=company,
        name=name[:120],
        description=(data.get('description') or '').strip(),
        trigger_category=(data.get('trigger_category') or '').strip() or None,
        trigger_priority=(data.get('trigger_priority') or '').strip() or None,
        trigger_template_name=(data.get('trigger_template_name') or '').strip() or None,
        approvers_json=json.dumps(clean_approvers),
        is_active=bool(data.get('is_active', True)),
        created_by_id=session['user_id']
    )
    db.session.add(w)
    db.session.commit()
    log_audit('approval_workflow_created', session['user_id'], 'approval_workflow', w.id,
              f"Workflow de aprobación creado: {name} ({company})")
    return jsonify({'success': True, 'workflow': _approvals_serialize_workflow(w)})


@app.route('/api/admin/approval-workflows/<int:wid>', methods=['PUT'])
def api_admin_workflow_update(wid):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    w = ApprovalWorkflow.query.get(wid)
    if not w:
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    if w.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    data = request.json or {}
    if 'name' in data:
        w.name = (data['name'] or '').strip()[:120]
    if 'description' in data:
        w.description = (data['description'] or '').strip()
    if 'trigger_category' in data:
        w.trigger_category = (data['trigger_category'] or '').strip() or None
    if 'trigger_priority' in data:
        w.trigger_priority = (data['trigger_priority'] or '').strip() or None
    if 'trigger_template_name' in data:
        w.trigger_template_name = (data['trigger_template_name'] or '').strip() or None
    if 'is_active' in data:
        w.is_active = bool(data['is_active'])
    if 'approvers' in data:
        approvers = data['approvers'] or []
        if not approvers:
            return jsonify({'success': False, 'error': 'Al menos un aprobador es requerido'}), 400
        clean = []
        for i, step in enumerate(approvers):
            uid = step.get('user_id')
            u = User.query.get(uid) if uid else None
            if not u or u.company != w.company:
                return jsonify({'success': False, 'error': f'Aprobador #{i+1} inválido'}), 400
            clean.append({
                'order': i + 1,
                'user_id': uid,
                'role_label': (step.get('role_label') or '').strip()[:120]
            })
        w.approvers_json = json.dumps(clean)
    w.updated_at = datetime.now()
    db.session.commit()
    log_audit('approval_workflow_updated', session['user_id'], 'approval_workflow', w.id, f"Workflow actualizado: {w.name}")
    return jsonify({'success': True, 'workflow': _approvals_serialize_workflow(w)})


@app.route('/api/admin/approval-workflows/<int:wid>', methods=['DELETE'])
def api_admin_workflow_delete(wid):
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    w = ApprovalWorkflow.query.get(wid)
    if not w:
        return jsonify({'success': False}), 404
    if w.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    name = w.name
    db.session.delete(w)
    db.session.commit()
    log_audit('approval_workflow_deleted', session['user_id'], 'approval_workflow', wid, f"Workflow eliminado: {name}")
    return jsonify({'success': True})


# Página para el aprobador (link del email)
@app.route('/approvals/decide/<token>')
def approvals_decide_page(token):
    approval = Approval.query.filter_by(token=token).first()
    if not approval:
        return "Enlace de aprobación inválido o expirado.", 404
    # Debe estar logueado como el aprobador (o admin master)
    if 'user_id' not in session:
        # Redirigir a login y volver acá
        return redirect(url_for('login') + f'?next={request.path}')
    user = User.query.get(session['user_id'])
    if not user or (user.id != approval.approver_id and not is_master_admin(user.company, user.role)):
        return "No autorizado. Solo el aprobador designado puede decidir.", 403
    ticket = approval.ticket
    return render_template('approvals/decide.html',
                           approval=approval,
                           ticket=ticket,
                           user=user,
                           company_info=COMPANY_COLORS.get(user.company, {}))


@app.route('/api/approvals/<token>/decision', methods=['POST'])
def api_approval_decision(token):
    approval = Approval.query.filter_by(token=token).first()
    if not approval:
        return jsonify({'success': False, 'error': 'Enlace inválido'}), 404
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Sesión requerida'}), 401
    user = User.query.get(session['user_id'])
    if user.id != approval.approver_id and not is_master_admin(user.company, user.role):
        return jsonify({'success': False, 'error': 'No sos el aprobador designado'}), 403
    if approval.status != 'pending':
        return jsonify({'success': False, 'error': f'Ya fue {approval.status} el {approval.decision_at}'}), 400

    # Solo permitir decidir si es el paso actual (no saltear)
    prior_pending = Approval.query.filter(
        Approval.ticket_id == approval.ticket_id,
        Approval.order < approval.order,
        Approval.status == 'pending'
    ).count()
    if prior_pending > 0:
        return jsonify({'success': False, 'error': 'Aún hay aprobadores previos pendientes'}), 400

    data = request.json or {}
    action = data.get('action')  # 'approve' | 'reject'
    comment = (data.get('comment') or '').strip()[:1000]

    if action not in ('approve', 'reject'):
        return jsonify({'success': False, 'error': 'Acción inválida'}), 400
    if action == 'reject' and not comment:
        return jsonify({'success': False, 'error': 'El rechazo requiere un comentario'}), 400

    approval.status = 'approved' if action == 'approve' else 'rejected'
    approval.decision_at = datetime.now()
    approval.comment = comment
    db.session.flush()

    # Actualizar cadena
    ticket = approval.ticket
    _finalize_approval_chain(ticket)
    db.session.commit()

    # Log
    log_audit(f'ticket_{approval.status}', user.id, 'approval', approval.id,
              f"Ticket {ticket.ticket_number} {approval.status} por {user.name} (paso {approval.order}): {comment[:200]}")

    # Notificar al creador si fue rechazado (o si la cadena completa aprobó)
    try:
        if ticket.status == 'rejected' and ticket.creator and ticket.creator.email:
            send_email(
                to_email=ticket.creator.email,
                subject=f"[DeskEli] Tu solicitud {ticket.ticket_number} fue rechazada",
                body=f"""<html><body style="font-family:Segoe UI;color:#1f2937;">
                    <div style="max-width:560px;margin:20px auto;padding:22px;background:#fef2f2;border-left:4px solid #dc2626;border-radius:8px;">
                        <h2 style="color:#991b1b;">Solicitud rechazada</h2>
                        <p>Tu ticket <strong>{ticket.ticket_number}</strong> ({ticket.title}) fue rechazado por {user.name} ({approval.approver_role_label or 'aprobador'}).</p>
                        <p><strong>Motivo:</strong> {comment}</p>
                    </div></body></html>""",
                company=ticket.company
            )
        elif ticket.status == 'open' and ticket.creator and ticket.creator.email:
            send_email(
                to_email=ticket.creator.email,
                subject=f"[DeskEli] Tu solicitud {ticket.ticket_number} fue aprobada",
                body=f"""<html><body style="font-family:Segoe UI;color:#1f2937;">
                    <div style="max-width:560px;margin:20px auto;padding:22px;background:#f0fdf4;border-left:4px solid #16a34a;border-radius:8px;">
                        <h2 style="color:#15803d;">Solicitud aprobada</h2>
                        <p>Tu ticket <strong>{ticket.ticket_number}</strong> ({ticket.title}) fue aprobado y pasa al equipo de TI para su atención.</p>
                    </div></body></html>""",
                company=ticket.company
            )
    except Exception:
        pass

    return jsonify({'success': True, 'ticket_status': ticket.status, 'action': approval.status})


@app.route('/api/approvals/pending')
def api_approvals_my_pending():
    """Mis aprobaciones pendientes (widget en dashboards)."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    approvals = Approval.query.filter_by(approver_id=session['user_id'], status='pending').all()
    # Filtrar solo los que son el paso activo (no hay previos pendientes)
    active = []
    for a in approvals:
        prior = Approval.query.filter(
            Approval.ticket_id == a.ticket_id,
            Approval.order < a.order,
            Approval.status == 'pending'
        ).count()
        if prior == 0:
            active.append(a)
    active.sort(key=lambda x: x.created_at or datetime.min, reverse=True)
    return jsonify({
        'success': True,
        'count': len(active),
        'approvals': [_approvals_serialize_approval(a, include_token=True) for a in active]
    })


@app.route('/api/kb/article/<int:article_id>/feedback', methods=['POST'])
def api_kb_feedback(article_id):
    """El usuario marca útil/no útil un artículo (idempotente por usuario)."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    a = KnowledgeArticle.query.get(article_id)
    if not a:
        return jsonify({'success': False}), 404
    data = request.json or {}
    is_helpful = bool(data.get('helpful'))
    comment = (data.get('comment') or '').strip()[:500]

    # ¿Ya dio feedback antes? Actualizarlo en vez de duplicar
    existing = KnowledgeArticleFeedback.query.filter_by(
        article_id=article_id, user_id=session['user_id']
    ).first()

    if existing:
        # Ajustar contadores si cambia el signo
        if existing.is_helpful != is_helpful:
            if existing.is_helpful:
                a.helpful_count = max(0, (a.helpful_count or 0) - 1)
            else:
                a.not_helpful_count = max(0, (a.not_helpful_count or 0) - 1)
            if is_helpful:
                a.helpful_count = (a.helpful_count or 0) + 1
            else:
                a.not_helpful_count = (a.not_helpful_count or 0) + 1
        existing.is_helpful = is_helpful
        existing.comment = comment
    else:
        fb = KnowledgeArticleFeedback(
            article_id=article_id,
            user_id=session['user_id'],
            is_helpful=is_helpful,
            comment=comment,
            ip_addr=request.remote_addr
        )
        db.session.add(fb)
        if is_helpful:
            a.helpful_count = (a.helpful_count or 0) + 1
        else:
            a.not_helpful_count = (a.not_helpful_count or 0) + 1

    db.session.commit()
    return jsonify({
        'success': True,
        'helpful_count': a.helpful_count,
        'not_helpful_count': a.not_helpful_count
    })


# ═════════════════════════════════════════════════════════════════════════════
# API ENDPOINTS
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/health')
def api_health():
    return jsonify({
        'status': 'healthy',
        'app': 'DeskEli',
        'version': '2.1.0',
        'timestamp': datetime.now().isoformat()
    })


@app.route('/api/session/ping')
def api_session_ping():
    """Keep-alive de sesión. Devuelve 200 si autenticado, 401 si expiró.

    El frontend llama a este endpoint cada 5 min para que la cookie de sesión
    se renueve automáticamente (SESSION_REFRESH_EACH_REQUEST=True por default)
    y para detectar temprano que la sesión murió.
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'authenticated': False}), 401
    session.permanent = True  # asegurar que se renueva la cookie
    return jsonify({
        'success': True,
        'authenticated': True,
        'user_id': session.get('user_id'),
        'role': session.get('role'),
        'company': session.get('company'),
        'server_time': datetime.now().isoformat(timespec='seconds')
    })

# Detección de categoría a partir del texto del usuario.
# Cada categoría tiene términos fuertes (high-confidence) y débiles (apoyo).
CATEGORY_HINTS = {
    'SAP': {
        'strong': ['sap', 'abap', 'transaccion', 'transacción', 'basis', 's/4hana', 's4hana', 'spool', 'mandante', 'mandantes', 'autorización sap', 'autorizacion sap', 'sap gui', 'gui sap', 'logon sap', 'sap logon', 'sm21', 'st22', 'su01', 'pfcg', 'me21n', 'va01', 'mb51', 'fb01', 'rsync sap'],
        'weak':   ['fiori', 'transports', 'mandante', 'modulo sap', 'módulo sap', 'fi', 'mm', 'sd', 'pp', 'hcm']
    },
    'Servidores': {
        'strong': ['servidor', 'server', 'windows server', 'linux server', 'iis', 'apache', 'nginx', 'serv. archivos', 'file server', 'dns', 'dhcp', 'active directory', 'controlador de dominio', 'domain controller', 'gpo', 'directiva de grupo', 'hyper-v', 'vmware', 'esxi', 'vcenter', 'cluster', 'failover'],
        'weak':   ['cae', 'cayo', 'caido', 'caída', 'ping', 'no responde', 'reboot', 'reinicio', 'reiniciar', 'restart', 'servicio detenido']
    },
    'Red': {
        'strong': ['red', 'wifi', 'wi-fi', 'lan', 'wan', 'switch', 'router', 'firewall', 'vpn', 'cisco', 'fortinet', 'mikrotik', 'cableado', 'puerto ethernet', 'rj45', 'ssid', 'banda ancha', 'fibra', 'isp', 'proveedor de internet'],
        'weak':   ['internet', 'conexion', 'conexión', 'conectar', 'conectividad', 'latencia', 'lento internet', 'sin red', 'sin internet']
    },
    'Email': {
        'strong': ['correo', 'email', 'outlook', 'mail', 'exchange', 'imap', 'smtp', 'pop3', 'firma de correo', 'distribución', 'buzón', 'buzon', 'casilla', 'spam', 'phishing', 'antispam', 'office 365 mail', 'm365 mail'],
        'weak':   ['enviar', 'recibir', 'adjunto', 'destinatario']
    },
    'Accesos': {
        'strong': ['contraseña', 'contrasena', 'password', 'clave de acceso', 'usuario bloqueado', 'cuenta bloqueada', 'olvide mi clave', 'olvidé mi clave', 'resetear clave', 'reset password', 'permisos', 'autorizacion', 'autorización', 'sso', 'mfa', '2fa', 'doble factor', 'token', 'pin'],
        'weak':   ['acceder', 'no me deja entrar', 'no puedo ingresar', 'bloqueo', 'desbloqueo']
    },
    'Hardware': {
        'strong': ['hardware', 'monitor', 'teclado', 'mouse', 'pantalla', 'cpu', 'computador', 'computadora', 'pc', 'laptop', 'portatil', 'portátil', 'memoria ram', 'disco duro', 'ssd', 'fuente de poder', 'cargador', 'bateria', 'batería', 'usb', 'hdmi', 'webcam', 'audífonos', 'audifonos'],
        'weak':   ['equipo', 'maquina', 'máquina']
    },
    'Software': {
        'strong': ['instalar programa', 'reinstalar', 'desinstalar', 'actualizar windows', 'actualizar software', 'licencia', 'aplicacion no abre', 'aplicación no abre', 'crashea', 'cierra solo', 'pantalla azul', 'bsod', 'driver', 'controlador'],
        'weak':   ['software', 'programa', 'aplicacion', 'aplicación']
    },
    'Impresoras': {
        'strong': ['impresora', 'impresion', 'impresión', 'imprimir', 'toner', 'cartucho', 'atasco', 'cola de impresión', 'driver impresora', 'plotter'],
        'weak':   ['papel', 'hoja']
    },
    'Telefonia': {
        'strong': ['telefono', 'teléfono', 'telefonia', 'telefonía', 'anexo', 'extension telefonica', 'extensión telefónica', 'voip', 'central telefónica', 'pbx', 'ip-pbx', 'teams phone', 'zoom phone', 'softphone'],
        'weak':   ['llamada', 'llamar', 'marcar']
    },
    'Office': {
        'strong': ['word', 'excel', 'powerpoint', 'onedrive', 'sharepoint', 'teams', 'office 365', 'm365', 'macro', 'planilla', 'documento word', 'hoja de calculo', 'hoja de cálculo'],
        'weak':   ['oficina']
    },
    'Seguridad': {
        'strong': ['antivirus', 'malware', 'virus', 'ransomware', 'troyano', 'edr', 'crowdstrike', 'defender', 'kaspersky', 'symantec', 'phishing', 'ataque', 'incidente de seguridad'],
        'weak':   ['seguridad']
    }
}


def detect_question_category(question_lower):
    """Detecta la categoría dominante en la pregunta del usuario.
    Retorna (category, confidence). Confidence: 0-100.
    Si no puede decidir o varias categorías compiten, retorna (None, 0)."""
    if not question_lower:
        return None, 0
    scores = {}
    for cat, hints in CATEGORY_HINTS.items():
        score = 0
        # Strong hints (peso 10) — usar word boundaries para evitar falsos positivos
        for term in hints['strong']:
            # Si el término tiene espacio o es largo (>=4), usar 'in' (substring)
            # Para términos cortos como "sap", "ad", exigir word boundary
            if ' ' in term or len(term) >= 4:
                if term in question_lower:
                    score += 10
            else:
                # word boundary check
                pattern = r'\b' + re.escape(term) + r'\b'
                if re.search(pattern, question_lower):
                    score += 10
        # Weak hints (peso 3)
        for term in hints['weak']:
            if ' ' in term or len(term) >= 4:
                if term in question_lower:
                    score += 3
            else:
                pattern = r'\b' + re.escape(term) + r'\b'
                if re.search(pattern, question_lower):
                    score += 3
        if score > 0:
            scores[cat] = score

    if not scores:
        return None, 0

    # Ordenar por puntaje y verificar margen de confianza
    sorted_scores = sorted(scores.items(), key=lambda x: -x[1])
    best_cat, best_sc = sorted_scores[0]
    second_sc = sorted_scores[1][1] if len(sorted_scores) > 1 else 0

    # Si la segunda categoría está muy cerca, no confiamos lo suficiente
    if best_sc < 10:
        return None, 0  # ni el mejor llegó al umbral fuerte
    if second_sc > 0 and (best_sc - second_sc) < 4:
        # Empate aproximado, retornar la mejor pero con baja confianza
        return best_cat, 40

    confidence = min(100, best_sc * 5)
    return best_cat, confidence


def search_kb_match(question_text):
    """Busca la mejor coincidencia en la Base de Conocimiento del bot.
    Retorna (kb_entry, score) o (None, 0) si no hay match.

    Estrategia:
    1. Detectar categoría de la pregunta (SAP, Servidores, Red, etc.)
    2. Si hay categoría con alta confianza, FILTRAR las KB a esa categoría
    3. Si no, buscar en todas pero penalizar cross-category fuerte
    4. Scoring por keyword (word boundary) + título KB + pregunta KB
    """
    q_lower = question_text.lower().strip()
    if len(q_lower) < 3:
        return None, 0

    all_kb = BotKnowledge.query.all()
    if not all_kb:
        return None, 0

    # PASO 1: Detectar categoría dominante en la pregunta
    detected_cat, confidence = detect_question_category(q_lower)
    if detected_cat:
        print(f"[BOT] Categoría detectada en pregunta: '{detected_cat}' (confianza: {confidence}%)")

    # PASO 2: Filtrar o priorizar KB por categoría
    candidate_kb = all_kb
    if detected_cat and confidence >= 50:
        # Alta confianza: limitar a esa categoría (case-insensitive)
        same_cat = [k for k in all_kb if (k.category or '').lower() == detected_cat.lower()]
        if same_cat:
            candidate_kb = same_cat
            print(f"[BOT] Filtrando a {len(same_cat)} KBs de categoría '{detected_cat}'")
        # Si no hay ninguna KB de esa categoría, seguimos con todas pero penalizaremos cross-cat

    stopwords = {'que', 'qué', 'como', 'cómo', 'el', 'la', 'los', 'las', 'un', 'una',
                 'de', 'del', 'al', 'en', 'a', 'con', 'por', 'para', 'mi', 'me',
                 'tu', 'te', 'es', 'son', 'no', 'si', 'sí', 'y', 'o', 'pero',
                 'lo', 'le', 'se', 'su', 'sus', 'esta', 'este', 'eso', 'esto',
                 'puedo', 'puede', 'puedes', 'tengo', 'tiene', 'hay', 'hace',
                 'esta', 'estoy', 'estan', 'están', 'fue', 'fueron', 'mas', 'más',
                 'ya', 'sin', 'aqui', 'aquí', 'ahi', 'ahí', 'pero', 'porque'}

    def tokenize(text):
        text = re.sub(r'[?¿.,;:!¡()\[\]]', ' ', text)
        return set(w for w in text.split() if len(w) > 2 and w not in stopwords)

    q_words = tokenize(q_lower)
    best_match = None
    best_score = 0

    for kb in candidate_kb:
        score = 0
        kb_cat = (kb.category or '').lower()

        # 1. Match de keywords con word-boundary para términos cortos
        keywords = [k.strip().lower() for k in (kb.keywords or '').split(',') if k.strip()]
        for kw in keywords:
            if not kw:
                continue
            # Keyword larga o multi-word: substring
            if ' ' in kw or len(kw) >= 5:
                if kw in q_lower:
                    score += 6
            elif len(kw) >= 3:
                # Word boundary para términos cortos
                pattern = r'\b' + re.escape(kw) + r'\b'
                if re.search(pattern, q_lower):
                    score += 5
            # Ignorar keywords de 1-2 caracteres (causan falsos positivos)

        # 2. Palabras de la pregunta vs pregunta KB
        kb_q_words = tokenize((kb.question or '').lower())
        common_words = q_words & kb_q_words
        score += len(common_words) * 2

        # 3. Palabras de la pregunta vs palabras de keywords del KB
        kb_kw_words = set()
        for kw in keywords:
            for w in kw.split():
                if len(w) >= 3:
                    kb_kw_words.add(w)
        score += len(q_words & kb_kw_words) * 2

        # 4. Bonus si la categoría del KB coincide con la detectada
        if detected_cat and kb_cat == detected_cat.lower():
            score += 8
        # Penalización fuerte si el KB es de otra categoría y la pregunta es claramente de otra
        elif detected_cat and confidence >= 60 and kb_cat and kb_cat != detected_cat.lower():
            score = int(score * 0.3)  # reducir 70% el score cross-category

        if score > best_score:
            best_score = score
            best_match = kb

    # Umbral dinámico: más estricto si la pregunta es corta
    min_threshold = 6 if len(q_words) >= 3 else 8
    if best_score >= min_threshold:
        cat_info = f" (cat: {best_match.category})" if best_match.category else ""
        print(f"[BOT] ✓ Mejor match: '{best_match.question}'{cat_info} score={best_score} (umbral: {min_threshold})")
        return best_match, best_score
    print(f"[BOT] Sin match suficiente (mejor score={best_score}, umbral: {min_threshold})")
    return None, 0


@app.route('/api/bot/ask', methods=['POST'])
def api_bot_ask():
    """Bot de soporte inteligente: primero busca en KB, luego Claude"""
    try:
        data = request.get_json(force=True) if request.data else {}
        question = data.get('question', '').strip()

        if not question or len(question) < 3:
            return jsonify({'success': False, 'error': 'Pregunta muy corta'}), 400

        answer = None
        source = None
        kb_match = None
        kb_score = 0

        # PASO 1: BUSCAR EN BASE DE CONOCIMIENTO (Soluciones Bot)
        try:
            kb_match, kb_score = search_kb_match(question)
            if kb_match:
                answer = kb_match.answer
                source = 'kb'
                print(f"[BOT] ✓ Match en KB: '{kb_match.question}' (score: {kb_score})")
        except Exception as e:
            print(f"[BOT] Error búsqueda KB: {e}")

        # PASO 2: Si no hay match en KB, usar CLAUDE
        if not answer:
            print(f"[BOT] No hay match en KB, llamando a Claude...")
            try:
                if CLAUDE_API_KEY and CLAUDE_API_KEY.startswith('sk-'):
                    import anthropic
                    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)

                    # Detectar categoría para dar contexto correcto a Claude
                    detected_cat_for_claude, _ = detect_question_category(question.lower())
                    cat_context = ''
                    if detected_cat_for_claude:
                        cat_context = f"\nContexto: la pregunta es sobre **{detected_cat_for_claude}**. NO asumas que es SAP u otra categoría — responde específicamente sobre {detected_cat_for_claude}."

                    # Pasar SOLO ejemplos de la categoría detectada para no contaminar
                    context_examples = ''
                    try:
                        if detected_cat_for_claude:
                            sample_kbs = BotKnowledge.query.filter(
                                BotKnowledge.category.ilike(detected_cat_for_claude)
                            ).limit(3).all()
                        else:
                            sample_kbs = []
                        if sample_kbs:
                            context_examples = '\n\nEjemplos de respuestas relevantes:\n'
                            for s in sample_kbs:
                                context_examples += f'- {s.question}\n  → {s.answer[:200]}\n'
                    except Exception:
                        pass

                    msg = client.messages.create(
                        model="claude-haiku-4-5",
                        max_tokens=400,
                        messages=[{
                            "role": "user",
                            "content": (
                                f"Eres un agente de soporte IT corporativo. Responde en español, brevemente (3-5 líneas) "
                                f"con pasos prácticos para el usuario final. {cat_context}\n\n"
                                f"Pregunta del usuario: {question}{context_examples}"
                            )
                        }]
                    )
                    answer = msg.content[0].text.strip() if msg.content else None
                    source = 'claude'
                    print(f"[BOT] Respuesta de Claude (cat: {detected_cat_for_claude or 'no detectada'})")
            except Exception as e:
                print(f"[BOT] Error Claude: {type(e).__name__}: {e}")

        # Si todo falla, respuesta genérica
        if not answer:
            answer = f"He registrado tu pregunta sobre '{question[:30]}...'. Un especialista te contactará pronto."
            source = 'fallback'

        # PASO 2: OBTENER USUARIO BOT
        bot_user = User.query.filter_by(username='bot_system').first()
        if not bot_user:
            try:
                bot_user = User(
                    username='bot_system',
                    name='Bot DeskEli',
                    email='bot@DeskEli.local',
                    role='admin',
                    company='Soporte',
                    password_hash='system',
                    is_active=True
                )
                db.session.add(bot_user)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                bot_user = None

        # PASO 3: OBTENER ESPECIALISTA
        specialist = User.query.filter_by(role='technician').first()
        specialist_name = specialist.name if specialist else 'especialista disponible'

        # PASO 4: CREAR TICKET (con microsegundos para evitar duplicados)
        import uuid
        ticket_number = f'AUTO-{datetime.now().strftime("%Y%m%d%H%M%S%f")}-{uuid.uuid4().hex[:4]}'
        try:
            ticket = Ticket(
                ticket_number=ticket_number,
                title=f'Consulta: {question[:50]}',
                description=f'Pregunta: {question}\n\n🤖 Bot: {answer}',
                priority='medium',
                category='General',
                company='Soporte',
                creator_id=bot_user.id if bot_user else 1,
                assignee_id=specialist.id if specialist else None,
                status='open',
                sla_minutes=240
            )
            db.session.add(ticket)
            db.session.flush()

            if bot_user:
                msg_obj = Message(
                    ticket_id=ticket.id,
                    user_id=bot_user.id,
                    text=answer
                )
                db.session.add(msg_obj)

            db.session.commit()
        except Exception as e:
            print(f"[BOT] Error creando ticket: {e}")
            db.session.rollback()

        # Si la respuesta viene de KB, el ticket es resolved (auto-cerrado)
        resolved = (source == 'kb')
        source_label = {
            'kb': '📚 Base de Conocimiento',
            'claude': '🤖 IA (Claude)',
            'fallback': '📝 Sin respuesta automática'
        }.get(source, source)

        ticket_msg = (
            f'✓ Respondido desde {source_label}. Ticket #{ticket_number} {"cerrado" if resolved else f"asignado a {specialist_name}"}'
        )

        return jsonify({
            'success': True,
            'answer': answer,
            'resolved': resolved,
            'source': source,
            'source_label': source_label,
            'kb_question': kb_match.question if kb_match else None,
            'kb_category': kb_match.category if kb_match else None,
            'ticket_number': ticket_number,
            'ticket_message': ticket_msg,
            'assigned_to': specialist_name
        })

    except Exception as e:
        print(f"[BOT] Error general: {e}")
        return jsonify({
            'success': False,
            'error': f'Error: {str(e)}'
        }), 500

@app.route('/api/bot/ticket-from-chat', methods=['POST'])
def api_bot_ticket_from_chat():
    """Crear ticket desde la conversación con el bot.
    - Si resolved=true: ticket cerrado automáticamente (registro KB)
    - Si resolved=false: asigna a especialista (técnico con menor carga)
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autenticado'}), 401

    try:
        data = request.get_json()
        title = (data.get('title') or 'Consulta del bot').strip()[:200]
        conversation = (data.get('conversation') or '').strip()
        resolved = bool(data.get('resolved', False))
        category = (data.get('category') or 'General').strip()
        priority = (data.get('priority') or 'medium').strip()
        company = session['company']
        user_id = session['user_id']

        if not conversation:
            return jsonify({'success': False, 'error': 'Conversación vacía'}), 400

        # Generar número de ticket
        count = Ticket.query.filter_by(company=company).count() + 1
        ticket_number = f'TKT-{company.upper()}-{count:05d}'

        # Si está resuelto: status=resolved, sin asignar
        # Si NO está resuelto: status=open, asignar al técnico con menor carga
        assignee_id = None
        assignee_name = None

        if not resolved:
            # Buscar técnico con menor carga en esta empresa
            from sqlalchemy import func
            technicians = User.query.filter_by(company=company, role='technician', is_active=True).all()
            if technicians:
                # Contar tickets activos por técnico
                best_tech = None
                min_load = float('inf')
                for t in technicians:
                    load = Ticket.query.filter(
                        Ticket.assignee_id == t.id,
                        Ticket.status.in_(['open', 'in_progress'])
                    ).count()
                    if load < min_load:
                        min_load = load
                        best_tech = t
                if best_tech:
                    assignee_id = best_tech.id
                    assignee_name = best_tech.name

        # SLA por prioridad
        sla_map = {'low': 1440, 'medium': 480, 'high': 240, 'critical': 60}
        sla_min = sla_map.get(priority, 480)

        # Descripción con la conversación
        description = (
            f"🤖 TICKET CREADO DESDE EL CHAT CON ELI\n"
            f"{'='*50}\n\n"
            f"Estado: {'✅ RESUELTO en el chat' if resolved else '⚠️ REQUIERE ESPECIALISTA'}\n"
            f"Categoría: {category}\n\n"
            f"--- CONVERSACIÓN ---\n\n"
            f"{conversation}\n\n"
            f"--- FIN CONVERSACIÓN ---"
        )

        # Crear el ticket
        ticket = Ticket(
            ticket_number=ticket_number,
            title=title,
            description=description,
            status='resolved' if resolved else ('in_progress' if assignee_id else 'open'),
            priority=priority,
            company=company,
            creator_id=user_id,
            assignee_id=assignee_id,
            category=category,
            sla_minutes=sla_min,
            sla_deadline=compute_sla_deadline(datetime.now(), sla_min, company)
        )
        if resolved:
            ticket.resolved_at = datetime.now()

        db.session.add(ticket)
        db.session.commit()

        # Audit log
        log_audit(
            'create_ticket_from_chat',
            user_id,
            'ticket',
            ticket.id,
            f'Ticket {ticket_number} creado desde chat - Resuelto: {resolved}'
        )

        # Notificar por email al técnico asignado automáticamente
        if assignee_id and not resolved:
            try:
                tech_to_notify = User.query.get(assignee_id)
                if tech_to_notify:
                    notify_ticket_assigned(
                        ticket=ticket,
                        new_assignee=tech_to_notify,
                        assigned_by_name='Asignación automática (IA por carga de trabajo)',
                        reason='Técnico con menor carga activa en la empresa'
                    )
            except Exception as e:
                print(f'[WARN] Notificación email: {e}')

        # Emitir WebSocket
        try:
            emit_ticket_event(company, 'ticket_created', {
                'ticket_number': ticket_number,
                'title': title,
                'priority': priority,
                'created_by': session.get('name', 'Usuario'),
                'resolved': resolved,
                'assigned_to': assignee_name
            })
        except Exception as e:
            print(f'[WARN] WebSocket emit: {e}')

        return jsonify({
            'success': True,
            'ticket_number': ticket_number,
            'ticket_id': ticket.id,
            'status': ticket.status,
            'resolved': resolved,
            'assigned_to': assignee_name,
            'message': (
                f'Ticket {ticket_number} ' +
                ('cerrado automáticamente' if resolved else
                 (f'asignado a {assignee_name}' if assignee_name else 'creado - pendiente de asignación'))
            )
        })

    except Exception as e:
        db.session.rollback()
        print(f'[ERROR ticket-from-chat] {e}')
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/search', methods=['GET'])
def api_search():
    """Búsqueda global (Ctrl+K)"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    query = request.args.get('q', '').lower()
    if len(query) < 2:
        return jsonify({'success': True, 'results': []})

    company = session['company']
    tickets = Ticket.query.filter_by(company=company).filter(
        (Ticket.title.ilike(f'%{query}%')) |
        (Ticket.description.ilike(f'%{query}%')) |
        (Ticket.ticket_number.ilike(f'%{query}%'))
    ).limit(20).all()

    results = [{
        'id': t.id,
        'ticket_number': t.ticket_number,
        'title': t.title,
        'status': t.status,
        'priority': t.priority
    } for t in tickets]

    return jsonify({'success': True, 'results': results})

# ════════════════════════════════════════════════════════════════════
# EXPORT DE TICKETS — Cabecera estándar de 12 columnas
# ════════════════════════════════════════════════════════════════════

# Encabezados oficiales del reporte
EXPORT_HEADERS = [
    'NRO.TICKET', 'ASUNTO', 'CATEGORIA', 'CLIENTE FINAL', 'ESPECIALISTA',
    'GRUPO ESPECIALISTA', 'ESTADO', 'FECHA DE REGISTRO', 'FECHA DE CIERRE',
    'SLA', 'TIPO DE CASO', 'TIPO DE REGISTRO'
]

# Mapeos legibles
_STATUS_LABEL = {
    'open': 'Abierto', 'in_progress': 'En Progreso',
    'resolved': 'Resuelto', 'closed': 'Cerrado'
}
_PRIORITY_LABEL = {
    'critical': 'Crítica', 'high': 'Alta',
    'medium': 'Media', 'low': 'Baja'
}


def _ticket_grupo_especialista(ticket):
    """Determina el 'grupo' del especialista a partir de:
    1. role_label del assignee (ej. 'Supervisor de Mesa')
    2. Primer subrol asignado al técnico
    3. Categoría del ticket como fallback
    """
    if not ticket.assignee:
        return 'Sin asignar'
    a = ticket.assignee
    if a.role_label:
        return a.role_label
    # Subroles via UserSubrole
    try:
        usr = UserSubrole.query.filter_by(user_id=a.id).first()
        if usr:
            sr = Subrole.query.get(usr.subrole_id)
            if sr and sr.name:
                return sr.name
    except Exception:
        pass
    # Categoría como fallback
    return (ticket.category or 'General').strip() or 'General'


def _ticket_sla_status(ticket):
    """Estado del SLA: 'Cumplido', 'Vencido', 'En plazo', 'N/A'."""
    if not ticket.sla_deadline:
        return 'N/A'
    now = datetime.now()
    if ticket.status in ('resolved', 'closed'):
        # Si tenemos resolved_at, comparar; si no, asumir cumplido
        ref = ticket.resolved_at or ticket.updated_at
        if ref and ticket.sla_deadline:
            return 'Cumplido' if ref <= ticket.sla_deadline else 'Vencido'
        return 'Cumplido'
    # Ticket activo
    return 'Vencido' if ticket.sla_deadline < now else 'En plazo'


def _ticket_tipo_caso(ticket):
    """Tipo del caso según prioridad. Si la categoría es Servidores → 'Infraestructura'."""
    cat = (ticket.category or '').lower()
    if cat in ('servidores', 'infraestructura', 'red'):
        return 'Incidente Infraestructura'
    if cat in ('accesos', 'seguridad'):
        return 'Solicitud de Acceso'
    if cat == 'sap':
        return 'Incidente SAP'
    # Default según prioridad
    if ticket.priority == 'critical':
        return 'Incidente Crítico'
    return 'Incidente'


def _ticket_tipo_registro(ticket):
    """Origen del ticket:
    - 'Automático' si número empieza con AUTO- o ALARMA
    - 'Bot/IA' si el creador es bot_system o título empieza con 'Consulta:'
    - 'Email' si fue creado vía buzón IMAP (revisar audit/log es complejo, usamos heurística)
    - 'Empleado' si el creador es role=employee
    - 'Técnico TI' si el creador es role=technician/admin
    """
    tn = ticket.ticket_number or ''
    if tn.startswith('AUTO-') or 'ALARMA' in (ticket.title or '').upper():
        return 'Automático (Sistema)'
    if (ticket.title or '').startswith('Consulta:'):
        return 'Bot/IA (Eli)'
    creator = ticket.creator if hasattr(ticket, 'creator') else None
    if creator:
        if (creator.username or '').lower() == 'bot_system':
            return 'Bot/IA (Eli)'
        if creator.role == 'employee':
            return 'Empleado'
        if creator.role in ('technician', 'admin'):
            return 'Técnico TI'
    return 'Manual'


def _build_export_row(ticket):
    """Devuelve la fila con las 12 columnas oficiales para Excel/CSV."""
    creator_name = ticket.creator.name if (hasattr(ticket, 'creator') and ticket.creator) else ''
    assignee_name = ticket.assignee.name if ticket.assignee else 'Sin asignar'
    return [
        ticket.ticket_number,
        ticket.title,
        ticket.category or 'General',
        creator_name,
        assignee_name,
        _ticket_grupo_especialista(ticket),
        _STATUS_LABEL.get(ticket.status, ticket.status or ''),
        ticket.created_at.strftime('%d/%m/%Y %H:%M') if ticket.created_at else '',
        ticket.resolved_at.strftime('%d/%m/%Y %H:%M') if ticket.resolved_at else '',
        _ticket_sla_status(ticket),
        _ticket_tipo_caso(ticket),
        _ticket_tipo_registro(ticket),
    ]


@app.route('/api/export/excel', methods=['GET'])
def api_export_excel():
    """Exporta tickets a Excel con 3 hojas:
    1. Resumen Ejecutivo (KPIs)
    2. Top Técnicos (ranking)
    3. Tickets (las 12 columnas oficiales)
    """
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']
    # Empresa display name
    company_obj = Company.query.filter_by(code=company).first()
    company_name = company_obj.name if company_obj else company.title()

    # Excluir tickets internos (DMs, chats)
    tickets = Ticket.query.filter(
        Ticket.company == company,
        ~Ticket.ticket_number.like('DM-%'),
        ~Ticket.ticket_number.like('CHAT-%'),
    ).order_by(Ticket.id.desc()).all()

    # ===== CÁLCULO DE MÉTRICAS =====
    now = datetime.now()
    total = len(tickets)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    resolved_today = sum(1 for t in tickets
                        if t.status == 'resolved' and t.resolved_at and t.resolved_at >= today_start)
    by_status = {'open': 0, 'in_progress': 0, 'resolved': 0, 'closed': 0}
    by_priority = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    by_category = {}
    sla_cumplidos = 0
    sla_vencidos = 0
    ratings = []
    resolution_hours = []
    for t in tickets:
        by_status[t.status or 'open'] = by_status.get(t.status or 'open', 0) + 1
        by_priority[t.priority or 'medium'] = by_priority.get(t.priority or 'medium', 0) + 1
        cat = (t.category or 'General').strip()
        by_category[cat] = by_category.get(cat, 0) + 1
        if t.status in ('resolved', 'closed') and t.sla_deadline and t.resolved_at:
            if t.resolved_at <= t.sla_deadline:
                sla_cumplidos += 1
            else:
                sla_vencidos += 1
        if t.rating:
            ratings.append(t.rating)
        if t.resolved_at and t.created_at:
            h = (t.resolved_at - t.created_at).total_seconds() / 3600
            if h >= 0:
                resolution_hours.append(h)

    sla_total = sla_cumplidos + sla_vencidos
    sla_pct = round((sla_cumplidos / sla_total) * 100, 1) if sla_total else 0
    avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else 0
    avg_resolution_h = round(sum(resolution_hours) / len(resolution_hours), 1) if resolution_hours else 0

    # Ranking de técnicos
    tech_stats_map = {}
    for t in tickets:
        if not t.assignee_id:
            continue
        if t.assignee_id not in tech_stats_map:
            u = t.assignee
            tech_stats_map[t.assignee_id] = {
                'name': u.name if u else f'Usuario {t.assignee_id}',
                'username': u.username if u else '',
                'assigned': 0,
                'resolved': 0,
                'avg_resolution_h': [],
                'critical': 0,
                'high': 0,
            }
        s = tech_stats_map[t.assignee_id]
        s['assigned'] += 1
        if t.status in ('resolved', 'closed'):
            s['resolved'] += 1
            if t.resolved_at and t.created_at:
                s['avg_resolution_h'].append((t.resolved_at - t.created_at).total_seconds() / 3600)
        if t.priority == 'critical': s['critical'] += 1
        elif t.priority == 'high': s['high'] += 1
    tech_stats = []
    for tid, s in tech_stats_map.items():
        rate = round((s['resolved'] / s['assigned']) * 100, 1) if s['assigned'] else 0
        avg_h = round(sum(s['avg_resolution_h']) / len(s['avg_resolution_h']), 1) if s['avg_resolution_h'] else 0
        tech_stats.append({
            'name': s['name'], 'username': s['username'],
            'assigned': s['assigned'], 'resolved': s['resolved'],
            'rate': rate, 'avg_resolution_h': avg_h,
            'critical': s['critical'], 'high': s['high'],
        })
    tech_stats.sort(key=lambda x: -x['resolved'])

    # ===== CONSTRUIR EXCEL =====
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Estilos comunes
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    subheader_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    title_font = Font(bold=True, size=18, color='1E40AF')
    section_font = Font(bold=True, size=14, color='1E40AF')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    status_labels = {'open': 'Abierto', 'in_progress': 'En Progreso', 'resolved': 'Resuelto', 'closed': 'Cerrado'}
    priority_labels = {'critical': '🔴 Crítica', 'high': '🟠 Alta', 'medium': '🟡 Media', 'low': '🟢 Baja'}

    # ==== HOJA 1: RESUMEN EJECUTIVO ====
    ws = wb.active
    ws.title = 'Resumen Ejecutivo'

    ws['A1'] = '📊 RESUMEN EJECUTIVO - Reporte de Tickets'
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A2'] = f'Empresa: {company_name}'
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:D2')
    ws['A3'] = f'Generado: {now.strftime("%d/%m/%Y %H:%M")}'
    ws['A3'].font = Font(italic=True, color='6B7280')
    ws.merge_cells('A3:D3')
    ws['A4'] = f'Tickets analizados: {total}  (DMs/Chats internos excluidos)'
    ws['A4'].font = Font(italic=True, color='6B7280')
    ws.merge_cells('A4:D4')

    # KPIs principales
    ws['A6'] = 'INDICADOR'
    ws['B6'] = 'VALOR'
    for col_letter in ('A', 'B'):
        ws[f'{col_letter}6'].font = header_font
        ws[f'{col_letter}6'].fill = header_fill
        ws[f'{col_letter}6'].border = border
        ws[f'{col_letter}6'].alignment = header_align

    kpis = [
        ('Tickets Totales', total),
        ('Resueltos Hoy', resolved_today),
        ('En Progreso', by_status.get('in_progress', 0)),
        ('Abiertos (sin asignar/pendientes)', by_status.get('open', 0)),
        ('Resueltos (total)', by_status.get('resolved', 0)),
        ('Cerrados (total)', by_status.get('closed', 0)),
        ('Cumplimiento SLA (%)', f'{sla_pct}%'),
        ('SLA Cumplidos', sla_cumplidos),
        ('SLA Vencidos', sla_vencidos),
        ('Tiempo Promedio Resolución (h)', avg_resolution_h),
        ('Calificación Promedio Cliente', f'{avg_rating}/5' if avg_rating else 'Sin calificaciones'),
        ('Tickets Críticos (total)', by_priority.get('critical', 0)),
    ]
    for i, (label, value) in enumerate(kpis, start=7):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
        if i % 2 == 0:
            ws[f'A{i}'].fill = subheader_fill
            ws[f'B{i}'].fill = subheader_fill

    # Distribución por estado
    row = len(kpis) + 9
    ws[f'A{row}'] = 'DISTRIBUCIÓN POR ESTADO'
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    ws[f'A{row}'] = 'Estado'; ws[f'B{row}'] = 'Cantidad'
    for col_letter in ('A', 'B'):
        ws[f'{col_letter}{row}'].font = header_font
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
        ws[f'{col_letter}{row}'].alignment = header_align
    for k, v in by_status.items():
        row += 1
        ws[f'A{row}'] = status_labels.get(k, k)
        ws[f'B{row}'] = v
        ws[f'A{row}'].border = border; ws[f'B{row}'].border = border

    # Distribución por prioridad
    row += 3
    ws[f'A{row}'] = 'DISTRIBUCIÓN POR PRIORIDAD'
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    ws[f'A{row}'] = 'Prioridad'; ws[f'B{row}'] = 'Cantidad'
    for col_letter in ('A', 'B'):
        ws[f'{col_letter}{row}'].font = header_font
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
        ws[f'{col_letter}{row}'].alignment = header_align
    for k in ('critical', 'high', 'medium', 'low'):
        row += 1
        ws[f'A{row}'] = priority_labels[k]
        ws[f'B{row}'] = by_priority.get(k, 0)
        ws[f'A{row}'].border = border; ws[f'B{row}'].border = border

    # Top categorías
    row += 3
    ws[f'A{row}'] = 'TOP CATEGORÍAS'
    ws[f'A{row}'].font = section_font
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    ws[f'A{row}'] = 'Categoría'; ws[f'B{row}'] = 'Cantidad'
    for col_letter in ('A', 'B'):
        ws[f'{col_letter}{row}'].font = header_font
        ws[f'{col_letter}{row}'].fill = header_fill
        ws[f'{col_letter}{row}'].border = border
        ws[f'{col_letter}{row}'].alignment = header_align
    sorted_cats = sorted(by_category.items(), key=lambda x: -x[1])[:10]
    for cat, count in sorted_cats:
        row += 1
        ws[f'A{row}'] = cat
        ws[f'B{row}'] = count
        ws[f'A{row}'].border = border; ws[f'B{row}'].border = border

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 22

    # ==== HOJA 2: TOP TÉCNICOS ====
    ws2 = wb.create_sheet('Top Técnicos')
    ws2['A1'] = '🏆 RANKING DE TÉCNICOS'
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:H1')
    ws2['A2'] = f'Empresa: {company_name}  ·  Tickets analizados: {total}'
    ws2['A2'].font = Font(italic=True, color='6B7280')
    ws2.merge_cells('A2:H2')

    tech_headers = ['#', 'Técnico', 'Usuario', 'Asignados', 'Resueltos', 'Tasa Resolución',
                    'Tiempo Prom. (h)', 'Críticos', 'Altas']
    for col, h in enumerate(tech_headers, start=1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = header_align

    if not tech_stats:
        ws2.cell(row=5, column=1, value='Sin tickets asignados a técnicos en el período')
        ws2.cell(row=5, column=1).font = Font(italic=True, color='6B7280')
        ws2.merge_cells('A5:I5')
    else:
        for i, tech in enumerate(tech_stats[:15], start=5):
            ws2.cell(row=i, column=1, value=i - 4).border = border
            ws2.cell(row=i, column=2, value=tech['name']).border = border
            ws2.cell(row=i, column=3, value=tech['username']).border = border
            ws2.cell(row=i, column=4, value=tech['assigned']).border = border
            ws2.cell(row=i, column=5, value=tech['resolved']).border = border
            ws2.cell(row=i, column=6, value=f"{tech['rate']}%").border = border
            ws2.cell(row=i, column=7, value=tech['avg_resolution_h'] or '—').border = border
            ws2.cell(row=i, column=8, value=tech['critical']).border = border
            ws2.cell(row=i, column=9, value=tech['high']).border = border
            # Color por tasa
            if tech['rate'] >= 80:
                ws2.cell(row=i, column=6).fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
            elif tech['rate'] >= 50:
                ws2.cell(row=i, column=6).fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            else:
                ws2.cell(row=i, column=6).fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for col_letter, w in zip(['A','B','C','D','E','F','G','H','I'], [6, 32, 22, 14, 14, 18, 18, 12, 12]):
        ws2.column_dimensions[col_letter].width = w
    ws2.freeze_panes = 'A5'

    # ==== HOJA 3: TICKETS (12 columnas oficiales) ====
    ws3 = wb.create_sheet('Tickets')

    # Header
    for col_idx, h in enumerate(EXPORT_HEADERS, start=1):
        c = ws3.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    # Color por estado del SLA
    sla_colors = {
        'Vencido': 'FECACA', 'En plazo': 'D1FAE5',
        'Cumplido': 'BFDBFE', 'N/A': 'F3F4F6'
    }
    for row_num, ticket in enumerate(tickets, start=2):
        row_data = _build_export_row(ticket)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_num, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=False)
        sla_status = row_data[9]
        if sla_status in sla_colors:
            ws3.cell(row=row_num, column=10).fill = PatternFill(
                start_color=sla_colors[sla_status], end_color=sla_colors[sla_status], fill_type='solid'
            )

    widths = [18, 45, 16, 22, 22, 22, 14, 18, 18, 12, 22, 22]
    for col_idx, w in enumerate(widths, start=1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
    ws3.freeze_panes = 'A2'

    # ==== HOJA 4: GRÁFICAS Y ANÁLISIS ====
    try:
        from openpyxl.drawing.image import Image as XLImage
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from io import BytesIO as _BIO

        ws4 = wb.create_sheet('Gráficas y Análisis')
        ws4['A1'] = '📊 ANÁLISIS VISUAL — Resultados del Período'
        ws4['A1'].font = title_font
        ws4.merge_cells('A1:H1')
        ws4['A2'] = f'Empresa: {company_name}  ·  Tickets analizados: {total}  ·  Generado: {now.strftime("%d/%m/%Y %H:%M")}'
        ws4['A2'].font = Font(italic=True, color='6B7280', size=10)
        ws4.merge_cells('A2:H2')

        # Colores corporativos
        c_primary = '#2563eb'
        c_success = '#16a34a'
        c_warn = '#f59e0b'
        c_danger = '#dc2626'
        c_purple = '#7c3aed'

        # Helper: genera la imagen y la inserta + interpretación + recomendaciones
        def _add_chart_section(title, png_buf, interpretation, recommendations, anchor_row, anchor_col='A'):
            # Título de sección
            ws4.cell(row=anchor_row, column=1, value=title).font = section_font
            ws4.merge_cells(start_row=anchor_row, start_column=1, end_row=anchor_row, end_column=8)

            # Insertar imagen
            try:
                img = XLImage(png_buf)
                img.width = 640
                img.height = 320
                img.anchor = f'{anchor_col}{anchor_row + 1}'
                ws4.add_image(img)
            except Exception as e:
                ws4.cell(row=anchor_row + 1, column=1, value=f'[Error al insertar gráfica: {e}]')

            current = anchor_row + 18  # 16 filas para la gráfica + 2 de aire

            # ─── Bloque INTERPRETACIÓN ───
            interp_title = ws4.cell(row=current, column=1, value='💡 INTERPRETACIÓN')
            interp_title.font = Font(bold=True, size=11, color='1E40AF')
            interp_title.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
            ws4.merge_cells(start_row=current, start_column=1, end_row=current, end_column=8)
            current += 1
            for line in interpretation.split('\n'):
                if line.strip():
                    c = ws4.cell(row=current, column=1, value=line.strip())
                    c.font = Font(size=10, color='374151')
                    c.alignment = Alignment(wrap_text=True, vertical='top')
                    ws4.merge_cells(start_row=current, start_column=1, end_row=current, end_column=8)
                    current += 1

            current += 1  # aire entre bloques

            # ─── Bloque RECOMENDACIONES ───
            reco_title = ws4.cell(row=current, column=1, value='🎯 RECOMENDACIONES')
            reco_title.font = Font(bold=True, size=11, color='065F46')
            reco_title.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
            ws4.merge_cells(start_row=current, start_column=1, end_row=current, end_column=8)
            current += 1
            for line in recommendations.split('\n'):
                if line.strip():
                    c = ws4.cell(row=current, column=1, value=line.strip())
                    c.font = Font(size=10, color='065F46')
                    c.alignment = Alignment(wrap_text=True, vertical='top')
                    c.fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
                    ws4.merge_cells(start_row=current, start_column=1, end_row=current, end_column=8)
                    current += 1

            return current + 3  # próxima sección

        # ───── 1. Distribución por Estado (donut) ─────
        fig, ax = plt.subplots(figsize=(8, 4.5))
        status_data = [(status_labels[k], v) for k, v in by_status.items() if v > 0]
        if status_data:
            labels_s = [s[0] for s in status_data]
            values_s = [s[1] for s in status_data]
            colors_s = ['#3b82f6', '#f59e0b', '#16a34a', '#6b7280']
            ax.pie(values_s, labels=labels_s, colors=colors_s[:len(values_s)],
                   autopct='%1.1f%%', startangle=90, wedgeprops={'width': 0.4, 'edgecolor': 'white'})
            ax.set_title('Distribución por Estado', fontweight='bold', fontsize=14)
        else:
            ax.text(0.5, 0.5, 'Sin datos', ha='center', va='center')
            ax.axis('off')
        buf1 = _BIO(); fig.savefig(buf1, format='png', dpi=110, bbox_inches='tight'); plt.close(fig); buf1.seek(0)

        pct_open = round(by_status.get('open', 0) / total * 100, 1) if total else 0
        pct_progress = round(by_status.get('in_progress', 0) / total * 100, 1) if total else 0
        pct_resolved = round((by_status.get('resolved', 0) + by_status.get('closed', 0)) / total * 100, 1) if total else 0
        interp1 = (
            f'• De los {total} tickets analizados, el {pct_resolved}% ya está resuelto/cerrado.\n'
            f'• El {pct_progress}% sigue en progreso y el {pct_open}% está abierto sin atender.\n'
            f'• Tickets activos (abiertos + en progreso): {by_status.get("open", 0) + by_status.get("in_progress", 0)}.'
        )
        recos1 = []
        if pct_open > 25:
            recos1.append('✓ URGENTE: Reasignar tickets abiertos a técnicos con menor carga (>25% sin atender es alto).')
            recos1.append('✓ Revisar el bot de auto-asignación: puede estar fallando o sin perfiles configurados.')
        if pct_progress > 40:
            recos1.append('✓ Investigar tickets que llevan mucho tiempo en progreso — posibles bloqueos.')
            recos1.append('✓ Implementar reuniones diarias de 15 min con el equipo TI para destrabar casos.')
        if pct_resolved < 60:
            recos1.append('✓ Tasa de resolución baja: revisar tiempos por categoría y reforzar el equipo de mayor demanda.')
        if pct_resolved >= 80:
            recos1.append('✓ EXCELENTE rendimiento. Mantener procesos actuales y documentar buenas prácticas.')
        if by_status.get('open', 0) > 0 and by_priority.get('critical', 0) > 0:
            recos1.append('✓ Verificar si hay tickets críticos abiertos sin asignar — escalarlos manualmente.')
        if not recos1:
            recos1.append('✓ Distribución de estados sana. Continuar monitoreando.')
        reco1_str = '\n'.join(recos1)

        next_row = _add_chart_section('1. Distribución por Estado', buf1, interp1, reco1_str, 4)

        # ───── 2. Distribución por Prioridad (barras) ─────
        fig, ax = plt.subplots(figsize=(8, 4.5))
        prio_order = ['critical', 'high', 'medium', 'low']
        prio_lbls = ['🔴 Crítica', '🟠 Alta', '🟡 Media', '🟢 Baja']
        prio_vals = [by_priority.get(p, 0) for p in prio_order]
        prio_clrs = [c_danger, '#ea580c', c_warn, c_success]
        bars = ax.bar(prio_lbls, prio_vals, color=prio_clrs)
        for b, v in zip(bars, prio_vals):
            if v > 0:
                ax.text(b.get_x() + b.get_width()/2, b.get_height() + 0.3, str(v), ha='center', fontweight='bold')
        ax.set_title('Distribución por Prioridad', fontweight='bold', fontsize=14)
        ax.set_ylabel('Cantidad de tickets')
        ax.grid(True, axis='y', alpha=0.3)
        buf2 = _BIO(); fig.savefig(buf2, format='png', dpi=110, bbox_inches='tight'); plt.close(fig); buf2.seek(0)

        crit_count = by_priority.get('critical', 0)
        high_count = by_priority.get('high', 0)
        med_count = by_priority.get('medium', 0)
        low_count = by_priority.get('low', 0)
        pct_crit_high = round((crit_count + high_count) / total * 100, 1) if total else 0
        interp2 = (
            f'• Críticas: {crit_count}  |  Altas: {high_count}  |  Medias: {med_count}  |  Bajas: {low_count}\n'
            f'• El {pct_crit_high}% del volumen es de prioridad Crítica o Alta.\n'
            f'• Distribución ideal en una empresa sana: 5-10% críticas, 15-25% altas, 50-60% medias, 15-25% bajas.'
        )
        recos2 = []
        if pct_crit_high > 30:
            recos2.append('✓ Revisar la guía de priorización con los usuarios — muchos casos marcados como críticos/altos sugiere uso indebido o infraestructura inestable.')
            recos2.append('✓ Capacitar al personal sobre cuándo usar cada nivel: Crítica = caída de producción, Alta = bloqueo de trabajo, Media = molestia, Baja = mejora.')
            recos2.append('✓ Auditar los tickets críticos del mes: ¿realmente fueron emergencias?')
        if crit_count > 0:
            recos2.append(f'✓ Verificar que los {crit_count} tickets críticos fueron resueltos dentro de su SLA (1 hora típico).')
        if low_count == 0 and total >= 10:
            recos2.append('✓ No hay tickets de prioridad Baja: probable que los usuarios suban la prioridad para ser atendidos antes. Revisar SLA de prioridad Baja.')
        if pct_crit_high <= 30 and crit_count > 0:
            recos2.append('✓ Distribución equilibrada. Continuar monitoreando que las críticas reales se atiendan a tiempo.')
        if not recos2:
            recos2.append('✓ Distribución de prioridades sana. Sin acciones requeridas.')
        reco2_str = '\n'.join(recos2)

        next_row = _add_chart_section('2. Distribución por Prioridad', buf2, interp2, reco2_str, next_row)

        # ───── 3. Top Categorías (barras horizontales) ─────
        fig, ax = plt.subplots(figsize=(8, 5))
        top_cats = sorted(by_category.items(), key=lambda x: -x[1])[:8]
        if top_cats:
            cats = [c[0][:22] for c in top_cats]
            vals = [c[1] for c in top_cats]
            bars = ax.barh(cats[::-1], vals[::-1], color=c_primary)
            for b, v in zip(bars, vals[::-1]):
                ax.text(b.get_width() + 0.2, b.get_y() + b.get_height()/2, str(v), va='center', fontweight='bold')
        ax.set_title('Top 8 Categorías Más Reportadas', fontweight='bold', fontsize=14)
        ax.set_xlabel('Cantidad')
        ax.grid(True, axis='x', alpha=0.3)
        buf3 = _BIO(); fig.savefig(buf3, format='png', dpi=110, bbox_inches='tight'); plt.close(fig); buf3.seek(0)

        top1 = top_cats[0] if top_cats else ('Ninguna', 0)
        top1_pct = round(top1[1] / total * 100, 1) if total else 0
        top1_cat_lower = top1[0].lower()
        interp3 = (
            f'• Categoría más reportada: {top1[0]} con {top1[1]} tickets ({top1_pct}% del total).\n'
            + (f'• Top 3: {", ".join(f"{c[0]} ({c[1]})" for c in top_cats[:3])}\n' if len(top_cats) >= 3 else '')
            + f'• Las primeras 3 categorías concentran el {round(sum(c[1] for c in top_cats[:3]) / total * 100, 1) if total else 0}% del volumen total.'
        )
        recos3 = []
        if top1_pct > 30:
            recos3.append(f'✓ La categoría "{top1[0]}" concentra >30% del volumen. Es una OPORTUNIDAD de automatización o mejora estructural.')
        # Recomendaciones específicas según las categorías top
        for cat_name, cat_count in top_cats[:5]:
            cl = cat_name.lower()
            cat_pct = round(cat_count / total * 100, 1) if total else 0
            if 'acceso' in cl or 'contrase' in cl or 'password' in cl:
                recos3.append(f'✓ "{cat_name}" ({cat_pct}%): Implementar self-service de reseteo de contraseñas. Ahorra 60-80% de estos tickets.')
            elif 'hardware' in cl:
                recos3.append(f'✓ "{cat_name}" ({cat_pct}%): Revisar plan de renovación de equipos y contratos de soporte.')
            elif 'sap' in cl:
                recos3.append(f'✓ "{cat_name}" ({cat_pct}%): Capacitar a key users para reducir consultas básicas; involucrar al equipo BASIS para temas técnicos.')
            elif 'red' in cl or 'wifi' in cl or 'conectiv' in cl:
                recos3.append(f'✓ "{cat_name}" ({cat_pct}%): Auditar capacidad de red, revisar cableado y planes de proveedor de internet.')
            elif 'impresora' in cl:
                recos3.append(f'✓ "{cat_name}" ({cat_pct}%): Considerar Managed Print Services y autoatención de tóner.')
            elif 'email' in cl or 'correo' in cl or 'outlook' in cl:
                recos3.append(f'✓ "{cat_name}" ({cat_pct}%): Crear FAQ con problemas comunes (firma, OOO, archivado).')
            elif 'software' in cl:
                recos3.append(f'✓ "{cat_name}" ({cat_pct}%): Estandarizar catálogo de software y automatizar instalaciones (SCCM/Intune).')
        if len(top_cats) >= 5:
            recos3.append('✓ Crear base de conocimiento (KB) con soluciones documentadas para las 5 categorías top — Eli (bot) las usará para resolver al instante.')
        if not recos3:
            recos3.append('✓ Sin categorías dominantes. Mantener el monitoreo regular.')
        reco3_str = '\n'.join(recos3)

        next_row = _add_chart_section('3. Top Categorías', buf3, interp3, reco3_str, next_row)

        # ───── 4. Top Técnicos por Productividad ─────
        fig, ax = plt.subplots(figsize=(8, 4.5))
        top_tech = tech_stats[:10]
        if top_tech:
            names = [t['name'][:22] for t in top_tech]
            asignados = [t['assigned'] for t in top_tech]
            resueltos = [t['resolved'] for t in top_tech]
            y_pos = list(range(len(names)))
            ax.barh(y_pos, asignados, color='#94a3b8', label='Asignados')
            ax.barh(y_pos, resueltos, color=c_success, label='Resueltos')
            ax.set_yticks(y_pos)
            ax.set_yticklabels(names)
            ax.invert_yaxis()
            ax.legend(loc='lower right')
        else:
            ax.text(0.5, 0.5, 'Sin técnicos con tickets', ha='center', va='center')
            ax.axis('off')
        ax.set_title('Ranking de Técnicos por Productividad', fontweight='bold', fontsize=14)
        ax.set_xlabel('Tickets')
        ax.grid(True, axis='x', alpha=0.3)
        buf4 = _BIO(); fig.savefig(buf4, format='png', dpi=110, bbox_inches='tight'); plt.close(fig); buf4.seek(0)

        if tech_stats:
            best = tech_stats[0]
            total_res = sum(t['resolved'] for t in tech_stats)
            best_pct = round(best['resolved'] / total_res * 100, 1) if total_res else 0
            avg_per_tech = round(total_res / len(tech_stats), 1) if tech_stats else 0
            low_performers = [t for t in tech_stats if t['rate'] < 50 and t['assigned'] >= 3]
            interp4 = (
                f'• Mejor técnico del período: {best["name"]} con {best["resolved"]} tickets resueltos ({best["rate"]}% tasa).\n'
                f'• {best["name"]} concentra el {best_pct}% del total resuelto del equipo.\n'
                f'• Total técnicos activos: {len(tech_stats)}. Promedio: {avg_per_tech} tickets resueltos por técnico.\n'
                + (f'• {len(low_performers)} técnico(s) con tasa de resolución <50%.' if low_performers else '• Todo el equipo con tasa de resolución sobre 50%.')
            )
            recos4 = []
            if best_pct > 40:
                recos4.append(f'✓ {best["name"]} carga el {best_pct}% del trabajo. Revisar si está sobrecargado y rebalancear con otros técnicos.')
                recos4.append('✓ Configurar perfiles de habilidades en otros técnicos para que la auto-asignación distribuya mejor.')
            if low_performers:
                low_names = ', '.join(t['name'] for t in low_performers[:3])
                recos4.append(f'✓ Técnicos con baja resolución: {low_names}. Investigar bloqueos o reasignar a otra categoría.')
            if best['avg_resolution_h'] and best['avg_resolution_h'] > 24:
                recos4.append(f'✓ Tiempo promedio de resolución alto ({best["avg_resolution_h"]}h). Revisar si los SLA son realistas o hay procesos pesados.')
            if len(tech_stats) < 3 and total > 10:
                recos4.append('✓ Pocos técnicos para el volumen actual. Evaluar contratación o redistribución de cargas.')
            top3 = tech_stats[:3]
            if all(t['rate'] >= 80 for t in top3) and len(top3) >= 3:
                recos4.append('✓ Top 3 técnicos con resolución >80% — destacar logros, considerar incentivos o premios.')
            if total_res > 0 and avg_per_tech >= 5:
                recos4.append('✓ Equipo productivo. Considerar capacitaciones avanzadas en categorías de mayor volumen.')
            if not recos4:
                recos4.append('✓ Equipo equilibrado. Mantener el monitoreo.')
            reco4_str = '\n'.join(recos4)
        else:
            interp4 = '• No hay técnicos con tickets asignados en el período.'
            reco4_str = ('✓ Verificar configuración del Agent Assignor — los tickets deberían asignarse automáticamente.\n'
                         '✓ Revisar perfiles de técnicos en la sección "Equipo".')

        next_row = _add_chart_section('4. Ranking de Técnicos', buf4, interp4, reco4_str, next_row)

        # ───── 5. Cumplimiento SLA (donut) ─────
        fig, ax = plt.subplots(figsize=(6, 4.5))
        if sla_total > 0:
            sizes = [sla_cumplidos, sla_vencidos]
            labels_sla = ['Cumplidos', 'Vencidos']
            ax.pie(sizes, labels=labels_sla, colors=[c_success, c_danger],
                   autopct='%1.1f%%', startangle=90, wedgeprops={'width': 0.4, 'edgecolor': 'white'})
            ax.text(0, 0, f'{sla_pct}%', ha='center', va='center', fontsize=22, fontweight='bold',
                    color=c_success if sla_pct >= 85 else c_danger)
        else:
            ax.text(0.5, 0.5, 'Sin tickets con SLA cerrado', ha='center', va='center')
            ax.axis('off')
        ax.set_title('Cumplimiento de SLA', fontweight='bold', fontsize=14)
        buf5 = _BIO(); fig.savefig(buf5, format='png', dpi=110, bbox_inches='tight'); plt.close(fig); buf5.seek(0)

        sla_estado = 'EXCELENTE' if sla_pct >= 95 else ('BUENO' if sla_pct >= 85 else ('REGULAR' if sla_pct >= 70 else 'CRÍTICO'))
        interp5 = (
            f'• Cumplimiento SLA: {sla_pct}% ({sla_cumplidos} cumplidos vs {sla_vencidos} vencidos).\n'
            f'• Estado: {sla_estado}.\n'
            f'• Meta esperada por la industria: ≥ 95% (excelente), 85-95% (bueno), 70-85% (regular), < 70% (crítico).\n'
            f'• Tiempo promedio de resolución: {avg_resolution_h} horas.'
        )
        recos5 = []
        if sla_pct < 70:
            recos5.append('✓ CRÍTICO: SLA por debajo del 70%. Acción inmediata requerida.')
            recos5.append('✓ Auditar los tickets vencidos: ¿son del mismo técnico? ¿misma categoría? ¿prioridad mal asignada?')
            recos5.append('✓ Aumentar capacidad del equipo o revisar los plazos configurados — podrían ser irrealistas.')
            recos5.append('✓ Implementar alertas automáticas al supervisor cuando un ticket pase del 80% del SLA sin movimiento.')
        elif sla_pct < 85:
            recos5.append('✓ Cumplimiento REGULAR. Identificar las categorías con peor SLA y reforzar.')
            recos5.append('✓ Configurar el Escalator del Orchestrator para que avise antes de vencer.')
            recos5.append('✓ Capacitar al equipo en gestión de tiempo y priorización.')
        elif sla_pct < 95:
            recos5.append('✓ Buen cumplimiento. Para llegar al 95% identificar los pocos casos vencidos y atacar la causa raíz.')
            recos5.append('✓ Revisar si los vencidos son por causas externas (esperando proveedor, usuario, etc.).')
        else:
            recos5.append('✓ EXCELENTE rendimiento de SLA. Mantener procesos actuales.')
            recos5.append('✓ Considerar reducir SLAs si están sobre-cumplidos para subir el estándar.')
            recos5.append('✓ Documentar y compartir buenas prácticas con el resto del equipo.')

        if avg_resolution_h > 48:
            recos5.append(f'✓ Tiempo de resolución promedio alto ({avg_resolution_h}h). Revisar procesos lentos y tickets bloqueados.')
        if sla_vencidos > sla_cumplidos and total > 5:
            recos5.append('✓ Más tickets vencidos que cumplidos. Reasignar carga y revisar la asignación automática.')
        reco5_str = '\n'.join(recos5)

        _add_chart_section('5. Cumplimiento de SLA', buf5, interp5, reco5_str, next_row)

        # Ajustar ancho columnas de la hoja 4
        for col_letter in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'):
            ws4.column_dimensions[col_letter].width = 16

    except Exception as e:
        print(f'[export_excel] Error generando hoja de gráficas: {e}')
        # Crear hoja simple con mensaje de error pero no romper todo el reporte
        try:
            ws4 = wb.create_sheet('Gráficas y Análisis')
            ws4['A1'] = 'Hoja de gráficas no disponible'
            ws4['A2'] = f'Error: {str(e)[:200]}'
            ws4['A3'] = 'Verificar que matplotlib esté instalado: pip install matplotlib'
        except Exception:
            pass

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    log_audit('export_tickets_xlsx', session['user_id'], 'report', None,
              f'Exportados {len(tickets)} tickets a Excel (4 hojas con gráficas)')

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'Reporte_Tickets_{company}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')


@app.route('/api/export/csv', methods=['GET'])
def api_export_csv():
    """Exporta tickets a CSV con las 12 columnas oficiales (UTF-8 con BOM para Excel)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']
    tickets = Ticket.query.filter(
        Ticket.company == company,
        ~Ticket.ticket_number.like('DM-%'),
        ~Ticket.ticket_number.like('CHAT-%'),
    ).order_by(Ticket.id.desc()).all()

    output = StringIO()
    # BOM para que Excel detecte UTF-8 con tildes
    output.write('﻿')
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_HEADERS)
    for ticket in tickets:
        writer.writerow(_build_export_row(ticket))

    output.seek(0)
    log_audit('export_tickets_csv', session['user_id'], 'report', None,
              f'Exportados {len(tickets)} tickets a CSV')

    # Necesitamos enviar bytes (UTF-8) — convertir StringIO a BytesIO
    out_bytes = BytesIO(output.getvalue().encode('utf-8'))

    return send_file(out_bytes,
                     mimetype='text/csv; charset=utf-8',
                     as_attachment=True,
                     download_name=f'Reporte_Tickets_{company}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv')

@app.route('/api/config/sla', methods=['POST'])
def api_config_sla():
    """Configura SLA por prioridad"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    data = request.get_json()

    for priority, minutes in data.items():
        config = Config.query.filter_by(key=f'sla_{priority}').first()
        if config:
            config.value = str(minutes)
        else:
            config = Config(key=f'sla_{priority}', value=str(minutes))
            db.session.add(config)

    db.session.commit()
    log_audit('update_config', session['user_id'], 'config', None, 'Configuración de SLA actualizada')

    return jsonify({'success': True})


@app.route('/api/config/business-hours', methods=['GET', 'POST'])
def api_config_business_hours():
    """Horario laboral por empresa: días, horas y feriados que NO cuentan para el SLA."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    company = session['company']

    if request.method == 'GET':
        def _get(key, default=''):
            c = Config.query.filter_by(key=f'biz_{key}_{company}').first()
            return c.value if c else default

        # Construir schedule por día (con fallback a formato viejo)
        schedule = {}
        old_days_csv = _get('days', '1,2,3,4,5')
        try:
            old_days_set = {int(d.strip()) for d in old_days_csv.split(',') if d.strip()}
        except Exception:
            old_days_set = {1,2,3,4,5}
        for iso in range(1, 8):
            day_enabled_cfg = _get(f'day{iso}_enabled', None)
            if day_enabled_cfg is None:
                # Backward compat
                enabled = iso in old_days_set
                schedule[iso] = {
                    'enabled': enabled,
                    'start': _get('start', '08:00') if enabled else '08:00',
                    'end': _get('end', '18:00') if enabled else '18:00'
                }
            else:
                schedule[iso] = {
                    'enabled': day_enabled_cfg == '1',
                    'start': _get(f'day{iso}_start', '08:00'),
                    'end': _get(f'day{iso}_end', '18:00')
                }
        return jsonify({
            'success': True,
            'company': company,
            'enabled': _get('enabled', '0') == '1',
            'schedule': schedule,
            'holidays': _get('holidays', '')
        })

    # POST: guardar
    try:
        import re
        data = request.get_json() or {}
        def _set(key, val):
            full_key = f'biz_{key}_{company}'
            c = Config.query.filter_by(key=full_key).first()
            if c:
                c.value = str(val)
            else:
                db.session.add(Config(key=full_key, value=str(val)))

        _set('enabled', '1' if data.get('enabled') else '0')

        # Schedule por día
        schedule = data.get('schedule') or {}
        time_re = re.compile(r'^\d{1,2}:\d{2}$')
        enabled_count = 0
        for iso in range(1, 8):
            day = schedule.get(str(iso)) or schedule.get(iso) or {}
            is_en = bool(day.get('enabled'))
            start = (day.get('start') or '08:00').strip()
            end = (day.get('end') or '18:00').strip()
            if is_en:
                if not time_re.match(start) or not time_re.match(end):
                    return jsonify({'success': False, 'error': f'Hora inválida en día {iso}. Usa HH:MM'}), 400
            _set(f'day{iso}_enabled', '1' if is_en else '0')
            _set(f'day{iso}_start', start)
            _set(f'day{iso}_end', end)
            if is_en:
                enabled_count += 1

        # Feriados
        holidays = data.get('holidays') or ''
        if isinstance(holidays, list):
            holidays = ','.join(holidays)
        valid_holidays = []
        for h in holidays.split(','):
            h = h.strip()
            if not h:
                continue
            try:
                datetime.strptime(h, '%Y-%m-%d')
                valid_holidays.append(h)
            except Exception:
                return jsonify({'success': False, 'error': f'Fecha inválida: {h}'}), 400
        _set('holidays', ','.join(valid_holidays))

        db.session.commit()
        cache_delete(f'biz_cfg:{company}')  # invalidar cache
        log_audit('update_business_hours', session['user_id'], 'config', None,
                  f'Horario laboral por día: {enabled_count} días laborales, feriados={len(valid_holidays)}')
        return jsonify({'success': True, 'message': 'Configuración guardada (horario por día)'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/theme', methods=['POST'])
def api_config_theme():
    """Cambia tema visual"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    data = request.get_json()
    theme = data.get('theme')

    if theme not in THEMES:
        return jsonify({'success': False, 'message': 'Tema no válido'}), 400

    # Guardar tema POR EMPRESA del admin que lo cambia
    company = session.get('company', 'default')
    key = f"theme_{company}"

    config = Config.query.filter_by(key=key).first()
    if config:
        config.value = theme
    else:
        config = Config(key=key, value=theme)
        db.session.add(config)

    db.session.commit()
    log_audit('change_theme', session['user_id'], 'config', None, f'Tema cambiado a "{theme}" para empresa {company}')
    cache_delete(f'theme:{company}')  # invalidar cache para que tome efecto inmediato
    return jsonify({'success': True, 'theme': theme, 'company': company})

def _get_db_server_version():
    """Devuelve la version del server de BD activo (util para mostrar en admin)."""
    try:
        from sqlalchemy import text
        dialect = db.engine.dialect.name
        if dialect == 'postgresql':
            row = db.session.execute(text('SELECT version()')).fetchone()
            return row[0][:80] if row else ''
        elif dialect == 'sqlite':
            row = db.session.execute(text('SELECT sqlite_version()')).fetchone()
            return f'SQLite {row[0]}' if row else ''
    except Exception:
        pass
    return ''


@app.route('/api/config/database', methods=['GET', 'POST'])
def api_config_database():
    """GET: retorna config actual de BD. POST: guarda nueva configuración.
    Restringido a admin master (Manufacturas Eliot) — Pash y Primatela
    NO pueden ver/cambiar la BD porque es una config global del sistema."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    if not is_master_admin():
        return jsonify({'success': False, 'error': 'Solo el admin master (Manufacturas Eliot) puede gestionar la BD del sistema'}), 403

    if request.method == 'GET':
        # Retornar configuración REAL basada en el engine activo, no el env var.
        # Esto asegura que si Coolify inyecta DATABASE_URL, la UI refleja el motor real.
        engine_url = db.engine.url
        dialect = db.engine.dialect.name  # 'postgresql', 'sqlite', 'mysql', etc.
        managed_by_env = bool(os.getenv('DATABASE_URL'))

        if dialect == 'postgresql':
            return jsonify({
                'success': True,
                'db_type': 'postgresql',
                'host': engine_url.host or '',
                'port': engine_url.port or 5432,
                'name': engine_url.database or '',
                'user': engine_url.username or '',
                'managed_by_env': managed_by_env,
                'server_version': _get_db_server_version(),
            })
        elif dialect == 'sqlite':
            return jsonify({
                'success': True,
                'db_type': 'sqlite',
                'db_path': engine_url.database or '',
                'managed_by_env': managed_by_env,
            })
        else:
            return jsonify({
                'success': True,
                'db_type': dialect,
                'db_path': str(engine_url),
                'managed_by_env': managed_by_env,
            })

    elif request.method == 'POST':
        # Guardar nueva configuración
        data = request.get_json()
        db_type = data.get('db_type', 'sqlite')

        if db_type == 'sqlite':
            db_url = f'sqlite:///{data.get("db_path", db_path)}'
        elif db_type == 'postgresql':
            host = data.get('host', 'localhost')
            port = data.get('port', 5432)
            name = data.get('name')
            user = data.get('user')
            password = data.get('password')
            db_url = f'postgresql://{user}:{password}@{host}:{port}/{name}'

        # Guardar en archivo .env (de manera simple, reemplazar línea)
        env_file = os.path.join(os.path.dirname(__file__), '.env')
        try:
            with open(env_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # Reemplazar DATABASE_URL existente o agregar nueva
            lines = content.split('\n')
            found = False
            for i, line in enumerate(lines):
                if line.startswith('DATABASE_URL='):
                    lines[i] = f'DATABASE_URL={db_url}'
                    found = True
                    break

            if not found:
                lines.append(f'DATABASE_URL={db_url}')

            with open(env_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))

            log_audit('update_config', session['user_id'], 'config', None,
                     f'Base de datos cambiada a {db_type}')

            return jsonify({'success': True, 'message': 'Configuración guardada. Por favor reinicia el servidor.'})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/config/database/test', methods=['GET'])
def api_config_database_test():
    """Prueba la conexión a la BD configurada. Solo master admin."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    if not is_master_admin():
        return jsonify({'success': False, 'error': 'Solo el admin master puede testear la BD'}), 403

    import time
    start = time.time()
    try:
        from sqlalchemy import create_engine
        db_url = os.getenv('DATABASE_URL', '')
        if db_url.startswith('postgres://'):
            db_url = db_url.replace('postgres://', 'postgresql://', 1)

        engine = create_engine(db_url)
        with engine.connect() as conn:
            conn.execute('SELECT 1')

        latency = int((time.time() - start) * 1000)
        return jsonify({
            'success': True,
            'message': 'Conexión exitosa',
            'latency_ms': latency
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/admin/orchestrator/status', methods=['GET'])
def api_orchestrator_status():
    """Estado del orchestrator + acciones por período.
    Acepta ?period=day|week|month|year|all (default: all).
    Devuelve contadores filtrados por período, recent_actions, trend para gráfica.
    """
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    user = User.query.get(session['user_id'])
    company = user.company
    scope = admin_companies_scope() or [company]

    # Filtro por empresa: master admin puede pedir una empresa especifica
    # via ?company=<code>. Non-master queda restringido a su empresa (ignora el param).
    requested_company = (request.args.get('company') or '').strip().lower()
    if is_master_admin() and requested_company:
        if requested_company in ('all', ''):
            pass  # scope se queda con todas
        elif requested_company in scope:
            scope = [requested_company]
        # Si piden una que no esta en scope, se ignora silenciosamente

    period = (request.args.get('period') or 'all').lower()
    now = datetime.now()

    # Definir rango temporal
    if period == 'day':
        period_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Hoy'
        bucket_format = '%H:00'
    elif period == 'week':
        period_start = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Últimos 7 días'
        bucket_format = '%d/%m'
    elif period == 'month':
        period_start = (now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Últimos 30 días'
        bucket_format = '%d/%m'
    elif period == 'year':
        period_start = (now - timedelta(days=365)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Últimos 12 meses'
        bucket_format = '%Y-%m'
    else:
        period_start = None
        period_label = 'Histórico completo'
        bucket_format = '%Y-%m'

    data = {}
    orch = app.config.get('orchestrator')

    # 1. Intentar el método del orchestrator (no acepta período aún, así que solo lo usamos para status_by_agent)
    try:
        if orch is not None:
            base_data = orch.get_dashboard_data(company)
            # Tomar status_by_agent y escalator_running (no dependen del período)
            data['status_by_agent'] = base_data.get('status_by_agent', {})
            data['escalator_running'] = base_data.get('escalator_running', False)
            data['use_llm'] = base_data.get('use_llm', False)
            data['mode'] = base_data.get('mode', 'desconocido')
    except Exception as e:
        print(f'[orchestrator/status] Error en get_dashboard_data: {type(e).__name__}: {e}')

    # 2. Contadores filtrados por período
    try:
        base_q = AgentAction.query.filter(AgentAction.company.in_(scope))
        if period_start:
            base_q = base_q.filter(AgentAction.created_at >= period_start)

        actions = base_q.order_by(AgentAction.created_at.desc()).all()

        # Contadores por agente
        tc = ta = tr = te = 0
        success_count = fail_count = 0
        llm_count = rules_count = 0
        confidences = []

        for a in actions:
            if a.agent_name == 'classifier': tc += 1
            elif a.agent_name == 'assignor': ta += 1
            elif a.agent_name == 'responder': tr += 1
            elif a.agent_name == 'escalator': te += 1
            if a.success: success_count += 1
            else: fail_count += 1
            if a.used_llm: llm_count += 1
            else: rules_count += 1
            if a.confidence: confidences.append(a.confidence)

        # Recent actions (las 20 más recientes del período)
        recent = actions[:20]

        # Tendencia: agrupar por bucket
        trend_buckets = {}
        trend_by_agent = {'classifier': {}, 'assignor': {}, 'responder': {}, 'escalator': {}}
        for a in actions:
            if not a.created_at:
                continue
            key = a.created_at.strftime(bucket_format)
            trend_buckets[key] = trend_buckets.get(key, 0) + 1
            if a.agent_name in trend_by_agent:
                trend_by_agent[a.agent_name][key] = trend_by_agent[a.agent_name].get(key, 0) + 1

        trend = [{'label': k, 'count': v} for k, v in sorted(trend_buckets.items())]
        all_buckets = sorted(trend_buckets.keys())
        trend_stacked = {
            'labels': all_buckets,
            'classifier': [trend_by_agent['classifier'].get(b, 0) for b in all_buckets],
            'assignor':   [trend_by_agent['assignor'].get(b, 0) for b in all_buckets],
            'responder':  [trend_by_agent['responder'].get(b, 0) for b in all_buckets],
            'escalator':  [trend_by_agent['escalator'].get(b, 0) for b in all_buckets],
        }

        pending_review = Ticket.query.filter(
            Ticket.company.in_(scope),
            Ticket.status == 'open',
            Ticket.assignee_id.isnot(None)
        ).count()

        avg_conf = int(sum(confidences) / len(confidences)) if confidences else 0
        success_rate = round((success_count / len(actions)) * 100, 1) if actions else 0
        llm_pct = round((llm_count / len(actions)) * 100, 1) if actions else 0

        data.update({
            'period': period,
            'period_label': period_label,
            'period_start': period_start.isoformat() if period_start else None,
            'scope': scope,
            'company_filter': requested_company or 'all',
            'is_master': is_master_admin(),
            'total_classified': tc,
            'total_assigned': ta,
            'total_responded': tr,
            'total_escalated': te,
            'total_actions': len(actions),
            'pending_review': pending_review,
            'success_count': success_count,
            'fail_count': fail_count,
            'success_rate': success_rate,
            'llm_count': llm_count,
            'rules_count': rules_count,
            'llm_pct': llm_pct,
            'avg_confidence': avg_conf,
            'recent_actions': [{
                'id': a.id,
                'ticket_id': a.ticket_id,
                'agent_name': a.agent_name,
                'action_type': a.action_type,
                'agent': a.agent_name,
                'action': a.action_type,
                'confidence': a.confidence,
                'used_llm': a.used_llm,
                'success': a.success,
                'duration_ms': a.duration_ms,
                'details': f"Ticket #{a.ticket_id} · {('OK' if a.success else 'FAIL')}{(' · ' + str(a.confidence) + '%') if a.confidence else ''}",
                'created_at': a.created_at.strftime('%Y-%m-%d %H:%M:%S') if a.created_at else '',
            } for a in recent],
            'trend': trend,
            'trend_stacked': trend_stacked,
            'orchestrator_available': orch is not None,
        })
    except Exception as e:
        print(f'[orchestrator/status] Error: {e}')

    return jsonify({'success': True, 'data': data})

@app.route('/api/admin/technician-profile/<int:user_id>', methods=['GET', 'POST'])
def api_technician_profile(user_id):
    """CRUD de perfil de habilidades de técnico"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    user = User.query.get(session['user_id'])
    tech = User.query.filter_by(id=user_id, company=user.company, role='technician').first()
    if not tech:
        return jsonify({'success': False, 'error': 'Técnico no encontrado'}), 404

    if request.method == 'GET':
        profile = TechnicianProfile.query.filter_by(user_id=user_id).first()
        if not profile:
            profile = TechnicianProfile(user_id=user_id, company=user.company)
            db.session.add(profile)
            db.session.commit()

        return jsonify({'success': True, 'profile': {
            'user_id': profile.user_id,
            'skills': profile.get_skills_list(),
            'skill_levels': json.loads(profile.skill_levels or '{}'),
            'max_tickets': profile.max_tickets,
            'is_available': profile.is_available,
            'avg_resolution_minutes': profile.avg_resolution_minutes,
            'tickets_resolved_total': profile.tickets_resolved_total,
        }})

    elif request.method == 'POST':
        data = request.get_json()
        profile = TechnicianProfile.query.filter_by(user_id=user_id).first()
        if not profile:
            profile = TechnicianProfile(user_id=user_id, company=user.company)

        profile.skills = ','.join(data.get('skills', []))
        profile.skill_levels = json.dumps(data.get('skill_levels', {}))
        profile.max_tickets = int(data.get('max_tickets', 5))
        profile.is_available = bool(data.get('is_available', True))

        db.session.add(profile)
        db.session.commit()

        log_audit('update_technician_profile', session['user_id'], 'technician', user_id,
                  f'Perfil de habilidades de {tech.name} actualizado')

        return jsonify({'success': True, 'profile': {
            'user_id': profile.user_id,
            'skills': profile.get_skills_list(),
            'skill_levels': json.loads(profile.skill_levels or '{}'),
            'max_tickets': profile.max_tickets,
            'is_available': profile.is_available,
        }})

# ═════════════════════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════════════════════
# PING AUTOMÁTICO DE SERVIDORES (RF-03-11)
# ═════════════════════════════════════════════════════════════════════════════

def ping_server(server_id):
    """Hacer ping a servidor y crear ticket si cae (RF-03-11)"""
    with app.app_context():
        server = Server.query.get(server_id)
        if not server:
            return

        try:
            # Intentar conectar al puerto 80
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2)
            result = sock.connect_ex((server.ip_address, 80))
            sock.close()

            was_online = server.is_online
            server.is_online = result == 0
            server.last_ping = datetime.now()
            db.session.commit()

            # Si cambió de ONLINE a OFFLINE, crear ticket de alerta automático
            if was_online and not server.is_online:
                creator = User.query.filter_by(company=server.company, role='admin').first()
                if creator:
                    ticket = Ticket(
                        ticket_number=get_next_ticket_number(server.company),
                        title=f'ALERTA: {server.name} desconectado',
                        description=f'El servidor {server.name} ({server.ip_address}) dejó de responder',
                        priority='critical',
                        creator_id=creator.id,
                        company=server.company,
                        category='Infraestructura',
                        sla_minutes=60,
                        sla_deadline=datetime.now() + timedelta(minutes=60)
                    )
                    db.session.add(ticket)
                    db.session.commit()

                    log_audit('server_down', creator.id, 'server', server_id, f'Servidor {server.name} desconectado')

                    # Enviar a Teams si hay webhook
                    send_teams_webhook(server.company, 'server_down', ticket)

        except (socket.timeout, socket.error, ConnectionRefusedError) as e:
            log_audit('ping_error', None, 'server', server_id, f'Error ping (socket): {str(e)}')
        except Exception as e:
            log_audit('ping_error', None, 'server', server_id, f'Error ping (general): {str(e)}')

def assign_ticket_auto(ticket):
    """
    Asignación automática basada en carga y perfil (RF-03-06).

    Asigna el ticket al técnico con menor carga de trabajo activo,
    con bonus si su especialidad coincide con la categoría.
    """
    if ticket.assignee_id:
        return  # Ya asignado

    available_technicians = User.query.filter_by(
        role='technician',
        company=ticket.company,
        is_active=True
    ).all()

    if not available_technicians:
        return

    technician_scores = {}
    for technician in available_technicians:
        active_ticket_count = len([
            t for t in technician.assigned_tickets
            if t.status in ['open', 'in_progress']
        ])
        base_score = active_ticket_count
        category_bonus = -5 if ticket.category.lower() in technician.name.lower() else 0
        technician_scores[technician.id] = base_score + category_bonus

    best_technician_id = min(technician_scores, key=technician_scores.get)
    ticket.assignee_id = best_technician_id

    log_audit(
        'auto_assign',
        None,
        'ticket',
        ticket.id,
        f'Ticket asignado automáticamente a técnico {best_technician_id}'
    )

    # Registrar en AgentAction para que cuente en el Orchestrator dashboard
    try:
        db.session.add(AgentAction(
            ticket_id=ticket.id,
            company=ticket.company,
            agent_name='assignor',
            action_type='assignor',
            confidence=80,
            used_llm=False,
            success=True,
            output_data=json.dumps({'technician_id': best_technician_id, 'method': 'rules_load+category'})
        ))
    except Exception:
        pass

    # Notificar al técnico asignado por email
    try:
        tech_to_notify = User.query.get(best_technician_id)
        if tech_to_notify:
            notify_ticket_assigned(
                ticket=ticket,
                new_assignee=tech_to_notify,
                assigned_by_name='Asignación automática (IA por carga + categoría)',
                reason=f'Técnico con menor carga para categoría "{ticket.category or "General"}"'
            )
    except Exception as e:
        print(f'[WARN] Notificación email: {e}')


def start_server_monitoring():
    """Iniciar monitoreo de servidores cada 5 minutos"""
    def monitor():
        from time import sleep
        while True:
            try:
                sleep(300)  # Cada 5 minutos
                # CRÍTICO: cualquier query en threads background requiere app_context
                with app.app_context():
                    servers = Server.query.filter_by(is_online=True).all()
                    for server in servers:
                        ping_server(server.id)
            except Exception as e:
                print(f'Error en monitoreo: {e}')

    thread = Thread(target=monitor, daemon=True)
    thread.start()

# ═════════════════════════════════════════════════════════════════════════════
# BACKUP AUTOMÁTICO RNF-02-05 - Diario comprimido, retención 30 copias
# ═════════════════════════════════════════════════════════════════════════════

def _get_db_file_path():
    """Devuelve el Path real al archivo SQLite que está usando la app."""
    uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
    if uri.startswith('sqlite:///'):
        raw = uri.replace('sqlite:///', '', 1)
        # Si es absoluto, usar tal cual; si es relativo, Flask usa instance_path
        p = Path(raw)
        if not p.is_absolute():
            p = Path(app.instance_path) / raw
        return p
    return None


def create_backup(user_id=None):
    """Crear backup comprimido y cifrado de la BD (RNF-02-05).

    Si DB_ENCRYPTION_KEY está disponible: genera .db.gz.enc (gzip + Fernet).
    Si no: cae a .db.gz plano (legacy, solo dev sin clave). Devuelve Path o None.
    """
    db_path = _get_db_file_path()
    if not db_path or not db_path.exists():
        log_audit('backup_failed', user_id, 'backup', None, f'BD no encontrada en {db_path}')
        return None

    backup_dir = Path('backups')
    backup_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    use_encryption = crypto_has_key()
    suffix = '.db.gz.enc' if use_encryption else '.db.gz'
    backup_file = backup_dir / f'ticketdesk_backup_{timestamp}{suffix}'

    try:
        # 1) Comprimir BD a bytes
        with open(db_path, 'rb') as f_in:
            raw = f_in.read()
        gz_buf = BytesIO()
        with gzip.GzipFile(fileobj=gz_buf, mode='wb') as gz:
            gz.write(raw)
        gz_bytes = gz_buf.getvalue()

        # 2) Cifrar (si hay clave) y escribir
        if use_encryption:
            payload = encrypt_bytes(gz_bytes)
        else:
            payload = gz_bytes
        with open(backup_file, 'wb') as f_out:
            f_out.write(payload)

        # 3) Purgar backups antiguos (mantener 30 más recientes, ambos formatos)
        all_backups = sorted(
            list(backup_dir.glob('ticketdesk_backup_*.db.gz')) +
            list(backup_dir.glob('ticketdesk_backup_*.db.gz.enc'))
        )
        removed = 0
        for old_backup in all_backups[:-30]:
            try:
                old_backup.unlink()
                removed += 1
            except Exception:
                pass

        log_audit('backup_created', user_id, 'backup', None,
                  f'Backup {backup_file.name} ({backup_file.stat().st_size} bytes) '
                  f'{"cifrado" if use_encryption else "PLANO"}, {removed} antiguos eliminados')
        return backup_file
    except Exception as e:
        log_audit('backup_failed', user_id, 'backup', None, f'Error backup: {str(e)}')
        return None

def start_backup_scheduler():
    """Iniciar scheduler de backups automáticos cada 24h"""
    def backup_loop():
        while True:
            try:
                with app.app_context():
                    create_backup()
                time.sleep(86400)  # 24 horas
            except Exception as e:
                print(f'[Backup Error] {e}')
                time.sleep(3600)  # Reintentar en 1 hora

    thread = Thread(target=backup_loop, daemon=True)
    thread.start()


# ═════════════════════════════════════════════════════════════════════════════
# ENDPOINTS DE BACKUP (admin)
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/backups', methods=['GET'])
def api_backups_list():
    """Listar backups disponibles en disco."""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    backup_dir = Path('backups')
    if not backup_dir.exists():
        return jsonify({'success': True, 'backups': []})

    # Incluir tanto .db.gz (legacy) como .db.gz.enc (cifrados)
    all_files = list(backup_dir.glob('ticketdesk_backup_*.db.gz')) + \
                list(backup_dir.glob('ticketdesk_backup_*.db.gz.enc'))
    backups = []
    for f in sorted(all_files, key=lambda p: p.stat().st_mtime, reverse=True):
        try:
            stat = f.stat()
            backups.append({
                'name': f.name,
                'encrypted': f.name.endswith('.enc'),
                'size_bytes': stat.st_size,
                'size_human': f"{stat.st_size / 1024:.1f} KB" if stat.st_size < 1024*1024 else f"{stat.st_size / 1024 / 1024:.2f} MB",
                'created_at': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
            })
        except Exception:
            continue

    db_path = _get_db_file_path()
    db_size = db_path.stat().st_size if db_path and db_path.exists() else 0

    return jsonify({
        'success': True,
        'backups': backups,
        'db_path': str(db_path) if db_path else None,
        'db_size_bytes': db_size,
        'db_size_human': f"{db_size / 1024:.1f} KB" if db_size < 1024*1024 else f"{db_size / 1024 / 1024:.2f} MB"
    })


@app.route('/api/admin/backups/create', methods=['POST'])
def api_backups_create():
    """Crear backup manual on-demand."""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    backup_file = create_backup(user_id=session['user_id'])
    if not backup_file:
        return jsonify({'success': False, 'error': 'No se pudo crear el backup. Revisa la consola del servidor.'}), 500
    stat = backup_file.stat()
    return jsonify({
        'success': True,
        'message': f'Backup creado: {backup_file.name}',
        'name': backup_file.name,
        'size_bytes': stat.st_size
    })


@app.route('/api/admin/backups/<path:filename>/download', methods=['GET'])
def api_backups_download(filename):
    """Descargar un backup."""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    # Validar que el nombre no tenga traversal
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'error': 'Nombre inválido'}), 400
    backup_dir = Path('backups')
    target = backup_dir / filename
    if not target.exists() or not target.is_file():
        return jsonify({'success': False, 'error': 'Backup no encontrado'}), 404
    log_audit('backup_download', session['user_id'], 'backup', None, f'Descargó {filename}')
    return send_file(str(target), as_attachment=True, download_name=filename)


@app.route('/api/admin/backups/<path:filename>', methods=['DELETE'])
def api_backups_delete(filename):
    """Eliminar un backup específico."""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'error': 'Nombre inválido'}), 400
    backup_dir = Path('backups')
    target = backup_dir / filename
    if not target.exists():
        return jsonify({'success': False, 'error': 'Backup no encontrado'}), 404
    try:
        target.unlink()
        log_audit('backup_delete', session['user_id'], 'backup', None, f'Eliminó {filename}')
        return jsonify({'success': True, 'message': f'{filename} eliminado'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/backups/<path:filename>/restore', methods=['POST'])
def api_backups_restore(filename):
    """Restaurar BD desde un backup. PELIGROSO - sobrescribe la BD actual."""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'error': 'Nombre inválido'}), 400

    backup_dir = Path('backups')
    target = backup_dir / filename
    if not target.exists():
        return jsonify({'success': False, 'error': 'Backup no encontrado'}), 404

    db_path = _get_db_file_path()
    if not db_path:
        return jsonify({'success': False, 'error': 'No se pudo determinar la ruta de la BD actual'}), 500

    # Antes de restaurar, hacer un backup de seguridad de la BD actual
    safety_backup = create_backup(user_id=session['user_id'])

    try:
        # Cerrar conexiones SQLAlchemy
        db.session.close()
        db.engine.dispose()

        # Restaurar: si el archivo está cifrado (.enc), descifrar primero
        with open(target, 'rb') as f_enc:
            payload = f_enc.read()
        if target.name.endswith('.enc'):
            if not crypto_has_key():
                return jsonify({'success': False,
                                'error': 'Backup cifrado pero DB_ENCRYPTION_KEY no disponible.'}), 500
            try:
                payload = decrypt_bytes(payload)
            except Exception as ex:
                return jsonify({'success': False,
                                'error': f'No se pudo descifrar el backup (clave incorrecta?): {ex}'}), 500
        # payload ahora son los bytes gzip
        with gzip.GzipFile(fileobj=BytesIO(payload), mode='rb') as f_in:
            with open(db_path, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)

        log_audit('backup_restore', session['user_id'], 'backup', None,
                  f'Restaurada BD desde {filename}. Safety backup: {safety_backup.name if safety_backup else "ninguno"}')

        return jsonify({
            'success': True,
            'message': f'BD restaurada desde {filename}. Se creó un backup de seguridad: {safety_backup.name if safety_backup else "fallido"}. RECARGA LA PÁGINA.',
            'safety_backup': safety_backup.name if safety_backup else None
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error restaurando: {e}'}), 500


def cleanup_token_blacklist():
    """SECURITY FIX 7: Limpieza automática de token blacklist expirados"""
    with app.app_context():
        try:
            # Eliminar tokens expirados
            expired_tokens = TokenBlacklist.query.filter(
                TokenBlacklist.expires_at < datetime.utcnow()
            ).delete()
            db.session.commit()
            if expired_tokens > 0:
                print(f'[Security] Removed {expired_tokens} expired tokens from blacklist')
        except Exception as e:
            print(f'[Security] Error cleaning token blacklist: {e}')

# ═════════════════════════════════════════════════════════════════════════════
# EMAIL → TICKET (IMAP polling)
# ═════════════════════════════════════════════════════════════════════════════

def _get_oauth_token_for_imap(tenant_id, client_id, client_secret):
    """Obtiene un access token de Microsoft mediante Client Credentials Flow,
    para autenticar IMAP vía SASL XOAUTH2.
    Devuelve (token, None) o (None, error_msg).
    """
    import requests as _r
    if not (tenant_id and client_id and client_secret):
        return None, 'Faltan credenciales OAuth (tenant_id, client_id, client_secret)'
    url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    data = {
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://outlook.office365.com/.default',
        'grant_type': 'client_credentials',
    }
    try:
        r = _r.post(url, data=data, timeout=15)
        if r.status_code != 200:
            try:
                err = r.json()
                msg = err.get('error_description', err.get('error', r.text))
            except Exception:
                msg = r.text[:300]
            return None, f'Token request failed ({r.status_code}): {msg}'
        token = r.json().get('access_token')
        if not token:
            return None, 'Respuesta sin access_token'
        return token, None
    except Exception as e:
        return None, f'Excepción obteniendo token: {e}'


def _imap_xoauth2_authenticate(conn, user, token):
    """Autentica una conexión IMAP usando SASL XOAUTH2 con token de Microsoft.
    Lanza imaplib.IMAP4.error si falla.

    Importante: imaplib.authenticate() aplica base64 al retorno del callback
    internamente. Si devolvemos ya base64-encoded, se hace doble encoding y
    Office 365 responde con 'AUTHENTICATE command error: BAD Command Argument
    Error. 12'. Hay que retornar los BYTES CRUDOS del SASL auth string."""
    auth_string = f'user={user}\x01auth=Bearer {token}\x01\x01'
    conn.authenticate('XOAUTH2', lambda x: auth_string.encode('utf-8'))


def _imap_connect_and_login(mb):
    """Abre la conexión IMAP del buzón y autentica según auth_type.
    Devuelve (conn, None) o (None, error_msg)."""
    import imaplib
    try:
        if mb.use_ssl:
            conn = imaplib.IMAP4_SSL(mb.imap_host, mb.imap_port)
        else:
            conn = imaplib.IMAP4(mb.imap_host, mb.imap_port)
    except Exception as e:
        return None, f'No se pudo conectar a {mb.imap_host}:{mb.imap_port}: {e}'

    auth_type = (mb.auth_type or 'password').lower()
    try:
        if auth_type == 'oauth2':
            token, err = _get_oauth_token_for_imap(
                mb.oauth_tenant_id, mb.oauth_client_id, decrypt_secret(mb.oauth_client_secret)
            )
            if not token:
                try: conn.logout()
                except Exception: pass
                return None, f'OAuth: {err}'
            _imap_xoauth2_authenticate(conn, mb.imap_user, token)
        else:
            conn.login(mb.imap_user, decrypt_secret(mb.imap_password) or '')
        return conn, None
    except imaplib.IMAP4.error as e:
        try: conn.logout()
        except Exception: pass
        return None, f'IMAP error: {e}'
    except Exception as e:
        try: conn.logout()
        except Exception: pass
        return None, f'Error de autenticación: {e}'


def fetch_emails_from_mailbox(mailbox_id):
    """Conecta al buzón IMAP, lee correos NUEVOS y crea un ticket por cada uno.
    Devuelve (created_count, error_msg_or_None)."""
    import imaplib
    import email as _emaillib
    from email.header import decode_header

    with app.app_context():
        mb = MailboxConfig.query.get(mailbox_id)
        if not mb or not mb.is_active:
            return 0, 'Buzón no encontrado o inactivo'

        try:
            # Conectar con auth flexible (password o OAuth2)
            conn, conn_err = _imap_connect_and_login(mb)
            if conn_err:
                return 0, conn_err
            conn.select(mb.folder)

            # Buscar UNSEEN (no leídos)
            typ, msg_ids = conn.search(None, 'UNSEEN')
            if typ != 'OK':
                conn.logout()
                return 0, 'Error al buscar correos no leídos'

            ids = msg_ids[0].split()
            if not ids:
                mb.last_check_at = datetime.now()
                mb.last_status = 'ok'
                mb.last_error = None
                db.session.commit()
                conn.logout()
                return 0, None

            created_count = 0
            for msg_id in ids:
                typ, msg_data = conn.fetch(msg_id, '(RFC822)')
                if typ != 'OK':
                    continue
                raw_msg = msg_data[0][1]
                em = _emaillib.message_from_bytes(raw_msg)

                # Decodificar campos
                def _dec(s):
                    if not s: return ''
                    parts = decode_header(s)
                    out = ''
                    for txt, enc in parts:
                        if isinstance(txt, bytes):
                            try:
                                out += txt.decode(enc or 'utf-8', errors='replace')
                            except Exception:
                                out += txt.decode('utf-8', errors='replace')
                        else:
                            out += txt
                    return out.strip()

                subject = _dec(em.get('Subject') or '(sin asunto)')
                sender = _dec(em.get('From') or 'desconocido')
                message_id = em.get('Message-ID') or f'no-msgid-{msg_id.decode()}'

                # Evitar duplicados
                existing = MailboxEmail.query.filter_by(mailbox_id=mb.id, message_id=message_id).first()
                if existing:
                    continue

                # Extraer body (texto plano si está disponible)
                body_text = ''
                if em.is_multipart():
                    for part in em.walk():
                        ctype = part.get_content_type()
                        if ctype == 'text/plain' and 'attachment' not in str(part.get('Content-Disposition', '')):
                            try:
                                body_text = part.get_payload(decode=True).decode(part.get_content_charset() or 'utf-8', errors='replace')
                                break
                            except Exception:
                                pass
                else:
                    try:
                        body_text = em.get_payload(decode=True).decode(em.get_content_charset() or 'utf-8', errors='replace')
                    except Exception:
                        body_text = str(em.get_payload())

                # Buscar al sender en User (por email)
                sender_email = sender
                if '<' in sender:
                    sender_email = sender.split('<')[-1].rstrip('>').strip()
                creator = User.query.filter_by(email=sender_email, company=mb.company).first()
                if not creator:
                    creator = User.query.filter_by(company=mb.company, role='admin').first()

                # Crear ticket
                sla_min = get_sla_minutes_for_priority(mb.default_priority)
                ticket = Ticket(
                    ticket_number=get_next_ticket_number(mb.company),
                    title=subject[:200],
                    description=(
                        f"📧 **Correo recibido en {mb.name}**\n\n"
                        f"**De:** {sender}\n"
                        f"**Asunto:** {subject}\n"
                        f"**Fecha:** {em.get('Date', '')}\n\n"
                        f"---\n\n{(body_text or '(cuerpo vacío)').strip()[:5000]}"
                    ),
                    category=mb.default_category or 'Email',
                    priority=mb.default_priority,
                    status='open',
                    creator_id=creator.id if creator else 1,
                    company=mb.company,
                    sla_minutes=sla_min,
                    sla_deadline=compute_sla_deadline(datetime.now(), sla_min, mb.company)
                )
                try:
                    assign_ticket_auto(ticket)
                except Exception:
                    pass
                db.session.add(ticket)
                db.session.flush()

                # Registrar correo procesado
                db.session.add(MailboxEmail(
                    mailbox_id=mb.id,
                    message_id=message_id[:500],
                    subject=subject[:500],
                    sender=sender[:200],
                    ticket_id=ticket.id
                ))

                # Marcar como leído
                conn.store(msg_id, '+FLAGS', '\\Seen')
                created_count += 1

            # Actualizar mailbox
            mb.last_check_at = datetime.now()
            mb.last_status = 'ok'
            mb.last_error = None
            mb.tickets_created = (mb.tickets_created or 0) + created_count
            db.session.commit()

            conn.close()
            conn.logout()

            log_audit('mailbox_fetch', None, 'mailbox', mb.id,
                      f'{mb.name}: {created_count} tickets creados desde correo')
            return created_count, None

        except imaplib.IMAP4.error as e:
            mb.last_check_at = datetime.now()
            mb.last_status = 'error'
            mb.last_error = f'IMAP error: {str(e)[:300]}'
            db.session.commit()
            return 0, str(e)
        except Exception as e:
            mb.last_check_at = datetime.now()
            mb.last_status = 'error'
            mb.last_error = str(e)[:300]
            db.session.commit()
            return 0, str(e)


def start_mailbox_poller():
    """Scheduler que revisa todos los buzones activos según su poll_interval."""
    def loop():
        while True:
            try:
                with app.app_context():
                    now = datetime.now()
                    boxes = MailboxConfig.query.filter_by(is_active=True).all()
                    for mb in boxes:
                        if mb.last_check_at:
                            elapsed = (now - mb.last_check_at).total_seconds() / 60
                            if elapsed < (mb.poll_interval_minutes or 5):
                                continue
                        try:
                            count, err = fetch_emails_from_mailbox(mb.id)
                            if count > 0:
                                print(f'[mailbox] {mb.name}: {count} tickets creados')
                        except Exception as e:
                            print(f'[mailbox-error] {mb.name}: {e}')
                time.sleep(60)  # check cada minuto si hay buzones que correspondan
            except Exception as e:
                print(f'[mailbox-poller] Error: {e}')
                time.sleep(60)

    thread = Thread(target=loop, daemon=True)
    thread.start()
    print('[mailbox-poller] Scheduler iniciado')


# ═════════════════════════════════════════════════════════════════════════════
# ALERTAS DE SLA POR CORREO (30%, 60%, 100% de vencimiento)
# ═════════════════════════════════════════════════════════════════════════════

SLA_ALERT_THRESHOLDS = [30, 60, 100]  # porcentaje de SLA consumido

def _sla_alert_email(ticket, threshold_pct, assignee):
    """Construye el HTML del correo de alerta SLA."""
    now = datetime.now()
    if ticket.sla_deadline:
        remaining_min = int((ticket.sla_deadline - now).total_seconds() / 60)
        deadline_str = ticket.sla_deadline.strftime('%Y-%m-%d %H:%M')
    else:
        remaining_min = 0
        deadline_str = '—'

    # Severidad y colores según threshold
    if threshold_pct >= 150:
        color = '#7f1d1d'
        bg = '#fecaca'
        icon = '🆘'
        subject_tag = '🆘 SLA +100% VENCIDO'
        urgency_msg = '<strong>El SLA está MUY vencido (más del 100% del tiempo).</strong> Atiende este caso inmediatamente o escálalo a un superior.'
    elif threshold_pct >= 100:
        color = '#dc2626'
        bg = '#fee2e2'
        icon = '⏰'
        subject_tag = '🚨 SLA VENCIDO'
        urgency_msg = '<strong>El SLA ya está VENCIDO.</strong> Atiende este caso de manera inmediata.'
    elif threshold_pct >= 60:
        color = '#d97706'
        bg = '#fef3c7'
        icon = '⚠️'
        subject_tag = '⚠ SLA al 60%'
        urgency_msg = '<strong>El ticket ha consumido el 60% de su SLA.</strong> Atiéndelo pronto para evitar el incumplimiento.'
    else:
        color = '#2563eb'
        bg = '#dbeafe'
        icon = '🔔'
        subject_tag = '🔔 SLA al 30%'
        urgency_msg = '<strong>El ticket ha consumido el 30% de su SLA.</strong> Iníciate cuanto antes para mantener el cumplimiento.'

    subject = f'[DeskEli] {subject_tag} · {ticket.ticket_number} · {ticket.title[:60]}'

    base_url = get_public_base_url()
    ticket_url = f'{base_url}/technician/ticket/{ticket.id}'

    body = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
        <div style="background:{color};color:white;padding:18px;border-radius:8px 8px 0 0;">
            <h2 style="margin:0;">{icon} Alerta de SLA — {threshold_pct}%</h2>
        </div>
        <div style="padding:22px;background:#f9fafb;border:1px solid #e5e7eb;border-top:none;">
            <p>Hola <strong>{assignee.name}</strong>,</p>
            <div style="background:{bg};border-left:4px solid {color};padding:12px;margin:14px 0;border-radius:4px;">
                {urgency_msg}
            </div>

            <table style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:14px;">
                <tr><td style="padding:6px 8px;background:#e5e7eb;font-weight:bold;width:35%;">Ticket</td>
                    <td style="padding:6px 8px;border:1px solid #e5e7eb;">{ticket.ticket_number}</td></tr>
                <tr><td style="padding:6px 8px;background:#e5e7eb;font-weight:bold;">Título</td>
                    <td style="padding:6px 8px;border:1px solid #e5e7eb;">{ticket.title}</td></tr>
                <tr><td style="padding:6px 8px;background:#e5e7eb;font-weight:bold;">Prioridad</td>
                    <td style="padding:6px 8px;border:1px solid #e5e7eb;">{ticket.priority.upper()}</td></tr>
                <tr><td style="padding:6px 8px;background:#e5e7eb;font-weight:bold;">Estado</td>
                    <td style="padding:6px 8px;border:1px solid #e5e7eb;">{ticket.status}</td></tr>
                <tr><td style="padding:6px 8px;background:#e5e7eb;font-weight:bold;">SLA Total</td>
                    <td style="padding:6px 8px;border:1px solid #e5e7eb;">{ticket.sla_minutes or '—'} minutos</td></tr>
                <tr><td style="padding:6px 8px;background:#e5e7eb;font-weight:bold;">Vencimiento</td>
                    <td style="padding:6px 8px;border:1px solid #e5e7eb;">{deadline_str}</td></tr>
                <tr><td style="padding:6px 8px;background:#e5e7eb;font-weight:bold;">Tiempo restante</td>
                    <td style="padding:6px 8px;border:1px solid #e5e7eb;color:{color};font-weight:bold;">
                        {('VENCIDO' if remaining_min <= 0 else f'{remaining_min} minutos')}
                    </td></tr>
            </table>

            <p style="text-align:center;margin:20px 0;">
                <a href="{ticket_url}" style="display:inline-block;background:#2563eb;color:white;padding:12px 28px;border-radius:6px;text-decoration:none;font-weight:bold;">
                    🎫 Abrir Ticket
                </a>
            </p>

            <p style="font-size:11px;color:#6b7280;margin-top:20px;border-top:1px solid #e5e7eb;padding-top:10px;">
                Este es un mensaje automático del sistema DeskEli. No respondas a este correo.
                Las alertas se envían al alcanzar 30%, 60% y 100% del SLA.
            </p>
        </div>
    </body></html>
    """
    return subject, body


def _send_sla_alert(ticket, threshold_pct):
    """Envía el correo de alerta al asignado del ticket."""
    if not ticket.assignee_id:
        return False
    assignee = User.query.get(ticket.assignee_id)
    if not assignee or not assignee.email:
        return False

    subject, body = _sla_alert_email(ticket, threshold_pct, assignee)
    try:
        sent = send_email(assignee.email, subject, body)
    except Exception as e:
        print(f'[sla-alert] Error enviando email: {e}')
        sent = False

    # Marcar como enviada aunque falle SMTP (evita reintentos infinitos)
    current = (ticket.sla_alerts_sent or '').strip()
    sent_list = [s for s in current.split(',') if s]
    if str(threshold_pct) not in sent_list:
        sent_list.append(str(threshold_pct))
    ticket.sla_alerts_sent = ','.join(sent_list)

    log_audit('sla_alert_sent', None, 'ticket', ticket.id,
              f'Alerta SLA {threshold_pct}% enviada a {assignee.email} ({"OK" if sent else "fallo SMTP"})')

    # Emitir evento real-time también
    try:
        emit_ticket_event(ticket.company, 'sla_alert', {
            'ticket_id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'threshold_pct': threshold_pct,
            'assignee_email': assignee.email
        })
    except Exception:
        pass
    return sent


def _sla_event_enabled(threshold_pct):
    """Lee de Config si el evento de SLA de ese threshold está habilitado.
    Map: 30→evt_sla_30, 60→evt_sla_60, 100→evt_sla_100, >100→evt_sla_overdue"""
    if threshold_pct >= 100:
        key = 'email_evt_sla_100'
    elif threshold_pct >= 60:
        key = 'email_evt_sla_60'
    elif threshold_pct >= 30:
        key = 'email_evt_sla_30'
    else:
        return False
    c = Config.query.filter_by(key=key).first()
    return c.value == '1' if c else True  # default activo


def check_sla_alerts():
    """Revisa tickets activos y envía alertas en 30%, 60%, 100%, +100% del SLA según config."""
    with app.app_context():
        now = datetime.now()
        tickets = Ticket.query.filter(
            Ticket.status.in_(['open', 'in_progress']),
            Ticket.sla_deadline.isnot(None),
            Ticket.assignee_id.isnot(None)
        ).all()

        # Verificar también el evento "+100% / vencido" extra
        overdue_cfg = Config.query.filter_by(key='email_evt_sla_overdue').first()
        overdue_enabled = overdue_cfg.value == '1' if overdue_cfg else True

        sent_count = 0
        for t in tickets:
            if not t.created_at or not t.sla_deadline or not t.sla_minutes:
                continue

            total_seconds = t.sla_minutes * 60
            elapsed_seconds = (now - t.created_at).total_seconds()
            if total_seconds <= 0:
                continue
            pct = (elapsed_seconds / total_seconds) * 100

            already_sent = set([s for s in (t.sla_alerts_sent or '').split(',') if s])

            # 30, 60, 100
            for threshold in SLA_ALERT_THRESHOLDS:
                if pct >= threshold and str(threshold) not in already_sent:
                    if not _sla_event_enabled(threshold):
                        # Marcar como "enviado" para no acumular el threshold a futuro
                        already_sent.add(str(threshold))
                        t.sla_alerts_sent = ','.join(sorted(already_sent, key=int) if all(s.lstrip('-').isdigit() for s in already_sent) else already_sent)
                        continue
                    try:
                        _send_sla_alert(t, threshold)
                        sent_count += 1
                    except Exception as e:
                        print(f'[sla-alert] Error en ticket {t.ticket_number}: {e}')

            # +100% (vencido — recordatorio adicional al 150% si el toggle está activo)
            if overdue_enabled and pct >= 150 and 'overdue' not in already_sent:
                try:
                    _send_sla_alert(t, 150)  # 150 = label "+100%"
                    already_sent.add('overdue')
                    sent_list = [s for s in (t.sla_alerts_sent or '').split(',') if s]
                    if 'overdue' not in sent_list:
                        sent_list.append('overdue')
                    t.sla_alerts_sent = ','.join(sent_list)
                    sent_count += 1
                except Exception as e:
                    print(f'[sla-alert] Error en +100 ticket {t.ticket_number}: {e}')

        if sent_count > 0:
            db.session.commit()
            print(f'[sla-alert] {sent_count} alertas enviadas')


def start_sla_alert_scheduler():
    """Thread que revisa SLA cada 2 minutos y envía correos en los thresholds."""
    def loop():
        while True:
            try:
                check_sla_alerts()
            except Exception as e:
                print(f'[sla-alert-scheduler] Error: {e}')
            time.sleep(120)  # cada 2 minutos
    thread = Thread(target=loop, daemon=True)
    thread.start()
    print('[sla-alert] Scheduler iniciado (revisión cada 2 min)')


def purge_audit_logs():
    """Elimina entradas de audit_logs más antiguas que AUDIT_LOG_RETENTION_DAYS (default 365).
    Configurable vía variable de entorno o config 'audit_retention_days'."""
    with app.app_context():
        try:
            retention_days = int(os.getenv('AUDIT_LOG_RETENTION_DAYS', '365'))
            cfg = Config.query.filter_by(key='audit_retention_days').first()
            if cfg and cfg.value.isdigit():
                retention_days = int(cfg.value)
            if retention_days <= 0:
                return  # Disabled
            cutoff = datetime.now() - timedelta(days=retention_days)
            deleted = AuditLog.query.filter(AuditLog.created_at < cutoff).delete()
            if deleted > 0:
                db.session.commit()
                print(f'[audit-purge] {deleted} logs antiguos (>{retention_days}d) eliminados')
        except Exception as e:
            print(f'[audit-purge] Error: {e}')


def start_audit_log_purge_scheduler():
    """Scheduler diario que purga logs viejos."""
    def loop():
        # Esperar 60s al arrancar para no bloquear
        time.sleep(60)
        while True:
            try:
                purge_audit_logs()
            except Exception as e:
                print(f'[audit-purge-scheduler] Error: {e}')
            time.sleep(86400)  # Cada 24h
    thread = Thread(target=loop, daemon=True)
    thread.start()
    print('[audit-purge] Scheduler iniciado (purga diaria, retención 365 días)')


def start_token_cleanup_scheduler():
    """SECURITY FIX 7: Iniciar scheduler de limpieza de token blacklist cada hora"""
    def cleanup_loop():
        while True:
            try:
                cleanup_token_blacklist()
                time.sleep(3600)  # Cada hora
            except Exception as e:
                print(f'[Token Cleanup Error] {e}')
                time.sleep(1800)  # Reintentar en 30 min

    thread = Thread(target=cleanup_loop, daemon=True)
    thread.start()

# ═════════════════════════════════════════════════════════════════════════════
# WATCHDOG AUTO-RESTART RNF-02-03 - Reiniciar ante 3 fallos consecutivos
# ═════════════════════════════════════════════════════════════════════════════

db_failure_count = 0
restart_threshold = 3

# ═════════════════════════════════════════════════════════════════════════════
# EMAIL SMTP (Opcional) - Envío de notificaciones por correo
# ═════════════════════════════════════════════════════════════════════════════

def _get_smtp_config(company=None):
    """Lee la configuración SMTP. Prioriza en este orden:
    1. Si se pasa `company`, busca en Company.smtp_* (config por empresa)
    2. Cae a Config global (tabla Config con keys email_smtp_*) — fallback
    3. Cae a variables de entorno (.env) — último recurso

    Retorna dict: server, port, user, password, from_addr, security, source.
    """
    cfg = {}
    source = 'env'

    # 1. Config por empresa (si la empresa la tiene definida)
    if company:
        try:
            c_obj = Company.query.filter_by(code=company).first()
            if c_obj and (c_obj.smtp_host or c_obj.smtp_user):
                cfg = {
                    'smtp_host': c_obj.smtp_host,
                    'smtp_port': str(c_obj.smtp_port) if c_obj.smtp_port else None,
                    'smtp_user': c_obj.smtp_user,
                    'smtp_pass': decrypt_secret(c_obj.smtp_password),
                    'smtp_from': c_obj.smtp_from,
                    'smtp_security': c_obj.smtp_security,
                }
                source = f'company:{company}'
        except Exception:
            pass

    # 2. Si no hay por empresa, leer global desde Config
    if not cfg.get('smtp_host'):
        try:
            cfg = {}
            for key in ('smtp_host', 'smtp_port', 'smtp_user', 'smtp_pass', 'smtp_from', 'smtp_security'):
                c = Config.query.filter_by(key=f'email_{key}').first()
                if c and c.value:
                    cfg[key] = c.value
            if cfg.get('smtp_host'):
                source = 'global'
        except Exception:
            pass

    # 3. Fallback al .env
    server = cfg.get('smtp_host') or os.getenv('SMTP_SERVER', 'localhost')
    try:
        port = int(cfg.get('smtp_port') or os.getenv('SMTP_PORT', 587))
    except (ValueError, TypeError):
        port = 587
    user = cfg.get('smtp_user') or os.getenv('SMTP_USER', '')
    password = cfg.get('smtp_pass') or os.getenv('SMTP_PASSWORD', '')
    from_addr = cfg.get('smtp_from') or os.getenv('SMTP_FROM', user)
    security = (cfg.get('smtp_security') or '').lower()
    return {
        'server': server, 'port': port, 'user': user, 'password': password,
        'from_addr': from_addr, 'security': security, 'source': source
    }


def send_email(to_email, subject, body, attachments=None, company=None, cc_emails=None):
    """Enviar email vía SMTP. Soporta SSL (465) y STARTTLS (587).
    Si se pasa `company`, usa la config SMTP de esa empresa (con fallback a global).
    Devuelve True/False. Loguea diagnóstico claro si SMTP rechaza la conexión.

    attachments: lista de tuplas (filename, bytes_or_bytesio, mime_subtype).
    cc_emails: lista de emails a copiar (CC), o string único.
    """
    import socket
    from email.mime.multipart import MIMEMultipart as _MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders

    cfg = _get_smtp_config(company=company)
    smtp_server = cfg['server']
    smtp_port = cfg['port']
    smtp_user = cfg['user']
    smtp_password = cfg['password']
    smtp_from = cfg['from_addr']
    print(f'[send_email] Usando SMTP de "{cfg["source"]}" ({smtp_server}:{smtp_port}) para enviar a {to_email}')

    if not smtp_user or not smtp_password:
        print('[send_email] SMTP no configurado (faltan SMTP_USER/SMTP_PASSWORD). Saltando envío.')
        return False

    msg = _MIMEMultipart()
    msg['From'] = smtp_from
    msg['To'] = to_email
    # Normalizar cc_emails a lista limpia (sin vacíos ni duplicados con TO)
    cc_list = []
    if cc_emails:
        raw_cc = cc_emails if isinstance(cc_emails, (list, tuple)) else [cc_emails]
        for e in raw_cc:
            e = (e or '').strip()
            if e and e.lower() != (to_email or '').lower() and e not in cc_list:
                cc_list.append(e)
    if cc_list:
        msg['Cc'] = ', '.join(cc_list)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    # Adjuntos
    if attachments:
        for att in attachments:
            try:
                filename, data, subtype = att if len(att) == 3 else (att[0], att[1], None)
                # Normalizar data a bytes
                if hasattr(data, 'read'):
                    data.seek(0)
                    payload = data.read()
                else:
                    payload = data
                main_type = 'application'
                sub_type = subtype or 'octet-stream'
                part = MIMEBase(main_type, sub_type)
                part.set_payload(payload)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
                msg.attach(part)
            except Exception as e:
                print(f'[send_email] No pude adjuntar {att[0] if att else "?"}: {e}')

    # Decidir SSL vs STARTTLS: respeta la elección de la UI; si no, usa el puerto
    use_ssl = (cfg.get('security') == 'ssl') or (smtp_port == 465)
    try:
        if use_ssl:
            import ssl as _ssl
            ctx = _ssl.create_default_context()
            with smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=20, context=ctx) as server:
                server.ehlo()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(smtp_server, smtp_port, timeout=20) as server:
                server.ehlo()
                server.starttls()
                server.ehlo()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
        return True

    except smtplib.SMTPAuthenticationError as e:
        print(f'[send_email] Autenticación SMTP falló: {e}')
        print('  → Office 365: tu cuenta puede tener MFA activo. Usa una App Password.')
        print('  → También verifica que "Authenticated SMTP" esté habilitado en el mailbox.')
        return False
    except (ConnectionResetError, socket.error, smtplib.SMTPServerDisconnected) as e:
        print(f'[send_email] El servidor SMTP cerró la conexión: {e}')
        print(f'  → Host: {smtp_server}:{smtp_port} (use_ssl={use_ssl})')
        print('  → Causa más común en Office 365: SMTP AUTH deshabilitado en el tenant o mailbox.')
        print('  → Soluciones:')
        print('    1) Habilita "Authenticated SMTP" en el mailbox del usuario remitente:')
        print('       Admin Center MS365 → Usuarios → seleccionar usuario → Correo → Administrar aplicaciones de correo electrónico → marcar "SMTP autenticado".')
        print('    2) O via PowerShell de Exchange Online:')
        print(f'       Set-CASMailbox -Identity {smtp_user} -SmtpClientAuthenticationDisabled $false')
        print('    3) Verifica que el tenant no tenga "Authenticated SMTP" deshabilitado globalmente:')
        print('       Get-TransportConfig | Format-List SmtpClientAuthenticationDisabled')
        print('    4) Si tienes MFA, genera una App Password (https://mysignins.microsoft.com/security-info → "Contraseñas de aplicación").')
        print('    5) Alternativa: usar puerto 465 con SSL en lugar de 587 con STARTTLS.')
        print('    6) Otra alternativa: usar Microsoft Graph API o un servicio como SendGrid.')
        return False
    except Exception as e:
        print(f'[Email Error] {e}')
        return False


def is_email_event_enabled(event_name):
    """Verifica si un evento de email está habilitado en la config admin.
    Por defecto activo si no hay registro."""
    try:
        c = Config.query.filter_by(key=f'email_evt_{event_name}').first()
        if c is None:
            return True  # default activo
        return c.value != '0'
    except Exception:
        return True


def notify_ticket_assigned(ticket, new_assignee, assigned_by_name='Sistema', reason=''):
    """Envía email al técnico cuando se le asigna un ticket.
    Respeta el flag email_evt_ticket_assigned y verifica que el técnico tenga email."""
    if not is_email_event_enabled('ticket_assigned'):
        print(f'[notify] Evento ticket_assigned deshabilitado en config, skip')
        return False
    if not new_assignee or not new_assignee.email:
        print(f'[notify] Técnico sin email, skip')
        return False

    # Construir URL del ticket (usa PUBLIC_URL env var o ALLOWED_ORIGINS como fallback)
    base_url = ''
    try:
        c = Config.query.filter_by(key='general_base_url').first()
        if c and c.value:
            base_url = c.value.rstrip('/')
    except Exception:
        pass
    if not base_url:
        base_url = get_public_base_url()

    ticket_url = f'{base_url}/technician/ticket/{ticket.id}'

    # Mapeo de prioridad → color + label
    prio_meta = {
        'critical': {'icon': '🔴', 'label': 'CRÍTICA', 'color': '#dc2626'},
        'high':     {'icon': '🟠', 'label': 'ALTA',    'color': '#ea580c'},
        'medium':   {'icon': '🟡', 'label': 'MEDIA',   'color': '#d97706'},
        'low':      {'icon': '🟢', 'label': 'BAJA',    'color': '#16a34a'},
    }
    pm = prio_meta.get(ticket.priority or 'medium', prio_meta['medium'])
    sla_str = ''
    if ticket.sla_deadline:
        sla_str = ticket.sla_deadline.strftime('%d/%m/%Y %H:%M')

    creator_name = ''
    try:
        creator = User.query.get(ticket.creator_id) if ticket.creator_id else None
        if creator:
            creator_name = f'{creator.name} ({creator.email or creator.username})'
    except Exception:
        pass

    # Resumen de descripción (primeros 500 chars sin HTML peligroso)
    desc_short = (ticket.description or '')
    if len(desc_short) > 500:
        desc_short = desc_short[:500] + '...'
    desc_safe = desc_short.replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')

    subject = f'[DeskEli] Te asignaron el ticket {ticket.ticket_number} — {ticket.title[:50]}'

    body = f"""
    <html><body style="font-family: Segoe UI, Arial, sans-serif; color: #1f2937; max-width: 680px; margin: 0 auto;">
        <div style="background: linear-gradient(135deg, #1f2937, #7c3aed); color: white; padding: 24px; border-radius: 10px 10px 0 0;">
            <h1 style="margin: 0; font-size: 22px;">🔔 Nuevo Ticket Asignado</h1>
            <p style="margin: 6px 0 0; opacity: 0.95;">{ticket.ticket_number} · {ticket.company.upper()}</p>
        </div>
        <div style="background: white; padding: 22px; border: 1px solid #e5e7eb; border-radius: 0 0 10px 10px;">
            <p>Hola <strong>{new_assignee.name}</strong>,</p>
            <p>Te asignaron un ticket. Te dejo el resumen:</p>

            <table style="width: 100%; border-collapse: collapse; margin: 14px 0;">
                <tr>
                    <td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700; width: 35%;">Número</td>
                    <td style="padding: 9px 12px; background: #ffffff;"><strong>{ticket.ticket_number}</strong></td>
                </tr>
                <tr>
                    <td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">Título</td>
                    <td style="padding: 9px 12px; background: #ffffff;">{ticket.title}</td>
                </tr>
                <tr>
                    <td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">Prioridad</td>
                    <td style="padding: 9px 12px; background: #ffffff;">
                        <span style="background: {pm['color']}; color: white; padding: 3px 10px; border-radius: 4px; font-weight: 700; font-size: 12px;">{pm['icon']} {pm['label']}</span>
                    </td>
                </tr>
                <tr>
                    <td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">Categoría</td>
                    <td style="padding: 9px 12px; background: #ffffff;">{ticket.category or 'General'}</td>
                </tr>
                <tr>
                    <td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">Estado</td>
                    <td style="padding: 9px 12px; background: #ffffff;">{ticket.status}</td>
                </tr>
                {f'<tr><td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">SLA vence</td><td style="padding: 9px 12px; background: #ffffff;">⏰ <strong>{sla_str}</strong></td></tr>' if sla_str else ''}
                {f'<tr><td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">Solicitante</td><td style="padding: 9px 12px; background: #ffffff;">{creator_name}</td></tr>' if creator_name else ''}
                <tr>
                    <td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">Asignado por</td>
                    <td style="padding: 9px 12px; background: #ffffff;">{assigned_by_name}</td>
                </tr>
                {f'<tr><td style="padding: 9px 12px; background: #f3f4f6; font-weight: 700;">Razón</td><td style="padding: 9px 12px; background: #ffffff;"><em>{reason}</em></td></tr>' if reason else ''}
            </table>

            <div style="background: #f9fafb; border-left: 4px solid #7c3aed; padding: 12px 14px; border-radius: 4px; margin-top: 14px;">
                <div style="font-size: 12px; font-weight: 700; color: #5b21b6; margin-bottom: 6px;">📝 DESCRIPCIÓN</div>
                <div style="font-size: 13px; color: #374151; line-height: 1.5;">{desc_safe or '<em>(Sin descripción)</em>'}</div>
            </div>

            <div style="text-align: center; margin-top: 22px;">
                <a href="{ticket_url}" style="display: inline-block; padding: 12px 28px; background: linear-gradient(135deg, #7c3aed, #2563eb); color: white; text-decoration: none; border-radius: 8px; font-weight: 700;">
                    🔧 Abrir ticket en DeskEli
                </a>
            </div>

            <p style="font-size: 11px; color: #6b7280; margin-top: 22px; text-align: center;">
                Este correo fue enviado automáticamente por DeskEli cuando se te asignó el ticket.
            </p>
        </div>
    </body></html>
    """

    ok = send_email(new_assignee.email, subject, body, company=ticket.company)
    status_str = 'OK' if ok else 'FALLO'
    print(f'[notify] Email asignacion {ticket.ticket_number} -> {new_assignee.email}: {status_str}')
    return ok


def start_watchdog():
    """Monitorear BD y reiniciar si hay 3 fallos consecutivos (RNF-02-03)"""
    global db_failure_count

    def watchdog_loop():
        global db_failure_count
        while True:
            try:
                with app.app_context():
                    # Intentar consulta simple
                    db.session.execute(db.text('SELECT 1'))
                    db_failure_count = 0  # Reset en éxito
                time.sleep(30)  # Verificar cada 30s
            except Exception as e:
                db_failure_count += 1
                error_type = type(e).__name__
                print(f'[Watchdog] DB Error #{db_failure_count} ({error_type}): {str(e)[:100]}')

                if db_failure_count >= restart_threshold:
                    try:
                        with app.app_context():
                            log_audit('watchdog_restart', None, 'watchdog', None,
                                      f'Watchdog iniciando restart tras {db_failure_count} fallos: {error_type}')
                    except Exception as audit_err:
                        print(f'[Watchdog] Could not log restart: {audit_err}')
                    print(f'[Watchdog] {restart_threshold} fallos detectados, reiniciando servicio...')
                    db_failure_count = 0
                    time.sleep(60)

    thread = Thread(target=watchdog_loop, daemon=True)
    thread.start()

# Empresa "master" que ve todas las demás. Configurable a futuro vía Config.
MASTER_COMPANY = 'eliot'

def is_master_admin(user_company=None, role=None):
    """¿El usuario actual (o el dado) es admin de la empresa master?"""
    if user_company is None:
        user_company = session.get('company')
    if role is None:
        role = session.get('role')
    return role == 'admin' and user_company == MASTER_COMPANY


def admin_companies_scope(user_company=None, role=None):
    """Lista de empresas que el usuario puede ver en el panel admin.
    - Admin de la empresa master ('eliot') → todas las empresas.
    - Otros (admin de pash/primatela, técnicos, empleados) → solo la suya.
    """
    if user_company is None:
        user_company = session.get('company')
    if role is None:
        role = session.get('role')
    if is_master_admin(user_company, role):
        return [c.code for c in Company.query.all()] or [user_company]
    return [user_company]


def can_user_access_ticket(user, ticket):
    """Determina si `user` tiene permiso para VER/EDITAR este ticket.
    Reglas:
    - Si el ticket es de la misma empresa que el user, si.
    - Si el user es master admin, si (Eliot admin ve todas las empresas).
    - Si el ticket esta asignado a cualquiera de las identidades espejo del user
      (misma persona replicada en otras empresas), si — para consolidar la vista
      del especialista al entrar a Eliot.
    """
    if not user or not ticket:
        return False
    if ticket.company == user.company:
        return True
    if is_master_admin(user.company, user.role):
        return True
    if ticket.assignee_id in get_user_identity_ids(user):
        return True
    return False


def get_user_identity_ids(user):
    """Devuelve todos los user.id que representan a la misma persona a traves
    de las empresas espejadas.
    - Si user es un ESPEJO (mirrored_from_id != NULL): devuelve source + todos
      los espejos hermanos + el user actual.
    - Si user es un ORIGEN (o usuario normal sin espejos): devuelve el propio id
      + todos los espejos que apuntan a el.
    Uso: consolidar la vista de tickets/subtareas asignados a cualquier
    identidad del especialista en el dashboard de una empresa."""
    if not user:
        return set()
    ids = {user.id}
    if user.mirrored_from_id:
        source_id = user.mirrored_from_id
        ids.add(source_id)
        for m in User.query.filter_by(mirrored_from_id=source_id).all():
            ids.add(m.id)
    else:
        for m in User.query.filter_by(mirrored_from_id=user.id).all():
            ids.add(m.id)
    return ids


def get_my_group_user_ids(user):
    """IDs de tecnicos/admins de la misma empresa que comparten AL MENOS 1 subrol
    con `user`. Incluye al propio user en el resultado. Si el user no tiene subroles,
    devuelve solo su propio id (grupo unipersonal).
    Usado para el filtro 'De mis grupos' en el dashboard del tecnico."""
    if not user:
        return set()
    my_subrole_ids = [a.subrole_id for a in UserSubrole.query.filter_by(user_id=user.id).all()]
    if not my_subrole_ids:
        return {user.id}
    # Otros usuarios que tienen al menos 1 de mis subroles
    peer_ids = {
        row.user_id for row in UserSubrole.query
        .filter(UserSubrole.subrole_id.in_(my_subrole_ids))
        .all()
    }
    peer_ids.add(user.id)
    # Restringir a misma empresa + roles operativos
    peers = User.query.filter(
        User.id.in_(peer_ids),
        User.company == user.company,
        User.role.in_(['technician', 'admin'])
    ).with_entities(User.id).all()
    return {row.id for row in peers}


def get_ticket_assignment_info(ticket):
    """Devuelve un dict {by, source, when} describiendo quién/qué hizo la última asignación.
    Busca primero en audit_logs (manual o auto_assign), luego en agent_actions (orchestrator IA)."""
    info = {'by': None, 'source': None, 'when': None, 'detail': None}
    if not ticket or not ticket.assignee_id:
        return info
    # 1. ¿Hubo reasignación manual en audit_logs?
    last_audit = AuditLog.query.filter_by(
        entity_type='ticket', entity_id=ticket.id
    ).filter(AuditLog.action.in_(['reassign_ticket', 'auto_assign'])).order_by(AuditLog.created_at.desc()).first()
    if last_audit:
        info['when'] = last_audit.created_at.strftime('%Y-%m-%d %H:%M') if last_audit.created_at else None
        info['detail'] = last_audit.description
        if last_audit.action == 'auto_assign':
            info['source'] = 'sistema'
            info['by'] = 'Sistema (auto-asignación por reglas)'
            return info
        # reassign_ticket → por un usuario
        if last_audit.user_id:
            u = User.query.get(last_audit.user_id)
            info['by'] = u.name if u else f'Usuario #{last_audit.user_id}'
            info['source'] = 'manual'
            return info
    # 2. ¿Asignación por orquestador IA?
    try:
        ag_action = AgentAction.query.filter_by(
            ticket_id=ticket.id, agent_name='assignor', success=True
        ).order_by(AgentAction.created_at.desc()).first()
        if ag_action:
            info['by'] = 'Sistema (Agent Orchestrator - IA)' if ag_action.used_llm else 'Sistema (Agent Orchestrator - reglas)'
            info['source'] = 'orchestrator'
            info['when'] = ag_action.created_at.strftime('%Y-%m-%d %H:%M') if ag_action.created_at else None
            info['detail'] = ag_action.output_data
            return info
    except Exception:
        pass
    # 3. Sin registro: fue creado ya con assignee_id por el creador
    creator = User.query.get(ticket.creator_id) if ticket.creator_id else None
    info['by'] = creator.name if creator else 'Desconocido'
    info['source'] = 'creation'
    info['when'] = ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else None
    info['detail'] = 'Asignado al momento de crear el ticket'
    return info


def get_company_theme(company):
    """Devuelve el tema visual configurado para una empresa. Cacheado 5min."""
    cache_key = f'theme:{company or "default"}'
    cached = cache_get(cache_key)
    if cached is not None:
        return cached
    if company:
        c = Config.query.filter_by(key=f"theme_{company}").first()
        if c and c.value in THEMES:
            cache_set(cache_key, c.value, ttl_seconds=300)
            return c.value
    legacy = Config.query.filter_by(key='theme').first()
    if legacy and legacy.value in THEMES:
        cache_set(cache_key, legacy.value, ttl_seconds=300)
        return legacy.value
    cache_set(cache_key, 'blue', ttl_seconds=300)
    return 'blue'


# ═════════════════════════════════════════════════════════════════════════════
# HORARIO LABORAL Y CÁLCULO DE SLA RESPETANDO BUSINESS HOURS
# ═════════════════════════════════════════════════════════════════════════════

def get_business_config(company):
    """Devuelve la configuración de horario laboral con horarios POR DÍA.
    Soporta horarios diferentes para cada día (ej: Sáb 9-13). Cacheado 5min.
    schedule: dict {iso_weekday: {start_h, start_m, end_h, end_m} | None (no laboral)}
    """
    if not company:
        return {'enabled': False}
    cache_key = f'biz_cfg:{company}'
    cached = cache_get(cache_key)
    if cached is not None:
        return cached
    def _get(key, default=None):
        c = Config.query.filter_by(key=f'biz_{key}_{company}').first()
        return c.value if c else default

    enabled = _get('enabled', '0') == '1'
    if not enabled:
        return {'enabled': False}

    # Construir schedule por día (compatibilidad con formato viejo)
    def _parse_hm(s, default_h, default_m):
        try:
            h, m = map(int, s.split(':'))
            return h, m
        except Exception:
            return default_h, default_m

    # Por día: keys biz_day{N}_start_{company}, biz_day{N}_end_{company}, biz_day{N}_enabled_{company}
    schedule = {}
    for iso in range(1, 8):
        day_enabled_cfg = _get(f'day{iso}_enabled', None)
        if day_enabled_cfg is None:
            # Fallback al formato viejo: usar start/end globales si el día está en days CSV
            days_csv = _get('days', '1,2,3,4,5')
            try:
                in_old_days = iso in {int(d.strip()) for d in days_csv.split(',') if d.strip()}
            except Exception:
                in_old_days = iso in {1,2,3,4,5}
            if in_old_days:
                sh, sm = _parse_hm(_get('start', '08:00'), 8, 0)
                eh, em = _parse_hm(_get('end', '18:00'), 18, 0)
                schedule[iso] = {'start_h': sh, 'start_m': sm, 'end_h': eh, 'end_m': em}
            else:
                schedule[iso] = None  # No laboral
        else:
            if day_enabled_cfg != '1':
                schedule[iso] = None
            else:
                sh, sm = _parse_hm(_get(f'day{iso}_start', '08:00'), 8, 0)
                eh, em = _parse_hm(_get(f'day{iso}_end', '18:00'), 18, 0)
                schedule[iso] = {'start_h': sh, 'start_m': sm, 'end_h': eh, 'end_m': em}

    holidays_csv = _get('holidays', '')
    holidays = set()
    for d in holidays_csv.split(','):
        d = d.strip()
        if d:
            try:
                holidays.add(datetime.strptime(d, '%Y-%m-%d').date())
            except Exception:
                pass

    result = {
        'enabled': True,
        'schedule': schedule,
        'holidays': holidays,
    }
    cache_set(cache_key, result, ttl_seconds=300)
    return result


def _is_working_day(dt, cfg):
    if dt.date() in cfg['holidays']:
        return False
    return cfg['schedule'].get(dt.isoweekday()) is not None


def _day_window(dt, cfg):
    """Devuelve (start_dt, end_dt) del horario laboral del día específico de dt.
    Cada día puede tener su propio horario."""
    day_cfg = cfg['schedule'].get(dt.isoweekday())
    if not day_cfg:
        return None, None
    start = dt.replace(hour=day_cfg['start_h'], minute=day_cfg['start_m'], second=0, microsecond=0)
    end = dt.replace(hour=day_cfg['end_h'], minute=day_cfg['end_m'], second=0, microsecond=0)
    return start, end


def _next_day_start(current_dt, cfg):
    """Avanza al inicio del día laboral siguiente respetando schedule por día."""
    nxt = current_dt + timedelta(days=1)
    nxt = nxt.replace(hour=0, minute=0, second=0, microsecond=0)
    # Buscar siguiente día laboral (con horario)
    for _ in range(370):
        if _is_working_day(nxt, cfg):
            day_cfg = cfg['schedule'][nxt.isoweekday()]
            return nxt.replace(hour=day_cfg['start_h'], minute=day_cfg['start_m'])
        nxt += timedelta(days=1)
    return current_dt


def business_minutes_add(start_dt, minutes_to_add, company):
    """Avanza `minutes_to_add` desde start_dt, contando SOLO horas laborales y días laborales.
    Si el horario laboral está deshabilitado, hace timedelta normal."""
    cfg = get_business_config(company)
    if not cfg['enabled']:
        return start_dt + timedelta(minutes=minutes_to_add)

    remaining = int(minutes_to_add)
    if remaining <= 0:
        return start_dt

    current = start_dt
    safety = 0
    max_iterations = 366 * 2  # protección contra loops infinitos (~2 años)

    while remaining > 0 and safety < max_iterations:
        safety += 1
        if not _is_working_day(current, cfg):
            current = _next_day_start(current, cfg)
            continue

        day_start, day_end = _day_window(current, cfg)
        if day_start is None:
            current = _next_day_start(current, cfg)
            continue
        if current < day_start:
            current = day_start
        if current >= day_end:
            current = _next_day_start(current, cfg)
            continue

        avail = int((day_end - current).total_seconds() / 60)
        if avail >= remaining:
            return current + timedelta(minutes=remaining)
        else:
            remaining -= avail
            current = _next_day_start(current, cfg)

    return current


def compute_sla_deadline(start_dt, sla_minutes, company):
    """Helper público: devuelve el datetime de vencimiento del SLA considerando
    el horario laboral configurado en la empresa."""
    if not sla_minutes:
        return None
    return business_minutes_add(start_dt, sla_minutes, company)


def get_sla_minutes_for_priority(priority):
    """Lee minutos de SLA configurados para una prioridad (config admin)."""
    sla_config = Config.query.filter_by(key=f"sla_{priority}").first()
    if sla_config:
        try:
            return int(sla_config.value)
        except (ValueError, TypeError):
            pass
    defaults = {'critical': 60, 'high': 240, 'medium': 480, 'low': 1440}
    return defaults.get(priority, 120)


def get_next_subtask_number(ticket):
    """Genera próximo código de subtarea independiente: SUB-ELIOT-00001"""
    company = (ticket.company or 'GEN').upper()
    prefix = f"SUB-{company}-"
    last = Subtask.query.join(Ticket, Subtask.ticket_id == Ticket.id).filter(
        Ticket.company == ticket.company,
        Subtask.subtask_number.like(f"{prefix}%")
    ).order_by(Subtask.id.desc()).first()
    next_num = 1
    if last and last.subtask_number:
        try:
            next_num = int(last.subtask_number.split('-')[-1]) + 1
        except (ValueError, IndexError):
            next_num = (Subtask.query.join(Ticket).filter(Ticket.company == ticket.company).count() + 1)
    return f"{prefix}{next_num:05d}"


def backfill_subtask_numbers():
    """Asigna SUB-COMPANY-XXXXX a subtareas existentes sin número."""
    orphans = Subtask.query.filter(
        (Subtask.subtask_number == None) | (Subtask.subtask_number == '')
    ).order_by(Subtask.id.asc()).all()
    if not orphans:
        return
    # Agrupar por empresa para numerar correlativo
    from collections import defaultdict
    by_company = defaultdict(list)
    for s in orphans:
        t = Ticket.query.get(s.ticket_id)
        if t:
            by_company[t.company].append((s, t))
    for company, items in by_company.items():
        comp_upper = (company or 'GEN').upper()
        prefix = f"SUB-{comp_upper}-"
        # Buscar el último número usado en esa empresa
        existing_last = Subtask.query.join(Ticket).filter(
            Ticket.company == company,
            Subtask.subtask_number.like(f"{prefix}%")
        ).all()
        max_num = 0
        for s in existing_last:
            try:
                n = int(s.subtask_number.split('-')[-1])
                if n > max_num:
                    max_num = n
            except (ValueError, IndexError):
                pass
        for sub, _ticket in items:
            max_num += 1
            sub.subtask_number = f"{prefix}{max_num:05d}"
    db.session.commit()
    print(f"[backfill_subtask_numbers] {len(orphans)} subtareas numeradas")


def migrate_users_schema():
    """Si users.username tiene UNIQUE en una sola columna (legado), recrear la tabla
    con la constraint compuesta (username, company)."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'users' not in inspector.get_table_names():
        return
    # SQLite expone UNIQUE column-level como constraint, no como index
    unique_constraints = inspector.get_unique_constraints('users')
    has_bad_unique = any(
        uc.get('column_names') == ['username']
        for uc in unique_constraints
    )
    if not has_bad_unique:
        return
    print("[migrate_users] Detectado UNIQUE legado en users.username, recreando tabla...")
    with db.engine.begin() as conn:
        conn.execute(text("ALTER TABLE users RENAME TO users_old_legacy"))
        conn.execute(text("""
            CREATE TABLE users (
                id INTEGER NOT NULL PRIMARY KEY,
                username VARCHAR(80) NOT NULL,
                name VARCHAR(120) NOT NULL,
                email VARCHAR(120) NOT NULL,
                role VARCHAR(20) NOT NULL,
                company VARCHAR(20) NOT NULL,
                password_hash VARCHAR(255),
                is_active BOOLEAN,
                last_login DATETIME,
                created_at DATETIME,
                CONSTRAINT _user_company_uc UNIQUE (username, company)
            )
        """))
        conn.execute(text("CREATE INDEX ix_users_company ON users (company)"))
        conn.execute(text("""
            INSERT INTO users (id, username, name, email, role, company, password_hash, is_active, last_login, created_at)
            SELECT id, username, name, email, role, company, password_hash, is_active, last_login, created_at FROM users_old_legacy
        """))
        conn.execute(text("DROP TABLE users_old_legacy"))
    print("[migrate_users] Tabla users recreada con UNIQUE compuesto (username, company)")


def migrate_report_recipients_cc():
    """Agrega columna cc_user_ids a report_recipients si no existe."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'report_recipients' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('report_recipients')}
    if 'cc_user_ids' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE report_recipients ADD COLUMN cc_user_ids TEXT"))
            print("[migrate_report_recipients] Columna cc_user_ids agregada")
        except Exception as e:
            print(f"[migrate_report_recipients] error cc_user_ids: {e}")


def migrate_report_recipients_monday_stuck():
    """Agrega columna send_monday_stuck a report_recipients si no existe."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'report_recipients' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('report_recipients')}
    if 'send_monday_stuck' not in existing_cols:
        try:
            postgresql = db.engine.dialect.name == 'postgresql'
            bool_default = 'FALSE' if postgresql else '0'
            with db.engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE report_recipients ADD COLUMN send_monday_stuck BOOLEAN DEFAULT {bool_default}"))
            print("[migrate_report_recipients] Columna send_monday_stuck agregada")
        except Exception as e:
            print(f"[migrate_report_recipients] error send_monday_stuck: {e}")


def migrate_report_recipients_team():
    """Agrega columna team_user_ids a report_recipients si no existe."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'report_recipients' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('report_recipients')}
    if 'team_user_ids' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE report_recipients ADD COLUMN team_user_ids TEXT"))
            print("[migrate_report_recipients] Columna team_user_ids agregada")
        except Exception as e:
            print(f"[migrate_report_recipients] error: {e}")


def migrate_users_role_label():
    """Agrega columnas a users si no existen."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'users' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('users')}
    if 'role_label' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN role_label VARCHAR(80)"))
            print("[migrate_users] Columna role_label agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando role_label: {e}")
    if 'force_logout_at' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN force_logout_at DATETIME"))
            print("[migrate_users] Columna force_logout_at agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando force_logout_at: {e}")
    if 'extra_role_labels' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN extra_role_labels TEXT"))
            print("[migrate_users] Columna extra_role_labels agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando extra_role_labels: {e}")
    if 'must_change_password' not in existing_cols:
        try:
            # PostgreSQL requiere FALSE, SQLite acepta 0
            _bool_default = 'FALSE' if db.engine.dialect.name == 'postgresql' else '0'
            with db.engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE users ADD COLUMN must_change_password BOOLEAN DEFAULT {_bool_default}"))
            print("[migrate_users] Columna must_change_password agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando must_change_password: {e}")
    if 'failed_login_attempts' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN failed_login_attempts INTEGER DEFAULT 0"))
            print("[migrate_users] Columna failed_login_attempts agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando failed_login_attempts: {e}")
    if 'locked_until' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN locked_until DATETIME"))
            print("[migrate_users] Columna locked_until agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando locked_until: {e}")
    if 'microsoft_object_id' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN microsoft_object_id VARCHAR(100)"))
                conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_users_microsoft_object_id ON users(microsoft_object_id)"))
            print("[migrate_users] Columna microsoft_object_id agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando microsoft_object_id: {e}")
    # Campos de contacto (área, ubicación, teléfono) en users
    for col_name, col_type in (('area', 'VARCHAR(120)'), ('location', 'VARCHAR(120)'), ('phone', 'VARCHAR(40)')):
        if col_name not in existing_cols:
            try:
                with db.engine.begin() as conn:
                    conn.execute(text(f"ALTER TABLE users ADD COLUMN {col_name} {col_type}"))
                print(f"[migrate_users] Columna {col_name} agregada")
            except Exception as e:
                print(f"[migrate_users] error agregando {col_name}: {e}")
    # Espejo de tecnicos de Eliot en Pash/Primatela: referencia al usuario origen
    if 'mirrored_from_id' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE users ADD COLUMN mirrored_from_id INTEGER"))
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_users_mirrored_from_id ON users(mirrored_from_id)"))
            print("[migrate_users] Columna mirrored_from_id agregada")
        except Exception as e:
            print(f"[migrate_users] error agregando mirrored_from_id: {e}")


def migrate_mailbox_oauth():
    """Agrega columnas OAuth a mailbox_configs si no existen."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'mailbox_configs' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('mailbox_configs')}
    additions = [
        ('auth_type', "VARCHAR(20) DEFAULT 'password'"),
        ('oauth_tenant_id', 'VARCHAR(100)'),
        ('oauth_client_id', 'VARCHAR(100)'),
        ('oauth_client_secret', 'VARCHAR(500)'),
    ]
    with db.engine.begin() as conn:
        for col_name, col_type in additions:
            if col_name not in existing_cols:
                try:
                    conn.execute(text(f"ALTER TABLE mailbox_configs ADD COLUMN {col_name} {col_type}"))
                    print(f"[migrate_mailbox] Columna {col_name} agregada")
                except Exception as e:
                    print(f"[migrate_mailbox] error agregando {col_name}: {e}")


def migrate_companies_smtp():
    """Agrega columnas SMTP a companies si no existen."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'companies' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('companies')}
    # Detectar dialecto para usar DEFAULT correcto (SQLite acepta 0, PostgreSQL requiere FALSE)
    is_postgres = db.engine.dialect.name == 'postgresql'
    bool_default = 'FALSE' if is_postgres else '0'

    additions = [
        ('smtp_host', 'VARCHAR(255)'),
        ('smtp_port', 'INTEGER'),
        ('smtp_user', 'VARCHAR(255)'),
        ('smtp_password', 'VARCHAR(255)'),
        ('smtp_from', 'VARCHAR(255)'),
        ('smtp_security', 'VARCHAR(10)'),
        # Microsoft Entra ID
        ('microsoft_tenant_id', 'VARCHAR(100)'),
        ('microsoft_client_id', 'VARCHAR(100)'),
        ('microsoft_client_secret', 'VARCHAR(500)'),
        ('microsoft_enabled', f'BOOLEAN DEFAULT {bool_default}'),
    ]
    # CRITICO: usar transacción aislada por columna. En PostgreSQL, si UNA falla
    # todas las siguientes fallan (transaction aborted state).
    for col_name, col_type in additions:
        if col_name in existing_cols:
            continue
        try:
            with db.engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE companies ADD COLUMN {col_name} {col_type}"))
            print(f"[migrate_companies] Columna {col_name} agregada")
        except Exception as e:
            print(f"[migrate_companies] error agregando {col_name}: {e}")


def migrate_messages_schema():
    """Agrega subtask_id a la tabla messages si no existe."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'messages' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('messages')}
    if 'subtask_id' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE messages ADD COLUMN subtask_id INTEGER REFERENCES subtasks(id)"))
                # Índice para mejorar queries
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_messages_subtask_id ON messages(subtask_id)"))
            print("[migrate_messages] Columna subtask_id agregada")
        except Exception as e:
            print(f"[migrate_messages] no se pudo agregar subtask_id: {e}")


def migrate_tickets_schema():
    """Agrega columnas nuevas a 'tickets' si no existen."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'tickets' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('tickets')}
    additions = [
        ('sla_alerts_sent', "VARCHAR(20) DEFAULT ''"),
        ('user_area', "VARCHAR(120)"),
        ('user_location', "VARCHAR(120)"),
        ('user_phone', "VARCHAR(40)"),
        # CSAT extendido
        ('rating_comment', "TEXT"),
        ('rating_nps', "INTEGER"),
        ('rating_at', "TIMESTAMP"),
        ('reminder_sent_at', "TIMESTAMP"),
    ]
    with db.engine.begin() as conn:
        for col_name, col_type in additions:
            if col_name not in existing_cols:
                try:
                    conn.execute(text(f"ALTER TABLE tickets ADD COLUMN {col_name} {col_type}"))
                    print(f"[migrate_tickets] Columna {col_name} agregada")
                except Exception as e:
                    print(f"[migrate_tickets] no se pudo agregar {col_name}: {e}")


def migrate_templates_schema():
    """Agrega columnas nuevas a 'templates' si no existen."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'templates' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('templates')}
    if 'form_fields' not in existing_cols:
        try:
            with db.engine.begin() as conn:
                conn.execute(text("ALTER TABLE templates ADD COLUMN form_fields TEXT"))
            print("[migrate_templates] Columna form_fields agregada")
        except Exception as e:
            print(f"[migrate_templates] error: {e}")


def migrate_subtasks_schema():
    """Agrega columnas nuevas a 'subtasks' si la tabla ya existe sin ellas (SQLite)."""
    from sqlalchemy import inspect, text
    inspector = inspect(db.engine)
    if 'subtasks' not in inspector.get_table_names():
        return
    existing_cols = {c['name'] for c in inspector.get_columns('subtasks')}
    additions = [
        ('subtask_number', 'VARCHAR(40)'),
        ('category', "VARCHAR(100) DEFAULT 'General'"),
        ('priority', "VARCHAR(20) DEFAULT 'medium'"),
        ('sla_minutes', 'INTEGER'),
        ('sla_deadline', 'DATETIME'),
        ('time_worked_seconds', 'INTEGER DEFAULT 0'),
        ('updated_at', 'DATETIME'),
        ('resolved_at', 'DATETIME'),
    ]
    with db.engine.begin() as conn:
        for col_name, col_type in additions:
            if col_name not in existing_cols:
                try:
                    conn.execute(text(f"ALTER TABLE subtasks ADD COLUMN {col_name} {col_type}"))
                except Exception as e:
                    print(f"[migrate_subtasks] no se pudo agregar {col_name}: {e}")


def _form_fields_for_category(category):
    """Devuelve un set de campos de formulario adecuado según la categoría del template."""
    cat = (category or 'General').lower().strip()
    if 'red' in cat or 'network' in cat:
        return [
            {'name': 'ubicacion', 'label': '📍 Ubicación / Piso / Oficina', 'type': 'text', 'required': True, 'placeholder': 'Ej: Piso 3 - Oficina 305'},
            {'name': 'tipo_conexion', 'label': '🔌 Tipo de conexión', 'type': 'select', 'required': True, 'options': ['WiFi', 'Cable Ethernet', 'VPN', 'Ambas']},
            {'name': 'cuando_inicio', 'label': '🕐 ¿Cuándo comenzó?', 'type': 'text', 'required': True, 'placeholder': 'Ej: Hoy a las 9:30 AM'},
            {'name': 'personas_afectadas', 'label': '👥 Personas afectadas', 'type': 'select', 'required': True, 'options': ['Solo yo', '2-5 personas', '6-20 personas', 'Más de 20 / todo el piso']},
            {'name': 'sintomas', 'label': '⚠ Síntomas observados', 'type': 'textarea', 'required': True, 'placeholder': 'Ej: No carga ninguna página, error DNS, lento...'},
        ]
    if 'hardware' in cat or 'rendimiento' in cat:
        return [
            {'name': 'equipo', 'label': '💻 Equipo afectado', 'type': 'text', 'required': True, 'placeholder': 'Ej: PC-CONTAB-05 / mi laptop'},
            {'name': 'sintoma', 'label': '⚠ Síntoma principal', 'type': 'select', 'required': True,
                'options': ['Muy lento', 'No enciende', 'Se reinicia / pantalla azul', 'Pantalla dañada', 'Ruido / sobrecalentamiento', 'Teclado / mouse no responden', 'Disco lleno', 'Otro']},
            {'name': 'cuando_inicio', 'label': '🕐 ¿Desde cuándo?', 'type': 'text', 'required': True, 'placeholder': 'Ej: Desde ayer / desde el último Windows Update'},
            {'name': 'frecuencia', 'label': '🔁 Frecuencia', 'type': 'select', 'required': False, 'options': ['Constante', 'Varias veces al día', 'Una vez al día', 'Esporádico']},
            {'name': 'detalles', 'label': '📝 Detalles adicionales', 'type': 'textarea', 'required': False},
        ]
    if 'software' in cat:
        return [
            {'name': 'aplicacion', 'label': '💻 Aplicación / programa', 'type': 'text', 'required': True, 'placeholder': 'Ej: Excel, Chrome, SAP GUI'},
            {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True,
                'options': ['No abre / se cierra', 'Error al usar', 'Funcionalidad faltante', 'Instalación / actualización', 'Configuración', 'Otro']},
            {'name': 'mensaje_error', 'label': '💬 Mensaje de error (si aplica)', 'type': 'textarea', 'required': False, 'placeholder': 'Copia/pega el mensaje exacto'},
            {'name': 'pasos', 'label': '📝 Pasos para reproducir', 'type': 'textarea', 'required': False, 'placeholder': '1. Abrir... 2. Click en... 3. Aparece error'},
        ]
    if 'email' in cat or 'correo' in cat or 'outlook' in cat:
        return [
            {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True,
                'options': ['No abre Outlook / se cierra', 'No envío correos', 'No recibo correos', 'Buzón lleno', 'Configurar firma', 'Respuesta automática', 'Otro']},
            {'name': 'cuenta', 'label': '📧 Cuenta afectada', 'type': 'text', 'required': True, 'placeholder': 'Tu email corporativo'},
            {'name': 'mensaje_error', 'label': '💬 Mensaje de error (si aplica)', 'type': 'text', 'required': False},
            {'name': 'detalles', 'label': '📝 Detalles', 'type': 'textarea', 'required': False},
        ]
    if 'impres' in cat or 'printer' in cat:
        return [
            {'name': 'impresora', 'label': '🖨️ Nombre / modelo de impresora', 'type': 'text', 'required': True, 'placeholder': 'Ej: HP-LaserJet-Piso3'},
            {'name': 'ubicacion', 'label': '📍 Ubicación', 'type': 'text', 'required': True},
            {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True,
                'options': ['No imprime', 'Imprime mal', 'Atasco de papel', 'Sin tóner / tinta', 'Instalar', 'Otro']},
            {'name': 'mensaje_error', 'label': '💬 Mensaje de error en pantalla', 'type': 'text', 'required': False},
        ]
    if 'sap' in cat:
        return [
            {'name': 'transaccion', 'label': '🔤 Transacción / Módulo SAP', 'type': 'text', 'required': True, 'placeholder': 'Ej: ME21N, FB60, VA01, Módulo FI'},
            {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True,
                'options': ['Error al ejecutar', 'Sin autorización', 'Sistema lento', 'Resultado incorrecto', 'Solicitud de acceso', 'Dump ABAP']},
            {'name': 'mensaje_error', 'label': '💬 Mensaje de error exacto', 'type': 'textarea', 'required': False},
            {'name': 'pasos', 'label': '📝 Pasos ejecutados', 'type': 'textarea', 'required': True, 'placeholder': '1. Abrir TX...\n2. Ingresar...\n3. Click en...'},
            {'name': 'mandante', 'label': '🌐 Mandante / Sistema (SID)', 'type': 'text', 'required': False, 'placeholder': 'Ej: PRD-100'},
        ]
    if 'acceso' in cat or 'access' in cat or 'contras' in cat or 'password' in cat:
        return [
            {'name': 'sistema', 'label': '🎯 Sistema / aplicación / recurso', 'type': 'text', 'required': True, 'placeholder': 'Ej: SAP módulo FI, carpeta de red, SharePoint'},
            {'name': 'tipo_acceso', 'label': '🔐 Tipo de acceso', 'type': 'select', 'required': True, 'options': ['Solo lectura', 'Lectura y escritura', 'Administrador', 'Reset de contraseña', 'Otro']},
            {'name': 'justificacion', 'label': '📋 Justificación', 'type': 'textarea', 'required': True, 'placeholder': 'Explica para qué necesitas este acceso'},
            {'name': 'aprobador', 'label': '✅ Jefe / aprobador', 'type': 'text', 'required': True, 'placeholder': 'Nombre y cargo de tu jefe'},
            {'name': 'duracion', 'label': '⏳ Duración', 'type': 'select', 'required': False, 'options': ['Permanente', 'Temporal (indicar fecha en detalles)']},
        ]
    if 'seguridad' in cat or 'security' in cat:
        return [
            {'name': 'tipo_incidente', 'label': '🛡️ Tipo de incidente', 'type': 'select', 'required': True,
                'options': ['Phishing / correo sospechoso', 'Virus / malware detectado', 'Posible robo de credenciales', 'Acceso no autorizado', 'Archivo cifrado (ransomware)', 'USB sospechoso', 'Otro']},
            {'name': 'sistema_afectado', 'label': '💻 Sistema / equipo afectado', 'type': 'text', 'required': True, 'placeholder': 'Ej: PC del usuario X, cuenta de correo, servidor'},
            {'name': 'cuando_detectado', 'label': '🕐 ¿Cuándo lo detectaste?', 'type': 'text', 'required': True},
            {'name': 'acciones_tomadas', 'label': '🚨 Acciones que tomaste', 'type': 'textarea', 'required': False, 'placeholder': 'Ej: Desconecté la red, cerré sesión, no hice click...'},
            {'name': 'detalles', 'label': '📝 Detalles del incidente', 'type': 'textarea', 'required': True},
        ]
    if 'servidor' in cat or 'server' in cat or 'infraestructura' in cat:
        return [
            {'name': 'servidor', 'label': '🖧 Servidor / sistema afectado', 'type': 'text', 'required': True, 'placeholder': 'Ej: SAP-PRD, AD, Exchange'},
            {'name': 'sintoma', 'label': '⚠ Síntoma observado', 'type': 'select', 'required': True, 'options': ['Caído / no responde', 'Lento', 'Errores intermitentes', 'No accesible desde red', 'Otro']},
            {'name': 'cuando_inicio', 'label': '🕐 ¿Cuándo comenzó?', 'type': 'text', 'required': True},
            {'name': 'usuarios_afectados', 'label': '👥 Cantidad de usuarios afectados', 'type': 'text', 'required': False, 'placeholder': 'Estimado'},
            {'name': 'detalles', 'label': '📝 Detalles', 'type': 'textarea', 'required': False},
        ]
    if 'telefon' in cat or 'phone' in cat:
        return [
            {'name': 'extension', 'label': '📞 Extensión / número', 'type': 'text', 'required': True, 'placeholder': 'Ej: 1234'},
            {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True, 'options': ['No tiene tono', 'No suena', 'Calidad de audio mala', 'No me llaman / no me escuchan', 'Configurar desvío', 'Buzón de voz', 'Otro']},
            {'name': 'cuando_inicio', 'label': '🕐 ¿Cuándo comenzó?', 'type': 'text', 'required': True},
            {'name': 'detalles', 'label': '📝 Detalles', 'type': 'textarea', 'required': False},
        ]
    if 'backup' in cat or 'respaldo' in cat:
        return [
            {'name': 'tipo_solicitud', 'label': '🗄 Tipo de solicitud', 'type': 'select', 'required': True, 'options': ['Restaurar archivos eliminados', 'Restaurar buzón / correos', 'Backup de equipo', 'Verificar que existe backup', 'Otro']},
            {'name': 'archivos_carpeta', 'label': '📁 Archivos / carpeta afectada', 'type': 'text', 'required': True, 'placeholder': 'Ruta completa o descripción'},
            {'name': 'cuando_perdio', 'label': '🕐 ¿Cuándo se perdió / borró?', 'type': 'text', 'required': True},
            {'name': 'detalles', 'label': '📝 Detalles', 'type': 'textarea', 'required': False},
        ]
    # Default genérico (General + cualquier no mapeada)
    return [
        {'name': 'que_pasa', 'label': '❓ ¿Qué está pasando?', 'type': 'textarea', 'required': True, 'placeholder': 'Describe el problema con tus palabras'},
        {'name': 'equipo_sistema', 'label': '💻 Equipo / sistema afectado', 'type': 'text', 'required': False, 'placeholder': 'Ej: Mi laptop, SAP, impresora del piso 2'},
        {'name': 'cuando_inicio', 'label': '🕐 ¿Cuándo comenzó?', 'type': 'text', 'required': False, 'placeholder': 'Ej: Hoy a las 9am / desde ayer'},
        {'name': 'pasos_intentados', 'label': '🔄 ¿Qué intentaste para resolverlo?', 'type': 'textarea', 'required': False},
        {'name': 'impacto', 'label': '📊 ¿Cómo te afecta?', 'type': 'select', 'required': False, 'options': ['Me bloquea totalmente', 'Puedo seguir trabajando con limitaciones', 'Es molestia pero no crítico', 'Solo es una consulta']},
    ]


def _parse_description_to_fields(description_template, category=None):
    """Parsea una description_template estructurada y extrae form_fields.
    Detecta:
    - "Etiqueta:" → campo de texto
    - "Etiqueta: [hint]" → campo de texto con placeholder
    - "[ ] opción1\\n[ ] opción2" → campo select con esas opciones (la pregunta es la línea anterior)
    - "¿pregunta? [Sí/No]" → campo select Sí/No
    """
    import re
    if not description_template:
        return None  # caer al genérico

    lines = description_template.split('\n')
    fields = []
    used_names = set()

    def slug(label):
        s = re.sub(r'[^a-zA-Z0-9áéíóúñ ]', '', label).strip().lower()
        s = re.sub(r'\s+', '_', s)[:40]
        if not s or s in used_names:
            s = (s + '_' + str(len(fields))) if s else f'campo_{len(fields)}'
        used_names.add(s)
        return s

    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue

        # Caso 1: ¿pregunta? [Sí/No]
        m_sino = re.match(r'^(.+?\?)\s*\[\s*S[íi]\s*/?\s*No\s*\]\s*$', line, re.IGNORECASE)
        if m_sino:
            label = m_sino.group(1).strip()
            fields.append({
                'name': slug(label), 'label': label, 'type': 'select',
                'required': False, 'options': ['Sí', 'No']
            })
            i += 1
            continue

        # Caso 2: línea termina con ":" y la siguiente empieza con "[ ]" → es un select multi-opción
        if line.endswith(':') and i + 1 < len(lines):
            next_line = lines[i+1].strip()
            if next_line.startswith('[ ]') or next_line.startswith('[]'):
                label = line.rstrip(':').strip()
                options = []
                j = i + 1
                while j < len(lines):
                    candidate = lines[j].strip()
                    if not candidate:
                        j += 1
                        break
                    if candidate.startswith('[ ]') or candidate.startswith('[]'):
                        opt = re.sub(r'^\[\s*\]\s*', '', candidate).strip()
                        # Limpiar opciones tipo "Otro: " → "Otro"
                        opt = re.sub(r':\s*$', '', opt).strip()
                        if opt:
                            options.append(opt)
                        j += 1
                    else:
                        break
                if options:
                    fields.append({
                        'name': slug(label), 'label': label, 'type': 'select',
                        'required': False, 'options': options
                    })
                    i = j
                    continue

        # Caso 3: "Etiqueta: [placeholder]" o "Etiqueta:" o "Etiqueta: valor"
        m_kv = re.match(r'^([^:\[]+?):\s*(\[(.+?)\])?\s*(.*)$', line)
        if m_kv:
            label = m_kv.group(1).strip()
            placeholder = (m_kv.group(3) or '').strip()
            if not placeholder and m_kv.group(4):
                placeholder = m_kv.group(4).strip()
            # Saltar si el label es muy genérico/decorativo
            if len(label) < 2 or label.startswith('-'):
                i += 1
                continue
            # Detectar si parece textarea (palabras como "descripción", "detalle", "mensaje", "pasos", "motivo")
            label_lower = label.lower()
            is_textarea = any(w in label_lower for w in ['descripción','descripcion','detalle','mensaje','pasos','motivo','observacion','observación','comentario','justificación','justificacion','síntoma','sintoma'])
            # Detectar si parece fecha
            is_date = 'fecha' in label_lower and 'desde' not in label_lower
            # Detectar si la línea es decorativa o título (encabezado tipo "Descripción del problema:")
            if not placeholder and label.endswith('problema') and i + 1 < len(lines):
                # Es un título de sección, no un campo
                i += 1
                continue
            ftype = 'textarea' if is_textarea else ('date' if is_date else 'text')
            fields.append({
                'name': slug(label),
                'label': label,
                'type': ftype,
                'required': False,
                'placeholder': placeholder if placeholder and placeholder != label else ''
            })
            i += 1
            continue

        # Caso 4: pregunta abierta sin estructura clara → la ignoramos (irá al campo libre)
        i += 1

    # Heurística: si extrajimos menos de 2 campos, mejor usar el genérico de categoría
    if len(fields) < 2:
        return None

    # Marcar el primer campo como required (suele ser el más importante)
    if fields and not any(f.get('required') for f in fields):
        fields[0]['required'] = True

    return fields


def convert_legacy_templates_to_forms(force=False):
    """Convierte todas las plantillas existentes en plantillas tipo formulario.
    Primero intenta parsear el description_template de cada plantilla para extraer campos
    específicos. Si no hay suficiente estructura, cae al genérico por categoría.
    Idempotente: por defecto solo toca las que NO tienen form_fields ya cargados.
    Si force=True, regenera todas las plantillas is_system."""
    import json
    if force:
        legacy = Template.query.filter(Template.is_system == True).all()
    else:
        legacy = Template.query.filter(
            (Template.form_fields == None) | (Template.form_fields == '') | (Template.form_fields == '[]')
        ).all()
    if not legacy:
        return 0

    parsed_count = 0
    fallback_count = 0
    for t in legacy:
        # Intentar parseo específico desde description_template
        parsed = _parse_description_to_fields(t.description_template, t.category)
        if parsed:
            t.form_fields = json.dumps(parsed, ensure_ascii=False)
            parsed_count += 1
        else:
            # Fallback genérico por categoría
            fields = _form_fields_for_category(t.category)
            t.form_fields = json.dumps(fields, ensure_ascii=False)
            fallback_count += 1
    db.session.commit()
    print(f"[init_db] {parsed_count + fallback_count} plantillas con formulario "
          f"({parsed_count} parseadas específicamente, {fallback_count} genéricas por categoría)")
    return parsed_count + fallback_count


def seed_default_templates():
    """Pre-carga plantillas de ticket tipo formulario para cada empresa (idempotente)."""
    import json
    templates_defs = [
        {
            'name': '🌐 Problema de Red / Internet',
            'description': 'Reportar fallas de conectividad',
            'title_template': 'Sin conexión a red — {ubicacion}',
            'category': 'Red',
            'priority': 'high',
            'form_fields': [
                {'name': 'ubicacion', 'label': '📍 Ubicación / Piso / Oficina', 'type': 'text', 'required': True, 'placeholder': 'Ej: Piso 3 — Oficina 305'},
                {'name': 'tipo_conexion', 'label': '🔌 Tipo de conexión', 'type': 'select', 'required': True,
                    'options': ['WiFi', 'Cable Ethernet', 'VPN', 'Ambas']},
                {'name': 'cuando_inicio', 'label': '🕐 ¿Cuándo comenzó el problema?', 'type': 'text', 'required': True, 'placeholder': 'Ej: Hoy a las 9:30 AM'},
                {'name': 'cuantas_personas', 'label': '👥 ¿Cuántas personas afectadas?', 'type': 'select', 'required': True,
                    'options': ['Solo yo', '2-5 personas', '6-20 personas', 'Más de 20 / todo el piso']},
                {'name': 'sintomas', 'label': '⚠ Síntomas observados', 'type': 'textarea', 'required': True, 'placeholder': 'Ej: No carga ninguna página, error DNS, lento intermitente...'},
                {'name': 'pasos_intentados', 'label': '🔄 ¿Qué intentaste para resolverlo?', 'type': 'textarea', 'required': False, 'placeholder': 'Ej: Reiniciar el equipo, olvidar la red WiFi...'},
            ]
        },
        {
            'name': '🖨️ Problema con Impresora',
            'description': 'Reportar fallas o configuración de impresora',
            'title_template': 'Problema con impresora — {impresora}',
            'category': 'Impresoras',
            'priority': 'medium',
            'form_fields': [
                {'name': 'impresora', 'label': '🖨️ Nombre / modelo de la impresora', 'type': 'text', 'required': True, 'placeholder': 'Ej: HP-LaserJet-Piso3 / Xerox WorkCentre 7855'},
                {'name': 'ubicacion', 'label': '📍 Ubicación de la impresora', 'type': 'text', 'required': True},
                {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True,
                    'options': ['No imprime', 'Imprime mal (manchas, líneas, colores)', 'Atasco de papel', 'Sin tóner / tinta', 'Necesito instalarla', 'Otra']},
                {'name': 'mensaje_error', 'label': '💬 Mensaje de error en pantalla', 'type': 'text', 'required': False, 'placeholder': 'Si aparece algún error, copialo aquí'},
                {'name': 'urgencia', 'label': '⏱ ¿Es urgente?', 'type': 'textarea', 'required': False, 'placeholder': 'Si bloquea un proceso importante, explica brevemente'},
            ]
        },
        {
            'name': '📊 Problema en SAP',
            'description': 'Reportar errores o solicitar acceso en SAP',
            'title_template': 'SAP — {tipo_problema} en {transaccion}',
            'category': 'SAP',
            'priority': 'high',
            'form_fields': [
                {'name': 'transaccion', 'label': '🔤 Transacción o Módulo SAP', 'type': 'text', 'required': True, 'placeholder': 'Ej: ME21N, FB60, VA01, Módulo FI'},
                {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True,
                    'options': ['Error al ejecutar', 'Sin autorización', 'Sistema lento', 'Resultado incorrecto', 'Solicitud de acceso', 'Dump ABAP']},
                {'name': 'mensaje_error', 'label': '💬 Mensaje de error exacto (copia/pega)', 'type': 'textarea', 'required': False, 'placeholder': 'Ej: SP_ERROR / Number range exhausted / Access denied for object S_TCODE'},
                {'name': 'pasos', 'label': '📝 Pasos que ejecutaste', 'type': 'textarea', 'required': True, 'placeholder': '1. Abrir TX VA01\n2. Ingresar tipo de pedido OR\n3. Click en continuar → aparece error'},
                {'name': 'mandante', 'label': '🌐 Mandante / Sistema (SID)', 'type': 'text', 'required': False, 'placeholder': 'Ej: PRD-100, QAS-200'},
            ]
        },
        {
            'name': '🔑 Solicitud de Acceso',
            'description': 'Solicitar acceso a un sistema, carpeta o aplicación',
            'title_template': 'Solicitud de acceso a {sistema}',
            'category': 'Accesos',
            'priority': 'medium',
            'form_fields': [
                {'name': 'sistema', 'label': '🎯 Sistema / aplicación / recurso', 'type': 'text', 'required': True, 'placeholder': 'Ej: SAP módulo FI, carpeta \\\\server\\finanzas, sitio SharePoint'},
                {'name': 'tipo_acceso', 'label': '🔐 Tipo de acceso requerido', 'type': 'select', 'required': True,
                    'options': ['Solo lectura', 'Lectura y escritura', 'Administrador', 'Otro']},
                {'name': 'justificacion', 'label': '📋 Justificación de negocio', 'type': 'textarea', 'required': True, 'placeholder': 'Explica para qué necesitas este acceso'},
                {'name': 'aprobador', 'label': '✅ Jefe / aprobador', 'type': 'text', 'required': True, 'placeholder': 'Nombre y cargo de tu jefe inmediato'},
                {'name': 'duracion', 'label': '⏳ Duración', 'type': 'select', 'required': True,
                    'options': ['Permanente', 'Temporal (especificar fecha fin en descripción)']},
            ]
        },
        {
            'name': '💻 Problema con Computador / Laptop',
            'description': 'Hardware / desempeño del equipo',
            'title_template': 'Equipo lento o con falla — {usuario_equipo}',
            'category': 'Hardware',
            'priority': 'medium',
            'form_fields': [
                {'name': 'usuario_equipo', 'label': '💻 Nombre del equipo o usuario', 'type': 'text', 'required': True, 'placeholder': 'Ej: PC-CONTAB-05 / mi laptop personal'},
                {'name': 'sintoma_principal', 'label': '⚠ Síntoma principal', 'type': 'select', 'required': True,
                    'options': ['Muy lento al iniciar / trabajar', 'Pantalla azul / se reinicia', 'No enciende', 'Disco lleno', 'Ruido extraño / sobrecalentamiento', 'Pantalla quebrada / dañada', 'Teclado / mouse no responden', 'Otro']},
                {'name': 'frecuencia', 'label': '🔁 Frecuencia', 'type': 'select', 'required': True,
                    'options': ['Constante', 'Varias veces al día', 'Una vez al día', 'Esporádico']},
                {'name': 'cuando_inicio', 'label': '🕐 ¿Desde cuándo?', 'type': 'text', 'required': True, 'placeholder': 'Ej: Desde ayer / desde el último Windows Update'},
                {'name': 'detalles', 'label': '📝 Detalles adicionales', 'type': 'textarea', 'required': False},
            ]
        },
        {
            'name': '📧 Problema con Email / Outlook',
            'description': 'Fallas de correo o configuración',
            'title_template': 'Outlook — {tipo_problema}',
            'category': 'Email',
            'priority': 'medium',
            'form_fields': [
                {'name': 'tipo_problema', 'label': '⚠ Tipo de problema', 'type': 'select', 'required': True,
                    'options': ['No abre Outlook / se cierra', 'No envío correos (se quedan en bandeja salida)', 'No recibo correos', 'Buzón lleno', 'Configurar firma', 'Configurar respuesta automática', 'Otro']},
                {'name': 'cuenta', 'label': '📧 Cuenta afectada', 'type': 'text', 'required': True, 'placeholder': 'Tu email corporativo'},
                {'name': 'mensaje_error', 'label': '💬 Mensaje de error (si aplica)', 'type': 'text', 'required': False},
                {'name': 'detalles', 'label': '📝 Detalles', 'type': 'textarea', 'required': False, 'placeholder': '¿Cuándo empezó? ¿Qué intentaste?'},
            ]
        },
    ]

    # Crear las plantillas para cada empresa (excepto la master que las comparte)
    companies = [c.code for c in Company.query.all()] or ['eliot', 'pash', 'primatela']
    created_total = 0
    for company in companies:
        for tdef in templates_defs:
            exists = Template.query.filter_by(name=tdef['name'], company=company).first()
            if exists:
                # Si existe pero no tiene form_fields, actualizar
                if not exists.form_fields and tdef.get('form_fields'):
                    exists.form_fields = json.dumps(tdef['form_fields'], ensure_ascii=False)
                continue
            db.session.add(Template(
                name=tdef['name'],
                description=tdef.get('description', ''),
                title_template=tdef['title_template'],
                description_template='',  # Se construirá desde form_fields al enviar
                category=tdef.get('category', 'General'),
                priority=tdef.get('priority', 'medium'),
                company=company,
                is_system=True,
                form_fields=json.dumps(tdef.get('form_fields', []), ensure_ascii=False)
            ))
            created_total += 1
    if created_total > 0:
        db.session.commit()
        print(f"[init_db] {created_total} plantillas tipo formulario creadas")


def seed_default_subroles():
    """Pre-carga el catálogo de subroles del sistema (idempotente)."""
    defaults = [
        ('Infraestructura', '🖧', 'Servidores, red, almacenamiento, virtualización'),
        ('SAP ABAP', '👨‍💻', 'Desarrollo ABAP y custom development en SAP'),
        ('Seguridad Tecnológica', '🛡️', 'Seguridad informática, accesos, auditoría'),
        ('Procesos SAP - Módulo SD', '🛒', 'Ventas y distribución'),
        ('Procesos SAP - Módulo MM', '📦', 'Materiales y compras'),
        ('Procesos SAP - Módulo PM', '🔧', 'Mantenimiento de planta'),
        ('Procesos SAP - Módulo CO', '📊', 'Controlling / contabilidad de costos'),
        ('Procesos SAP - Módulo WM', '🏬', 'Gestión de almacén'),
        ('Procesos SAP - Módulo EWM', '🏭', 'Extended Warehouse Management'),
        ('Procesos SAP - Módulo TI', '🚚', 'Transportation Integration'),
        ('Procesos SAP - Módulo FI', '💰', 'Finanzas'),
        ('Procesos SAP - Módulo TM', '🚛', 'Transportation Management'),
        ('Soporte', '🎧', 'Mesa de ayuda y atención al usuario'),
        ('Procesos IA', '🤖', 'Inteligencia artificial y agentes automáticos'),
    ]
    created = 0
    for name, icon, desc in defaults:
        if not Subrole.query.filter_by(name=name, company=None).first():
            db.session.add(Subrole(
                name=name, icon=icon, description=desc,
                company=None, is_system=True, is_active=True
            ))
            created += 1
    if created > 0:
        db.session.commit()
        print(f"[init_db] {created} subroles del sistema creados")


def init_db():
    """Inicializa la base de datos"""
    with app.app_context():
        db.create_all()
        migrate_users_schema()
        migrate_users_role_label()
        migrate_companies_smtp()
        migrate_mailbox_oauth()
        migrate_tickets_schema()
        migrate_templates_schema()
        try:
            migrate_messages_schema()
        except Exception as _e:
            print(f"[migrate] messages_schema: {_e}")
        try:
            migrate_report_recipients_team()
        except Exception as _e:
            print(f"[migrate] report_recipients_team: {_e}")
        try:
            migrate_report_recipients_cc()
        except Exception as _e:
            print(f"[migrate] report_recipients_cc: {_e}")
        try:
            migrate_report_recipients_monday_stuck()
        except Exception as _e:
            print(f"[migrate] report_recipients_monday_stuck: {_e}")
        migrate_subtasks_schema()
        backfill_subtask_numbers()
        seed_default_subroles()
        seed_default_templates()
        convert_legacy_templates_to_forms()

        # Crear empresas si no existen
        if not Company.query.first():
            companies = [
                Company(
                    code='eliot',
                    name='Manufacturas Eliot',
                    icon='🏭',
                    primary_color='#2563eb',
                    secondary_color='#1e40af'
                ),
                Company(
                    code='pash',
                    name='Pash Technologies',
                    icon='🏢',
                    primary_color='#7c3aed',
                    secondary_color='#6d28d9'
                ),
                Company(
                    code='primatela',
                    name='Primatela Solutions',
                    icon='🌴',
                    primary_color='#059669',
                    secondary_color='#047857'
                ),
            ]
            for company in companies:
                db.session.add(company)
            db.session.commit()

        # Crear base de conocimiento del bot si no existe (antes de retornar)
        if not BotKnowledge.query.first():
            bot_data = [
                ('wifi, red, conectar, conexion', '¿No tienes conexión WiFi?',
                 'Intenta lo siguiente:\n1. Reinicia tu router (apágalo 30 segundos)\n2. Verifica que estés escribiendo bien la contraseña\n3. Acércate más al router\n4. Reinicia tu dispositivo\nSi persiste, crea un ticket para que nuestro equipo revise.',
                 'Red', 'medium'),
                ('contraseña, olvide, reset, acceso', '¿Olvidaste tu contraseña?',
                 'Puedes resetear tu contraseña directamente:\n1. Ve a la página de login\n2. Haz clic en "¿Olvidaste tu contraseña?"\n3. Sigue las instrucciones enviadas a tu correo\nSi no recibiste el email, abre un ticket.',
                 'Autenticación', 'low'),
                ('impresora, no imprime, driver', '¿Problemas con la impresora?',
                 'Soluciona problemas de impresión:\n1. Verifica que la impresora esté encendida\n2. Abre "Dispositivos" > "Impresoras" y ve si aparece\n3. Intenta imprimir un documento de prueba\n4. Si falta el driver, descárgalo del sitio del fabricante\nSi sigue sin funcionar, crea un ticket.',
                 'Hardware', 'medium'),
                ('slow, lento, rendimiento, lag', '¿Tu computadora está lenta?',
                 'Mejora el rendimiento:\n1. Cierra programas que no uses\n2. Libera espacio en disco (elimina archivos temporales)\n3. Reinicia tu equipo\n4. Verifica que tu antivirus esté actualizado\n5. Abre el Administrador de tareas y cierra procesos innecesarios\nSi persiste después de reiniciar, crea un ticket.',
                 'Rendimiento', 'medium'),
                ('email, correo, outlook', '¿Problemas con el correo?',
                 'Soluciona problemas de email:\n1. Verifica tu conexión a internet\n2. Reinicia Outlook o tu cliente de correo\n3. Sincroniza manualmente tu bandeja\n4. Verifica que tu contraseña sea correcta\nSi el problema continúa, crea un ticket con detalles del error.',
                 'Email', 'medium'),
                ('vpn, remoto, conexion remota', '¿No puedes conectarte a la VPN?',
                 'Conectarse a la VPN:\n1. Asegúrate de tener la VPN instalada\n2. Abre el cliente VPN\n3. Introduce tu usuario y contraseña de dominio\n4. Si falla, reinicia el cliente VPN\n5. Verifica que tu conexión a internet sea estable\nPara problemas de VPN, contacta al equipo de infraestructura.',
                 'Conectividad', 'high'),
                ('software, aplicacion, programa', '¿No funciona una aplicación?',
                 'Pasos para resolver problemas de software:\n1. Reinicia la aplicación\n2. Reinicia tu computadora completamente\n3. Verifica que tengas la versión más reciente instalada\n4. Actualiza los drivers de tu tarjeta gráfica\nSi persiste, crea un ticket incluyendo el nombre del programa y el error exacto.',
                 'Software', 'medium'),
                ('actualizacion, update, patch', '¿Necesitas actualizar tu sistema?',
                 'Actualizar Windows:\n1. Ve a Configuración > Actualización y seguridad\n2. Haz clic en "Buscar actualizaciones"\n3. Descarga e instala las actualizaciones disponibles\n4. Reinicia tu equipo si es necesario\nNo desactives las actualizaciones automáticas; es importante para la seguridad.',
                 'Sistema', 'medium'),
            ]

            for keywords, question, answer, category, priority in bot_data:
                kb = BotKnowledge(keywords=keywords, question=question, answer=answer,
                                category=category, priority=priority)
                db.session.add(kb)
            db.session.commit()

        # Crear usuarios demo de las 3 empresas si faltan (idempotente por empresa)
        user_template = [
            ('john', 'John', 'john', 'employee'),
            ('carlos', 'Carlos', 'carlos', 'technician'),
            ('ana', 'Ana', 'ana', 'admin'),
        ]

        companies_data = [
            ('eliot', 'Smith'),
            ('pash', 'Pash'),
            ('primatela', 'Primatela'),
        ]

        created_count = 0
        for company, lastname in companies_data:
            for username, firstname, email_prefix, role in user_template:
                # Verificar si ya existe ese usuario en esa empresa específica
                existing = User.query.filter_by(username=username, company=company).first()
                if existing:
                    continue
                # Hash de la contraseña temporal — el usuario debe cambiarla en su primer login
                _bootstrap_pw = os.getenv('BOOTSTRAP_PASSWORD', 'DeskEli2026!')
                _bootstrap_hash = hashlib.pbkdf2_hmac(
                    'sha256', _bootstrap_pw.encode(), username.encode(), 100000
                ).hex()
                # Dominio configurable; default: dominio corporativo PatPrimo
                _seed_domain = os.getenv('SEED_EMAIL_DOMAIN', 'patprimo.com.co')
                db.session.add(User(
                    username=username,
                    name=f'{firstname} {lastname}',
                    email=f'{email_prefix}@{_seed_domain}',
                    role=role,
                    company=company,
                    password_hash=_bootstrap_hash,
                    must_change_password=True,
                    is_active=True,
                ))
                created_count += 1

        if created_count > 0:
            db.session.commit()
            print(f"[init_db] {created_count} usuarios demo creados")

        # Si los usuarios ya existían todos, saltar la creación de tickets demo
        if created_count == 0:
            return

        # Config env para saltearse los tickets demo en produccion
        # (defualt true para bootstrap inicial; setear false despues del primer deploy)
        if os.getenv('SEED_DEMO_TICKETS', 'true').lower() in ('false', '0', 'no'):
            print('[init_db] SEED_DEMO_TICKETS=false → no se crean tickets demo')
            return

        # Doble guarda: si ya existen tickets en alguna empresa, no re-seedear
        # (evita que tickets demo re-aparezcan tras un wipe manual)
        if Ticket.query.count() > 0:
            print('[init_db] Ya hay tickets en la BD → skip demo tickets')
            return

        # Configuración inicial
        defaults = [
            Config(key='sla_low', value='480'),
            Config(key='sla_medium', value='240'),
            Config(key='sla_high', value='120'),
            Config(key='sla_critical', value='60'),
            Config(key='theme', value='blue')
        ]

        for config in defaults:
            if not Config.query.filter_by(key=config.key).first():
                db.session.add(config)
        db.session.commit()

        # Crear tickets de ejemplo (3 empresas)
        tickets_data = [
            # Eliot
            Ticket(ticket_number='TKT-ELIOT-00001', title='WiFi no conecta', description='La red WiFi no funciona',
                   priority='high', creator_id=1, assignee_id=2, status='in_progress', company='eliot',
                   sla_minutes=120, sla_deadline=datetime.now() + timedelta(minutes=120)),
            Ticket(ticket_number='TKT-ELIOT-00002', title='Driver de impresora', description='No puedo imprimir',
                   priority='medium', creator_id=1, status='open', company='eliot',
                   sla_minutes=240, sla_deadline=datetime.now() + timedelta(minutes=240)),
            # Pash
            Ticket(ticket_number='TKT-PASH-00001', title='Base de datos lenta', description='La BD responde lentamente',
                   priority='high', creator_id=4, assignee_id=5, status='in_progress', company='pash',
                   sla_minutes=120, sla_deadline=datetime.now() + timedelta(minutes=120)),
            Ticket(ticket_number='TKT-PASH-00002', title='Contraseña olvidada', description='No puedo acceder a mi cuenta',
                   priority='low', creator_id=4, status='open', company='pash',
                   sla_minutes=480, sla_deadline=datetime.now() + timedelta(minutes=480)),
            # Primatela
            Ticket(ticket_number='TKT-PRIMATELA-00001', title='Licencia software expirada', description='La licencia de Office expiró',
                   priority='critical', creator_id=7, assignee_id=8, status='in_progress', company='primatela',
                   sla_minutes=60, sla_deadline=datetime.now() + timedelta(minutes=60)),
            Ticket(ticket_number='TKT-PRIMATELA-00002', title='Backup fallido', description='El respaldo no se completó',
                   priority='high', creator_id=7, status='open', company='primatela',
                   sla_minutes=120, sla_deadline=datetime.now() + timedelta(minutes=120)),
        ]

        for ticket in tickets_data:
            if not Ticket.query.filter_by(ticket_number=ticket.ticket_number).first():
                db.session.add(ticket)
        db.session.commit()

        # Crear plantillas predefinidas (8 sistema + personalizables)
        templates_data = [
            ('Problema de acceso', 'Usuario no puede acceder al sistema', 'No puedo ingresar a [SISTEMA]', 'Usuario reporta que no puede acceder', 'Acceso', 'medium', 'eliot'),
            ('Problema de velocidad', 'Sistema lento o con lag', '[SISTEMA] está muy lento', 'El sistema responde lentamente', 'Rendimiento', 'medium', 'eliot'),
            ('Error de licencia', 'Licencia de software expirada', 'Licencia expirada de [SOFTWARE]', 'Necesito renovar la licencia', 'Software', 'high', 'eliot'),
            ('Backup fallido', 'Respaldo de datos no completado', 'Backup de [SISTEMA] falló', 'El respaldo no se completó correctamente', 'Backup', 'high', 'eliot'),
            ('Hardware defectuoso', 'Problema con equipo físico', '[EQUIPO] no funciona', 'El equipo está defectuoso', 'Hardware', 'high', 'eliot'),
            ('Configuración de red', 'Problema de conectividad', 'No tengo conexión a [RED]', 'No puedo conectarme a la red', 'Red', 'medium', 'eliot'),
            ('Reseteo de contraseña', 'Usuario olvidó su clave', 'Necesito resetear mi contraseña', 'Olvidé mi contraseña de [SISTEMA]', 'Autenticación', 'low', 'eliot'),
            ('Reporte de problema', 'Reporte genérico de problema', 'Problema con [ÁREA]', 'Describe el problema que enfrentas', 'General', 'low', 'eliot'),
        ]

        for name, desc, title, desc_templ, cat, prio, comp in templates_data:
            if not Template.query.filter_by(name=name, company=comp).first():
                template = Template(name=name, description=desc, title_template=title,
                                  description_template=desc_templ, category=cat,
                                  priority=prio, company=comp, is_system=True)
                db.session.add(template)
        db.session.commit()

# ═════════════════════════════════════════════════════════════════════════════
# ADMIN ENDPOINTS - CRUD DE TICKETS
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/tickets/create', methods=['POST'])
def api_admin_create_ticket():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    data = request.get_json()
    user = User.query.get(session['user_id'])

    sla_config = Config.query.filter_by(key=f"sla_{data.get('priority', 'medium')}").first()
    sla_minutes = int(sla_config.value) if sla_config else 120

    ticket = Ticket(
        ticket_number=get_next_ticket_number(user.company),
        title=data.get('title'),
        description=data.get('description'),
        category=data.get('category', 'General'),
        priority=data.get('priority', 'medium'),
        creator_id=session['user_id'],
        company=user.company,
        sla_minutes=sla_minutes,
        sla_deadline=compute_sla_deadline(datetime.now(), sla_minutes, user.company)
    )

    if data.get('assignee_id'):
        ticket.assignee_id = data.get('assignee_id')

    db.session.add(ticket)
    db.session.commit()

    # Hook del Agent Orchestrator
    orch = app.config.get('orchestrator')
    if orch is not None:
        try:
            orch.process_new_ticket(ticket)
        except Exception as e:
            print(f'[Orchestrator Hook admin_create] Fallo: {e}')

    # Fallback robusto: si el orchestrator no asignó (sin API, falló o no estaba) usar lógica clásica
    db.session.refresh(ticket)
    if not ticket.assignee_id and not data.get('assignee_id'):
        try:
            assign_ticket_auto(ticket)
            if ticket.assignee_id and ticket.status == 'open':
                ticket.status = 'in_progress'
            db.session.commit()
        except Exception as e:
            print(f'[admin_create fallback-assign] Error: {e}')

    log_audit('create_ticket', session['user_id'], 'ticket', ticket.id, f"Ticket {ticket.ticket_number} creado por admin")

    # Emitir evento real-time RT-02
    emit_ticket_event(user.company, 'ticket_created', {
        'id': ticket.id,
        'ticket_number': ticket.ticket_number,
        'title': ticket.title,
        'priority': ticket.priority,
        'status': ticket.status,
        'created_by': user.name
    })

    # Notificar al técnico si fue asignado al crear (manual o auto)
    try:
        if ticket.assignee_id:
            tech_to_notify = User.query.get(ticket.assignee_id)
            if tech_to_notify and tech_to_notify.id != session['user_id']:
                notify_ticket_assigned(
                    ticket=ticket,
                    new_assignee=tech_to_notify,
                    assigned_by_name=user.name,
                    reason='Ticket creado y asignado por admin'
                )
    except Exception as e:
        print(f'[WARN] Notificación email admin_create: {e}')

    return jsonify({'success': True, 'ticket_id': ticket.id, 'ticket_number': ticket.ticket_number})

@app.route('/api/admin/tickets/<int:ticket_id>/edit', methods=['POST'])
def api_admin_edit_ticket(ticket_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    ticket = Ticket.query.get_or_404(ticket_id)

    # SECURITY FIX 9: Validar company_id en endpoint sensible
    if not validate_company_access(ticket):
        return jsonify({'success': False, 'message': 'No autorizado'}), 403

    data = request.get_json()

    # SECURITY FIX 3: Sanitizar inputs antes de guardar
    ticket.title = sanitize_input(data.get('title', ticket.title), max_length=200)
    ticket.description = sanitize_html(data.get('description', ticket.description)) if data.get('description') else ticket.description
    ticket.category = sanitize_input(data.get('category', ticket.category), max_length=100)
    ticket.priority = data.get('priority', ticket.priority)
    if ticket.priority not in ['low', 'medium', 'high', 'critical']:
        ticket.priority = 'medium'
    ticket.status = data.get('status', ticket.status)
    old_assignee_id = ticket.assignee_id
    ticket.assignee_id = data.get('assignee_id', ticket.assignee_id)

    db.session.commit()
    log_audit('edit_ticket', session['user_id'], 'ticket', ticket_id, f"Ticket {ticket.ticket_number} editado")

    # Si cambió el asignado en la edición, notificar al nuevo
    try:
        if ticket.assignee_id and ticket.assignee_id != old_assignee_id and ticket.assignee_id != session['user_id']:
            new_tech = User.query.get(ticket.assignee_id)
            if new_tech:
                editor = User.query.get(session['user_id'])
                notify_ticket_assigned(
                    ticket=ticket,
                    new_assignee=new_tech,
                    assigned_by_name=editor.name if editor else 'Administrador',
                    reason='Asignación modificada desde edición del ticket'
                )
    except Exception as e:
        print(f'[WARN] Notificación email edit_ticket: {e}')

    return jsonify({'success': True})

@app.route('/api/admin/tickets/<int:ticket_id>/delete', methods=['POST'])
def api_admin_delete_ticket(ticket_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.company != session['company']:
        return jsonify({'success': False}), 403

    ticket_num = ticket.ticket_number
    db.session.delete(ticket)
    db.session.commit()

    log_audit('delete_ticket', session['user_id'], 'ticket', ticket_id, f"Ticket {ticket_num} eliminado")

    return jsonify({'success': True})

# ═════════════════════════════════════════════════════════════════════════════
# PLANTILLAS
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/templates', methods=['GET'])
def api_list_templates():
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    templates = Template.query.filter_by(company=session['company']).all()
    result = [{
        'id': t.id,
        'name': t.name,
        'category': t.category,
        'priority': t.priority,
        'is_system': t.is_system
    } for t in templates]

    return jsonify({'success': True, 'templates': result})

# ═════════════════════════════════════════════════════════════════════════════
# REPORTES
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/reports/dashboard', methods=['GET'])
def api_reports_dashboard():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']
    tickets = Ticket.query.filter_by(company=company).all()

    total = len(tickets)
    resolved = len([t for t in tickets if t.status == 'resolved'])
    sla_compliance = sum([1 for t in tickets if t.sla_deadline and t.sla_deadline > datetime.now()]) / max(1, total)

    # Tiempo promedio de resolución
    resolved_tickets = [t for t in tickets if t.resolved_at]
    avg_time = sum([(t.resolved_at - t.created_at).total_seconds() for t in resolved_tickets]) / max(1, len(resolved_tickets)) / 3600

    # Carga por técnico
    technicians = User.query.filter_by(role='technician', company=company).all()
    tech_load = [{
        'name': t.name,
        'tickets_assigned': len([tk for tk in tickets if tk.assignee_id == t.id]),
        'tickets_resolved': len([tk for tk in tickets if tk.assignee_id == t.id and tk.status == 'resolved'])
    } for t in technicians]

    return jsonify({
        'success': True,
        'total_tickets': total,
        'resolved': resolved,
        'sla_compliance': round(sla_compliance * 100, 1),
        'avg_resolution_hours': round(avg_time, 1),
        'tech_load': tech_load
    })

# ═════════════════════════════════════════════════════════════════════════════
# ESCALACIÓN DE SLA
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/escalate-sla', methods=['POST'])
def api_escalate_sla():
    data = request.get_json()
    ticket_id = data.get('ticket_id')
    ticket = Ticket.query.get(ticket_id)

    if not ticket:
        return jsonify({'success': False}), 404

    if ticket.sla_deadline:
        # Calcular porcentaje SLA correctamente: (segundos transcurridos) / (total segundos SLA) * 100
        elapsed_seconds = (datetime.now() - ticket.created_at).total_seconds()
        sla_total_seconds = ticket.sla_minutes * 60
        elapsed_pct = elapsed_seconds / sla_total_seconds if sla_total_seconds > 0 else 0

        level = 0
        if elapsed_pct >= 2.0:
            level = 3  # 200%+
        elif elapsed_pct >= 1.0:
            level = 2  # 100%+
        elif elapsed_pct >= 0.5:
            level = 1  # 50%+

        return jsonify({'success': True, 'escalation_level': level, 'elapsed_pct': round(elapsed_pct * 100, 1)})

    return jsonify({'success': False})

# ═════════════════════════════════════════════════════════════════════════════
# SESIONES ACTIVAS
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/sessions', methods=['GET'])
def api_list_sessions():
    """Lista usuarios con actividad reciente (sesiones potencialmente activas).
    Una sesión se considera activa si last_login está dentro de los últimos 15 min
    (alineado con el timeout de sesión configurado)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    scope = admin_companies_scope()
    cutoff = datetime.now() - timedelta(minutes=15)

    users = User.query.filter(
        User.company.in_(scope),
        User.last_login.isnot(None)
    ).order_by(User.last_login.desc()).all()

    rows = []
    for u in users:
        is_active_session = u.last_login and u.last_login >= cutoff
        was_kicked = bool(u.force_logout_at and u.last_login and u.force_logout_at >= u.last_login)
        rows.append({
            'user_id': u.id,
            'name': u.name,
            'username': u.username,
            'email': u.email,
            'role': u.role,
            'role_label': u.role_label or '',
            'company': u.company,
            'last_login': u.last_login.isoformat() if u.last_login else None,
            'last_login_display': u.last_login.strftime('%d/%m/%Y %H:%M') if u.last_login else 'Nunca',
            'minutes_ago': int((datetime.now() - u.last_login).total_seconds() / 60) if u.last_login else None,
            'is_active_session': is_active_session,
            'was_kicked': was_kicked,
            'force_logout_at': u.force_logout_at.strftime('%d/%m %H:%M') if u.force_logout_at else None,
        })

    return jsonify({'success': True, 'sessions': rows})


@app.route('/api/admin/sessions/<int:user_id>/kick', methods=['POST'])
def api_kick_user(user_id):
    """Expulsa de verdad: actualiza force_logout_at. En la próxima request del usuario
    expulsado, su sesión se invalida automáticamente."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    user = User.query.get(user_id)
    if not user or user.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404

    if user.id == session['user_id']:
        return jsonify({'success': False, 'error': 'No puedes expulsarte a vos mismo'}), 400

    user.force_logout_at = datetime.now()
    db.session.commit()
    _refresh_force_logout_cache()

    log_audit('kick_user', session['user_id'], 'user', user_id,
              f"Usuario {user.username}@{user.company} expulsado de sesión")
    return jsonify({'success': True, 'message': f'{user.name} fue expulsado. Su sesión queda inválida.'})

# ═════════════════════════════════════════════════════════════════════════════
# MICROSOFT TEAMS WEBHOOKS (RF-03-12)
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/webhooks', methods=['GET'])
def api_list_webhooks():
    """Listar webhooks de Teams configurados"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    webhooks = Webhook.query.filter_by(company=session['company']).all()
    result = [{
        'id': w.id,
        'url': w.url,
        'events': w.events,
        'is_active': w.is_active
    } for w in webhooks]

    return jsonify({'success': True, 'webhooks': result})

@app.route('/api/admin/webhooks', methods=['POST'])
def api_add_webhook():
    """Agregar webhook de Teams (RF-03-12)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    data = request.get_json()
    webhook = Webhook(
        company=session['company'],
        url=data.get('url'),
        events=data.get('events', 'ticket_created,ticket_resolved'),
        is_active=True
    )

    db.session.add(webhook)
    db.session.commit()

    log_audit('add_webhook', session['user_id'], 'webhook', webhook.id, f'Webhook Teams agregado')

    return jsonify({'success': True, 'webhook_id': webhook.id})

@app.route('/api/admin/webhooks/<int:webhook_id>', methods=['DELETE'])
def api_delete_webhook(webhook_id):
    """Eliminar webhook"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    webhook = Webhook.query.get_or_404(webhook_id)
    if webhook.company != session['company']:
        return jsonify({'success': False}), 403

    db.session.delete(webhook)
    db.session.commit()

    log_audit('delete_webhook', session['user_id'], 'webhook', webhook_id, 'Webhook Teams eliminado')

    return jsonify({'success': True})

def send_teams_webhook(company, event, ticket):
    """Enviar notificación a Teams webhook (RF-03-12)"""
    webhooks = Webhook.query.filter_by(company=company, is_active=True).all()

    for webhook in webhooks:
        if event not in webhook.events:
            continue

        import requests
        message = {
            'title': f'Ticket {event.upper()}: {ticket.ticket_number}',
            'text': f'{ticket.title}\nPrioridad: {ticket.priority}',
            'sections': [{
                'facts': [
                    {'name': 'Número', 'value': ticket.ticket_number},
                    {'name': 'Estado', 'value': ticket.status},
                    {'name': 'Prioridad', 'value': ticket.priority}
                ]
            }],
            'potentialAction': [{
                'name': 'Ver Ticket',
                'targets': [{'os': 'default', 'uri': f'http://localhost:5050/employee/ticket/{ticket.id}'}]
            }]
        }

        try:
            requests.post(webhook.url, json={'@type': 'MessageCard', '@context': 'https://schema.org/extensions', **message})
        except Exception as e:
            log_audit('webhook_error', None, 'webhook', webhook.id, f'Error enviando webhook: {str(e)}')

# ═════════════════════════════════════════════════════════════════════════════
# MONITOR DE SERVIDORES
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/servers', methods=['GET'])
def api_list_servers():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    servers = Server.query.filter_by(company=session['company']).order_by(Server.is_critical.desc(), Server.name).all()
    result = [{
        'id': s.id,
        'name': s.name,
        'host': s.ip_address,
        'ip_address': s.ip_address,
        'port': s.port or 443,
        'description': s.description or '',
        'is_critical': bool(s.is_critical),
        'is_active': bool(s.is_active),
        'last_status': s.last_status or 'unknown',
        'last_check': s.last_ping.strftime('%d/%m %H:%M') if s.last_ping else 'Nunca',
        'is_online': s.is_online
    } for s in servers]

    return jsonify({'success': True, 'servers': result})

@app.route('/api/admin/servers', methods=['POST'])
def api_add_server():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    try:
        data = request.get_json()
        name = (data.get('name') or '').strip()
        host = (data.get('host') or data.get('ip_address') or '').strip()
        port = int(data.get('port', 443))
        description = (data.get('description') or '').strip()
        is_critical = bool(data.get('is_critical', False))

        if not name or not host:
            return jsonify({'success': False, 'error': 'Nombre y host son requeridos'}), 400

        server = Server(
            name=name,
            ip_address=host,
            port=port,
            description=description,
            is_critical=is_critical,
            is_active=True,
            company=session['company']
        )

        db.session.add(server)
        db.session.commit()

        log_audit('add_server', session['user_id'], 'server', server.id, f"Servidor {server.name} agregado ({host}:{port})")

        return jsonify({'success': True, 'id': server.id, 'message': 'Servidor agregado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/servers/<int:server_id>', methods=['PUT'])
def api_update_server(server_id):
    """Actualizar configuracion de un servidor"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    server = Server.query.get(server_id)
    if not server or server.company != session['company']:
        return jsonify({'success': False, 'error': 'No encontrado'}), 404

    try:
        data = request.get_json()
        old_data = f"{server.name} ({server.ip_address}:{server.port})"

        if 'name' in data:
            name = (data.get('name') or '').strip()
            if not name:
                return jsonify({'success': False, 'error': 'Nombre requerido'}), 400
            server.name = name

        if 'host' in data or 'ip_address' in data:
            host = (data.get('host') or data.get('ip_address') or '').strip()
            if not host:
                return jsonify({'success': False, 'error': 'Host requerido'}), 400
            server.ip_address = host

        if 'port' in data:
            server.port = int(data.get('port', 443))

        if 'description' in data:
            server.description = (data.get('description') or '').strip()

        if 'is_critical' in data:
            server.is_critical = bool(data.get('is_critical'))

        if 'is_active' in data:
            server.is_active = bool(data.get('is_active'))
            # Si se desactiva el servidor, limpiar alarma
            if not server.is_active:
                server.alarm_active = False
                server.consecutive_failures = 0

        db.session.commit()

        log_audit('update_server', session['user_id'], 'server', server.id,
                  f'Servidor actualizado: {old_data} -> {server.name} ({server.ip_address}:{server.port})')

        return jsonify({
            'success': True,
            'message': 'Servidor actualizado',
            'server': {
                'id': server.id,
                'name': server.name,
                'host': server.ip_address,
                'port': server.port,
                'description': server.description or '',
                'is_critical': bool(server.is_critical),
                'is_active': bool(server.is_active)
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/servers/<int:server_id>', methods=['DELETE'])
def api_delete_server(server_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    server = Server.query.get(server_id)
    if not server or server.company != session['company']:
        return jsonify({'success': False}), 403

    name = server.name
    db.session.delete(server)
    db.session.commit()

    log_audit('delete_server', session['user_id'], 'server', server_id, f"Servidor {name} eliminado")

    return jsonify({'success': True, 'message': 'Servidor eliminado'})


@app.route('/api/admin/servers/<int:server_id>/simulate', methods=['POST'])
def api_simulate_server_outage(server_id):
    """Simula la caída (o recuperación) de un servidor.
    Body: {action: 'down' | 'up'}
    Si es 'down' y no hay ticket abierto, crea uno automáticamente."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    server = Server.query.get(server_id)
    if not server or server.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Servidor no encontrado'}), 404

    data = request.get_json() or {}
    action = (data.get('action') or 'down').lower()

    if action not in ('down', 'up'):
        return jsonify({'success': False, 'error': 'Acción inválida (down|up)'}), 400

    user = User.query.get(session['user_id'])
    ticket_created = None

    if action == 'down':
        # Marcar el servidor como caído
        server.is_online = False
        server.last_status = 'down'
        server.last_ping = datetime.now()
        server.consecutive_failures = (server.consecutive_failures or 0) + 1
        server.alarm_active = True

        # ¿Hay ya un ticket abierto para este servidor? (buscar por título)
        ticket_title_prefix = f"⚠ Servidor caído: {server.name}"
        existing_ticket = Ticket.query.filter(
            Ticket.company == server.company,
            Ticket.title.like(f"%{server.name}%"),
            Ticket.status.in_(['open', 'in_progress'])
        ).first()

        if not existing_ticket:
            # Crear ticket automáticamente
            priority = 'critical' if server.is_critical else 'high'
            sla_minutes = get_sla_minutes_for_priority(priority)

            new_ticket = Ticket(
                ticket_number=get_next_ticket_number(server.company),
                title=ticket_title_prefix,
                description=(
                    f"Se detectó la caída del servidor monitoreado.\n\n"
                    f"**Servidor:** {server.name}\n"
                    f"**IP / Host:** {server.ip_address}\n"
                    f"**Puerto:** {server.port}\n"
                    f"**Descripción:** {server.description or '(sin descripción)'}\n"
                    f"**Crítico:** {'Sí' if server.is_critical else 'No'}\n"
                    f"**Fallos consecutivos:** {server.consecutive_failures}\n"
                    f"**Detectado por:** {user.name} (simulación manual)\n"
                    f"**Hora:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                    "Acción sugerida:\n"
                    "1. Verificar estado físico/red del servidor.\n"
                    "2. Validar servicios y procesos críticos.\n"
                    "3. Notificar a usuarios afectados."
                ),
                category='Servidores',
                priority=priority,
                status='open',
                creator_id=user.id,
                company=server.company,
                sla_minutes=sla_minutes,
                sla_deadline=compute_sla_deadline(datetime.now(), sla_minutes, server.company)
            )
            # Asignar automáticamente
            assign_ticket_auto(new_ticket)
            db.session.add(new_ticket)
            db.session.commit()
            ticket_created = {
                'id': new_ticket.id,
                'ticket_number': new_ticket.ticket_number,
                'priority': new_ticket.priority,
                'assignee': new_ticket.assignee.name if new_ticket.assignee else None
            }
            log_audit('server_outage_simulated', user.id, 'server', server.id,
                      f"Caída simulada de {server.name} → ticket {new_ticket.ticket_number} creado")

            # Emitir evento WebSocket
            try:
                emit_ticket_event(server.company, 'server_down', {
                    'server_id': server.id,
                    'server_name': server.name,
                    'ticket_id': new_ticket.id,
                    'ticket_number': new_ticket.ticket_number
                })
            except Exception:
                pass
        else:
            log_audit('server_outage_simulated', user.id, 'server', server.id,
                      f"Caída simulada de {server.name} (ticket {existing_ticket.ticket_number} ya abierto)")
            ticket_created = {
                'id': existing_ticket.id,
                'ticket_number': existing_ticket.ticket_number,
                'existing': True
            }

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Servidor {server.name} marcado como caído',
            'server_status': 'down',
            'ticket': ticket_created
        })

    else:  # up
        # Restaurar
        server.is_online = True
        server.last_status = 'online'
        server.last_ping = datetime.now()
        server.consecutive_failures = 0
        server.alarm_active = False
        db.session.commit()

        # Buscar ticket abierto del outage y cerrarlo automáticamente (opcional)
        related = Ticket.query.filter(
            Ticket.company == server.company,
            Ticket.title.like(f"%{server.name}%"),
            Ticket.category == 'Servidores',
            Ticket.status.in_(['open', 'in_progress'])
        ).first()
        resolved_ticket = None
        if related:
            related.status = 'resolved'
            related.resolved_at = datetime.now()
            related.updated_at = datetime.now()
            db.session.add(Message(
                ticket_id=related.id,
                user_id=user.id,
                text=f'✅ SOLUCIÓN: Servidor {server.name} recuperado (simulación). Conexión restablecida.'
            ))
            db.session.commit()
            resolved_ticket = {'id': related.id, 'ticket_number': related.ticket_number}

        log_audit('server_recovery_simulated', user.id, 'server', server.id,
                  f"Recuperación simulada de {server.name}")

        return jsonify({
            'success': True,
            'message': f'Servidor {server.name} restaurado',
            'server_status': 'online',
            'resolved_ticket': resolved_ticket
        })


# ═════════════════════════════════════════════════════════════════════════════
# WEBSOCKET SOCKET.IO RT-01 A RT-07 - PROPAGACIÓN REAL-TIME
# ═════════════════════════════════════════════════════════════════════════════

@socketio.on('connect')
def handle_connect():
    """RT-01: Usuario conectado"""
    if 'user_id' not in session:
        return False
    company = session.get('company')
    if company:
        join_room(f'company_{company}')
    emit('user_connected', {
        'user': session.get('name'),
        'company': company,
        'timestamp': datetime.now().isoformat()
    })

@socketio.on('disconnect')
def handle_disconnect():
    """RT-01: Usuario desconectado"""
    if 'user_id' in session:
        company = session.get('company')
        if company:
            leave_room(f'company_{company}')

def emit_ticket_event(company, event_type, ticket_data):
    """Emitir evento a todos en la company (RT-02, RT-03, RT-04, RT-06)"""
    socketio.emit(event_type, {
        'ticket': ticket_data,
        'timestamp': datetime.now().isoformat()
    }, room=f'company_{company}')

# ═════════════════════════════════════════════════════════════════════════════
# CRONÓMETRO RF-02-05 - TIEMPO TRABAJADO
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/ticket/<int:ticket_id>/time', methods=['GET'])
def api_get_ticket_time(ticket_id):
    """Obtener tiempo trabajado en ticket"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    ticket = Ticket.query.get(ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    return jsonify({
        'success': True,
        'seconds': ticket.time_worked_seconds or 0,
        'ticket_number': ticket.ticket_number
    })

@app.route('/api/ticket/<int:ticket_id>/history', methods=['GET'])
def api_get_ticket_history(ticket_id):
    """Obtener historial de cambios del ticket (RF-01-08)"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    ticket = Ticket.query.get(ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    # Obtener audit logs del ticket
    history = AuditLog.query.filter_by(
        entity_type='ticket',
        entity_id=ticket_id
    ).order_by(AuditLog.created_at.desc()).all()

    history_data = [{
        'action': h.action.upper(),
        'description': h.description,
        'created_at': h.created_at.isoformat()
    } for h in history]

    return jsonify({
        'success': True,
        'history': history_data
    })

@app.route('/api/ticket/<int:ticket_id>/reopen', methods=['POST'])
def api_reopen_ticket(ticket_id):
    """El usuario final reabre un ticket resuelto explicando por qué la solución no le sirvió.
    Reglas:
    - Solo el creador del ticket (o admin) puede reabrirlo
    - Debe estar en estado resolved (no se reabren cerrados/closed para evitar abuso)
    - Requiere reason de al menos 10 caracteres
    - Resetea status a in_progress, conserva asignado, agrega mensaje a la conversación
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({'success': False, 'error': 'Ticket no encontrado'}), 404

    # Solo el creador o admin (con scope) puede reabrir
    user_id = session['user_id']
    is_creator = ticket.creator_id == user_id
    is_admin_scope = session.get('role') == 'admin' and ticket.company in admin_companies_scope()
    if not (is_creator or is_admin_scope):
        return jsonify({'success': False, 'error': 'Solo el creador del ticket puede reabrirlo.'}), 403

    if ticket.status != 'resolved':
        return jsonify({
            'success': False,
            'error': f'Solo se pueden reabrir tickets en estado "Resuelto" (actual: {ticket.status}).'
        }), 400

    data = request.json or {}
    reason = (data.get('reason') or '').strip()
    if len(reason) < 10:
        return jsonify({'success': False, 'error': 'Debes explicar el motivo de la reapertura (mínimo 10 caracteres).'}), 400

    user = User.query.get(user_id)
    user_name = user.name if user else 'Usuario'

    # Cambiar estado y limpiar campos de resolución
    ticket.status = 'in_progress'
    ticket.resolved_at = None
    ticket.updated_at = datetime.now()
    # Limpiar rating al reabrir para que pueda calificar la nueva solución
    ticket.rating = None
    # Resetear alertas SLA para que se recalcule contra el nuevo plazo
    ticket.sla_alerts_sent = ''
    # Si querés extender el SLA al reabrir, descomentá:
    # ticket.sla_deadline = datetime.now() + timedelta(minutes=ticket.sla_minutes or 240)

    # Mensaje en la conversación
    db.session.add(Message(
        ticket_id=ticket.id,
        user_id=user_id,
        text=f'🔄 TICKET REABIERTO por {user_name}\n\nMotivo: {reason}'
    ))
    db.session.commit()

    log_audit('ticket_reopened', user_id, 'ticket', ticket.id,
              f'Ticket {ticket.ticket_number} reabierto por {user_name}. Motivo: {reason[:200]}')

    # Notificar via WebSocket
    try:
        emit_ticket_event(ticket.company, 'ticket_reopened', {
            'ticket_id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'reopened_by': user_name,
            'assignee_id': ticket.assignee_id
        })
    except Exception:
        pass

    # Enviar correo al técnico asignado
    try:
        if ticket.assignee_id:
            assignee = User.query.get(ticket.assignee_id)
            if assignee and assignee.email:
                subject = f'[DeskEli] 🔄 Ticket reabierto · {ticket.ticket_number}'
                body = f"""
                <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
                    <div style="background:#f59e0b;color:white;padding:18px;border-radius:8px 8px 0 0;">
                        <h2 style="margin:0;">🔄 Ticket Reabierto</h2>
                    </div>
                    <div style="padding:22px;background:#fffbeb;border:1px solid #fde68a;border-top:none;">
                        <p>Hola <strong>{assignee.name}</strong>,</p>
                        <p>El usuario <strong>{user_name}</strong> reabrió el ticket <strong>{ticket.ticket_number}</strong> porque
                        la solución entregada no resolvió su problema.</p>
                        <div style="background:white;border-left:4px solid #f59e0b;padding:12px;margin:14px 0;border-radius:4px;">
                            <div style="font-weight:bold;color:#92400e;margin-bottom:6px;">💬 Motivo del usuario:</div>
                            <div style="white-space:pre-wrap;color:#374151;">{reason}</div>
                        </div>
                        <p>El ticket volvió al estado <strong>En Progreso</strong> y sigue asignado a vos.</p>
                    </div>
                </body></html>
                """
                send_email(assignee.email, subject, body)
    except Exception as e:
        print(f'[reopen] No se pudo enviar correo: {e}')

    return jsonify({
        'success': True,
        'message': 'Ticket reabierto correctamente. El técnico fue notificado.',
        'new_status': 'in_progress'
    })


@app.route('/api/ticket/<int:ticket_id>/resolve', methods=['POST'])
def api_resolve_ticket(ticket_id):
    """Técnico/admin cierra el ticket documentando la solución. El usuario califica después."""
    if 'user_id' not in session or session.get('role') not in ('technician', 'admin'):
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({'success': False, 'error': 'Ticket no encontrado'}), 404
    if ticket.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': f'No tienes acceso a tickets de la empresa "{ticket.company}".'}), 403

    data = request.json or {}
    resolution = (data.get('resolution') or '').strip()

    if not resolution or len(resolution) < 10:
        return jsonify({
            'success': False,
            'error': 'Debes informar cómo resolviste el problema (mínimo 10 caracteres) antes de cerrar el ticket.'
        }), 400

    # VALIDACIÓN: no permitir cerrar si tiene subtareas pendientes
    pending_subtasks = Subtask.query.filter(
        Subtask.ticket_id == ticket_id,
        Subtask.status.notin_(['resolved', 'closed', 'cancelled'])
    ).all()
    if pending_subtasks:
        pending_list = [f"#{s.subtask_number or s.id}: {s.title[:60]} ({s.status})"
                        for s in pending_subtasks[:5]]
        extra = f' (+ {len(pending_subtasks) - 5} más)' if len(pending_subtasks) > 5 else ''
        return jsonify({
            'success': False,
            'error': f'No se puede cerrar el ticket: tiene {len(pending_subtasks)} subtarea(s) pendiente(s). Resolvé primero:\n\n• ' + '\n• '.join(pending_list) + extra,
            'pending_subtasks': [{
                'id': s.id,
                'subtask_number': s.subtask_number,
                'title': s.title,
                'status': s.status,
                'assignee': s.assignee.name if getattr(s, 'assignee', None) else None
            } for s in pending_subtasks]
        }), 400

    # Registrar solución como mensaje
    db.session.add(Message(
        ticket_id=ticket_id,
        user_id=session['user_id'],
        text='✅ SOLUCIÓN: ' + resolution
    ))

    # Marcar como resuelto (sin rating: lo calificará el usuario)
    ticket.status = 'resolved'
    ticket.resolved_at = datetime.now()
    ticket.updated_at = datetime.now()
    db.session.commit()

    log_audit('ticket_resolved', session['user_id'], 'ticket', ticket_id,
              f"Ticket resuelto · solución: {resolution[:120]}")

    # Emitir evento WebSocket
    try:
        emit_ticket_event(ticket.company, 'ticket_resolved', {
            'ticket_id': ticket.id, 'ticket_number': ticket.ticket_number
        })
    except Exception:
        pass

    # Enviar correo al creador con la solución para que califique
    email_sent = False
    try:
        creator = User.query.get(ticket.creator_id)
        if creator and creator.email:
            base_url = get_public_base_url()
            ticket_url = f"{base_url}/employee/ticket/{ticket.id}"
            subject = f"[DeskEli] Ticket {ticket.ticket_number} resuelto - califica la solución"
            body = f"""
            <html><body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="background: #16a34a; color: white; padding: 20px; border-radius: 8px 8px 0 0;">
                    <h2 style="margin: 0;">✅ Tu ticket fue resuelto</h2>
                </div>
                <div style="padding: 24px; background: #f9fafb; border: 1px solid #e5e7eb;">
                    <p>Hola <strong>{creator.name}</strong>,</p>
                    <p>Tu ticket <strong>{ticket.ticket_number}</strong> — <em>{ticket.title}</em> — ha sido marcado como resuelto.</p>

                    <div style="background: white; border-left: 4px solid #16a34a; padding: 14px; margin: 16px 0; border-radius: 4px;">
                        <div style="font-weight: bold; color: #065f46; margin-bottom: 6px;">💡 Solución aplicada:</div>
                        <div style="white-space: pre-wrap; color: #374151;">{resolution}</div>
                    </div>

                    <p>Por favor califica esta solución para ayudarnos a mejorar:</p>
                    <p style="text-align: center;">
                        <a href="{ticket_url}" style="display: inline-block; background: #2563eb; color: white; padding: 12px 28px; border-radius: 6px; text-decoration: none; font-weight: bold;">
                            ⭐ Calificar Solución
                        </a>
                    </p>
                    <p style="font-size: 12px; color: #6b7280;">Si la solución no resolvió tu problema, puedes reabrir el ticket desde el portal.</p>
                </div>
                <div style="text-align: center; padding: 12px; font-size: 11px; color: #9ca3af;">
                    DeskEli · {ticket.company.title()}
                </div>
            </body></html>
            """
            send_email(creator.email, subject, body, company=ticket.company)
            email_sent = True
    except Exception as e:
        print(f"[resolve] No se pudo enviar correo: {e}")

    return jsonify({
        'success': True,
        'message': 'Ticket resuelto. El usuario fue notificado para calificarlo.',
        'email_sent': email_sent
    })


@app.route('/api/ticket/<int:ticket_id>/rating', methods=['POST'])
def api_save_ticket_rating(ticket_id):
    """El usuario final califica un ticket ya resuelto (RF-01-09)."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    ticket = Ticket.query.get(ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    data = request.json or {}
    rating = data.get('rating', 0)
    comment = (data.get('comment') or '').strip()[:2000]
    nps_raw = data.get('nps')

    if not (1 <= rating <= 5):
        return jsonify({'success': False, 'error': 'Rating inválido'}), 400

    # NPS opcional: 0-10 si viene
    nps_score = None
    if nps_raw is not None and nps_raw != '':
        try:
            nps_score = int(nps_raw)
            if not (0 <= nps_score <= 10):
                return jsonify({'success': False, 'error': 'NPS debe estar entre 0 y 10'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'NPS inválido'}), 400

    # La calificación debe darla el creador (usuario final), no el técnico
    if ticket.creator_id != session['user_id'] and session.get('role') != 'admin':
        return jsonify({
            'success': False,
            'error': 'Solo el usuario que creó el ticket puede calificarlo.'
        }), 403

    ticket.rating = rating
    if comment:
        ticket.rating_comment = comment
    if nps_score is not None:
        ticket.rating_nps = nps_score
    ticket.rating_at = datetime.now()
    ticket.updated_at = datetime.now()
    # Si aún no estaba resuelto, marcarlo (caso raro: usuario califica antes de cierre)
    if ticket.status != 'resolved':
        ticket.status = 'resolved'
        ticket.resolved_at = datetime.now()
    db.session.commit()

    log_audit('ticket_rated', session['user_id'], 'ticket', ticket_id,
              f"Ticket calificado con {rating}/5"
              + (f" (NPS: {nps_score})" if nps_score is not None else '')
              + (" con comentario" if comment else ''))

    return jsonify({
        'success': True,
        'message': f'¡Gracias por tu calificación de {rating}/5!'
    })

def _time_entry_serialize(e):
    return {
        'id': e.id,
        'ticket_id': e.ticket_id,
        'user_id': e.user_id,
        'user_name': e.user.name if e.user else '(usuario eliminado)',
        'started_at': e.started_at.isoformat() if e.started_at else None,
        'ended_at': e.ended_at.isoformat() if e.ended_at else None,
        'duration_seconds': e.duration_seconds or 0,
        'is_running': e.ended_at is None,
        'notes': e.notes or '',
        'is_manual': e.is_manual,
        'company': e.company,
        'created_at': e.created_at.isoformat() if e.created_at else None,
    }


def _sync_ticket_total_time(ticket):
    """Recalcula time_worked_seconds del ticket sumando las entries cerradas."""
    total = db.session.query(db.func.coalesce(db.func.sum(TimeEntry.duration_seconds), 0)) \
        .filter(TimeEntry.ticket_id == ticket.id, TimeEntry.ended_at.isnot(None)).scalar()
    ticket.time_worked_seconds = int(total or 0)


@app.route('/api/tickets/<int:ticket_id>/time-entries', methods=['GET'])
def api_time_entries_list(ticket_id):
    """Lista las entradas de tiempo del ticket."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({'success': False, 'error': 'Ticket no encontrado'}), 404
    # Empleados solo ven sus propios tickets; técnicos/admin ven los de su empresa (respetando espejos)
    if not can_user_access_ticket(User.query.get(session['user_id']), ticket):
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    entries = TimeEntry.query.filter_by(ticket_id=ticket_id).order_by(TimeEntry.started_at.asc()).all()
    total_seconds = sum(e.duration_seconds or 0 for e in entries if e.ended_at)
    # Si hay una corriendo, sumar el tiempo transcurrido en vivo
    running = next((e for e in entries if e.ended_at is None), None)
    if running:
        total_seconds += int((datetime.now() - running.started_at).total_seconds())
    return jsonify({
        'success': True,
        'entries': [_time_entry_serialize(e) for e in entries],
        'total_seconds': total_seconds,
        'running_entry_id': running.id if running else None
    })


@app.route('/api/tickets/<int:ticket_id>/time-entries/start', methods=['POST'])
def api_time_entry_start(ticket_id):
    """Inicia el cronómetro para el usuario actual en este ticket.

    Reglas:
    - Solo técnicos y admin (los empleados no registran esfuerzo)
    - Un solo cronómetro activo por (usuario, ticket) a la vez
    - Si el usuario tiene otro cronómetro corriendo en OTRO ticket, se cierra
      automáticamente antes de abrir este (evita entries olvidadas)
    """
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    user = User.query.get(session['user_id'])
    if not user or user.role not in ('technician', 'admin'):
        return jsonify({'success': False, 'error': 'Solo técnicos y administradores pueden registrar tiempo'}), 403

    ticket = Ticket.query.get(ticket_id)
    if not ticket or not can_user_access_ticket(user, ticket):
        return jsonify({'success': False, 'error': 'Sin acceso al ticket'}), 403

    # ¿Ya hay uno activo en este mismo ticket? Devolver ese.
    existing = TimeEntry.query.filter_by(ticket_id=ticket_id, user_id=user.id, ended_at=None).first()
    if existing:
        return jsonify({'success': True, 'entry': _time_entry_serialize(existing), 'note': 'ya estaba corriendo'})

    # Cerrar automáticamente cualquier cronómetro del user en otros tickets
    other_running = TimeEntry.query.filter_by(user_id=user.id, ended_at=None).all()
    closed_others = []
    for e in other_running:
        e.ended_at = datetime.now()
        e.duration_seconds = int((e.ended_at - e.started_at).total_seconds())
        closed_others.append(e.ticket.ticket_number if e.ticket else str(e.id))
        # Actualizar totales del ticket cerrado
        if e.ticket:
            _sync_ticket_total_time(e.ticket)

    now = datetime.now()
    entry = TimeEntry(
        ticket_id=ticket_id,
        user_id=user.id,
        company=ticket.company,
        started_at=now,
        duration_seconds=0,
        is_manual=False
    )
    db.session.add(entry)
    db.session.commit()

    log_audit('time_start', user.id, 'time_entry', entry.id,
              f"Inició cronómetro en {ticket.ticket_number}"
              + (f" (auto-cerró: {', '.join(closed_others)})" if closed_others else ''))
    return jsonify({'success': True, 'entry': _time_entry_serialize(entry),
                    'closed_others': closed_others})


@app.route('/api/tickets/<int:ticket_id>/time-entries/stop', methods=['POST'])
def api_time_entry_stop(ticket_id):
    """Detiene el cronómetro activo del usuario en este ticket."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    user = User.query.get(session['user_id'])
    ticket = Ticket.query.get(ticket_id)
    if not ticket or not can_user_access_ticket(user, ticket):
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403

    running = TimeEntry.query.filter_by(ticket_id=ticket_id, user_id=user.id, ended_at=None).first()
    if not running:
        return jsonify({'success': False, 'error': 'No hay cronómetro activo en este ticket'}), 400

    data = request.json or {}
    notes = (data.get('notes') or '').strip()[:1000]
    running.ended_at = datetime.now()
    running.duration_seconds = int((running.ended_at - running.started_at).total_seconds())
    if notes:
        running.notes = notes
    _sync_ticket_total_time(ticket)
    db.session.commit()

    log_audit('time_stop', user.id, 'time_entry', running.id,
              f"Detuvo cronómetro en {ticket.ticket_number} ({running.duration_seconds}s)")
    return jsonify({'success': True, 'entry': _time_entry_serialize(running),
                    'ticket_total_seconds': ticket.time_worked_seconds})


@app.route('/api/tickets/<int:ticket_id>/time-entries/manual', methods=['POST'])
def api_time_entry_manual(ticket_id):
    """Ingresa una entrada manual con inicio/fin explícitos o duración en minutos."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    user = User.query.get(session['user_id'])
    if not user or user.role not in ('technician', 'admin'):
        return jsonify({'success': False, 'error': 'Solo técnicos y administradores'}), 403
    ticket = Ticket.query.get(ticket_id)
    if not ticket or not can_user_access_ticket(user, ticket):
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403

    data = request.json or {}
    started_raw = data.get('started_at')
    ended_raw = data.get('ended_at')
    duration_minutes = data.get('duration_minutes')
    notes = (data.get('notes') or '').strip()[:1000]

    # Modo 1: started_at + duration_minutes (más simple para "trabajé X min ayer")
    if started_raw and duration_minutes and not ended_raw:
        try:
            started = datetime.fromisoformat(started_raw)
            dur_min = int(duration_minutes)
            if dur_min <= 0 or dur_min > 60 * 24:
                return jsonify({'success': False, 'error': 'Duración debe ser 1-1440 minutos'}), 400
            ended = started + timedelta(minutes=dur_min)
            duration_sec = dur_min * 60
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Formato inválido'}), 400
    # Modo 2: started_at + ended_at explícitos
    elif started_raw and ended_raw:
        try:
            started = datetime.fromisoformat(started_raw)
            ended = datetime.fromisoformat(ended_raw)
            if ended <= started:
                return jsonify({'success': False, 'error': 'La hora de fin debe ser posterior al inicio'}), 400
            duration_sec = int((ended - started).total_seconds())
            if duration_sec > 60 * 60 * 24:
                return jsonify({'success': False, 'error': 'Una entrada no puede durar más de 24 horas'}), 400
        except (ValueError, TypeError):
            return jsonify({'success': False, 'error': 'Formato de fecha inválido'}), 400
    else:
        return jsonify({'success': False, 'error': 'Debés indicar inicio + fin o inicio + duración'}), 400

    entry = TimeEntry(
        ticket_id=ticket_id,
        user_id=user.id,
        company=ticket.company,
        started_at=started,
        ended_at=ended,
        duration_seconds=duration_sec,
        notes=notes,
        is_manual=True
    )
    db.session.add(entry)
    _sync_ticket_total_time(ticket)
    db.session.commit()

    log_audit('time_manual', user.id, 'time_entry', entry.id,
              f"Entrada manual en {ticket.ticket_number}: {duration_sec}s")
    return jsonify({'success': True, 'entry': _time_entry_serialize(entry),
                    'ticket_total_seconds': ticket.time_worked_seconds})


@app.route('/api/time-entries/<int:entry_id>', methods=['DELETE'])
def api_time_entry_delete(entry_id):
    """Elimina una entrada. Solo el dueño o un admin puede."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    user = User.query.get(session['user_id'])
    entry = TimeEntry.query.get(entry_id)
    if not entry:
        return jsonify({'success': False}), 404
    # Autorización: propietario o admin de la empresa del entry
    if entry.user_id != user.id and not (user.role == 'admin' and entry.company in admin_companies_scope()):
        return jsonify({'success': False, 'error': 'Sin permiso'}), 403
    ticket = entry.ticket
    db.session.delete(entry)
    if ticket:
        _sync_ticket_total_time(ticket)
    db.session.commit()
    log_audit('time_delete', user.id, 'time_entry', entry_id,
              f"Entrada de tiempo eliminada (ticket {ticket.ticket_number if ticket else '?'})")
    return jsonify({'success': True})


# Dashboard admin de esfuerzo
@app.route('/admin/time-tracking')
def admin_time_tracking():
    if 'user_id' not in session or session['role'] != 'admin':
        return redirect(url_for('login'))
    user = User.query.get(session['user_id'])
    is_master = is_master_admin()
    available_companies = []
    if is_master:
        available_companies = [
            {'code': c.code, 'name': c.name}
            for c in Company.query.filter_by(is_active=True).order_by(Company.name).all()
        ]
    return render_template('admin/time_tracking.html',
                           company_info=COMPANY_COLORS.get(user.company, {}),
                           is_master=is_master,
                           available_companies=available_companies,
                           current_company=user.company)


@app.route('/api/admin/time-tracking/summary')
def api_admin_time_summary():
    """Métricas agregadas de esfuerzo para el dashboard admin."""
    if 'user_id' not in session or session.get('role') != 'admin':
        return jsonify({'success': False}), 401

    company_filter = request.args.get('company', session.get('company'))
    try:
        days = int(request.args.get('days', 30))
    except (ValueError, TypeError):
        days = 30
    days = max(1, min(days, 365))
    since = datetime.now() - timedelta(days=days)
    scope = admin_companies_scope()

    q = TimeEntry.query.filter(
        TimeEntry.ended_at.isnot(None),
        TimeEntry.created_at >= since,
        TimeEntry.company.in_(scope)
    )
    if company_filter and company_filter != 'all':
        if company_filter not in scope:
            return jsonify({'success': False, 'error': 'Sin acceso'}), 403
        q = q.filter(TimeEntry.company == company_filter)

    entries = q.all()
    total_seconds = sum(e.duration_seconds or 0 for e in entries)
    total_hours = round(total_seconds / 3600, 2)
    distinct_tickets = len(set(e.ticket_id for e in entries))
    avg_per_ticket = round(total_seconds / distinct_tickets / 60, 1) if distinct_tickets else 0

    # Por técnico
    from collections import defaultdict
    by_user = defaultdict(lambda: {'seconds': 0, 'entries': 0, 'tickets': set()})
    for e in entries:
        by_user[e.user_id]['seconds'] += (e.duration_seconds or 0)
        by_user[e.user_id]['entries'] += 1
        by_user[e.user_id]['tickets'].add(e.ticket_id)

    by_user_list = []
    for uid, data in by_user.items():
        u = User.query.get(uid)
        if not u:
            continue
        by_user_list.append({
            'id': uid, 'name': u.name, 'company': u.company,
            'hours': round(data['seconds'] / 3600, 2),
            'entries': data['entries'],
            'tickets': len(data['tickets']),
            'avg_per_ticket_min': round(data['seconds'] / len(data['tickets']) / 60, 1) if data['tickets'] else 0
        })
    by_user_list.sort(key=lambda x: -x['hours'])

    # Por categoría
    by_category = defaultdict(lambda: {'seconds': 0, 'tickets': set()})
    for e in entries:
        if e.ticket:
            cat = e.ticket.category or 'Sin categoría'
            by_category[cat]['seconds'] += (e.duration_seconds or 0)
            by_category[cat]['tickets'].add(e.ticket_id)
    by_category_list = [
        {'category': c, 'hours': round(d['seconds'] / 3600, 2), 'tickets': len(d['tickets'])}
        for c, d in by_category.items()
    ]
    by_category_list.sort(key=lambda x: -x['hours'])

    # Tickets con más horas (top 15)
    tickets_seconds = defaultdict(int)
    for e in entries:
        tickets_seconds[e.ticket_id] += (e.duration_seconds or 0)
    top_tickets = []
    for tid, sec in sorted(tickets_seconds.items(), key=lambda x: -x[1])[:15]:
        t = Ticket.query.get(tid)
        if not t:
            continue
        top_tickets.append({
            'id': tid,
            'number': t.ticket_number,
            'title': t.title[:80],
            'status': t.status,
            'company': t.company,
            'hours': round(sec / 3600, 2)
        })

    # Tendencia diaria
    daily = defaultdict(int)
    for e in entries:
        if e.ended_at:
            day = e.ended_at.strftime('%Y-%m-%d')
            daily[day] += (e.duration_seconds or 0)
    trend = [{'date': d, 'hours': round(s / 3600, 2)}
             for d, s in sorted(daily.items())[-30:]]

    return jsonify({
        'success': True,
        'days': days,
        'total_hours': total_hours,
        'total_entries': len(entries),
        'distinct_tickets': distinct_tickets,
        'avg_per_ticket_min': avg_per_ticket,
        'by_technician': by_user_list[:30],
        'by_category': by_category_list[:15],
        'top_tickets': top_tickets,
        'trend': trend,
    })


# NOTA: el antiguo endpoint POST /api/ticket/<id>/time fue reemplazado por el
# nuevo sistema de TimeEntry (/api/tickets/<id>/time-entries/*). El GET sigue
# vigente para compatibilidad — devuelve el total acumulado del ticket.

@app.route('/api/ticket/<int:ticket_id>/request-info', methods=['POST'])
def api_ticket_request_info(ticket_id):
    """Enviar email al creador solicitando información adicional.
    TO: creador del ticket. CC: técnico asignado. También registra el mensaje
    como comentario público en el chat del ticket."""
    if 'user_id' not in session or session.get('role') not in ('technician', 'admin'):
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso a esta empresa'}), 403

    data = request.get_json() or {}
    message = (data.get('message') or '').strip()
    if not message or len(message) < 10:
        return jsonify({'success': False, 'error': 'Escribí un mensaje (mínimo 10 caracteres)'}), 400
    if len(message) > 3000:
        return jsonify({'success': False, 'error': 'Mensaje demasiado largo (máx 3000 caracteres)'}), 400

    # Sanitizar HTML (bleach)
    safe_message = sanitize_html(message) if 'sanitize_html' in globals() else message

    creator = User.query.get(ticket.creator_id) if ticket.creator_id else None
    if not creator or not creator.email:
        return jsonify({'success': False, 'error': 'El solicitante no tiene email registrado'}), 400

    assignee = User.query.get(ticket.assignee_id) if ticket.assignee_id else None
    current_user = User.query.get(session['user_id'])

    # 1) Registrar como mensaje público en el chat del ticket
    try:
        msg = Message(
            ticket_id=ticket.id,
            user_id=session['user_id'],
            text=f'📧 Solicitud de información enviada por email:\n\n{safe_message}',
        )
        db.session.add(msg)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f'[request-info] No pude guardar el mensaje: {e}')

    # 2) Cambiar estado a "waiting_user" opcionalmente (pausa SLA)
    change_status = bool(data.get('change_status', True))
    if change_status and ticket.status not in ('waiting_user', 'resolved', 'closed'):
        ticket.status = 'waiting_user'
        db.session.commit()

    # 3) Enviar email
    base_url = get_public_base_url()
    ticket_url = f'{base_url}/employee/ticket/{ticket.id}'

    subject = f'[DeskEli] Necesitamos más información · {ticket.ticket_number} · {ticket.title[:60]}'

    # Formatear el mensaje conservando saltos de linea
    message_html = safe_message.replace('\n', '<br>')

    body = f"""
    <html><body style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
        <div style="background:#f59e0b;color:white;padding:18px;border-radius:8px 8px 0 0;">
            <h2 style="margin:0;">📧 Solicitud de información</h2>
        </div>
        <div style="padding:22px;background:#f9fafb;border:1px solid #e5e7eb;border-top:none;">
            <p>Hola <strong>{creator.name or creator.username}</strong>,</p>
            <p>Sobre el ticket <strong>{ticket.ticket_number}</strong> — <em>"{ticket.title}"</em>:</p>

            <div style="background:white;border-left:4px solid #f59e0b;padding:14px 18px;margin:16px 0;border-radius:4px;">
                {message_html}
            </div>

            <p style="margin:16px 0;">Por favor respondé desde el ticket para agilizar la resolución.
            El técnico asignado está en copia de este correo.</p>

            <p style="text-align:center;margin:24px 0;">
                <a href="{ticket_url}" style="display:inline-block;background:#2563eb;color:white;padding:12px 28px;border-radius:6px;text-decoration:none;font-weight:bold;">
                    🎫 Responder en el ticket
                </a>
            </p>

            <div style="font-size:12px;color:#6b7280;margin-top:20px;border-top:1px solid #e5e7eb;padding-top:12px;">
                <strong>Detalles del ticket:</strong><br>
                Número: {ticket.ticket_number}<br>
                Prioridad: {ticket.priority.upper()}<br>
                Solicitó info: {current_user.name if current_user else 'Equipo TI'}<br>
                {f'Técnico asignado: {assignee.name}' if assignee else ''}
            </div>

            <p style="font-size:11px;color:#9ca3af;margin-top:14px;">
                Este es un mensaje automático de DeskEli. Podés responder desde el ticket usando el botón de arriba.
            </p>
        </div>
    </body></html>
    """

    cc_list = [assignee.email] if assignee and assignee.email and assignee.id != session['user_id'] else []
    # También CC al técnico que envía (si no es el mismo assignee)
    if current_user and current_user.email and current_user.id != (assignee.id if assignee else None):
        if current_user.email not in cc_list and current_user.email.lower() != creator.email.lower():
            cc_list.append(current_user.email)

    email_ok = False
    try:
        email_ok = send_email(
            to_email=creator.email,
            subject=subject,
            body=body,
            company=ticket.company,
            cc_emails=cc_list,
        )
    except Exception as e:
        print(f'[request-info] Error enviando email: {e}')

    log_audit(
        'request_info_from_user',
        session['user_id'],
        'ticket',
        ticket.id,
        f'Solicitud de info enviada por {current_user.email if current_user else "?"} '
        f'a {creator.email} (CC: {", ".join(cc_list) or "—"}) para ticket {ticket.ticket_number}. '
        f'Email: {"OK" if email_ok else "FALLO"}'
    )

    return jsonify({
        'success': True,
        'email_sent': email_ok,
        'to': creator.email,
        'cc': cc_list,
        'ticket_status': ticket.status,
        'message': ('✓ Solicitud enviada por email al usuario' if email_ok
                    else '⚠ Mensaje guardado en el ticket pero el email no pudo enviarse (revisar config SMTP)')
    })


@app.route('/api/ticket/<int:ticket_id>/message', methods=['POST'])
def api_add_message(ticket_id):
    """Agregar mensaje/comentario a un ticket (chat empleado-técnico)"""
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in (request.headers.get('Accept') or '')

    if 'user_id' not in session:
        if wants_json:
            return jsonify({'success': False, 'error': 'No autorizado'}), 401
        return redirect(url_for('login'))

    ticket = Ticket.query.get_or_404(ticket_id)

    # Validar acceso por empresa (admin master ve todas)
    if ticket.company not in admin_companies_scope():
        if wants_json:
            return jsonify({'success': False, 'error': f'No tienes acceso a tickets de "{ticket.company}".'}), 403
        # Volver al detalle del ticket, no a login
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        if session.get('role') == 'technician':
            return redirect(url_for('technician_dashboard'))
        return redirect(url_for('employee_dashboard'))

    # Helper para redirigir al detalle correcto según rol
    def _back_to_ticket():
        role = session.get('role')
        if role == 'admin':
            return redirect(url_for('admin_ticket_detail', ticket_id=ticket_id))
        if role == 'technician':
            return redirect(url_for('technician_ticket', ticket_id=ticket_id))
        return redirect(url_for('employee_ticket', ticket_id=ticket_id))

    # Obtener texto del mensaje (form POST)
    text = request.form.get('text', '').strip()
    if not text or len(text) < 1:
        if wants_json:
            return jsonify({'success': False, 'error': 'Mensaje vacío'}), 400
        return _back_to_ticket()

    text = sanitize_html(text[:2000])

    msg = Message(
        ticket_id=ticket_id,
        user_id=session['user_id'],
        text=text
    )
    db.session.add(msg)

    # Actualizar timestamp del ticket
    ticket.updated_at = datetime.now()
    db.session.commit()

    log_audit('add_message', session['user_id'], 'ticket', ticket_id,
              f'Mensaje agregado al ticket {ticket.ticket_number}')

    # Emitir por WebSocket
    try:
        emit_ticket_event(ticket.company, 'ticket_updated', {
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'new_message': True
        })
    except Exception:
        pass

    if wants_json:
        return jsonify({'success': True, 'message_id': msg.id})
    return _back_to_ticket()

# ═════════════════════════════════════════════════════════════════════════════
# REASIGNACIÓN CON MOTIVO RF-02-08 + ESCALAR A CRÍTICA RF-02-07
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/ticket/<int:ticket_id>/reassign', methods=['POST'])
def api_reassign_ticket(ticket_id):
    """Reasignar/asignar ticket (admins y técnicos)"""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({'success': False, 'error': 'Ticket no encontrado'}), 404

    scope = admin_companies_scope()
    if ticket.company not in scope:
        return jsonify({
            'success': False,
            'error': f'No puedes reasignar tickets de la empresa "{ticket.company}". Tu acceso está limitado a: {", ".join(scope)}.'
        }), 403

    # No permitir reasignar tickets cerrados o resueltos
    if ticket.status in ('resolved', 'closed'):
        status_label = 'cerrado' if ticket.status == 'closed' else 'resuelto'
        return jsonify({
            'success': False,
            'error': f'No se puede reasignar un ticket {status_label}. Reabre el ticket primero si necesitas cambiar el asignado.'
        }), 400

    data = request.json or {}
    new_assignee_id = data.get('assignee_id') or data.get('technician_id')
    reason = (data.get('reason') or 'Asignación directa').strip()

    if not new_assignee_id:
        return jsonify({'success': False, 'error': 'Debes seleccionar un técnico (assignee_id requerido)'}), 400

    try:
        new_assignee_id = int(new_assignee_id)
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'assignee_id inválido'}), 400

    # Verificar que el nuevo asignado existe
    new_assignee = User.query.get(new_assignee_id)
    if not new_assignee:
        return jsonify({'success': False, 'error': f'Usuario con id={new_assignee_id} no existe'}), 404

    # El nuevo asignado debe ser de la misma empresa que el TICKET (no del que reasigna)
    if new_assignee.company != ticket.company:
        return jsonify({
            'success': False,
            'error': f'El técnico "{new_assignee.name}" es de la empresa "{new_assignee.company}" y este ticket es de "{ticket.company}". No se puede asignar entre empresas.'
        }), 400

    # Debe ser técnico o admin (no asignar a un employee)
    if new_assignee.role not in ('technician', 'admin'):
        return jsonify({
            'success': False,
            'error': f'No se puede asignar a "{new_assignee.name}" — solo técnicos o admins (es rol "{new_assignee.role}").'
        }), 400

    if not new_assignee.is_active:
        return jsonify({
            'success': False,
            'error': f'El usuario "{new_assignee.name}" está inactivo.'
        }), 400

    old_assignee = ticket.assignee.name if ticket.assignee else 'Sin asignar'
    ticket.assignee_id = new_assignee_id
    ticket.updated_at = datetime.now()
    if ticket.status == 'open':
        ticket.status = 'in_progress'

    db.session.commit()

    log_audit('reassign_ticket', session['user_id'], 'ticket', ticket_id,
              f'Asignado de {old_assignee} a {new_assignee.name}: {reason}')

    # Emitir evento en tiempo real
    try:
        emit_ticket_event(ticket.company, 'ticket_assigned', {
            'ticket_number': ticket.ticket_number,
            'ticket_id': ticket.id,
            'assignee_id': new_assignee.id,
            'assignee_name': new_assignee.name,
            'assigned_by': session.get('name', 'Admin')
        })
    except Exception as e:
        print(f'[WARN] WebSocket emit: {e}')

    # Notificar por email al técnico asignado (no bloquea la respuesta si falla)
    try:
        notify_ticket_assigned(
            ticket=ticket,
            new_assignee=new_assignee,
            assigned_by_name=session.get('name', 'Administrador'),
            reason=reason if reason and reason != 'Asignación directa' else ''
        )
    except Exception as e:
        print(f'[WARN] No se pudo notificar asignación por email: {e}')

    return jsonify({
        'success': True,
        'message': f'Ticket asignado a {new_assignee.name}',
        'assignee_name': new_assignee.name,
        'new_status': ticket.status
    })

@app.route('/api/notifications/push-subscribe', methods=['POST'])
def api_push_subscribe():
    """Suscribirse a notificaciones push (BAJO)"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    data = request.json
    endpoint = data.get('endpoint')
    auth = data.get('auth')
    p256dh = data.get('p256dh')

    # Guardar suscripción (en producción: guardar en BD)
    # Por ahora: solo registrar
    log_audit('push_subscribe', session['user_id'], 'notification', None,
              f'Usuario suscrito a notificaciones push')

    return jsonify({
        'success': True,
        'message': 'Suscripción a notificaciones registrada'
    })

@app.route('/api/system/metrics', methods=['GET'])
def api_system_metrics():
    """Métricas del sistema (BAJO)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']

    # Contar tickets
    total = Ticket.query.filter_by(company=company).count()
    open_t = Ticket.query.filter_by(company=company, status='open').count()
    progress_t = Ticket.query.filter_by(company=company, status='in_progress').count()
    resolved_t = Ticket.query.filter_by(company=company, status='resolved').count()

    # Contar usuarios
    total_users = User.query.filter_by(company=company).count()
    technicians = User.query.filter_by(company=company, role='technician').count()

    # Uptime (desde inicio)
    uptime = datetime.now() - datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    return jsonify({
        'success': True,
        'tickets': {
            'total': total,
            'open': open_t,
            'in_progress': progress_t,
            'resolved': resolved_t
        },
        'users': {
            'total': total_users,
            'technicians': technicians
        },
        'system': {
            'uptime_hours': round(uptime.total_seconds() / 3600, 1),
            'timestamp': datetime.now().isoformat()
        }
    })

@app.route('/api/tickets/filter', methods=['POST'])
def api_filter_tickets():
    """Filtros avanzados para cola de técnicos (BAJO)"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    data = request.json
    company = session['company']

    query = Ticket.query.filter_by(company=company)

    # Filtro por estado
    if data.get('status'):
        query = query.filter_by(status=data['status'])

    # Filtro por prioridad
    if data.get('priority'):
        query = query.filter_by(priority=data['priority'])

    # Filtro por categoría
    if data.get('category'):
        query = query.filter_by(category=data['category'])

    # Filtro por técnico asignado
    if data.get('assignee_id'):
        query = query.filter_by(assignee_id=data['assignee_id'])

    # Filtro por SLA crítico
    if data.get('critical_sla'):
        now = datetime.now()
        query = query.filter(Ticket.sla_deadline < now + timedelta(minutes=30))

    tickets = query.all()

    return jsonify({
        'success': True,
        'count': len(tickets),
        'tickets': [{
            'id': t.id,
            'ticket_number': t.ticket_number,
            'title': t.title,
            'status': t.status,
            'priority': t.priority,
            'sla_remaining': t.sla_remaining
        } for t in tickets]
    })

@app.route('/api/ticket/<int:ticket_id>/escalate', methods=['POST'])
def api_escalate_priority(ticket_id):
    """Escalar ticket a prioridad Crítica (RF-02-07)"""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    ticket = Ticket.query.get(ticket_id)
    if not ticket:
        return jsonify({'success': False, 'error': 'Ticket no encontrado'}), 404
    if ticket.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': f'No tienes acceso a tickets de la empresa "{ticket.company}".'}), 403

    # No permitir escalar tickets cerrados o resueltos
    if ticket.status in ('resolved', 'closed'):
        status_label = 'cerrado' if ticket.status == 'closed' else 'resuelto'
        return jsonify({
            'success': False,
            'error': f'No se puede escalar un ticket {status_label}.'
        }), 400

    # Si ya está crítico, no hacer nada
    if ticket.priority == 'critical':
        return jsonify({'success': False, 'error': 'El ticket ya es crítico'}), 400

    data = request.get_json(silent=True) or {}
    reason = (data.get('reason') or 'Sin motivo especificado').strip()[:500]

    old_priority = ticket.priority
    ticket.priority = 'critical'
    ticket.sla_minutes = 60  # SLA crítico: 1 hora
    ticket.sla_deadline = datetime.now() + timedelta(minutes=60)
    ticket.updated_at = datetime.now()

    # Agregar motivo a la descripción
    escalation_note = f"\n\n--- 🚨 ESCALACIÓN ({datetime.now().strftime('%Y-%m-%d %H:%M')}) ---\nEscalado por: {session.get('name', 'Usuario')}\nDe: {old_priority.upper()} → CRITICAL\nMotivo: {reason}"
    ticket.description = (ticket.description or '') + escalation_note

    db.session.commit()

    log_audit('escalate_priority', session['user_id'], 'ticket', ticket_id,
              f'Escalado de {old_priority} a CRITICAL - Motivo: {reason}')

    # Emitir en tiempo real
    try:
        emit_ticket_event(ticket.company, 'ticket_escalated', {
            'ticket_number': ticket.ticket_number,
            'new_priority': 'critical',
            'old_priority': old_priority,
            'escalated_by': session.get('name', 'Usuario'),
            'reason': reason
        })
    except Exception as e:
        print(f'[WARN] WebSocket emit: {e}')

    return jsonify({'success': True, 'message': 'Ticket escalado a Crítico'})

@app.route('/api/ticket/<int:ticket_id>/status', methods=['PUT'])
def api_update_ticket_status(ticket_id):
    """Actualizar estado del ticket (para Kanban)"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    try:
        data = request.get_json()
        new_status = data.get('status', '').strip()

        if new_status not in ['open', 'in_progress', 'resolved', 'closed']:
            return jsonify({'success': False, 'error': 'Estado inválido'}), 400

        ticket = Ticket.query.get(ticket_id)
        if not ticket:
            return jsonify({'success': False, 'error': 'Ticket no encontrado'}), 404
        if ticket.company not in admin_companies_scope():
            return jsonify({'success': False, 'error': f'No tienes acceso a tickets de "{ticket.company}".'}), 403

        # VALIDACIÓN: no permitir pasar a resolved/closed si hay subtareas pendientes
        if new_status in ('resolved', 'closed') and ticket.status not in ('resolved', 'closed'):
            pending = Subtask.query.filter(
                Subtask.ticket_id == ticket_id,
                Subtask.status.notin_(['resolved', 'closed', 'cancelled'])
            ).all()
            if pending:
                pending_list = [f"#{s.subtask_number or s.id}: {s.title[:60]}" for s in pending[:5]]
                extra = f' (+ {len(pending) - 5} más)' if len(pending) > 5 else ''
                return jsonify({
                    'success': False,
                    'error': f'No se puede cerrar: tiene {len(pending)} subtarea(s) pendiente(s). Resolvé primero:\n\n• ' + '\n• '.join(pending_list) + extra,
                    'pending_subtasks_count': len(pending)
                }), 400

        old_status = ticket.status
        ticket.status = new_status
        ticket.updated_at = datetime.now()

        if new_status == 'resolved':
            ticket.resolved_at = datetime.now()

        db.session.commit()

        log_audit('update_status', session['user_id'], 'ticket', ticket_id,
                  f'Estado cambiado de {old_status} a {new_status}')

        # Emitir evento en tiempo real
        emit_ticket_event(ticket.company, 'ticket_status_changed', {
            'ticket_number': ticket.ticket_number,
            'old_status': old_status,
            'new_status': new_status,
            'changed_by': session.get('name', 'Usuario')
        })

        return jsonify({'success': True, 'message': f'Ticket movido a {new_status}'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# SUBTAREAS DE TICKETS (cases asociados al ticket padre)
# ============================================================

def _serialize_subtask(s):
    company_code = s.ticket.company if s.ticket else None
    company_info = COMPANY_COLORS.get(company_code, {}) if company_code else {}
    return {
        'id': s.id,
        'subtask_number': s.subtask_number,
        'ticket_id': s.ticket_id,
        'parent_number': s.ticket.ticket_number if s.ticket else None,
        'company': company_code,
        'company_label': company_info.get('name', company_code.title()) if company_code else '',
        'company_icon': company_info.get('icon', '🏢'),
        'title': s.title,
        'description': s.description or '',
        'category': s.category or 'General',
        'status': s.status,
        'priority': s.priority,
        'sla_minutes': s.sla_minutes,
        'sla_deadline': s.sla_deadline.strftime('%Y-%m-%d %H:%M') if s.sla_deadline else None,
        'sla_remaining': s.sla_remaining,
        'sla_expired': s.sla_expired,
        'time_worked_seconds': s.time_worked_seconds or 0,
        'assignee_id': s.assignee_id,
        'assignee_name': s.assignee.name if s.assignee else None,
        'created_by_name': s.created_by.name if s.created_by else None,
        'order_idx': s.order_idx,
        'created_at': s.created_at.strftime('%Y-%m-%d %H:%M') if s.created_at else None,
        'updated_at': s.updated_at.strftime('%Y-%m-%d %H:%M') if s.updated_at else None,
        'resolved_at': s.resolved_at.strftime('%Y-%m-%d %H:%M') if s.resolved_at else None,
        'attachments': [_serialize_attachment(a) for a in (s.attachments or [])],
        'attachment_count': len(s.attachments or [])
    }


@app.route('/api/ticket/<int:ticket_id>/subtasks', methods=['GET'])
def api_subtasks_list(ticket_id):
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.company != session['company']:
        return jsonify({'success': False}), 403

    subtasks = Subtask.query.filter_by(ticket_id=ticket_id).order_by(Subtask.order_idx.asc(), Subtask.id.asc()).all()
    result = [_serialize_subtask(s) for s in subtasks]

    total = len(result)
    done = sum(1 for s in result if s['status'] == 'resolved')
    progress = int((done / total) * 100) if total > 0 else 0

    technicians = User.query.filter_by(
        company=session['company'], role='technician', is_active=True
    ).order_by(User.name).all()
    technicians_list = [{'id': u.id, 'name': u.name} for u in technicians]

    return jsonify({
        'success': True,
        'subtasks': result,
        'total': total,
        'done': done,
        'progress': progress,
        'technicians': technicians_list
    })


@app.route('/api/ticket/<int:ticket_id>/subtasks', methods=['POST'])
def api_subtasks_create(ticket_id):
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.company != session['company']:
        return jsonify({'success': False}), 403

    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({'success': False, 'error': 'Título requerido'}), 400

    priority = data.get('priority') or ticket.priority or 'medium'
    if priority not in ['low', 'medium', 'high', 'critical']:
        priority = 'medium'
    category = (data.get('category') or ticket.category or 'General')[:100]
    sla_minutes = get_sla_minutes_for_priority(priority)

    last = Subtask.query.filter_by(ticket_id=ticket_id).order_by(Subtask.order_idx.desc()).first()
    next_idx = (last.order_idx + 1) if last else 0

    subtask = Subtask(
        ticket_id=ticket_id,
        subtask_number=get_next_subtask_number(ticket),
        title=title[:255],
        description=(data.get('description') or '').strip() or None,
        category=category,
        status='open',
        priority=priority,
        sla_minutes=sla_minutes,
        sla_deadline=compute_sla_deadline(datetime.now(), sla_minutes, ticket.company),
        assignee_id=data.get('assignee_id') or ticket.assignee_id,
        created_by_id=session['user_id'],
        order_idx=next_idx
    )
    db.session.add(subtask)
    db.session.commit()

    log_audit('subtask_create', session['user_id'], 'ticket', ticket_id,
              f'Subtarea {subtask.subtask_number} creada: {title[:60]}')

    try:
        emit_ticket_event(ticket.company, 'subtask_changed', {
            'ticket_id': ticket_id, 'subtask_id': subtask.id, 'action': 'created'
        })
    except Exception:
        pass

    return jsonify({'success': True, 'subtask': _serialize_subtask(subtask)})


@app.route('/api/technician/my-subtasks', methods=['GET'])
def api_my_subtasks():
    """Subtareas. scope=mine (default) → asignadas al técnico; scope=team → del grupo de subroles."""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    scope = request.args.get('scope', 'mine')
    user = User.query.get(session['user_id'])

    try:
        if scope == 'mine':
            # Incluye subtareas asignadas a mis identidades (self + espejos +
            # origen) sin importar la empresa. Consolida vista cross-company.
            identity_ids = get_user_identity_ids(user)
            query = Subtask.query.filter(Subtask.assignee_id.in_(identity_ids))
        elif scope == 'team':
            # "De mis grupos" = subtareas asignadas a tecnicos que comparten
            # al menos 1 subrol conmigo, dentro de mi empresa actual.
            group_ids = get_my_group_user_ids(user)
            query = Subtask.query.join(Ticket, Subtask.ticket_id == Ticket.id).filter(
                Ticket.company == session['company'],
                Subtask.assignee_id.in_(group_ids)
            )
        else:
            query = Subtask.query.join(Ticket, Subtask.ticket_id == Ticket.id).filter(
                Ticket.company == session['company']
            )
        subtasks = query.all()

        # Orden: resueltas al final, dentro de cada grupo por SLA ascendente (None al final)
        status_rank = {'in_progress': 0, 'open': 1, 'resolved': 2}
        far_future = datetime(2999, 12, 31)
        subtasks.sort(key=lambda s: (
            status_rank.get(s.status, 1),
            s.sla_deadline or far_future
        ))

        return jsonify({
            'success': True,
            'subtasks': [_serialize_subtask(s) for s in subtasks],
            'total': len(subtasks)
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/subtask/<int:subtask_id>', methods=['GET'])
def api_subtask_get(subtask_id):
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    subtask = Subtask.query.get_or_404(subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403
    return jsonify({'success': True, 'subtask': _serialize_subtask(subtask)})


@app.route('/api/subtask/<int:subtask_id>', methods=['PATCH'])
def api_subtask_update(subtask_id):
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    subtask = Subtask.query.get_or_404(subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    data = request.get_json() or {}

    if 'title' in data:
        new_title = (data['title'] or '').strip()
        if not new_title:
            return jsonify({'success': False, 'error': 'Título requerido'}), 400
        subtask.title = new_title[:255]
    if 'description' in data:
        subtask.description = (data['description'] or '').strip() or None
    if 'category' in data:
        subtask.category = (data['category'] or 'General')[:100]
    if 'assignee_id' in data:
        subtask.assignee_id = data['assignee_id'] or None
    if 'priority' in data:
        new_priority = data['priority']
        if new_priority not in ['low', 'medium', 'high', 'critical']:
            return jsonify({'success': False, 'error': 'Prioridad inválida'}), 400
        if new_priority != subtask.priority:
            subtask.priority = new_priority
            subtask.sla_minutes = get_sla_minutes_for_priority(new_priority)
            subtask.sla_deadline = datetime.now() + timedelta(minutes=subtask.sla_minutes)
    if 'time_worked_seconds' in data:
        try:
            subtask.time_worked_seconds = max(0, int(data['time_worked_seconds']))
        except (ValueError, TypeError):
            pass
    if 'status' in data:
        new_status = data['status']
        if new_status not in ['open', 'in_progress', 'resolved']:
            return jsonify({'success': False, 'error': 'Estado inválido'}), 400
        old_status = subtask.status
        subtask.status = new_status
        if new_status == 'resolved' and old_status != 'resolved':
            subtask.resolved_at = datetime.now()
            subtask.completed_at = datetime.now()
        elif new_status != 'resolved':
            subtask.resolved_at = None
            subtask.completed_at = None

    db.session.commit()

    log_audit('subtask_update', session['user_id'], 'ticket', subtask.ticket_id,
              f'Subtarea {subtask.subtask_number or subtask_id} actualizada')

    try:
        emit_ticket_event(ticket.company, 'subtask_changed', {
            'ticket_id': subtask.ticket_id, 'subtask_id': subtask.id, 'action': 'updated'
        })
    except Exception:
        pass

    return jsonify({'success': True, 'subtask': _serialize_subtask(subtask)})


@app.route('/api/subtask/<int:subtask_id>', methods=['DELETE'])
def api_subtask_delete(subtask_id):
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    subtask = Subtask.query.get_or_404(subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    ticket_id = subtask.ticket_id
    subtask_num = subtask.subtask_number
    db.session.delete(subtask)
    db.session.commit()

    log_audit('subtask_delete', session['user_id'], 'ticket', ticket_id,
              f'Subtarea {subtask_num or subtask_id} eliminada')

    try:
        emit_ticket_event(ticket.company, 'subtask_changed', {
            'ticket_id': ticket_id, 'subtask_id': subtask_id, 'action': 'deleted'
        })
    except Exception:
        pass

    return jsonify({'success': True})


# ============================================================
# ADJUNTOS DE SUBTAREAS
# ============================================================

def _serialize_attachment(a):
    return {
        'id': a.id,
        'subtask_id': a.subtask_id,
        'original_name': a.original_name,
        'mime_type': a.mime_type,
        'size_bytes': a.size_bytes or 0,
        'uploaded_by_name': a.uploaded_by.name if a.uploaded_by else None,
        'uploaded_at': a.uploaded_at.strftime('%Y-%m-%d %H:%M') if a.uploaded_at else None,
        'download_url': f'/api/subtask/attachment/{a.id}/download'
    }


@app.route('/api/subtask/<int:subtask_id>/messages', methods=['GET'])
def api_subtask_messages_list(subtask_id):
    """Listar mensajes/comentarios de una subtarea."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autenticado'}), 401
    subtask = Subtask.query.get_or_404(subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403

    messages = Message.query.filter_by(subtask_id=subtask_id).order_by(Message.created_at.asc()).all()
    return jsonify({
        'success': True,
        'messages': [{
            'id': m.id,
            'text': m.text,
            'user_id': m.user_id,
            'user_name': m.user.name if m.user else '—',
            'user_role': m.user.role if m.user else None,
            'created_at': m.created_at.strftime('%Y-%m-%d %H:%M') if m.created_at else None,
            'is_mine': m.user_id == session.get('user_id'),
        } for m in messages]
    })


@app.route('/api/subtask/<int:subtask_id>/messages', methods=['POST'])
def api_subtask_messages_create(subtask_id):
    """Agregar un mensaje/comentario a una subtarea."""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autenticado'}), 401
    subtask = Subtask.query.get_or_404(subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403

    data = request.get_json() or {}
    text = (data.get('text') or '').strip()
    if not text:
        return jsonify({'success': False, 'error': 'Mensaje vacío'}), 400
    if len(text) > 5000:
        return jsonify({'success': False, 'error': 'Mensaje demasiado largo (máx 5000 chars)'}), 400

    # Sanitizar HTML (los mensajes de admin/técnicos también pasan por bleach)
    if 'sanitize_html' in globals():
        text = sanitize_html(text)

    msg = Message(
        ticket_id=subtask.ticket_id,  # heredar del padre
        subtask_id=subtask_id,
        user_id=session['user_id'],
        text=text,
    )
    db.session.add(msg)
    db.session.commit()

    log_audit('subtask_message_added', session['user_id'], 'subtask', subtask_id,
              f'Mensaje agregado a subtarea {subtask.subtask_number or subtask_id}')

    # Broadcast websocket a los técnicos y admin de la empresa
    try:
        emit_ticket_event(ticket.company, 'subtask_message', {
            'subtask_id': subtask_id,
            'ticket_id': subtask.ticket_id,
            'message': {
                'id': msg.id,
                'text': msg.text,
                'user_name': msg.user.name if msg.user else '—',
                'created_at': msg.created_at.strftime('%Y-%m-%d %H:%M'),
            }
        })
    except Exception:
        pass

    return jsonify({
        'success': True,
        'message': {
            'id': msg.id,
            'text': msg.text,
            'user_id': msg.user_id,
            'user_name': msg.user.name if msg.user else '—',
            'user_role': msg.user.role if msg.user else None,
            'created_at': msg.created_at.strftime('%Y-%m-%d %H:%M'),
            'is_mine': True,
        }
    })


@app.route('/api/subtask/<int:subtask_id>/attachments', methods=['GET'])
def api_subtask_attachments_list(subtask_id):
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    subtask = Subtask.query.get_or_404(subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403
    attachments = SubtaskAttachment.query.filter_by(subtask_id=subtask_id).order_by(SubtaskAttachment.uploaded_at.desc()).all()
    return jsonify({
        'success': True,
        'attachments': [_serialize_attachment(a) for a in attachments]
    })


@app.route('/api/subtask/<int:subtask_id>/attachments', methods=['POST'])
def api_subtask_attachments_upload(subtask_id):
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    subtask = Subtask.query.get_or_404(subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    from werkzeug.utils import secure_filename

    if 'files' not in request.files and 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se recibieron archivos'}), 400

    files = request.files.getlist('files') if 'files' in request.files else [request.files['file']]
    saved = []
    errors = []

    for f in files:
        if not f or not f.filename:
            continue
        if not _allowed_attachment(f.filename):
            errors.append(f"{f.filename}: tipo de archivo no permitido")
            continue

        try:
            # Comprimir automáticamente si es imagen
            out_bytes, out_filename, out_mime, stats = compress_upload(f)
            safe = secure_filename(out_filename) or 'archivo'
            ext = safe.rsplit('.', 1)[1].lower() if '.' in safe else ''
            stored = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex
            path = os.path.join(app.config['UPLOAD_FOLDER'], stored)
            with open(path, 'wb') as fh:
                fh.write(out_bytes)
            size = stats['final_size']
            final_name = (out_filename or f.filename)[:255]
            final_mime = (out_mime or f.mimetype or '')[:120]
        except Exception as e:
            errors.append(f"{f.filename}: error al guardar ({e})")
            continue

        att = SubtaskAttachment(
            subtask_id=subtask_id,
            original_name=final_name,
            stored_name=stored,
            mime_type=final_mime,
            size_bytes=size,
            uploaded_by_id=session['user_id']
        )
        db.session.add(att)
        db.session.flush()
        saved.append(_serialize_attachment(att))

    db.session.commit()

    log_audit('subtask_attach', session['user_id'], 'ticket', subtask.ticket_id,
              f'Subtarea {subtask.subtask_number or subtask_id}: {len(saved)} adjunto(s)')

    return jsonify({
        'success': True,
        'attachments': saved,
        'errors': errors
    })


@app.route('/api/subtask/attachment/<int:att_id>/download', methods=['GET'])
def api_subtask_attachment_download(att_id):
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    att = SubtaskAttachment.query.get_or_404(att_id)
    subtask = Subtask.query.get(att.subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id) if subtask else None
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403
    path = os.path.join(app.config['UPLOAD_FOLDER'], att.stored_name)
    if not os.path.exists(path):
        return jsonify({'success': False, 'error': 'Archivo no encontrado en disco'}), 404
    return send_file(path, as_attachment=True, download_name=att.original_name)


@app.route('/api/subtask/attachment/<int:att_id>', methods=['DELETE'])
def api_subtask_attachment_delete(att_id):
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    att = SubtaskAttachment.query.get_or_404(att_id)
    subtask = Subtask.query.get(att.subtask_id)
    ticket = Ticket.query.get(subtask.ticket_id) if subtask else None
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    path = os.path.join(app.config['UPLOAD_FOLDER'], att.stored_name)
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass

    db.session.delete(att)
    db.session.commit()
    log_audit('subtask_attach_delete', session['user_id'], 'ticket', subtask.ticket_id if subtask else None,
              f'Adjunto eliminado: {att.original_name}')
    return jsonify({'success': True})


# ============================================================
# ADJUNTOS DE TICKETS
# ============================================================

def _serialize_ticket_attachment(a):
    return {
        'id': a.id,
        'ticket_id': a.ticket_id,
        'original_name': a.original_name,
        'mime_type': a.mime_type,
        'size_bytes': a.size_bytes or 0,
        'uploaded_by_name': a.uploaded_by.name if a.uploaded_by else None,
        'uploaded_at': a.uploaded_at.strftime('%Y-%m-%d %H:%M') if a.uploaded_at else None,
        'download_url': f'/api/ticket/attachment/{a.id}/download'
    }


@app.route('/api/ticket/<int:ticket_id>/attachments', methods=['GET'])
def api_ticket_attachments_list(ticket_id):
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.company != session['company']:
        return jsonify({'success': False}), 403
    attachments = TicketAttachment.query.filter_by(ticket_id=ticket_id).order_by(TicketAttachment.uploaded_at.desc()).all()
    return jsonify({
        'success': True,
        'attachments': [_serialize_ticket_attachment(a) for a in attachments]
    })


@app.route('/api/ticket/<int:ticket_id>/attachments', methods=['POST'])
def api_ticket_attachments_upload(ticket_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.company != session['company']:
        return jsonify({'success': False}), 403

    from werkzeug.utils import secure_filename

    if 'files' not in request.files and 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se recibieron archivos'}), 400

    files = request.files.getlist('files') if 'files' in request.files else [request.files['file']]
    saved = []
    errors = []

    for f in files:
        if not f or not f.filename:
            continue
        if not _allowed_attachment(f.filename):
            errors.append(f"{f.filename}: tipo de archivo no permitido")
            continue

        safe = secure_filename(f.filename) or 'archivo'
        ext = safe.rsplit('.', 1)[1].lower() if '.' in safe else ''
        stored = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex
        path = os.path.join(app.config['TICKET_UPLOAD_FOLDER'], stored)
        try:
            f.save(path)
            size = os.path.getsize(path)
        except Exception as e:
            errors.append(f"{f.filename}: error al guardar ({e})")
            continue

        att = TicketAttachment(
            ticket_id=ticket_id,
            original_name=f.filename[:255],
            stored_name=stored,
            mime_type=(f.mimetype or '')[:120],
            size_bytes=size,
            uploaded_by_id=session['user_id']
        )
        db.session.add(att)
        db.session.flush()
        saved.append(_serialize_ticket_attachment(att))

    db.session.commit()
    log_audit('ticket_attach', session['user_id'], 'ticket', ticket_id,
              f'{len(saved)} adjunto(s) en ticket {ticket.ticket_number}')

    return jsonify({'success': True, 'attachments': saved, 'errors': errors})


@app.route('/api/ticket/attachment/<int:att_id>/download', methods=['GET'])
def api_ticket_attachment_download(att_id):
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    att = TicketAttachment.query.get_or_404(att_id)
    ticket = Ticket.query.get(att.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403
    path = os.path.join(app.config['TICKET_UPLOAD_FOLDER'], att.stored_name)
    if not os.path.exists(path):
        return jsonify({'success': False, 'error': 'Archivo no encontrado en disco'}), 404
    return send_file(path, as_attachment=True, download_name=att.original_name)


@app.route('/api/ticket/attachment/<int:att_id>', methods=['DELETE'])
def api_ticket_attachment_delete(att_id):
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    att = TicketAttachment.query.get_or_404(att_id)
    ticket = Ticket.query.get(att.ticket_id)
    if not ticket or ticket.company not in admin_companies_scope():
        return jsonify({'success': False}), 403

    path = os.path.join(app.config['TICKET_UPLOAD_FOLDER'], att.stored_name)
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass

    db.session.delete(att)
    db.session.commit()
    log_audit('ticket_attach_delete', session['user_id'], 'ticket', att.ticket_id,
              f'Adjunto eliminado: {att.original_name}')
    return jsonify({'success': True})


# ============================================================
# NUEVAS FUNCIONES: Chat especialistas, Conectividad, Escalaciones
# ============================================================

def _dm_ticket_number(company, user_a_id, user_b_id):
    """Devuelve el ticket_number del canal DM 1-a-1.
    Ordena los IDs para garantizar un único canal por par."""
    a, b = sorted([int(user_a_id), int(user_b_id)])
    return f'DM-{company.upper()}-{a}-{b}'


def _get_or_create_chat_ticket(ticket_number, company, creator_id, title, description):
    """Helper: obtiene (o crea si no existe) el ticket-contenedor del chat."""
    t = Ticket.query.filter_by(ticket_number=ticket_number).first()
    if t:
        return t
    t = Ticket(
        ticket_number=ticket_number,
        title=title,
        description=description,
        status='open',
        priority='low',
        company=company,
        creator_id=creator_id,
        category='Chat Interno'
    )
    db.session.add(t)
    db.session.flush()
    return t


@app.route('/api/chat/specialists/unread-summary', methods=['GET'])
def api_chat_specialists_unread():
    """Resumen de todos los DMs en los que participa el usuario.
    Para cada DM devuelve: other_user_id, other_user_name, last_msg_id,
    last_msg_from_id, last_msg_text (preview), last_msg_time.

    El cliente compara last_msg_id con su última lectura local (localStorage)
    para saber cuáles tiene sin leer."""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    me_id = session['user_id']
    company = session['company']
    # Patrón ticket_number: DM-{COMPANY}-{minId}-{maxId}
    prefix = f'DM-{company.upper()}-'

    # Tickets DM que incluyen el ID del usuario al final O en medio
    dm_tickets = Ticket.query.filter(
        Ticket.ticket_number.like(prefix + '%')
    ).all()

    summary = []
    for t in dm_tickets:
        # Parsear los IDs del ticket_number
        try:
            parts = t.ticket_number.replace(prefix, '').split('-')
            if len(parts) != 2:
                continue
            a, b = int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            continue
        # ¿El usuario actual participa?
        if me_id not in (a, b):
            continue
        other_id = b if a == me_id else a
        # Último mensaje del ticket
        last = Message.query.filter_by(ticket_id=t.id).order_by(Message.id.desc()).first()
        if not last:
            continue
        other = User.query.get(other_id)
        if not other:
            continue
        summary.append({
            'other_user_id': other_id,
            'other_user_name': other.name,
            'other_user_role': other.role,
            'last_msg_id': last.id,
            'last_msg_from_id': last.user_id,
            'last_msg_preview': (last.text or '')[:80],
            'last_msg_time': last.created_at.strftime('%H:%M %d/%m') if last.created_at else '',
        })

    # Ordenar por last_msg_id descendente (más reciente primero)
    summary.sort(key=lambda x: -x['last_msg_id'])
    return jsonify({'success': True, 'dms': summary})


@app.route('/api/chat/specialists/messages', methods=['GET'])
def api_chat_specialists_get():
    """Obtener mensajes del chat.
    Si se pasa ?recipient_id=X → chat PRIVADO 1-a-1 entre el usuario actual y X (solo ellos lo ven).
    Sin recipient_id → chat grupal de especialistas (legacy, todos los técnicos+admins lo ven)."""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    company = session['company']
    me_id = session['user_id']
    recipient_id = request.args.get('recipient_id', type=int)

    if recipient_id and recipient_id != me_id:
        # DM 1-a-1: ambos participantes deben ser técnico/admin de la misma empresa
        other = User.query.get(recipient_id)
        if not other or other.company != company or other.role not in ('technician', 'admin'):
            return jsonify({'success': False, 'error': 'Destinatario inválido'}), 400
        ticket_number = _dm_ticket_number(company, me_id, recipient_id)
        chat_ticket = Ticket.query.filter_by(ticket_number=ticket_number).first()
        is_dm = True
    else:
        # Grupal (legacy)
        chat_ticket = Ticket.query.filter_by(
            company=company,
            ticket_number=f'CHAT-{company.upper()}-SPECIALISTS'
        ).first()
        is_dm = False

    if not chat_ticket:
        return jsonify({'success': True, 'messages': [], 'is_dm': is_dm})

    messages = Message.query.filter_by(ticket_id=chat_ticket.id).order_by(Message.created_at.asc()).limit(100).all()
    result = [{
        'id': m.id,
        'user_id': m.user_id,
        'user_name': m.user.name if m.user else 'Desconocido',
        'user_role': m.user.role if m.user else '',
        'content': m.text,
        'timestamp': m.created_at.strftime('%H:%M %d/%m')
    } for m in messages]

    return jsonify({'success': True, 'messages': result, 'is_dm': is_dm})


@app.route('/api/chat/specialists/send', methods=['POST'])
def api_chat_specialists_send():
    """Enviar mensaje al chat.
    Si body tiene recipient_id → chat PRIVADO 1-a-1; si no → grupal (legacy)."""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    try:
        data = request.get_json() or {}
        content = (data.get('content') or '').strip()
        if not content:
            return jsonify({'success': False, 'error': 'Mensaje vacío'}), 400

        company = session['company']
        me_id = session['user_id']
        recipient_id = data.get('recipient_id')
        try:
            recipient_id = int(recipient_id) if recipient_id else None
        except (TypeError, ValueError):
            recipient_id = None

        if recipient_id and recipient_id != me_id:
            # DM 1-a-1
            other = User.query.get(recipient_id)
            if not other or other.company != company or other.role not in ('technician', 'admin'):
                return jsonify({'success': False, 'error': 'Destinatario inválido'}), 400
            ticket_number = _dm_ticket_number(company, me_id, recipient_id)
            title = f'DM privado entre {company}'
            description = f'Conversación privada 1-a-1 — solo participantes la ven.'
            chat_ticket = _get_or_create_chat_ticket(ticket_number, company, me_id, title, description)
            is_dm = True
        else:
            # Grupal
            ticket_number = f'CHAT-{company.upper()}-SPECIALISTS'
            title = f'Chat interno de especialistas - {company}'
            description = 'Canal grupal entre técnicos y administradores de la empresa.'
            chat_ticket = _get_or_create_chat_ticket(ticket_number, company, me_id, title, description)
            is_dm = False

        msg = Message(
            ticket_id=chat_ticket.id,
            user_id=me_id,
            text=content
        )
        db.session.add(msg)
        db.session.commit()

        payload = {
            'user_id': me_id,
            'user_name': session.get('name', 'Usuario'),
            'user_role': session.get('role', ''),
            'content': content,
            'timestamp': datetime.now().strftime('%H:%M %d/%m'),
            'is_dm': is_dm,
        }

        # Emitir vía WebSocket
        try:
            if is_dm:
                # Solo a las salas personales de ambos participantes
                socketio.emit('chat_message', {**payload, 'dm_with': recipient_id},
                              room=f'user_dm_{me_id}')
                socketio.emit('chat_message', {**payload, 'dm_with': me_id},
                              room=f'user_dm_{recipient_id}')
            else:
                # Sala grupal por empresa
                socketio.emit('chat_message', payload, room=f'chat_specialists_{company}')
        except Exception as e:
            print(f"[WARN] No se pudo emitir evento chat: {e}")

        return jsonify({'success': True, 'message': payload, 'is_dm': is_dm})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


def ping_server(host, port, timeout=2):
    """Hace ping TCP a un servidor y retorna (status, latency_ms)"""
    import socket
    import time
    start = time.time()
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((host, port))
        sock.close()
        latency = round((time.time() - start) * 1000, 1)
        if result == 0:
            return ('online' if latency < 500 else 'slow', latency)
        return ('offline', 0)
    except Exception:
        return ('offline', 0)


def create_alarm_ticket(server, company):
    """Crea un ticket cuando un servidor cae. Prioridad crítica/alta según is_critical.

    Reglas anti-spam vs detección de eventos nuevos:
    - Si EXISTE un ticket abierto del mismo servidor creado en las últimas
      4 horas → reutiliza ese (evita crear 10 tickets en cadenas de fallos).
    - Si el ticket abierto es MÁS VIEJO que 4 horas → crea uno nuevo
      (es una caída nueva, en otro día/turno, debe registrarse aparte).
    """
    admin_user = User.query.filter_by(company=company, role='admin').first()
    if not admin_user:
        admin_user = User.query.filter_by(company=company).first()
    if not admin_user:
        return None

    DEDUPE_WINDOW_HOURS = 4
    cutoff = datetime.now() - timedelta(hours=DEDUPE_WINDOW_HOURS)

    # Buscar ticket reciente del mismo servidor
    recent_existing = Ticket.query.filter(
        Ticket.company == company,
        Ticket.title.like(f'%{server.name}%'),
        Ticket.category.in_(['Infraestructura', 'Servidores']),
        Ticket.status.in_(['open', 'in_progress']),
        Ticket.created_at >= cutoff
    ).order_by(Ticket.created_at.desc()).first()
    if recent_existing:
        # Existe uno reciente → no spam, sólo agregamos un mensaje de seguimiento
        try:
            db.session.add(Message(
                ticket_id=recent_existing.id,
                user_id=admin_user.id,
                text=f'🔁 El monitor detectó OTRA caída del servidor {server.name} a las '
                     f'{datetime.now().strftime("%H:%M:%S")} (sumando a este ticket existente).'
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()
        return recent_existing

    # Si hay un ticket abierto pero viejo (>4h), registramos en él y creamos uno NUEVO
    old_existing = Ticket.query.filter(
        Ticket.company == company,
        Ticket.title.like(f'%{server.name}%'),
        Ticket.category.in_(['Infraestructura', 'Servidores']),
        Ticket.status.in_(['open', 'in_progress'])
    ).order_by(Ticket.created_at.desc()).first()

    priority = 'critical' if server.is_critical else 'high'
    sla_minutes = get_sla_minutes_for_priority(priority)
    ticket_number = get_next_ticket_number(company)

    ticket = Ticket(
        ticket_number=ticket_number,
        title=f'🚨 ALARMA: {server.name} no responde',
        description=(
            f'**Servidor:** {server.name}\n'
            f'**Host:** {server.ip_address}:{server.port}\n'
            f'**Descripción:** {server.description or "N/A"}\n'
            f'**Crítico:** {"Sí" if server.is_critical else "No"}\n'
            f'**Fallas consecutivas:** {server.consecutive_failures}\n'
            f'**Última verificación:** {server.last_ping.strftime("%Y-%m-%d %H:%M:%S") if server.last_ping else "—"}\n\n'
            f'Este ticket fue generado **automáticamente** por el Monitor de Conectividad al detectar que el servidor no responde.\n\n'
            f'**Acciones sugeridas:**\n'
            f'1. Verificar estado físico / red del servidor.\n'
            f'2. Revisar servicios y procesos críticos.\n'
            f'3. Notificar a usuarios afectados si la caída persiste.'
        ),
        status='open',
        priority=priority,
        company=company,
        creator_id=admin_user.id,
        category='Servidores',
        sla_minutes=sla_minutes,
        sla_deadline=compute_sla_deadline(datetime.now(), sla_minutes, company)
    )
    # Auto-asignar al mejor técnico
    try:
        assign_ticket_auto(ticket)
    except Exception:
        pass
    db.session.add(ticket)
    db.session.commit()

    # Si existía un ticket abierto viejo (>4h), referenciar el nuevo en él
    if old_existing:
        try:
            db.session.add(Message(
                ticket_id=old_existing.id,
                user_id=admin_user.id,
                text=f'⚠️ Nueva caída detectada del servidor {server.name}. '
                     f'Se creó un ticket nuevo: **{ticket_number}** (el anterior quedó '
                     f'sin cerrar de hace más de {DEDUPE_WINDOW_HOURS}h). '
                     f'Revisar y cerrar este ticket si ya no aplica.'
            ))
            db.session.commit()
            log_audit('server_down', admin_user.id, 'server', server.id,
                      f'Nueva caída detectada, ticket nuevo {ticket_number} (anterior: {old_existing.ticket_number})')
        except Exception as e:
            db.session.rollback()
            print(f'[WARN] No se pudo agregar referencia en ticket viejo: {e}')

    try:
        socketio.emit('server_alarm', {
            'server_name': server.name,
            'host': server.ip_address,
            'ticket_number': ticket_number,
            'severity': 'critical',
            'message': f'Servidor {server.name} está OFFLINE'
        }, room=f'company_{company}')
    except Exception as e:
        print(f'[WARN] No se pudo emitir alarma: {e}')

    return ticket


@app.route('/api/monitor/connectivity', methods=['GET'])
def api_monitor_connectivity():
    """Monitor de conectividad - solo servidores de la empresa del usuario"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    company = session['company']
    servers = Server.query.filter_by(company=company, is_active=True).all()

    if not servers:
        return jsonify({
            'success': True,
            'servers': [],
            'summary': {'total': 0, 'online': 0, 'offline': 0, 'slow': 0, 'uptime_pct': 0, 'alarms_active': 0},
            'message': 'No hay servidores configurados. Agrega servidores en el panel de configuración.'
        })

    results = []
    new_alarms = []

    for s in servers:
        port = s.port or 443
        status, latency = ping_server(s.ip_address, port)

        s.last_status = status
        s.last_latency_ms = latency
        s.last_ping = datetime.now()
        s.is_online = (status != 'offline')

        if status == 'offline':
            s.consecutive_failures = (s.consecutive_failures or 0) + 1
            threshold = 1 if s.is_critical else 2
            if s.consecutive_failures >= threshold and not s.alarm_active:
                s.alarm_active = True
                # Crear ticket SIEMPRE — crítico o no.
                # La prioridad se ajusta dentro de create_alarm_ticket según is_critical.
                alarm_ticket = create_alarm_ticket(s, company)
                if alarm_ticket:
                    new_alarms.append({
                        'server_name': s.name,
                        'ticket_number': alarm_ticket.ticket_number,
                        'priority': alarm_ticket.priority,
                        'is_critical': bool(s.is_critical)
                    })
                log_audit('server_alarm', session.get('user_id'), 'server', s.id,
                          f'Alarma activada: {s.name} ({s.ip_address}:{port}) OFFLINE → ticket {alarm_ticket.ticket_number if alarm_ticket else "(no creado)"}')
        else:
            if s.alarm_active:
                # Servidor se recuperó → cerrar ticket relacionado si existe
                try:
                    related = Ticket.query.filter(
                        Ticket.company == company,
                        Ticket.title.like(f'%{s.name}%'),
                        Ticket.category.in_(['Infraestructura', 'Servidores']),
                        Ticket.status.in_(['open', 'in_progress'])
                    ).first()
                    if related:
                        related.status = 'resolved'
                        related.resolved_at = datetime.now()
                        related.updated_at = datetime.now()
                        creator = User.query.filter_by(company=company, role='admin').first()
                        db.session.add(Message(
                            ticket_id=related.id,
                            user_id=creator.id if creator else 1,
                            text=f'✅ Servidor {s.name} recuperado automáticamente. Conexión restablecida (latencia {latency}ms).'
                        ))
                except Exception as e:
                    print(f'[connectivity] Error cerrando ticket de recuperación: {e}')
                log_audit('server_recovered', session.get('user_id'), 'server', s.id,
                          f'Servidor recuperado: {s.name} ({s.ip_address}:{port})')
            s.consecutive_failures = 0
            s.alarm_active = False

        results.append({
            'id': s.id,
            'name': s.name,
            'host': s.ip_address,
            'port': port,
            'description': s.description or '',
            'is_critical': bool(s.is_critical),
            'status': status,
            'latency_ms': latency,
            'consecutive_failures': s.consecutive_failures or 0,
            'alarm_active': bool(s.alarm_active),
            'checked_at': datetime.now().strftime('%H:%M:%S')
        })

    db.session.commit()
    online_count = len([r for r in results if r['status'] == 'online'])

    return jsonify({
        'success': True,
        'servers': results,
        'new_alarms': new_alarms,
        'summary': {
            'total': len(results),
            'online': online_count,
            'offline': len([r for r in results if r['status'] == 'offline']),
            'slow': len([r for r in results if r['status'] == 'slow']),
            'alarms_active': len([r for r in results if r['alarm_active']]),
            'uptime_pct': round(online_count / len(results) * 100, 1) if results else 0
        }
    })


@app.route('/api/admin/companies', methods=['GET'])
def api_admin_companies_list():
    """Listar empresas con conteo de usuarios.
    Master admin (Eliot) ve todas; los demás solo ven la suya."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    scope = admin_companies_scope()
    companies = Company.query.filter(Company.code.in_(scope)).order_by(Company.name).all()
    result = []
    for c in companies:
        user_count = User.query.filter_by(company=c.code).count()
        active_users = User.query.filter_by(company=c.code, is_active=True).count()
        ticket_count = Ticket.query.filter_by(company=c.code).count()
        result.append({
            'id': c.id,
            'code': c.code,
            'name': c.name,
            'icon': c.icon or '🏢',
            'primary_color': c.primary_color or '#2563eb',
            'secondary_color': c.secondary_color or '#1e40af',
            'ldap_server': c.ldap_server or '',
            'ldap_base_dn': c.ldap_base_dn or '',
            'ldap_bind_user': c.ldap_bind_user or '',
            'microsoft_tenant_id': c.microsoft_tenant_id or '',
            'microsoft_client_id': c.microsoft_client_id or '',
            'microsoft_enabled': bool(c.microsoft_enabled),
            'is_active': bool(c.is_active),
            'user_count': user_count,
            'active_users': active_users,
            'ticket_count': ticket_count
        })
    return jsonify({
        'success': True,
        'companies': result,
        'is_master': is_master_admin(),
        'can_create': is_master_admin(),
        'can_delete': is_master_admin()
    })


@app.route('/api/admin/companies', methods=['POST'])
def api_admin_companies_create():
    """Crear empresa (solo admin master Eliot)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    if not is_master_admin():
        return jsonify({'success': False, 'error': 'Solo el admin de la empresa master puede crear empresas.'}), 403
    try:
        data = request.get_json()
        code = (data.get('code') or '').strip().lower()
        name = (data.get('name') or '').strip()

        if not code or not name:
            return jsonify({'success': False, 'error': 'Código y nombre son requeridos'}), 400

        if not code.replace('_', '').isalnum():
            return jsonify({'success': False, 'error': 'El código solo puede contener letras, números y guion bajo'}), 400

        existing = Company.query.filter_by(code=code).first()
        if existing:
            return jsonify({'success': False, 'error': f'Ya existe una empresa con código "{code}"'}), 400

        c = Company(
            code=code,
            name=name,
            icon=(data.get('icon') or '🏢').strip(),
            primary_color=(data.get('primary_color') or '#2563eb').strip(),
            secondary_color=(data.get('secondary_color') or '#1e40af').strip(),
            ldap_server=(data.get('ldap_server') or '').strip(),
            ldap_base_dn=(data.get('ldap_base_dn') or '').strip(),
            ldap_bind_user=(data.get('ldap_bind_user') or '').strip(),
            ldap_bind_password=encrypt_secret((data.get('ldap_bind_password') or '').strip() or None),
            is_active=bool(data.get('is_active', True))
        )
        db.session.add(c)
        db.session.commit()
        log_audit('create_company', session['user_id'], 'company', c.id, f'Empresa creada: {name} ({code})')
        return jsonify({'success': True, 'id': c.id, 'message': f'Empresa "{name}" creada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/companies/<int:company_id>', methods=['PUT'])
def api_admin_companies_update(company_id):
    """Actualizar empresa. Master admin actualiza cualquiera;
    el resto solo puede actualizar la propia."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    c = Company.query.get(company_id)
    if not c:
        return jsonify({'success': False, 'error': 'No encontrada'}), 404
    # Solo el admin master puede tocar empresas que no sean la suya
    if not is_master_admin() and c.code != session.get('company'):
        return jsonify({'success': False, 'error': 'No puedes modificar otras empresas.'}), 403
    try:
        data = request.get_json()
        if 'name' in data: c.name = data['name'].strip()
        if 'icon' in data: c.icon = data['icon'].strip()
        if 'primary_color' in data: c.primary_color = data['primary_color'].strip()
        if 'secondary_color' in data: c.secondary_color = data['secondary_color'].strip()
        if 'ldap_server' in data: c.ldap_server = data['ldap_server'].strip()
        if 'ldap_base_dn' in data: c.ldap_base_dn = data['ldap_base_dn'].strip()
        if 'ldap_bind_user' in data: c.ldap_bind_user = data['ldap_bind_user'].strip()
        if 'ldap_bind_password' in data and data['ldap_bind_password']:
            c.ldap_bind_password = encrypt_secret(data['ldap_bind_password'].strip())
        # Microsoft Entra ID
        if 'microsoft_tenant_id' in data:
            c.microsoft_tenant_id = (data['microsoft_tenant_id'] or '').strip()[:100] or None
        if 'microsoft_client_id' in data:
            c.microsoft_client_id = (data['microsoft_client_id'] or '').strip()[:100] or None
        if 'microsoft_client_secret' in data and data['microsoft_client_secret']:
            c.microsoft_client_secret = encrypt_secret(data['microsoft_client_secret'].strip())
        if 'microsoft_enabled' in data:
            c.microsoft_enabled = bool(data['microsoft_enabled'])
        if 'is_active' in data: c.is_active = bool(data['is_active'])
        db.session.commit()
        log_audit('update_company', session['user_id'], 'company', c.id, f'Empresa actualizada: {c.name}')
        return jsonify({'success': True, 'message': 'Empresa actualizada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/companies/<int:company_id>/test-ldap', methods=['POST'])
def api_admin_companies_test_ldap(company_id):
    """Prueba la conexión LDAP de una empresa.
    Permite enviar credenciales temporales en el body para test sin guardar."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    c = Company.query.get(company_id)
    if not c:
        return jsonify({'success': False, 'error': 'Empresa no encontrada'}), 404
    if not is_master_admin() and c.code != session.get('company'):
        return jsonify({'success': False, 'error': 'Sin permiso sobre esta empresa.'}), 403

    if not LDAP_AVAILABLE:
        return jsonify({'success': False, 'error': 'Librería ldap3 no disponible en el servidor.'}), 500

    # Permitir override del body para probar credenciales nuevas sin guardarlas
    data = request.get_json() or {}

    class _TempCompany:
        pass

    tc = _TempCompany()
    tc.ldap_server = (data.get('ldap_server') or c.ldap_server or '').strip()
    tc.ldap_base_dn = (data.get('ldap_base_dn') or c.ldap_base_dn or '').strip()
    tc.ldap_bind_user = (data.get('ldap_bind_user') or c.ldap_bind_user or '').strip()
    # Si el body trae password, se usa esa; si no, se descifra la guardada
    raw_pw = data.get('ldap_bind_password')
    if raw_pw:
        # Para no almacenarla cifrada en el objeto temp, ponemos un placeholder
        # y modificamos LdapConfig.from_company después.
        pass

    cfg = LdapConfig.from_company(tc, decrypt_secret)
    if not cfg:
        return jsonify({'success': False, 'error': 'Falta configurar el servidor LDAP.'}), 400
    # Si llegó password en body, la sobrescribimos en la config
    if raw_pw:
        cfg.bind_password = raw_pw

    ok, msg = test_ldap_connection(cfg)
    log_audit(
        'ldap_test', session['user_id'], 'company', c.id,
        f'Test LDAP para {c.code}: {"OK" if ok else "FALLO"} - {msg}'
    )
    return jsonify({'success': ok, 'message': msg, 'server': cfg.server, 'port': cfg.port, 'ssl': cfg.use_ssl})


@app.route('/api/admin/companies/<int:company_id>', methods=['DELETE'])
def api_admin_companies_delete(company_id):
    """Eliminar empresa (solo admin master Eliot, y solo si no tiene usuarios/tickets)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    if not is_master_admin():
        return jsonify({'success': False, 'error': 'Solo el admin de la empresa master puede eliminar empresas.'}), 403
    c = Company.query.get(company_id)
    if not c:
        return jsonify({'success': False, 'error': 'No encontrada'}), 404
    # No permitir eliminar la propia empresa master
    if c.code == MASTER_COMPANY:
        return jsonify({'success': False, 'error': 'No se puede eliminar la empresa master.'}), 400

    user_count = User.query.filter_by(company=c.code).count()
    ticket_count = Ticket.query.filter_by(company=c.code).count()
    if user_count > 0 or ticket_count > 0:
        return jsonify({
            'success': False,
            'error': f'No se puede eliminar. Tiene {user_count} usuarios y {ticket_count} tickets. Desactívala en su lugar.'
        }), 400

    name = c.name
    db.session.delete(c)
    db.session.commit()
    log_audit('delete_company', session['user_id'], 'company', company_id, f'Empresa eliminada: {name}')
    return jsonify({'success': True, 'message': 'Empresa eliminada'})


@app.route('/api/admin/tags', methods=['GET'])
def api_admin_tags_list():
    """Listar tags de la empresa"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    company = session['company']
    tags = Tag.query.filter_by(company=company).order_by(Tag.name).all()
    return jsonify({
        'success': True,
        'tags': [{
            'id': t.id,
            'name': t.name,
            'color': t.color or '#2563eb',
            'icon': t.icon or '🏷️',
            'description': t.description or '',
            'usage_count': t.usage_count or 0
        } for t in tags]
    })


@app.route('/api/admin/tags', methods=['POST'])
def api_admin_tags_create():
    """Crear tag"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json()
        name = (data.get('name') or '').strip()
        if not name:
            return jsonify({'success': False, 'error': 'Nombre requerido'}), 400

        # Verificar duplicado
        existing = Tag.query.filter_by(company=session['company'], name=name).first()
        if existing:
            return jsonify({'success': False, 'error': f'El tag "{name}" ya existe'}), 400

        tag = Tag(
            name=name,
            color=(data.get('color') or '#2563eb').strip(),
            icon=(data.get('icon') or '🏷️').strip(),
            description=(data.get('description') or '').strip(),
            company=session['company']
        )
        db.session.add(tag)
        db.session.commit()
        log_audit('create_tag', session['user_id'], 'tag', tag.id, f'Tag creado: {name}')
        return jsonify({'success': True, 'id': tag.id, 'message': f'Tag "{name}" creado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/tags/<int:tag_id>', methods=['PUT'])
def api_admin_tags_update(tag_id):
    """Actualizar tag"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    tag = Tag.query.get(tag_id)
    if not tag or tag.company != session['company']:
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    try:
        data = request.get_json()
        if 'name' in data: tag.name = data['name'].strip()
        if 'color' in data: tag.color = data['color'].strip()
        if 'icon' in data: tag.icon = data['icon'].strip()
        if 'description' in data: tag.description = data['description'].strip()
        db.session.commit()
        return jsonify({'success': True, 'message': 'Tag actualizado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/tags/<int:tag_id>', methods=['DELETE'])
def api_admin_tags_delete(tag_id):
    """Eliminar tag"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    tag = Tag.query.get(tag_id)
    if not tag or tag.company != session['company']:
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    name = tag.name
    db.session.delete(tag)
    db.session.commit()
    log_audit('delete_tag', session['user_id'], 'tag', tag_id, f'Tag eliminado: {name}')
    return jsonify({'success': True, 'message': 'Tag eliminado'})


@app.route('/api/admin/tags/seed', methods=['POST'])
def api_admin_tags_seed():
    """Crear tags predefinidos"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']
    seed_tags = [
        # Categorías técnicas
        {'name': 'Hardware', 'color': '#3b82f6', 'icon': '🖥️', 'description': 'Problemas físicos de equipos'},
        {'name': 'Software', 'color': '#8b5cf6', 'icon': '💻', 'description': 'Aplicaciones y programas'},
        {'name': 'Red', 'color': '#10b981', 'icon': '🌐', 'description': 'Conectividad y red'},
        {'name': 'WiFi', 'color': '#06b6d4', 'icon': '📶', 'description': 'Problemas de WiFi'},
        {'name': 'VPN', 'color': '#6366f1', 'icon': '🔒', 'description': 'Conexión VPN'},
        {'name': 'Email', 'color': '#0891b2', 'icon': '📧', 'description': 'Correo electrónico'},
        {'name': 'SAP', 'color': '#ea580c', 'icon': '📊', 'description': 'Sistema SAP'},
        {'name': 'Office', 'color': '#dc2626', 'icon': '📝', 'description': 'Microsoft Office'},
        {'name': 'Windows', 'color': '#0284c7', 'icon': '🪟', 'description': 'Sistema operativo Windows'},
        {'name': 'Linux', 'color': '#facc15', 'icon': '🐧', 'description': 'Sistema operativo Linux'},
        {'name': 'Impresora', 'color': '#7c3aed', 'icon': '🖨️', 'description': 'Impresoras y escáneres'},
        {'name': 'Telefonía', 'color': '#06b6d4', 'icon': '📞', 'description': 'Teléfonos y extensiones'},
        {'name': 'Móvil', 'color': '#d946ef', 'icon': '📱', 'description': 'Dispositivos móviles'},
        {'name': 'Servidores', 'color': '#475569', 'icon': '🖧', 'description': 'Servidores e infraestructura'},
        {'name': 'Base de Datos', 'color': '#0d9488', 'icon': '🗄️', 'description': 'Bases de datos'},
        {'name': 'Backup', 'color': '#65a30d', 'icon': '💾', 'description': 'Copias de seguridad'},
        # Estados
        {'name': 'Urgente', 'color': '#dc2626', 'icon': '🚨', 'description': 'Requiere atención inmediata'},
        {'name': 'En espera', 'color': '#f59e0b', 'icon': '⏳', 'description': 'Esperando respuesta'},
        {'name': 'En progreso', 'color': '#3b82f6', 'icon': '🔄', 'description': 'En desarrollo'},
        {'name': 'Resuelto', 'color': '#16a34a', 'icon': '✓', 'description': 'Ya solucionado'},
        {'name': 'Recurrente', 'color': '#dc2626', 'icon': '🔁', 'description': 'Problema que se repite'},
        {'name': 'Documentar', 'color': '#7c3aed', 'icon': '📋', 'description': 'Requiere documentación'},
        # Acceso y seguridad
        {'name': 'Accesos', 'color': '#f59e0b', 'icon': '🔑', 'description': 'Permisos y accesos'},
        {'name': 'Contraseña', 'color': '#fb7185', 'icon': '🔐', 'description': 'Cambio de contraseña'},
        {'name': 'Seguridad', 'color': '#991b1b', 'icon': '🛡️', 'description': 'Incidentes de seguridad'},
        {'name': 'Phishing', 'color': '#dc2626', 'icon': '🎣', 'description': 'Correos sospechosos'},
        {'name': 'Virus', 'color': '#991b1b', 'icon': '🦠', 'description': 'Malware o virus'},
        # Departamentos
        {'name': 'RRHH', 'color': '#ec4899', 'icon': '👥', 'description': 'Recursos Humanos'},
        {'name': 'Finanzas', 'color': '#16a34a', 'icon': '💰', 'description': 'Departamento financiero'},
        {'name': 'Ventas', 'color': '#3b82f6', 'icon': '💼', 'description': 'Equipo de ventas'},
        {'name': 'Producción', 'color': '#ea580c', 'icon': '🏭', 'description': 'Planta de producción'},
        {'name': 'Logística', 'color': '#0891b2', 'icon': '🚚', 'description': 'Almacén y logística'},
        # Tipos de solicitud
        {'name': 'Solicitud', 'color': '#6366f1', 'icon': '📥', 'description': 'Solicitud nueva'},
        {'name': 'Cambio', 'color': '#f59e0b', 'icon': '🔧', 'description': 'Cambio de configuración'},
        {'name': 'Mejora', 'color': '#16a34a', 'icon': '⭐', 'description': 'Mejora o sugerencia'},
        {'name': 'Capacitación', 'color': '#8b5cf6', 'icon': '🎓', 'description': 'Necesita capacitación'},
        {'name': 'Mantenimiento', 'color': '#64748b', 'icon': '🔨', 'description': 'Mantenimiento preventivo'},
    ]

    created, skipped = 0, 0
    for t in seed_tags:
        existing = Tag.query.filter_by(company=company, name=t['name']).first()
        if existing:
            skipped += 1
            continue
        tag = Tag(name=t['name'], color=t['color'], icon=t['icon'],
                  description=t['description'], company=company)
        db.session.add(tag)
        created += 1

    db.session.commit()
    log_audit('seed_tags', session['user_id'], 'tag', None, f'Tags iniciales: {created} creados, {skipped} omitidos')
    return jsonify({
        'success': True,
        'message': f'{created} tags creados ({skipped} ya existían)',
        'created': created,
        'skipped': skipped
    })


@app.route('/api/admin/bot-kb', methods=['GET'])
def api_admin_botkb_list():
    """Listar base de conocimiento del bot"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    kbs = BotKnowledge.query.order_by(BotKnowledge.category, BotKnowledge.question).all()
    return jsonify({
        'success': True,
        'knowledge': [{
            'id': k.id,
            'keywords': k.keywords or '',
            'question': k.question,
            'answer': k.answer,
            'category': k.category or 'General',
            'priority': k.priority or 'medium'
        } for k in kbs]
    })


@app.route('/api/admin/bot-kb', methods=['POST'])
def api_admin_botkb_create():
    """Crear entrada en base de conocimiento"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json()
        kb = BotKnowledge(
            keywords=(data.get('keywords') or '').strip(),
            question=(data.get('question') or '').strip(),
            answer=(data.get('answer') or '').strip(),
            category=(data.get('category') or 'General').strip(),
            priority=(data.get('priority') or 'medium').strip()
        )
        if not kb.question or not kb.answer:
            return jsonify({'success': False, 'error': 'Pregunta y respuesta son requeridas'}), 400
        db.session.add(kb)
        db.session.commit()
        log_audit('create_botkb', session['user_id'], 'botkb', kb.id, f'KB: {kb.question[:50]}')
        return jsonify({'success': True, 'id': kb.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/bot-kb/<int:kb_id>', methods=['PUT'])
def api_admin_botkb_update(kb_id):
    """Actualizar entrada KB"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    kb = BotKnowledge.query.get(kb_id)
    if not kb:
        return jsonify({'success': False}), 404
    try:
        data = request.get_json()
        for f in ['keywords', 'question', 'answer', 'category', 'priority']:
            if f in data: setattr(kb, f, str(data[f]).strip())
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/bot-kb/<int:kb_id>', methods=['DELETE'])
def api_admin_botkb_delete(kb_id):
    """Eliminar entrada KB"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    kb = BotKnowledge.query.get(kb_id)
    if not kb:
        return jsonify({'success': False}), 404
    db.session.delete(kb)
    db.session.commit()
    log_audit('delete_botkb', session['user_id'], 'botkb', kb_id, 'KB eliminado')
    return jsonify({'success': True})


@app.route('/api/admin/bot-kb/seed', methods=['POST'])
def api_admin_botkb_seed():
    """Cargar base de conocimiento amplia para el bot"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    kb_entries = [
        # CONTRASEÑAS Y CUENTAS
        {'category': 'Contraseñas', 'priority': 'medium', 'keywords': 'contraseña,password,olvide,resetear,olvidé',
         'question': '¿Cómo restablezco mi contraseña?',
         'answer': '1. Ve a https://passwordreset.empresa.local\n2. Ingresa tu usuario corporativo\n3. Responde tu pregunta de seguridad o usa tu correo personal verificado\n4. Crea una nueva contraseña que cumpla: mínimo 12 caracteres, mayúsculas, minúsculas, números y un símbolo\n5. Espera 5 minutos antes de iniciar sesión.\n\nSi no funciona, crea un ticket en categoría Accesos.'},
        {'category': 'Contraseñas', 'priority': 'high', 'keywords': 'cuenta,bloqueada,bloqueado,lock,locked',
         'question': '¿Mi cuenta está bloqueada, qué hago?',
         'answer': 'Las cuentas se bloquean tras 5 intentos fallidos consecutivos.\n\n1. Espera 30 minutos (desbloqueo automático)\n2. Si necesitas acceso urgente: contacta a soporte\n3. Reporta intentos sospechosos de acceso (puede ser ataque)\n\nNUNCA compartas tu contraseña con nadie del soporte.'},
        {'category': 'Contraseñas', 'priority': 'medium', 'keywords': 'cambiar,contraseña,expira,vencida,expiró',
         'question': '¿Cómo cambio mi contraseña?',
         'answer': 'Windows: Ctrl+Alt+Suprimir → "Cambiar contraseña"\nWeb: Settings → Cuenta → Cambiar contraseña\nMóvil: Outlook → Cuenta → Seguridad\n\nPolítica: las contraseñas expiran cada 90 días. Recibirás aviso 14 días antes.'},

        # EMAIL Y OUTLOOK
        {'category': 'Email', 'priority': 'medium', 'keywords': 'outlook,no abre,abre,inicia,inicio,crash',
         'question': 'Outlook no abre o se cierra solo',
         'answer': '1. Cierra Outlook completamente (Administrador de tareas → fin tarea)\n2. Inicia en modo seguro: Win+R → "outlook /safe"\n3. Si funciona en modo seguro, deshabilita complementos:\n   - Archivo → Opciones → Complementos → Ir → Desmarca todos\n4. Si el problema persiste, repara el perfil:\n   - Panel Control → Correo → Mostrar perfiles → Reparar\n5. Última opción: crear nuevo perfil de Outlook'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'correo,enviar,no envia,no envío,bandeja salida,stuck',
         'question': 'Mis correos no se envían (se quedan en Bandeja de salida)',
         'answer': '1. Verifica tu conexión a internet\n2. Outlook → Enviar/Recibir → Enviar todo\n3. Revisa que el correo no tenga adjuntos mayores a 25MB\n4. Si tiene adjuntos grandes, súbelos a OneDrive y comparte el link\n5. Revisa carpeta "Bandeja de salida": si hay correos atascados, ábrelos y reenvíalos\n6. Reinicia Outlook'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'no recibo,correos,emails,bandeja entrada,spam',
         'question': 'No estoy recibiendo correos',
         'answer': '1. Revisa carpeta Spam/Correo no deseado\n2. Verifica reglas: Archivo → Administrar reglas y alertas\n3. Asegúrate de no haber excedido tu cuota de buzón (límite: 50GB)\n4. Pide al remitente que verifique no estar en tu lista de bloqueados\n5. Revisa "Outlook elementos eliminados"\n6. Comprueba estado del servicio en: https://status.office.com'},
        {'category': 'Email', 'priority': 'low', 'keywords': 'firma,signature,configurar firma,correo',
         'question': '¿Cómo configuro mi firma de correo?',
         'answer': 'Outlook Desktop:\n1. Archivo → Opciones → Correo → Firmas\n2. Nuevo → Asigna nombre\n3. Edita la firma con tu nombre, cargo, teléfono\n4. Selecciónala como predeterminada\n\nOutlook Web:\n1. Settings (engranaje) → Ver toda la configuración\n2. Correo → Redactar y responder → Firma de correo\n\nFirma corporativa estándar disponible en: \\\\fileserver\\plantillas\\firmas'},
        {'category': 'Email', 'priority': 'low', 'keywords': 'fuera oficina,vacaciones,respuesta automatica,out of office',
         'question': '¿Cómo activo respuesta automática de fuera de oficina?',
         'answer': 'Outlook Desktop:\n1. Archivo → Información → Respuestas automáticas\n2. "Enviar respuestas automáticas"\n3. Marca rango de fechas\n4. Escribe mensaje para interno y externo (pestañas separadas)\n5. Aceptar\n\nIncluye:\n- Fechas de tu ausencia\n- Contacto alternativo\n- Cuándo regresarás\n- Mensaje cordial'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'adjunto,archivo,grande,grande,tamaño,size',
         'question': 'No puedo enviar archivos grandes por correo',
         'answer': 'Límite de adjuntos: 25 MB por mensaje.\n\nAlternativas:\n1. **OneDrive**: Sube el archivo, click derecho → Compartir → Copiar link → Pega en el correo\n2. **SharePoint**: Para archivos del equipo, sube ahí y comparte enlace\n3. **Comprimir**: Click derecho → Enviar a → Carpeta comprimida (.zip)\n4. **WeTransfer corporativo**: Para externos\n\nNunca envíes información confidencial por servicios públicos.'},

        # RED Y WIFI
        {'category': 'Red', 'priority': 'high', 'keywords': 'wifi,no conecta,sin internet,red,ssid',
         'question': 'No puedo conectarme al WiFi',
         'answer': '1. Olvida la red y vuelve a conectar:\n   - Configuración → Red e Internet → WiFi → Administrar redes conocidas\n   - Selecciona la red corporativa → Olvidar\n   - Vuelve a conectar e ingresa credenciales\n\n2. Reinicia adaptador WiFi:\n   - Win+R → "ncpa.cpl"\n   - Click derecho en WiFi → Deshabilitar → Habilitar\n\n3. Renueva IP:\n   - cmd como admin: "ipconfig /release" luego "ipconfig /renew"\n\n4. Reinicia el equipo\n\n5. Verifica que estés en cobertura del Access Point'},
        {'category': 'Red', 'priority': 'high', 'keywords': 'internet,lento,slow,conexion,red lenta',
         'question': 'Mi internet está muy lento',
         'answer': 'Diagnóstico rápido:\n\n1. Prueba velocidad en https://speedtest.net (debe ser >50 Mbps)\n2. Reinicia tu equipo y router\n3. Cierra apps que consumen ancho de banda:\n   - Streaming (YouTube, Netflix)\n   - Descargas en segundo plano\n   - OneDrive sincronizando\n   - Actualizaciones de Windows\n\n4. Cambia a cable Ethernet si es posible (más estable que WiFi)\n5. Verifica que no haya 20+ usuarios en el mismo AP\n\nSi persiste, reporta a soporte con la velocidad medida.'},
        {'category': 'Red', 'priority': 'high', 'keywords': 'vpn,no conecta,conectar vpn,vpn error,fortinet,cisco anyconnect',
         'question': 'No puedo conectarme a la VPN',
         'answer': '1. Verifica tu conexión a internet primero\n2. Cierra completamente el cliente VPN\n3. Reinicia el servicio VPN como administrador\n4. Verifica tus credenciales (mismas de Windows)\n5. Si usas MFA, asegúrate que tu app autenticadora funciona\n6. Algunos firewalls bloquean VPN. Prueba desde otra red\n7. Limpia caché DNS: cmd → "ipconfig /flushdns"\n\nClientes oficiales:\n- Cisco AnyConnect\n- FortiClient\n- Global Protect\n\nNo uses VPNs no aprobadas, comprometen la seguridad.'},
        {'category': 'Red', 'priority': 'medium', 'keywords': 'dns,no resuelve,navegar,paginas,no carga',
         'question': 'Las páginas web no cargan pero tengo internet',
         'answer': 'Problema típico de DNS:\n\n1. Limpia caché DNS:\n   - cmd como admin: "ipconfig /flushdns"\n\n2. Cambia el DNS:\n   - Win+R → ncpa.cpl\n   - Click derecho en tu conexión → Propiedades\n   - IPv4 → Propiedades\n   - DNS preferido: 8.8.8.8\n   - DNS alternativo: 8.8.4.4\n\n3. Reinicia el navegador\n4. Prueba en modo incógnito\n5. Limpia caché del navegador (Ctrl+Shift+Supr)'},
        {'category': 'Red', 'priority': 'medium', 'keywords': 'ping,no responde,timeout,conexion',
         'question': '¿Cómo verifico conectividad a un servidor?',
         'answer': 'Usa el comando PING:\n\n1. Abre CMD (Win+R → cmd)\n2. Escribe: ping nombre-servidor.empresa.local\n3. Interpretación:\n   - "Respuesta de..." = conectividad OK\n   - "Tiempo de espera agotado" = sin conexión\n   - "Host de destino inaccesible" = problema de ruta\n\nPara probar puerto específico:\n   - tnc nombre-servidor -p 443 (PowerShell)\n   - telnet nombre-servidor 443\n\nReporta tickets con los resultados del ping.'},

        # WINDOWS Y SISTEMA OPERATIVO
        {'category': 'Windows', 'priority': 'medium', 'keywords': 'lento,pc lento,computador lento,slow,rendimiento',
         'question': 'Mi computador está muy lento',
         'answer': 'Pasos de optimización:\n\n1. **Reinicia el equipo** (la mayoría de problemas se solucionan así)\n2. Cierra programas que no uses (mira Administrador de Tareas: Ctrl+Shift+Esc)\n3. Verifica espacio en disco C: (debe tener >20% libre)\n4. Desactiva programas de inicio:\n   - Win+R → "msconfig" → Inicio → Abrir Administrador de tareas\n   - Deshabilita los no esenciales\n5. Ejecuta limpieza de disco:\n   - Click derecho en C: → Propiedades → Liberar espacio\n6. Verifica si tu antivirus está escaneando\n\nSi persiste, escala a soporte. Puede requerir más RAM o SSD.'},
        {'category': 'Windows', 'priority': 'high', 'keywords': 'pantalla azul,bsod,blue screen,error,reinicia',
         'question': 'Pantalla azul (BSOD) en mi PC',
         'answer': '**Anota el código de error** (ej: PAGE_FAULT_IN_NONPAGED_AREA, IRQL_NOT_LESS_OR_EQUAL)\n\n1. Reinicia y observa si se repite\n2. Si pasa al iniciar Windows:\n   - Inicia en Modo Seguro (F8 al arrancar)\n   - Restaura sistema a una fecha anterior\n3. Causas comunes:\n   - Drivers desactualizados (especialmente GPU)\n   - RAM defectuosa\n   - Disco con sectores defectuosos\n   - Software incompatible\n\n4. Diagnóstico:\n   - cmd como admin: "sfc /scannow"\n   - cmd como admin: "DISM /Online /Cleanup-Image /RestoreHealth"\n\nReporta el código exacto del error.'},
        {'category': 'Windows', 'priority': 'medium', 'keywords': 'actualizar,windows update,actualizaciones,update',
         'question': '¿Cómo actualizo Windows?',
         'answer': '1. Inicio → Configuración → Windows Update\n2. Buscar actualizaciones\n3. Descargar e instalar\n4. Reiniciar cuando se solicite\n\nMejores prácticas:\n- Guarda tu trabajo antes de reiniciar\n- Conecta a corriente (laptops)\n- No apagues durante actualización\n- Programar fuera de horario laboral\n\nSi las actualizaciones fallan:\n- cmd como admin: "wuauclt /resetauthorization /detectnow"\n- Espera 10 minutos\n- Reintenta'},
        {'category': 'Windows', 'priority': 'medium', 'keywords': 'archivos,perdidos,recuperar,borrados,papelera',
         'question': '¿Cómo recupero archivos borrados?',
         'answer': '1. **Papelera de reciclaje**: Doble click → Buscar → Click derecho → Restaurar\n\n2. **Versiones anteriores**:\n   - Click derecho en la carpeta padre → Propiedades → Versiones anteriores\n\n3. **OneDrive**: Si lo tenías sincronizado:\n   - https://onedrive.com → Papelera → Restaurar\n   - Historial de versiones: archivo → click derecho → Historial de versiones\n\n4. **SharePoint**:\n   - Sitio → Configuración → Contenido del sitio → Papelera\n\n5. **Backup empresarial**: Crea ticket urgente. Los respaldos cubren últimos 30 días.\n\nNo escribas más datos al disco hasta intentar recuperarlos.'},
        {'category': 'Windows', 'priority': 'low', 'keywords': 'idioma,teclado,configuracion teclado,layout',
         'question': '¿Cómo cambio el idioma del teclado?',
         'answer': '1. Configuración → Hora e idioma → Idioma\n2. "Agregar idioma" si no aparece\n3. Click en idioma → Opciones → Agregar teclado\n4. Cambiar entre idiomas: Alt+Shift o Win+Espacio\n\nTeclados comunes:\n- Español (Internacional): la ñ y acentos\n- Español (Latinoamérica): teclas más comunes en LATAM\n- Inglés (EE.UU.): sin ñ, distribución QWERTY estándar'},

        # OFFICE
        {'category': 'Office', 'priority': 'medium', 'keywords': 'excel,formula,no funciona,calcula',
         'question': 'Las fórmulas de Excel no calculan',
         'answer': '1. Verifica que el cálculo automático esté activo:\n   - Fórmulas → Opciones para el cálculo → Automático\n\n2. La celda puede tener formato Texto:\n   - Selecciona la celda → Formato → Número\n   - Borra el contenido y vuelve a escribir la fórmula\n\n3. Verifica errores en la fórmula:\n   - #REF! = referencia inválida\n   - #VALUE! = tipo de dato incorrecto\n   - #DIV/0! = división por cero\n   - #NAME? = función mal escrita\n\n4. Presiona F9 para recalcular manualmente\n5. Ctrl+Alt+F9 fuerza recálculo completo'},
        {'category': 'Office', 'priority': 'medium', 'keywords': 'word,no guarda,perdido,documento,recuperar',
         'question': 'Perdí mi documento de Word sin guardar',
         'answer': '1. Abre Word\n2. Archivo → Información → Administrar documento\n3. "Recuperar libros no guardados" (Word) o documentos\n4. Busca el archivo\n\nUbicaciones de auto-guardado:\n- Word: %APPDATA%\\Microsoft\\Word\\\n- Excel: %APPDATA%\\Microsoft\\Excel\\\n\nPara el futuro:\n- Habilita AutoRecuperación cada 1 min: Archivo → Opciones → Guardar\n- Guarda directamente en OneDrive (auto-guarda)\n- Usa Ctrl+S frecuentemente'},
        {'category': 'Office', 'priority': 'low', 'keywords': 'teams,reunion,meeting,zoom',
         'question': '¿Cómo programo una reunión en Teams?',
         'answer': '1. Abre Teams\n2. Calendario (icono lateral)\n3. "Nueva reunión" (esquina superior derecha)\n4. Llena:\n   - Título\n   - Asistentes (escribe nombres)\n   - Fecha y hora\n   - Detalles\n5. Guardar\n\nLa reunión aparece en:\n- Tu calendario de Outlook\n- Calendarios de invitados\n- Pestaña "Próximas" en Teams\n\nPara unirse: click en el evento → "Unirse"'},
        {'category': 'Office', 'priority': 'low', 'keywords': 'sharepoint,onedrive,compartir,carpeta',
         'question': '¿Cómo comparto archivos en OneDrive/SharePoint?',
         'answer': '1. Click derecho sobre el archivo/carpeta → Compartir\n2. Escribe los correos de los destinatarios\n3. Elige permisos:\n   - **Puede editar**: pueden modificar\n   - **Puede ver**: solo lectura\n   - **Puede comentar**: lectura + comentarios\n4. Opcionales:\n   - Establecer fecha de expiración\n   - Contraseña para acceso\n   - Bloquear descargas\n5. Enviar\n\nSeguridad:\n- NO compartir con "Cualquier persona con el vínculo" información confidencial\n- Para externos, usa la opción "Personas específicas"'},

        # IMPRESORAS
        {'category': 'Impresoras', 'priority': 'medium', 'keywords': 'impresora,no imprime,impresion,print',
         'question': 'La impresora no imprime',
         'answer': '1. **Revisa físicamente**:\n   - Encendida y con papel\n   - Sin atascos visibles\n   - Tóner/tinta disponible\n   - Cable de red conectado\n\n2. **Cola de impresión**:\n   - Configuración → Dispositivos → Impresoras\n   - Click derecho en la impresora → Ver cola de impresión\n   - Cancela trabajos atascados\n\n3. **Reinicia el servicio de impresión**:\n   - Win+R → services.msc\n   - Busca "Cola de impresión" → Reiniciar\n\n4. **Reinstala la impresora**:\n   - Quitar dispositivo\n   - Agregar impresora → busca por nombre/IP\n\nNombre estándar: PRINTER-PISO-AREA (ej: PRINTER-3-CONTABILIDAD)'},
        {'category': 'Impresoras', 'priority': 'low', 'keywords': 'agregar impresora,instalar impresora,nueva',
         'question': '¿Cómo agrego una impresora?',
         'answer': '1. Configuración → Dispositivos → Impresoras y escáneres\n2. "Agregar impresora o escáner"\n3. Espera a que la encuentre por red\n4. Si no aparece: "La impresora deseada no aparece"\n5. "Agregar usando dirección TCP/IP":\n   - IP de la impresora (consulta a soporte)\n   - Tipo: TCP/IP estándar\n\nImpresoras corporativas comunes:\n- impresion.empresa.local\n\nCualquier usuario puede agregar impresoras. Si el driver no se instala, contacta soporte.'},
        {'category': 'Impresoras', 'priority': 'low', 'keywords': 'escanear,scan,digitalizar',
         'question': '¿Cómo escaneo un documento?',
         'answer': '**Desde impresora multifuncional**:\n1. Coloca el documento en el escáner\n2. En el panel de la impresora: "Escanear" o "Scan to email"\n3. Selecciona tu correo (o ingresa otra dirección)\n4. Configura: PDF (mejor compartir) o JPG (imágenes)\n5. Inicia escaneo\n6. Recibirás el archivo en tu correo\n\n**Desde Windows**:\n- Abre "Escáner de Windows" (Microsoft Store)\n- Sigue las instrucciones\n\n**Móvil**:\n- App "Microsoft Lens" o "Adobe Scan"\n- Captura con la cámara\n- Genera PDF automáticamente'},

        # SAP
        {'category': 'SAP', 'priority': 'critical', 'keywords': 'sap,no abre,no inicia,saplogon,gui',
         'question': 'SAP GUI no abre o da error al iniciar',
         'answer': '1. Verifica conexión a la red corporativa o VPN\n2. Verifica que el servicio esté disponible:\n   - Ping al servidor de aplicaciones\n3. Limpia caché de SAP:\n   - Cierra SAP completamente\n   - Borra: %APPDATA%\\SAP\\Common\n4. Verifica versión de SAP GUI (mínima 7.60)\n5. Verifica que el archivo "saplogon.ini" o "SAPUILandscape.xml" exista\n   - Ubicación: %APPDATA%\\SAP\\Common\n\nCódigos de error comunes:\n- "Partner not reached" = sin conexión al servidor\n- "Connection broken" = caída del servicio\n- "Logon not possible" = credenciales o mandante incorrecto'},
        {'category': 'SAP', 'priority': 'high', 'keywords': 'sap,bloqueado,locked,usuario,gui',
         'question': 'Mi usuario SAP está bloqueado',
         'answer': 'Causas comunes de bloqueo SAP:\n1. Múltiples intentos fallidos de login\n2. Inactividad por más de 90 días\n3. Cierre administrativo\n4. Vencimiento de contraseña\n\nSolución:\n- Crea ticket en categoría SAP\n- Indica: usuario SAP, mandante (cliente), ambiente (PRD/QAS/DEV)\n- Soporte BASIS desbloquea normalmente en < 1 hora\n\nMientras tanto:\n- No intentes más logins (extiende bloqueo)\n- Si es urgente, escala al supervisor BASIS'},
        {'category': 'SAP', 'priority': 'medium', 'keywords': 'sap,transaccion,error,abap dump',
         'question': 'Error en transacción SAP (ABAP Dump)',
         'answer': '1. **Anota la información del dump**:\n   - Nombre del programa\n   - Línea del error\n   - Mensaje técnico\n\n2. **Captura pantalla del error**\n\n3. **Pasos para reproducirlo**:\n   - Transacción exacta\n   - Datos ingresados\n   - Botones presionados\n\n4. **Información útil**:\n   - Mandante (consulta con SY-MANDT)\n   - Tu usuario SAP\n   - Hora exacta del error\n\nCrear ticket en categoría SAP con toda esta info. El equipo ABAP analizará el dump (transacción ST22).'},
        {'category': 'SAP', 'priority': 'medium', 'keywords': 'sap,reporte,lento,timeout,no genera',
         'question': 'Reporte SAP muy lento o da timeout',
         'answer': '1. **Limita el rango**:\n   - Filtra por fechas más pequeñas\n   - Filtra por sociedad/centro/material específico\n   - Usa variantes guardadas\n\n2. **Ejecuta en background (en segundo plano)**:\n   - Programa → Ejecutar en segundo plano (F9)\n   - Recibirás notificación cuando termine\n   - Ver con: SP01 o SMX\n\n3. **Si es crítico**:\n   - Reporta a soporte SAP\n   - Posible optimización por equipo BASIS\n   - Tabla con índices faltantes (transacción DBACOCKPIT)'},

        # SEGURIDAD
        {'category': 'Seguridad', 'priority': 'critical', 'keywords': 'phishing,sospechoso,fraude,correo extraño',
         'question': 'Recibí un correo sospechoso (posible phishing)',
         'answer': '**NO HAGAS CLICK en ningún link**\n\n1. **Identificadores de phishing**:\n   - Urgencia anormal ("actúa ahora!")\n   - Solicita contraseñas\n   - Errores ortográficos\n   - Remitente extraño o suplantado\n   - Links a dominios extraños\n\n2. **Acciones**:\n   - NO respondas\n   - NO hagas click en links\n   - NO descargues adjuntos\n   - Repórtalo: "Inicio → Reportar como phishing" (Outlook)\n   - Adjunta el correo a un ticket de Seguridad\n   - Elimínalo\n\n3. **Si hiciste click**: \n   - Cambia tu contraseña INMEDIATAMENTE\n   - Notifica a Seguridad\n   - Escanea tu equipo con antivirus'},
        {'category': 'Seguridad', 'priority': 'critical', 'keywords': 'virus,malware,ransomware,infectado,trojan',
         'question': 'Mi equipo parece tener un virus',
         'answer': '**DESCONECTA EL EQUIPO DE LA RED INMEDIATAMENTE** (saca cable Ethernet, apaga WiFi).\n\n1. Síntomas comunes:\n   - Lentitud extrema\n   - Ventanas emergentes constantes\n   - Archivos cifrados/renombrados (ransomware)\n   - Programas desconocidos\n   - Antivirus deshabilitado\n   - Navegador redirige a sitios extraños\n\n2. **Acciones**:\n   - NO ingreses contraseñas\n   - Apaga el equipo\n   - Llama a soporte: ext. 5000\n   - Reporta como CRÍTICO\n\n3. NO intentes "limpiarlo" tú mismo - puede empeorar\n\nEl equipo de Seguridad usará herramientas especializadas. Posible reinstalación completa.'},
        {'category': 'Seguridad', 'priority': 'high', 'keywords': 'usb,dispositivo,no reconoce,bloqueado',
         'question': '¿Por qué no puedo usar mi USB?',
         'answer': 'Por políticas de seguridad, los puertos USB están restringidos:\n\n- **USBs personales**: bloqueados (riesgo de virus)\n- **USBs corporativos cifrados**: permitidos\n- **Dispositivos médicos**: aprobación necesaria\n\nSi necesitas transferir datos:\n1. **Internamente**: usa OneDrive, SharePoint o carpetas de red\n2. **Externamente**: WeTransfer corporativo o link OneDrive con expiración\n3. **USB corporativo**: solicita uno en categoría Seguridad\n\nMover datos por USBs personales viola políticas y puede acarrear sanciones.'},
        {'category': 'Seguridad', 'priority': 'high', 'keywords': 'mfa,2fa,doble factor,authenticator,token',
         'question': '¿Cómo configuro el doble factor (MFA)?',
         'answer': '1. Instala app autenticadora:\n   - **Microsoft Authenticator** (recomendado)\n   - Google Authenticator\n\n2. Ve a https://mysignins.microsoft.com\n3. Información de seguridad → Agregar método → Aplicación autenticadora\n4. Escanea el código QR con la app\n5. Verifica con código de 6 dígitos\n\n**Importante**:\n- Habilita notificaciones push (más cómodo)\n- Guarda códigos de recuperación en lugar seguro\n- Si cambias de móvil, configura ANTES de deshacerte del viejo\n\nMFA es obligatorio para todos los empleados desde 2024.'},

        # HARDWARE
        {'category': 'Hardware', 'priority': 'high', 'keywords': 'no enciende,no prende,muerto,sin energia',
         'question': 'Mi computador no enciende',
         'answer': '**Laptop**:\n1. Conecta al cargador (puede estar sin batería)\n2. Espera 30 segundos y presiona botón encendido\n3. Si el LED de carga no se enciende: cargador o batería dañados\n4. Reset por hardware: mantén botón encendido 30 segundos sin cargador\n5. Pulsa el botón mientras conectas el cargador\n\n**Desktop**:\n1. Verifica enchufe a corriente\n2. Verifica interruptor trasero de la fuente\n3. Cable del monitor conectado al CPU (no a la placa madre)\n4. Probar con otra fuente de poder si tienes\n\nSi no hay LEDs ni ventiladores: contacta soporte (posible falla de fuente).'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'monitor,pantalla,no señal,no signal,negro',
         'question': 'Mi monitor no muestra imagen',
         'answer': '1. Verifica que el monitor esté encendido (LED frontal)\n2. Verifica cable de video (HDMI/DisplayPort/VGA):\n   - Bien conectado en ambos extremos\n   - Sin daños visibles\n3. Si tienes 2 cables, prueba el otro\n4. Cambia la entrada del monitor (botón "Source/Input")\n5. Si el CPU funciona pero el monitor no:\n   - Verifica que está conectado a la GPU (no a la placa base) en desktops\n   - Prueba con otro monitor\n6. Conexión correcta:\n   - Laptop: cable a salida HDMI/USB-C\n   - Desktop: cable a tarjeta gráfica (no a placa madre)\n\nWin+P para cambiar modo de pantalla (extender/duplicar)'},
        {'category': 'Hardware', 'priority': 'low', 'keywords': 'mouse,teclado,no funciona,inalambrico,wireless',
         'question': 'Mi teclado o mouse no funciona',
         'answer': '**Inalámbrico**:\n1. Verifica nivel de batería\n2. Verifica el receptor USB (cambia de puerto)\n3. Reinicia: apaga 10 segundos y enciende\n4. Re-empareja: presiona botón de pareo\n\n**Bluetooth**:\n1. Configuración → Dispositivos Bluetooth\n2. Quitar dispositivo\n3. Modo emparejamiento (botón especial 3-5 segundos)\n4. Volver a agregar\n\n**Cableado**:\n1. Cambia de puerto USB\n2. Prueba en otro equipo\n3. Si funciona en otro: problema del PC\n4. Si no funciona en otro: hardware dañado, reemplazar\n\nReinicia el equipo después de instalación.'},

        # MOVILES
        {'category': 'Móvil', 'priority': 'medium', 'keywords': 'correo movil,outlook movil,configurar correo telefono',
         'question': '¿Cómo configuro el correo corporativo en mi móvil?',
         'answer': '**Outlook App (recomendado)**:\n1. Descarga "Microsoft Outlook" desde tu tienda\n2. Abrir → Agregar cuenta\n3. Ingresa correo corporativo\n4. Acepta los requisitos de la empresa (MDM)\n5. Configura PIN o biométrico\n\n**Configuración manual (Exchange)**:\n- Servidor: outlook.office365.com\n- Dominio: empresa\n- Usuario: tu correo completo\n- Puerto: 443\n- SSL: Sí\n\n**Política de seguridad**:\n- Tu móvil debe tener PIN/contraseña\n- Cifrado activado\n- Si se pierde, IT puede borrar los datos corporativos remotamente'},
        {'category': 'Móvil', 'priority': 'low', 'keywords': 'teams movil,zoom movil,llamadas',
         'question': '¿Cómo uso Teams en mi móvil?',
         'answer': '1. Descarga "Microsoft Teams" desde App Store/Play Store\n2. Inicia sesión con tu cuenta corporativa\n3. Verifica MFA cuando se solicite\n\nFuncionalidades:\n- Chats y llamadas\n- Reuniones (puede unirse desde calendario)\n- Compartir pantalla del móvil\n- Notificaciones push\n\nConsejos:\n- Habilita notificaciones para no perder mensajes\n- Usa audio del móvil en reuniones de pie\n- Para reuniones largas, conecta auriculares\n- Modo "no molestar" en horarios fuera de trabajo'},

        # TELEFONIA
        {'category': 'Telefonía', 'priority': 'medium', 'keywords': 'telefono,extension,llamadas,no funciona',
         'question': 'Mi teléfono IP no funciona',
         'answer': '1. **Reinicia el teléfono**:\n   - Desconecta el cable de red 10 segundos\n   - Vuelve a conectar\n   - Espera que se inicie (1-2 min)\n\n2. **Verifica cable de red**:\n   - Conectado al puerto correcto\n   - LED del puerto encendido\n\n3. **Sin tono**:\n   - Verifica volumen del auricular\n   - Prueba auricular alternativo\n\n4. **No registra**:\n   - Indica error en pantalla\n   - Contacta soporte con número de extensión\n\nExtensiones por área:\n- 1000-1999: Administración\n- 2000-2999: Ventas\n- 3000-3999: Operaciones\n- 4000-4999: TI'},
        {'category': 'Telefonía', 'priority': 'low', 'keywords': 'desviar,llamadas,reenvio,forward',
         'question': '¿Cómo desvío mis llamadas?',
         'answer': 'En teléfono IP corporativo:\n\n1. Marca **\\*72** + número destino (siempre desvío)\n2. Marca **\\*73** para cancelar desvío\n3. **\\*92** + número = desvío ocupado\n4. **\\*94** + número = desvío no contesta\n\nPara desviar al móvil:\n- *72 + 9 + número celular (sin guiones)\n\nPara recibir mensajes de voz:\n- Marca **\\*86** desde tu teléfono\n- Ingresa PIN (defecto: 1234)\n\nMensajería de voz también llega a tu correo como archivo MP3.'},

        # SERVIDORES Y APLICACIONES
        {'category': 'Servidores', 'priority': 'high', 'keywords': 'no carga,aplicacion,sistema caido,no funciona',
         'question': 'Una aplicación interna no carga',
         'answer': '1. **Verifica estado del sistema**:\n   - Página de estado: status.empresa.local\n   - Pregunta a colegas si les pasa lo mismo\n\n2. **Si solo a ti**:\n   - Limpia caché del navegador (Ctrl+Shift+Supr)\n   - Prueba modo incógnito\n   - Prueba otro navegador (Chrome/Edge/Firefox)\n   - Reinicia tu sesión\n\n3. **Si a todos**:\n   - El servicio puede estar caído\n   - Crea ticket en categoría Servidores\n   - Indica: aplicación, hora, mensaje de error\n\n4. **Errores comunes**:\n   - 404: página no existe\n   - 500: error del servidor\n   - 502/503: servicio no disponible\n   - 401/403: sin permisos'},
        {'category': 'Servidores', 'priority': 'medium', 'keywords': 'carpeta red,compartido,no accede,acceso denegado',
         'question': 'No puedo acceder a una carpeta de red',
         'answer': '1. **Verifica conectividad**:\n   - Estás en la red corporativa o VPN\n   - Mapa de red: \\\\fileserver\\departamento\n\n2. **Reconectar unidad**:\n   - Win+R → "cmd"\n   - "net use Z: /delete"\n   - "net use Z: \\\\fileserver\\compartido /persistent:yes"\n\n3. **Permisos**:\n   - Pide acceso al gerente del área dueña de la carpeta\n   - Aprobación llega a soporte\n   - Tiempo de respuesta: 1-2 días\n\n4. **Acceso denegado**:\n   - Cierra sesión y vuelve a iniciar\n   - Esto refresca los permisos de grupo'},

        # GENERAL / INFO
        {'category': 'General', 'priority': 'low', 'keywords': 'horario,soporte,atencion,ayuda',
         'question': '¿Cuál es el horario de soporte TI?',
         'answer': '**Mesa de Ayuda**:\n- Lunes a Viernes: 7:00 AM - 6:00 PM\n- Sábados: 8:00 AM - 12:00 PM\n- Domingos y festivos: cerrado\n\n**Emergencias 24/7**:\n- Solo para incidentes CRÍTICOS\n- Caídas de servicios productivos\n- Ext. 5911 o +57 300 555 0911\n- WhatsApp soporte emergencias: +57 300 555 0911\n\n**Canales no-emergencia**:\n- DeskEli (este sistema): siempre disponible\n- Email: soporte@empresa.com\n- Teams: canal "Soporte TI"\n- Bot de soporte: 24/7 para preguntas frecuentes'},
        {'category': 'General', 'priority': 'low', 'keywords': 'sla,tiempo respuesta,cuanto tarda',
         'question': '¿Cuál es el tiempo de respuesta de soporte?',
         'answer': '**SLA por prioridad**:\n\n| Prioridad | Respuesta | Resolución |\n|-----------|-----------|------------|\n| Crítica | 15 min | 2 horas |\n| Alta | 1 hora | 4 horas |\n| Media | 4 horas | 1 día |\n| Baja | 1 día | 5 días |\n\n**Definiciones**:\n- **Crítica**: Servicio caído, afecta múltiples usuarios o producción\n- **Alta**: No puedes trabajar, sin alternativa\n- **Media**: Inconveniente pero puedes trabajar\n- **Baja**: Solicitud no urgente, mejora\n\nSi tu ticket no recibe respuesta en el SLA, escala a tu gerente.'},
        {'category': 'General', 'priority': 'low', 'keywords': 'capacitacion,curso,training,aprender',
         'question': '¿Hay capacitaciones disponibles?',
         'answer': 'Sí, ofrecemos capacitaciones mensuales sobre:\n\n- **Office 365**: Excel avanzado, PowerPoint, Outlook\n- **Teams**: Reuniones, canales, colaboración\n- **Seguridad**: Phishing, MFA, contraseñas\n- **SAP**: Por módulos (FI, MM, SD, etc.)\n- **Power BI**: Dashboards básicos\n\n**Para inscribirte**:\n1. Portal Capacita: capacita.empresa.local\n2. Calendario en Teams: canal "Capacitaciones TI"\n3. Solicitud al ticket en categoría General → Capacitación\n\nLas capacitaciones internas son gratis. Externas requieren aprobación de tu gerente.'},
        {'category': 'General', 'priority': 'low', 'keywords': 'inventario,activos,asignacion,equipo',
         'question': '¿Cómo consulto mis activos asignados?',
         'answer': '**Portal de activos**:\n1. activos.empresa.local\n2. Inicia sesión con tu cuenta corporativa\n3. "Mis activos" muestra:\n   - Equipos (PC, laptop, tablet)\n   - Periféricos (monitor, teclado, mouse)\n   - Móviles corporativos\n   - Licencias de software\n\n**Reportar daño**:\n- Click en el activo → "Reportar incidencia"\n- O crea ticket en categoría Hardware\n\n**Devolución**:\n- Al cambiar de equipo o salir de la empresa\n- Coordina con tu gerente y soporte\n- Firma acta de entrega/devolución'},

        # MIGRACIONES Y CAMBIOS
        {'category': 'General', 'priority': 'medium', 'keywords': 'nuevo equipo,cambio pc,migracion datos,traslado',
         'question': '¿Cómo traspaso mis datos a un equipo nuevo?',
         'answer': '**Antes de la migración**:\n1. Asegúrate que TODO esté en OneDrive (no solo Escritorio local)\n2. Exporta firmas de Outlook\n3. Anota tu lista de software instalado\n4. Backup de archivos personales en red\n\n**Durante la entrega del nuevo equipo**:\n1. Soporte instala configuración base\n2. Inicia sesión con tu cuenta corporativa\n3. OneDrive sincroniza automáticamente\n4. Outlook se configura solo\n5. Instala software adicional según necesidad\n\n**Verifica**:\n- [ ] Documentos sincronizados\n- [ ] Correos completos\n- [ ] Conexión VPN\n- [ ] SAP y otros sistemas\n- [ ] Impresoras agregadas\n\nLa migración típicamente toma 1-2 horas.'},

        # ========== 100 NUEVAS SOLUCIONES ==========

        # === CONTRASEÑAS / AUTENTICACIÓN (10) ===
        {'category': 'Contraseñas', 'priority': 'medium', 'keywords': 'cambiar pin,nuevo pin,resetear pin',
         'question': '¿Cómo cambio mi PIN de Windows Hello?',
         'answer': '1. Inicio → Configuración → Cuentas\n2. Opciones de inicio de sesión\n3. PIN (Windows Hello) → Cambiar\n4. Confirma contraseña actual\n5. Ingresa nuevo PIN (mínimo 4 dígitos)\n6. Confirma de nuevo'},
        {'category': 'Contraseñas', 'priority': 'high', 'keywords': 'cuenta hackeada,intrusion,acceso no autorizado',
         'question': 'Sospecho que mi cuenta fue hackeada',
         'answer': '**ACCIÓN INMEDIATA**:\n1. Cambia tu contraseña AHORA desde otro dispositivo\n2. Cierra todas las sesiones activas: mysignins.microsoft.com\n3. Revisa actividad reciente de inicios de sesión\n4. Activa MFA si no lo tienes\n5. Notifica a Seguridad: ext. 5911\n6. Revisa carpeta Enviados de correo (¿se enviaron mensajes que no fuiste tú?)'},
        {'category': 'Contraseñas', 'priority': 'low', 'keywords': 'gestor contraseñas,password manager,guardar credenciales',
         'question': '¿Puedo usar un gestor de contraseñas?',
         'answer': 'Sí, **se recomienda**. Opciones aprobadas:\n- Microsoft Authenticator (sincroniza con cuenta MS)\n- Bitwarden (open source)\n- 1Password (corporativo, consultar disponibilidad)\n\n**NO USAR**:\n- Guardar contraseñas en archivos .txt\n- Compartir por chat o correo\n- Usar la misma contraseña en varios sitios'},
        {'category': 'Contraseñas', 'priority': 'medium', 'keywords': 'expira,expirar,vencimiento,90 dias',
         'question': '¿Cada cuánto expira mi contraseña?',
         'answer': 'Las contraseñas corporativas expiran cada **90 días**.\n\nRecibirás:\n- Notificación 14 días antes\n- Recordatorio 7 días antes\n- Aviso diario los últimos 3 días\n\nPara cambiarla antes:\n- Ctrl+Alt+Suprimir → Cambiar contraseña\n- O desde passwordreset.empresa.local'},
        {'category': 'Contraseñas', 'priority': 'high', 'keywords': 'olvide pin,pin bloqueado',
         'question': 'Olvidé mi PIN de Windows',
         'answer': '1. En la pantalla de login, debajo del PIN: "¿Olvidaste tu PIN?"\n2. Inicia sesión con tu contraseña\n3. Verifica identidad (MFA si tienes)\n4. Configura un nuevo PIN\n\nSi no puedes:\n- Conecta el equipo a la red corporativa\n- Reinicia\n- Vuelve a intentar\n- Última opción: ticket en categoría Accesos'},
        {'category': 'Autenticación', 'priority': 'medium', 'keywords': 'nuevo dispositivo,authenticator,perdi celular',
         'question': 'Cambié de celular, ¿cómo migro mi Authenticator?',
         'answer': '**ANTES de descartar el teléfono viejo**:\n1. Microsoft Authenticator app → Configuración → Iniciar copia de seguridad en la nube\n2. En el nuevo teléfono: instala app, inicia sesión con misma cuenta personal\n3. Tap en "Recuperar desde copia de seguridad"\n\n**Si ya no tienes el teléfono viejo**:\n- Ve a https://mysignins.microsoft.com/security-info\n- Elimina el método "App de autenticación" antiguo\n- Agrega uno nuevo con el celular actual\n- Necesitarás MFA alternativo (SMS, llamada) para confirmar'},
        {'category': 'Autenticación', 'priority': 'low', 'keywords': 'codigos respaldo,recuperacion',
         'question': '¿Cómo genero códigos de respaldo de MFA?',
         'answer': '1. Ve a https://mysignins.microsoft.com/security-info\n2. Busca "Códigos de recuperación"\n3. Generar nuevos códigos (10 únicos)\n4. **Guárdalos en lugar SEGURO** (no en el celular)\n5. Usa uno cada vez que no tengas acceso a Authenticator\n\nCada código es de **un solo uso**.'},
        {'category': 'Autenticación', 'priority': 'high', 'keywords': 'mfa no funciona,no recibo codigo',
         'question': 'No me llegan códigos SMS de verificación',
         'answer': '1. Verifica el número registrado en https://mysignins.microsoft.com\n2. Asegúrate que el celular tenga señal\n3. Verifica que no esté en modo avión / no molestar\n4. Espera 5 minutos (a veces hay retrasos)\n5. Prueba "Llamarme" en lugar de SMS\n6. Si nada funciona: usa la app Authenticator\n\nÚltima opción: contactar Soporte para reset de MFA con verificación de identidad.'},
        {'category': 'Autenticación', 'priority': 'low', 'keywords': 'sso,single sign on,inicio sesion unico',
         'question': '¿Qué es SSO (Single Sign-On)?',
         'answer': '**SSO** = Inicio de Sesión Único.\n\nSignifica que al ingresar UNA vez con tu cuenta corporativa, accedes automáticamente a:\n- Correo (Outlook)\n- OneDrive\n- Teams\n- SharePoint\n- DeskEli\n- Aplicaciones internas\n\nSin volver a ingresar contraseña. Es más seguro y conveniente.'},
        {'category': 'Contraseñas', 'priority': 'medium', 'keywords': 'politica contraseñas,requisitos',
         'question': '¿Cuál es la política de contraseñas?',
         'answer': '**Requisitos mínimos**:\n- 12 caracteres mínimo\n- Al menos 1 mayúscula\n- Al menos 1 minúscula\n- Al menos 1 número\n- Al menos 1 símbolo especial\n- NO puede contener tu nombre/usuario\n- NO puede ser igual a las últimas 5\n\n**Recomendaciones**:\n- Usa frases largas (más seguras)\n- Ejemplo: "MiCafe2024SinAzucar!"\n- NO uses fechas personales\n- NO la repitas en otros sitios'},

        # === HARDWARE - LAPTOPS Y MONITORES (10) ===
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'bateria,no carga,laptop bateria',
         'question': 'Mi laptop no carga la batería',
         'answer': '1. Verifica que el cargador esté conectado correctamente\n2. Prueba con otro tomacorriente\n3. Inspecciona el cable del cargador (sin dobleces o cortes)\n4. Limpia el conector con aire comprimido\n5. Calibra la batería:\n   - Cárgala al 100%\n   - Desconecta y úsala hasta apagado\n   - Cárgala 100% sin interrupciones\n6. Si el LED no enciende al conectar: cargador defectuoso\n7. Reporta a soporte para reemplazo si persiste'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'laptop calienta,sobrecalentamiento,temperatura',
         'question': 'Mi laptop se calienta mucho',
         'answer': '1. Verifica que las rejillas de ventilación NO estén bloqueadas\n2. No la uses sobre cama o sofá (bloquean ventilación)\n3. Usa base con ventilador si trabajas largas horas\n4. Cierra apps pesadas innecesarias\n5. Verifica en Administrador de Tareas qué consume CPU\n6. Limpia ventiladores (cada 6 meses, soporte puede ayudar)\n7. Si pasa los 80°C continuamente: requiere mantenimiento'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'tactil,touchpad,no funciona',
         'question': 'El touchpad de mi laptop no funciona',
         'answer': '1. Verifica tecla función: Fn + F7 (o icono touchpad) para activar\n2. Configuración → Dispositivos → Panel táctil\n3. Verifica que "Panel táctil" esté ON\n4. Actualiza driver:\n   - Win+X → Administrador de dispositivos\n   - Dispositivos de interfaz humana → Panel táctil\n   - Click derecho → Actualizar driver\n5. Si tienes mouse externo conectado, desconéctalo y reinicia\n6. Reset de configuración táctil en Settings'},
        {'category': 'Hardware', 'priority': 'high', 'keywords': 'pantalla parpadea,flickering',
         'question': 'La pantalla parpadea o tiembla',
         'answer': '1. **Identifica si es la pantalla o el cable**:\n   - Mueve la pantalla → si cambia, es el cable\n   - Conecta a monitor externo → si externo OK, es pantalla\n2. Actualiza el driver de gráficos\n3. Cambia la tasa de refresco:\n   - Configuración → Pantalla → Avanzado\n   - Tasa: 60Hz estable\n4. Revisa cables si es desktop\n5. Si persiste: requiere reparación o reemplazo'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'monitor segundo,dual monitor,extender',
         'question': '¿Cómo conecto un segundo monitor?',
         'answer': '1. Conecta el monitor secundario por HDMI/DisplayPort/USB-C\n2. Win+P → Selecciona modo:\n   - **Solo pantalla del PC**: solo principal\n   - **Duplicar**: ambas iguales\n   - **Extender**: escritorio amplio (recomendado)\n   - **Solo segunda**: solo monitor externo\n3. Configuración → Sistema → Pantalla\n4. Arrastra las pantallas para definir posición\n5. Asigna resolución para cada una\n6. Si no detecta el segundo monitor: prueba otro cable o puerto'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'audio,sonido,no escucha',
         'question': 'No tengo sonido en mi computador',
         'answer': '1. Verifica volumen NO esté en silencio (icono altavoz)\n2. Click derecho icono altavoz → "Abrir mezclador de volumen"\n3. Sube volumen de la app específica\n4. Verifica el dispositivo de salida correcto:\n   - Click derecho altavoz → Sonidos → Reproducción\n5. Si usas headset, prueba con altavoces y viceversa\n6. Actualiza driver de audio (Administrador de dispositivos)\n7. Ejecuta solucionador: Configuración → Solución problemas → Audio'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'microfono,mic,no graba,no escucha',
         'question': 'Mi micrófono no funciona en reuniones',
         'answer': '1. Configuración → Sistema → Sonido\n2. Entrada → verifica dispositivo correcto seleccionado\n3. Prueba: habla y mira la barra de nivel\n4. Permisos:\n   - Configuración → Privacidad → Micrófono\n   - Activar "Permitir acceso al micrófono"\n   - Verificar apps específicas (Teams, Zoom)\n5. En Teams: Configuración → Dispositivos → seleccionar mic\n6. Reinicia la app o reinicia el PC'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'camara,webcam,no funciona,bloqueada',
         'question': 'Mi cámara web no funciona',
         'answer': '1. Verifica que la lente NO esté tapada (algunos laptops tienen interruptor físico)\n2. Permisos:\n   - Configuración → Privacidad → Cámara\n   - Activar acceso global y de apps\n3. Cierra otras apps que puedan usar la cámara (Teams, OBS)\n4. Actualiza driver de cámara\n5. Privacidad LED: si está rojo, otra app la está usando\n6. Prueba con app Cámara de Windows para descartar problema de hardware'},
        {'category': 'Hardware', 'priority': 'low', 'keywords': 'docking,dock,estacion',
         'question': 'Mi docking station no detecta dispositivos',
         'answer': '1. Desconecta y reconecta el cable USB-C/Thunderbolt principal\n2. Verifica que el dock tenga corriente (LED)\n3. Reinstala drivers del fabricante del dock\n4. En Win+X → Administrador de dispositivos: busca dispositivos con !\n5. Actualiza Windows (las nuevas versiones traen mejor soporte)\n6. Reinicia con dock desconectado, luego conecta'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'usb no detecta,puerto usb,reconoce',
         'question': 'Los puertos USB no reconocen dispositivos',
         'answer': '1. Prueba el dispositivo en OTRO puerto USB\n2. Prueba OTRO dispositivo en el mismo puerto\n3. Reinicia el PC\n4. Win+X → Administrador de dispositivos:\n   - Buscar "Controladoras USB"\n   - Click derecho → Desinstalar (no asustes, se reinstalan al reiniciar)\n5. Configuración de energía:\n   - Panel control → Opciones de energía → Cambiar plan → Avanzada\n   - USB → Suspensión selectiva → Deshabilitada\n6. Si todos los puertos fallan: requiere mantenimiento'},

        # === SOFTWARE / APPS (10) ===
        {'category': 'Software', 'priority': 'medium', 'keywords': 'chrome lento,navegador lento',
         'question': 'Chrome está muy lento',
         'answer': '1. Limpia caché: Ctrl+Shift+Supr\n2. Desactiva extensiones innecesarias: chrome://extensions\n3. Cierra pestañas que no uses (consumen memoria)\n4. Actualiza Chrome: chrome://settings/help\n5. Restablece configuración: chrome://settings/reset\n6. Verifica que no haya malware: chrome://settings/cleanup\n7. Última opción: desinstalar y reinstalar'},
        {'category': 'Software', 'priority': 'high', 'keywords': 'no instala,error instalacion',
         'question': 'No puedo instalar un programa',
         'answer': '1. Verifica si tu cuenta tiene permisos de administrador local\n2. Click derecho instalador → "Ejecutar como administrador"\n3. Verifica si el archivo no está bloqueado:\n   - Click derecho → Propiedades → "Desbloquear"\n4. Desactiva temporalmente antivirus (puede bloquear)\n5. Verifica espacio en disco C: (mínimo 5GB)\n6. Si requiere permisos elevados: solicita instalación por soporte'},
        {'category': 'Software', 'priority': 'medium', 'keywords': 'desinstalar programa,eliminar app',
         'question': '¿Cómo desinstalo un programa?',
         'answer': '1. Configuración → Aplicaciones → Aplicaciones instaladas\n2. Busca el programa\n3. Click en ⋯ (3 puntos) → Desinstalar\n4. Sigue el asistente\n\n**Si no se desinstala**:\n- Panel de control → Programas → Programas y características\n- Selecciona programa → Desinstalar\n\nPara programas que dejan residuos: usa **Revo Uninstaller** (consulta a soporte primero).'},
        {'category': 'Software', 'priority': 'low', 'keywords': 'office 365,version office,actualizar office',
         'question': '¿Qué versión de Office tengo?',
         'answer': '1. Abre Word, Excel o Outlook\n2. Archivo → Cuenta\n3. Verás:\n   - "Acerca de Word" → versión exacta\n   - Información del producto: versión\n   - Botón "Actualizar opciones" para verificar actualizaciones\n\nVersiones corporativas comunes:\n- Microsoft 365 (suscripción - siempre actualizada)\n- Office 2021 LTSC (perpetuo)\n- Office 2019 (legacy)'},
        {'category': 'Software', 'priority': 'medium', 'keywords': 'powerpoint,presentacion no abre',
         'question': 'PowerPoint no abre o presenta errores',
         'answer': '1. Cierra Office completamente\n2. Win+R → "powerpnt /safe" (modo seguro)\n3. Si funciona en modo seguro:\n   - Archivo → Opciones → Complementos\n   - Administrar: "Complementos COM" → Ir\n   - Desactivar uno a uno hasta encontrar el problemático\n4. Reparar Office:\n   - Configuración → Aplicaciones → Microsoft 365\n   - Modificar → Reparación rápida\n5. Última opción: Reparación en línea'},
        {'category': 'Office', 'priority': 'medium', 'keywords': 'outlook archivo grande,pst grande,bandeja llena',
         'question': 'Mi archivo PST de Outlook es muy grande',
         'answer': '1. Verifica tamaño:\n   - Archivo → Configuración de cuenta → Archivo de datos\n   - Click "Configuración del archivo" → Compactar\n2. Archiva correos antiguos:\n   - Archivo → Información → Herramientas de limpieza → Archivar\n   - Mover correos mayores a X meses\n3. Vacía Elementos eliminados\n4. Vacía Correo no deseado\n5. Considera mover adjuntos grandes a OneDrive\n\n**Límite recomendado**: < 50GB (puede afectar rendimiento si pasa).'},
        {'category': 'Office', 'priority': 'medium', 'keywords': 'pdf,convertir pdf,exportar pdf',
         'question': '¿Cómo convierto un documento a PDF?',
         'answer': '**Desde Word/Excel/PowerPoint**:\n1. Archivo → Guardar como\n2. Tipo: PDF (*.pdf)\n3. Opciones avanzadas:\n   - "Estándar" para calidad alta\n   - "Tamaño mínimo" para correo\n4. Guardar\n\n**Desde cualquier app (imprimir)**:\n1. Imprimir (Ctrl+P)\n2. Impresora: "Microsoft Print to PDF"\n3. Imprimir → elige nombre y ubicación'},
        {'category': 'Software', 'priority': 'low', 'keywords': 'captura pantalla,screenshot,recortes',
         'question': '¿Cómo tomo una captura de pantalla?',
         'answer': '**Métodos en Windows**:\n1. **PrtScn**: copia toda la pantalla al portapapeles\n2. **Alt + PrtScn**: solo ventana activa\n3. **Win + Shift + S**: herramienta de recorte (recomendada)\n4. **Win + PrtScn**: guarda automáticamente en Imágenes\n5. **App "Recortes"**: Win → buscar "Recortes"\n\nPara grabar pantalla: **Win + G** (Game Bar) o usa Teams para compartir y grabar.'},
        {'category': 'Software', 'priority': 'medium', 'keywords': 'modo oscuro,dark mode,tema oscuro',
         'question': '¿Cómo activo el modo oscuro?',
         'answer': '**Windows**:\n1. Configuración → Personalización → Colores\n2. "Elige tu modo" → Oscuro\n3. Aplica a apps de Microsoft y compatibles\n\n**Office**:\n1. Archivo → Cuenta\n2. Tema de Office → Negro / Oscuro\n\n**Chrome/Edge**: chrome://flags → "Force Dark Mode" o ajustes del navegador.\n\n**Teams**: Configuración (⋯) → General → Tema: Oscuro'},
        {'category': 'Software', 'priority': 'medium', 'keywords': 'asistente,cortana,desactivar',
         'question': '¿Cómo desactivo Cortana / Asistente?',
         'answer': '**Windows 10/11**:\n1. Click derecho barra tareas → Configuración barra tareas\n2. Buscar → Oculto\n3. Cortana → Oculto\n\nPara desactivar completamente:\n- Win+R → "gpedit.msc" (solo Pro)\n- Configuración del equipo → Plantillas administrativas → Componentes de Windows → Buscar\n- "Permitir Cortana" → Deshabilitado'},

        # === RED / CONECTIVIDAD (10) ===
        {'category': 'Red', 'priority': 'medium', 'keywords': 'ip,direccion ip,cual es mi ip',
         'question': '¿Cómo veo mi dirección IP?',
         'answer': '**IP local (red interna)**:\n1. Win+R → "cmd"\n2. Escribe: ipconfig\n3. Busca "Dirección IPv4"\n\n**IP pública (internet)**:\n- Google → "cuál es mi ip"\n- O visita: https://www.whatismyip.com\n\n**Detalles completos**:\n- ipconfig /all (en cmd)'},
        {'category': 'Red', 'priority': 'high', 'keywords': 'proxy,configurar proxy',
         'question': '¿Cómo configuro el proxy corporativo?',
         'answer': '**Automático (preferido)**:\n1. Configuración → Red e Internet → Proxy\n2. "Detectar automáticamente"\n\n**Manual** (si se requiere):\n1. Servidor: proxy.empresa.local\n2. Puerto: 8080\n3. Excluir direcciones internas: *.empresa.local;127.0.0.1\n\nUsuario/contraseña: usar tu cuenta corporativa.'},
        {'category': 'Red', 'priority': 'medium', 'keywords': 'wifi visitas,red invitados,visitantes',
         'question': '¿Cómo doy WiFi a visitantes?',
         'answer': 'Las visitas deben conectarse a la red **WiFi-Guest** (no a la corporativa).\n\n1. Recepción genera código de un día\n2. Conectar a SSID: `Guest-Empresa`\n3. Aceptar términos en el portal cautivo\n4. Ingresar código\n\n**Importante**:\n- La red Guest NO accede a recursos internos\n- Velocidad limitada\n- Sesión expira en 24h\n- NO compartas tu WiFi corporativo personal'},
        {'category': 'Red', 'priority': 'medium', 'keywords': 'red lenta dia especifico,horario',
         'question': 'La red está lenta solo en ciertos horarios',
         'answer': 'Causas comunes:\n\n1. **Hora pico** (9-11 AM, 2-4 PM): mucho uso simultáneo\n2. **Backups programados**: 12-2 PM y 10 PM\n3. **Actualizaciones masivas**: domingos\n4. **Streaming**: si hay eventos en vivo\n\nQué hacer:\n- Trabaja con tareas pesadas en horas valle\n- Programa descargas grandes para horarios nocturnos\n- Reporta si el lentitud es constante todo el día'},
        {'category': 'Red', 'priority': 'medium', 'keywords': 'cable red,ethernet no funciona',
         'question': 'El cable de red no funciona',
         'answer': '1. Verifica que el cable esté conectado a ambos extremos\n2. LED del puerto: debe estar encendido (verde/amarillo)\n3. Prueba con OTRO cable\n4. Prueba en OTRO puerto de pared\n5. ipconfig en cmd → verifica si hay IP\n6. Win+R → ncpa.cpl → Ethernet:\n   - Click derecho → Deshabilitar → Habilitar\n7. Si no hay IP: contactar soporte (puede ser configuración del puerto)'},
        {'category': 'Red', 'priority': 'high', 'keywords': 'firewall,bloqueado,no accede sitio',
         'question': 'No puedo acceder a un sitio web específico',
         'answer': '1. Verifica que el sitio funcione en otros computadores\n2. Prueba en modo incógnito\n3. Limpia caché DNS: cmd → "ipconfig /flushdns"\n4. El sitio puede estar **bloqueado por política**:\n   - Redes sociales en horario laboral\n   - Sitios de entretenimiento\n   - Sitios maliciosos detectados\n5. Si es necesario para trabajo: solicita desbloqueo en ticket de Red (justificar uso)'},
        {'category': 'Red', 'priority': 'medium', 'keywords': 'wifi se desconecta,intermitente',
         'question': 'Mi WiFi se desconecta cada cierto tiempo',
         'answer': '1. Verifica nivel de señal (icono WiFi). Si es bajo:\n   - Acércate al Access Point\n   - Cambia a banda 5GHz si tu equipo lo soporta\n2. Olvida la red y reconecta\n3. Actualiza driver WiFi:\n   - Administrador de dispositivos → Adaptadores de red\n4. Configuración avanzada de adaptador → "Roaming Aggressiveness": Medium\n5. Reporta a soporte con horarios de desconexión'},
        {'category': 'Red', 'priority': 'low', 'keywords': 'medir velocidad,test internet',
         'question': '¿Cómo mido la velocidad de internet?',
         'answer': '1. Cierra apps que consuman ancho de banda\n2. Ve a https://www.speedtest.net o https://fast.com\n3. Inicia la prueba\n4. Anota:\n   - Velocidad de bajada (download)\n   - Velocidad de subida (upload)\n   - Latencia (ping)\n\n**Estándares aceptables corporativos**:\n- Bajada: ≥ 50 Mbps\n- Subida: ≥ 10 Mbps\n- Ping: < 50 ms\n\nSi está muy bajo, reporta con captura de pantalla.'},
        {'category': 'Red', 'priority': 'medium', 'keywords': 'cisco,jabber,zoom',
         'question': 'No puedo conectarme a videollamadas',
         'answer': '1. Verifica internet estable (>5 Mbps)\n2. Cierra otras apps que consuman ancho de banda\n3. Permite acceso a cámara y micrófono\n4. Si usas VPN: prueba sin VPN (algunas calls van por túnel separado)\n5. Actualiza la app (Teams/Zoom/Webex)\n6. Reinicia el router si trabajas desde casa\n7. Conecta por cable Ethernet si tienes (más estable que WiFi)'},
        {'category': 'Red', 'priority': 'high', 'keywords': 'puerto bloqueado,firewall puerto',
         'question': 'Necesito que abran un puerto en firewall',
         'answer': 'Para solicitar apertura de puerto:\n\n1. **Crea ticket en categoría Red/Firewall** con:\n   - Aplicación que lo requiere\n   - Justificación de negocio\n   - Puerto exacto (TCP/UDP)\n   - IP destino\n   - Riesgo evaluado\n2. Aprobación necesaria de:\n   - Tu gerente\n   - Seguridad de TI\n3. Implementación en ventana de mantenimiento\n\n**No se aprueban** puertos genéricos (FTP 21 abierto, etc.) salvo casos especiales.'},

        # === SAP / SISTEMAS NEGOCIO (10) ===
        {'category': 'SAP', 'priority': 'medium', 'keywords': 'sap rendimiento,sap lento general',
         'question': 'SAP responde muy lento en general',
         'answer': '1. Verifica que estés en la red corporativa (no Guest WiFi)\n2. Cierra transacciones que no uses (cada una consume memoria)\n3. Borra caché SAP:\n   - Cerrar SAP completamente\n   - Borrar carpeta: %APPDATA%\\SAP\\Common\\Cache\n4. Reinicia SAP\n5. Si es lento en TODOS los usuarios: problema del servidor (reportar a BASIS)\n6. Si es solo en tu PC: problema local (driver, antivirus, RAM)'},
        {'category': 'SAP', 'priority': 'low', 'keywords': 'sap variantes,guardar variante',
         'question': '¿Cómo guardo una variante en SAP?',
         'answer': '1. Ingresa a la transacción\n2. Llena los filtros que quieres guardar\n3. Menú: Ir a → Variantes → Guardar como variante\n4. Asigna:\n   - Nombre de variante (sin espacios)\n   - Descripción\n   - Solo para mí / Universal\n5. Guardar\n\nPara cargar después: Variantes → Obtener... → seleccionar.'},
        {'category': 'SAP', 'priority': 'medium', 'keywords': 'sap exportar excel,descarga sap',
         'question': '¿Cómo exporto datos de SAP a Excel?',
         'answer': '1. Ejecuta el reporte o transacción\n2. Cuando muestre la grilla de datos:\n   - Lista → Exportar → Hoja de cálculo\n   - O click derecho → Exportar\n3. Selecciona:\n   - Excel (en MHTML format)\n   - Hoja de cálculo XML\n4. Asigna nombre y ubicación\n\nPara reportes muy grandes (>50K filas), considera **ejecutar en background** y descargar el archivo después.'},
        {'category': 'SAP', 'priority': 'high', 'keywords': 'sap mandante,cliente sap',
         'question': '¿Cómo sé en qué mandante SAP estoy?',
         'answer': '1. En la barra de estado inferior de SAP, busca:\n   - SID (sistema): ej PRD/QAS/DEV\n   - Mandante: ej 100, 200, 300\n   - Usuario: tu nombre de usuario SAP\n2. O ejecuta transacción: SU3 (datos propios)\n\n**Mandantes típicos**:\n- 100: Producción\n- 200: Quality\n- 300: Desarrollo\n\n**Atención**: si modificas en PRD-100 afecta el negocio en vivo. Verifica antes de cambios.'},
        {'category': 'SAP', 'priority': 'medium', 'keywords': 'sap autorizaciones,permisos sap',
         'question': 'Me sale "No autorizado" en una transacción SAP',
         'answer': '1. Anota el código exacto del error (S_TCODE, etc.)\n2. Verifica con tu gerente si DEBES tener acceso\n3. Si es así, crea ticket en categoría SAP con:\n   - Transacción solicitada\n   - Justificación de negocio\n   - Aprobador (gerente)\n   - Si es temporal o permanente\n4. El equipo de Roles/Authorizations procesará\n\nTiempo estimado: 24-48h después de aprobación.'},
        {'category': 'SAP', 'priority': 'low', 'keywords': 'sap favoritos,menu favoritos',
         'question': '¿Cómo agrego una transacción a Favoritos en SAP?',
         'answer': '1. En el menú principal SAP, click derecho en "Favoritos"\n2. Selecciona "Insertar transacción"\n3. Ingresa el código (ej. FB60, MM03)\n4. Asigna nombre descriptivo\n5. Aceptar\n\n**Tip**: Arrastra y suelta desde el menú estándar a Favoritos.\nReorganiza con click derecho → Mover hacia arriba/abajo.'},
        {'category': 'SAP', 'priority': 'medium', 'keywords': 'sap impresion,formulario',
         'question': 'No puedo imprimir desde SAP',
         'answer': '1. Verifica que tengas impresoras asignadas:\n   - Transacción SPAD\n2. Configura impresora predeterminada:\n   - SU3 → pestaña Defaults\n   - Dispositivo de salida: tu impresora\n3. En el momento de imprimir:\n   - Selecciona dispositivo en el spool\n   - Marca "Salida inmediata"\n4. Si no aparece: verifica nombre exacto de impresora con soporte\n5. Para PDF: imprime a "Microsoft Print to PDF"'},
        {'category': 'SAP', 'priority': 'medium', 'keywords': 'sap formulario imprimir mal,layout',
         'question': 'Un formulario SAP imprime con formato incorrecto',
         'answer': '1. Verifica que el form sea para tu impresora (LOCL local vs servidor)\n2. Configura impresora:\n   - Tipo: ZHPLJ (Hewlett Packard LaserJet)\n   - Densidad: 12\n3. Si es formulario SAPscript modificado:\n   - Reportar a desarrollo ABAP con captura\n   - Especificar transacción y formulario (ej. RVINVOICE01)\n4. Para SmartForms o Adobe Forms: similar, reportar a desarrollo'},
        {'category': 'SAP', 'priority': 'low', 'keywords': 'sap atajos,shortcuts,teclas',
         'question': '¿Cuáles son los atajos de SAP más útiles?',
         'answer': '**Navegación**:\n- F3: Volver atrás\n- F4: Lista de valores (búsqueda)\n- F1: Ayuda\n- F8: Ejecutar\n- F11: Guardar\n- /n: Nueva transacción\n- /o: Nueva ventana\n- /nex: Cerrar SAP\n\n**Edición**:\n- Ctrl+S: Guardar\n- Ctrl+P: Imprimir\n- Ctrl+Y: Copiar pantalla\n- Ctrl+A: Seleccionar todo'},
        {'category': 'SAP', 'priority': 'high', 'keywords': 'sap caido,sap down,inaccesible',
         'question': 'SAP no responde - sistema caído',
         'answer': 'Si SAP no responde:\n\n1. **Verifica si es solo tu PC** o todos:\n   - Pregunta a colegas\n   - Si solo tú: reinicia tu cliente SAP\n2. **Si es general**:\n   - NO hagas nada con datos en transacciones abiertas\n   - Cierra tu cliente sin guardar (datos se respaldan en BD)\n   - Reporta INMEDIATAMENTE a soporte (categoría SAP - URGENTE)\n   - Notifica a tu gerente del impacto\n3. Equipo BASIS investiga la causa\n4. Espera comunicado oficial para reconectarse'},

        # === SEGURIDAD (10) ===
        {'category': 'Seguridad', 'priority': 'critical', 'keywords': 'ransomware,archivos cifrados,extorsion',
         'question': 'Mis archivos aparecen cifrados / con extensión rara',
         'answer': '**¡ACCIÓN INMEDIATA - POSIBLE RANSOMWARE!**\n\n1. **DESCONECTA EL EQUIPO**: cable de red + WiFi\n2. **NO APAGUES** el computador (pierde evidencia)\n3. **NO PAGUES** ningún rescate\n4. Llama URGENTE a Seguridad: ext. 5911\n5. No intentes recuperar archivos solo\n6. Identifica el alcance:\n   - ¿Solo C: o también unidades de red?\n   - ¿Otros equipos cerca afectados?\n\nEl equipo de Seguridad activará el protocolo de respuesta a incidentes.'},
        {'category': 'Seguridad', 'priority': 'high', 'keywords': 'sospechoso link,url falsa',
         'question': '¿Cómo identifico un link sospechoso?',
         'answer': '**Señales de alerta**:\n1. Dominios extraños: paypa1.com (con 1 en vez de l)\n2. URLs largas y confusas\n3. Acortadores (bit.ly, tinyurl) en correos importantes\n4. HTTP sin S (sin cifrado)\n5. Mensaje de urgencia anormal\n6. Solicita información sensible\n7. Errores ortográficos en el dominio\n\n**Cómo verificar**:\n- Pasa el mouse SIN clickear → veras la URL real\n- Busca el sitio oficial en Google\n- Verifica con Seguridad si tienes duda'},
        {'category': 'Seguridad', 'priority': 'medium', 'keywords': 'wifi publica,starbucks,aeropuerto',
         'question': '¿Es seguro usar WiFi pública?',
         'answer': '**NO se recomienda para trabajo corporativo**.\n\nSi DEBES usar WiFi pública:\n1. ✅ **SIEMPRE conecta VPN corporativa primero**\n2. ✅ Solo accede a sitios HTTPS (candado verde)\n3. ❌ NO accedas a Bancos\n4. ❌ NO ingreses contraseñas sensibles\n5. ❌ NO transfieras archivos confidenciales\n\n**Mejor alternativa**: Hotspot de tu celular con datos móviles (cifrado).'},
        {'category': 'Seguridad', 'priority': 'high', 'keywords': 'computador robado,equipo perdido',
         'question': 'Perdí o me robaron mi laptop corporativa',
         'answer': '**ACCIÓN INMEDIATA** (no esperes a mañana):\n\n1. Notifica a Seguridad: ext. 5911 o WhatsApp 24/7\n2. Notifica a tu gerente\n3. Cambia contraseñas desde otro equipo:\n   - Email corporativo\n   - VPN\n   - SAP\n   - Otros sistemas\n4. Si fue robo: denuncia policial\n5. Seguridad activará:\n   - Borrado remoto (si MDM activo)\n   - Bloqueo de cuenta\n   - Revisión de accesos recientes\n6. Llena formato de pérdida de activo en RRHH'},
        {'category': 'Seguridad', 'priority': 'medium', 'keywords': 'antivirus alerta,deteccion,malware',
         'question': 'Mi antivirus detectó un virus',
         'answer': '1. **NO ignores** la alerta\n2. Permite que el antivirus **ponga en cuarentena** o elimine el archivo\n3. NO recuperes el archivo de cuarentena\n4. Ejecuta análisis completo del sistema\n5. Si detecta múltiples archivos: posible infección activa\n6. Reporta a Seguridad con captura de la alerta\n7. Indica de dónde obtuviste el archivo (correo, descarga, USB)\n\nNO uses el equipo para tareas sensibles hasta confirmar limpieza.'},
        {'category': 'Seguridad', 'priority': 'high', 'keywords': 'enviar info confidencial,datos privados',
         'question': '¿Cómo envío información confidencial de forma segura?',
         'answer': '**Métodos APROBADOS**:\n1. **OneDrive/SharePoint** con link de acceso limitado\n2. **Microsoft Teams** (cifrado end-to-end en chats privados)\n3. **Email corporativo** marcando como Confidencial:\n   - Outlook → Opciones → Cifrar → Solo cifrar\n\n**EVITAR**:\n- ❌ WhatsApp para datos corporativos\n- ❌ Gmail personal\n- ❌ USBs sin cifrado\n- ❌ Servicios como WeTransfer público\n\nPara documentos muy sensibles: agregar contraseña al archivo (Word/Excel → Cifrar con contraseña).'},
        {'category': 'Seguridad', 'priority': 'medium', 'keywords': 'permisos administrador,admin local',
         'question': '¿Por qué no tengo permisos de administrador local?',
         'answer': 'Por política de seguridad, los usuarios estándar NO tienen permisos admin para:\n- Prevenir instalación de malware\n- Mantener integridad del equipo\n- Cumplir normativas\n\n**Si NECESITAS instalar software**:\n1. Crea ticket en categoría Software\n2. Justifica para qué lo necesitas\n3. Soporte instala como admin\n\n**Excepciones**: roles que requieran admin (desarrolladores, técnicos TI) tienen permisos elevados separados.'},
        {'category': 'Seguridad', 'priority': 'low', 'keywords': 'bloqueo pantalla,protector pantalla',
         'question': '¿Cómo bloqueo mi pantalla rápido?',
         'answer': '**Win + L** = Bloquea inmediatamente.\n\n**Política corporativa**: el equipo se bloquea automáticamente tras 10 min de inactividad.\n\n**Buenas prácticas**:\n- Bloquea SIEMPRE al alejarte del PC\n- No dejes notas con contraseña pegadas\n- Si trabajas en lugares públicos: pantalla de privacidad\n\n**Recuerda**: dejar pantalla desbloqueada es una violación de seguridad que puede ser auditada.'},
        {'category': 'Seguridad', 'priority': 'medium', 'keywords': 'compartir contraseña,dar password',
         'question': 'Un compañero me pide mi contraseña, ¿qué hago?',
         'answer': '**NUNCA compartas tu contraseña**, ni con:\n- Compañeros\n- Tu gerente\n- Soporte técnico (legítimo NUNCA la pide)\n- Cualquier persona externa\n\nSi necesita acceso a información:\n1. Comparte el archivo específico (no la cuenta)\n2. Solicita permisos en la herramienta\n3. Que cree su propia cuenta\n\nSi alguien insiste: repórtalo como intento de ingeniería social.\n\n**Recuerda**: las acciones con tu cuenta quedan registradas a tu nombre.'},
        {'category': 'Seguridad', 'priority': 'high', 'keywords': 'datos personales,gdpr,proteccion',
         'question': '¿Qué datos personales puedo manejar?',
         'answer': 'Según política de Protección de Datos:\n\n**Permitido (con autorización)**:\n- Nombre, email corporativo, teléfono\n- Datos de empleados internos\n- Información de clientes con consentimiento\n\n**Sensible - requiere encriptación**:\n- Documentos de identidad\n- Datos médicos\n- Información financiera\n- Datos biométricos\n\n**Prohibido**:\n- Compartir con terceros sin autorización\n- Almacenar fuera de sistemas corporativos\n- Enviar por canales no seguros\n\nAnte dudas: consulta con Compliance.'},

        # === MOBILIARIO TECNOLÓGICO / DISPOSITIVOS (10) ===
        {'category': 'Móvil', 'priority': 'medium', 'keywords': 'celular corporativo,asignacion movil',
         'question': '¿Cómo solicito un celular corporativo?',
         'answer': 'Pasos:\n1. Tu cargo debe estar autorizado para celular (consulta RRHH)\n2. Llena formato de solicitud en portal RRHH\n3. Aprobación de tu gerente\n4. Aprobación del responsable presupuestal\n5. Soporte coordina entrega:\n   - Equipo asignado\n   - SIM con plan corporativo\n   - Configuración inicial\n6. Firmas acta de entrega\n\nTiempo total: 1-2 semanas según disponibilidad.'},
        {'category': 'Móvil', 'priority': 'medium', 'keywords': 'whatsapp business,whatsapp empresa',
         'question': '¿Puedo usar WhatsApp para trabajo?',
         'answer': 'Por política:\n\n**Permitido**:\n- Coordinación interna informal\n- Mensajes no confidenciales\n- WhatsApp Business si es tu rol (ventas, soporte cliente)\n\n**NO permitido**:\n- Compartir documentos confidenciales\n- Discutir datos sensibles de clientes\n- Reemplazar comunicación oficial (email/Teams)\n\n**Recomendación**: Usa Teams para temas laborales. Es más seguro y queda registro corporativo.'},
        {'category': 'Móvil', 'priority': 'low', 'keywords': 'instalar app,permitir aplicacion movil',
         'question': '¿Puedo instalar apps en mi celular corporativo?',
         'answer': '**Según política MDM (Mobile Device Management)**:\n\n- ✅ Apps de productividad estándar: aprobadas\n- ⚠️ Apps personales: en perfil personal (separado)\n- ❌ Apps de fuentes desconocidas: bloqueadas\n- ❌ Jailbreak/Root: prohibido\n\nLas apps corporativas (Outlook, Teams, Authenticator) se instalan automáticamente.\n\nPara solicitar app no estándar: ticket en Móvil con justificación.'},
        {'category': 'Móvil', 'priority': 'medium', 'keywords': 'backup movil,respaldo telefono',
         'question': '¿Cómo respaldo mi celular?',
         'answer': '**Android**:\n1. Configuración → Sistema → Copia de seguridad\n2. Habilita "Google Drive backup"\n3. Selecciona qué respaldar: apps, fotos, contactos, SMS\n\n**iPhone**:\n1. Configuración → [Tu nombre] → iCloud\n2. Copia en iCloud → activar\n3. "Hacer copia ahora"\n\n**Importante**:\n- Datos corporativos (correo, Teams) NO se respaldan en cuenta personal\n- Quedan en la nube de Microsoft separadamente'},
        {'category': 'Móvil', 'priority': 'medium', 'keywords': 'sincronizar contactos,outlook contactos',
         'question': 'Mis contactos de Outlook no se sincronizan al celular',
         'answer': '**iPhone**:\n1. Configuración → Outlook (app) → Contactos → Activar\n2. O en Outlook app: Settings → Tu cuenta → Contactos → Sincronizar\n\n**Android**:\n1. Outlook app → Settings → Sincronizar contactos\n2. Permitir permisos al sistema\n3. Configuración → Cuentas → Outlook → Sincronizar contactos\n\nPuede tardar hasta 24h en mostrar todos. Reinicia el celular si no aparece.'},
        {'category': 'Hardware', 'priority': 'low', 'keywords': 'mouse inalambrico,bateria mouse',
         'question': '¿Cómo cambio la batería de mi mouse inalámbrico?',
         'answer': '1. Verifica tipo de mouse:\n   - **Con pilas AA/AAA**: deslice tapa inferior\n   - **Recargable USB-C/Micro USB**: conecta cable\n   - **Bluetooth**: ver indicador en pantalla\n2. Reemplaza con baterías nuevas (alcalinas duran más)\n3. Para mouse Logitech: app Logi Options muestra nivel\n4. Si el mouse ya no enciende ni con baterías nuevas: solicita reemplazo'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'auriculares headset,no conecta',
         'question': 'Mis auriculares no se conectan',
         'answer': '**Con cable**:\n1. Verifica que esté conectado al puerto correcto (verde audio)\n2. Configuración → Sonido → Salida → seleccionar\n\n**Bluetooth**:\n1. Configuración → Dispositivos Bluetooth → Agregar\n2. Pon auriculares en modo emparejamiento (mantener botón 3s)\n3. Selecciona desde la lista\n4. Si ya estaban emparejados pero no conectan: olvidar y volver a agregar\n\n**USB**:\n1. Cambia de puerto USB\n2. Espera que Windows instale el driver\n3. Configurar como salida predeterminada'},
        {'category': 'Hardware', 'priority': 'medium', 'keywords': 'webcam externa,configurar camara',
         'question': '¿Cómo configuro una cámara web externa?',
         'answer': '1. Conecta la webcam al USB\n2. Espera 30s a que se instale automáticamente\n3. Verifica en Administrador de dispositivos:\n   - "Dispositivos de imágenes" → tu cámara\n4. Si tiene software del fabricante (Logitech, etc.): instálalo\n5. En Teams:\n   - Configuración → Dispositivos → Cámara → seleccionar\n6. En Zoom: Configuración → Vídeo → cámara\n\nPara mejor calidad: buena iluminación frontal, fondo neutral.'},
        {'category': 'Hardware', 'priority': 'low', 'keywords': 'limpiar teclado,mantenimiento pc',
         'question': '¿Cómo limpio mi teclado y pantalla?',
         'answer': '**Teclado**:\n1. Apaga el PC o desconecta el teclado\n2. Voltea boca abajo y golpea suavemente para sacar polvo\n3. Aire comprimido (lata) entre teclas\n4. Pasa paño microfibra con poco alcohol isopropílico\n\n**Pantalla**:\n1. Apaga el monitor\n2. Pasa paño microfibra SECO\n3. Si hay manchas: rocía paño (NUNCA la pantalla) con limpiador específico\n4. NUNCA uses alcohol, agua o productos con amoniaco\n\nLimpieza recomendada: semanal'},
        {'category': 'Hardware', 'priority': 'high', 'keywords': 'pantalla rota,dañada,quebrada',
         'question': 'Se me cayó el laptop y la pantalla está rota',
         'answer': '**Pasos**:\n1. NO uses el equipo (puede empeorar el daño)\n2. Apaga inmediatamente\n3. Reporta a soporte:\n   - Categoría: Hardware\n   - Prioridad: Alta\n   - Incluye fotos del daño\n4. Solicita equipo temporal mientras se repara\n5. Llena formato de incidente en RRHH (puede haber implicaciones de seguro)\n\nTiempo de reparación: 5-15 días según disponibilidad de repuestos.'},

        # === EMAIL AVANZADO (10) ===
        {'category': 'Email', 'priority': 'medium', 'keywords': 'reglas correo,filtros email',
         'question': '¿Cómo creo reglas para organizar correos?',
         'answer': '**Outlook Desktop**:\n1. Inicio → Reglas → Administrar reglas\n2. Nueva regla:\n   - Empezar desde plantilla\n   - O comenzar desde regla en blanco\n3. Define condición (ej: de "boss@empresa.com")\n4. Define acción (ej: mover a carpeta "Jefes")\n5. Activar\n\n**Outlook Web**:\n1. Configuración → Ver toda la configuración\n2. Correo → Reglas → Agregar regla\n\n**Ideas útiles**:\n- Mover newsletters a carpeta específica\n- Marcar correos VIP\n- Auto-responder fuera de horario'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'reenviar automatico,forward email',
         'question': '¿Cómo configuro reenvío automático?',
         'answer': '**Outlook Web** (mejor opción):\n1. Configuración → Ver toda la configuración\n2. Correo → Reenvío\n3. Activar reenvío\n4. Email destino\n5. (Opcional) "Conservar copia"\n6. Guardar\n\n**Limitaciones**:\n- Solo puedes reenviar a 1 dirección\n- No se pueden reenviar correos confidenciales (puede ser bloqueado)\n- Auditado por políticas DLP\n\n**Para vacaciones**: mejor usar "Respuestas automáticas".'},
        {'category': 'Email', 'priority': 'low', 'keywords': 'calendario compartido,delegar calendario',
         'question': '¿Cómo comparto mi calendario?',
         'answer': '**Outlook**:\n1. Calendario → click derecho en tu calendario\n2. Compartir → Compartir calendario\n3. Agregar permisos a personas:\n   - Puede ver disponibilidad\n   - Puede ver detalles\n   - Puede editar\n   - Delegado (responde por ti)\n4. Enviar invitación\n\n**Para delegar**:\n- Archivo → Configuración cuenta → Delegar acceso\n- Agregar persona y nivel'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'busqueda correos,buscar email',
         'question': '¿Cómo busco un correo viejo?',
         'answer': '**Outlook - búsqueda rápida**:\n1. Barra de búsqueda (arriba)\n2. Escribe texto, remitente, asunto\n\n**Búsqueda avanzada**:\n- `de:nombre@empresa.com` → de remitente\n- `asunto:reporte` → en el asunto\n- `tiene:adjunto` → con adjuntos\n- `recibido:>1/1/2024` → desde fecha\n- `tamaño:>5MB` → grandes\n\n**Ctrl+E**: enfoca el cuadro de búsqueda.\n**Búsqueda avanzada UI**: Ctrl+Alt+S'},
        {'category': 'Email', 'priority': 'low', 'keywords': 'plantilla correo,quick part',
         'question': '¿Cómo creo plantillas de correo?',
         'answer': '**Método 1 - Plantillas .oft**:\n1. Crea correo nuevo con el texto deseado\n2. Archivo → Guardar como\n3. Tipo: Plantilla Outlook (*.oft)\n4. Para usar: Nuevos elementos → Mas elementos → Elegir formulario\n\n**Método 2 - Mis plantillas (recomendado)**:\n1. En nuevo correo: Mensaje → Ver plantillas\n2. + Plantilla\n3. Título y contenido\n4. Guardar\n5. Click en la plantilla para insertar\n\nÚtil para respuestas frecuentes.'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'correo no leido,marcar leido',
         'question': '¿Cómo marco correos como leídos masivamente?',
         'answer': '**Marcar TODOS leídos en una carpeta**:\n1. Click derecho en la carpeta\n2. "Marcar todos como leídos"\n\n**Marcar selección**:\n1. Selecciona correos (Ctrl+click o Shift+click)\n2. Click derecho → "Marcar como leído"\n3. O atajo: **Ctrl+Q**\n\n**Marcar como no leído**: Ctrl+U\n\nÚtil cuando regresas de vacaciones con 500 correos.'},
        {'category': 'Email', 'priority': 'low', 'keywords': 'firmar archivo,certificado digital',
         'question': '¿Cómo firmo digitalmente un correo?',
         'answer': 'Si tienes certificado digital:\n\n1. **Outlook**:\n   - Archivo → Opciones → Centro de confianza → Configuración\n   - Seguridad del correo → Configuración\n   - Firmar todos los mensajes salientes\n2. **Para certificado nuevo**:\n   - Solicita a Seguridad en ticket\n   - Instala en certificado en tu cuenta\n   - Importar en Outlook\n\n**Beneficio**: el receptor sabe con certeza que el correo es tuyo y no fue alterado.'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'organizar bandeja entrada,carpetas',
         'question': '¿Cómo organizo mi bandeja de entrada?',
         'answer': '**Metodología recomendada**:\n\n1. **Inbox Zero**: meta = 0 correos al final del día\n2. Crea carpetas:\n   - 📁 Acción inmediata\n   - 📁 Esperando respuesta\n   - 📁 Para revisar\n   - 📁 Archivo\n3. Reglas automáticas para clasificar\n4. Usa categorías por color\n5. Marca con estrella los importantes\n6. Archiva (Ctrl+E) lo procesado\n\n**Tip**: dedica 2-3 momentos al día a procesar email, no constantemente.'},
        {'category': 'Email', 'priority': 'low', 'keywords': 'estadisticas correo,quien mas envia',
         'question': '¿Puedo ver estadísticas de mis correos?',
         'answer': '**Outlook Web**:\n1. Configuración → Ver toda la configuración\n2. General → Estadísticas\n\nMuestra:\n- Correos enviados/recibidos por día/semana\n- Top remitentes\n- Tiempo promedio de respuesta\n- Uso del buzón\n\n**Microsoft MyAnalytics**: insights de productividad personal (si está habilitado).'},
        {'category': 'Email', 'priority': 'medium', 'keywords': 'enviar a muchas personas,bcc cco',
         'question': '¿Debo usar CCO/BCC al enviar a muchas personas?',
         'answer': '**Sí, mejores prácticas**:\n\n**Para >5 destinatarios externos**:\n- Usa **CCO** (Bcc) para proteger emails de cada uno\n- Pon tu dirección en "Para"\n\n**Para listas internas**:\n- Crea una **lista de distribución**\n- Usa el nombre de la lista en "Para"\n\n**Evita**:\n- "Responder a todos" sin necesidad (genera ruido)\n- Listas de cientos en "Para" (expone direcciones)\n\n**Atajo**: Si no ves CCO, en "Para" → Opciones → CCO'},

        # === GENERAL / VARIOS (10) ===
        {'category': 'General', 'priority': 'low', 'keywords': 'que es DeskEli,sistema tickets',
         'question': '¿qué es DeskEli y para qué sirve?',
         'answer': '**DeskEli** es el sistema oficial de soporte TI de la empresa.\n\n**Sirve para**:\n- Reportar problemas técnicos\n- Solicitar servicios (instalar software, dar accesos)\n- Hacer consultas al equipo TI\n- Ver el estado de tus solicitudes\n- Recibir respuestas oficiales con trazabilidad\n\n**Ventajas vs llamadas/correo**:\n- Queda registro oficial\n- Sabes en qué etapa está\n- Recibes SLA garantizado\n- Métricas para mejorar el servicio'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'como crear ticket,nuevo caso',
         'question': '¿Cómo creo un ticket de soporte?',
         'answer': '1. ingresa a DeskEli con tu cuenta\n2. Click "Crear nuevo ticket"\n3. Llena:\n   - **Título**: descriptivo y corto\n   - **Categoría**: Hardware/Software/Red/etc.\n   - **Prioridad**: según urgencia real\n   - **Descripción**: detalle el problema\n4. Adjunta capturas si ayudan\n5. Crear\n\n**Tip**: Antes de crear, prueba con el bot (puede resolver al instante con base de conocimiento).'},
        {'category': 'General', 'priority': 'low', 'keywords': 'donde veo mis tickets,mis casos',
         'question': '¿Dónde veo mis tickets creados?',
         'answer': 'En DeskEli Portal de Empleados:\n\n1. Ingresa con tu usuario\n2. Verás tu **dashboard** con:\n   - Tus tickets abiertos\n   - Tickets en progreso\n   - Tickets resueltos\n3. Click en cualquier ticket → ves detalles + historial\n4. Puedes agregar comentarios\n5. Calificar el servicio cuando se cierre\n\nFiltros disponibles: estado, prioridad, fecha.'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'calificar ticket,encuesta',
         'question': '¿Cómo califico el servicio recibido?',
         'answer': '1. Cuando tu ticket sea resuelto, recibirás notificación\n2. Abre el ticket en DeskEli\n3. Verás opción: "Calificar este servicio"\n4. Asigna 1-5 estrellas:\n   - ⭐⭐⭐⭐⭐ Excelente\n   - ⭐ Muy malo\n5. (Opcional) Comentario\n6. Enviar\n\n**Tu feedback ayuda** a mejorar el servicio y reconocer al equipo.'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'reabrir ticket,no quedo resuelto',
         'question': 'Un ticket fue cerrado pero el problema no se resolvió',
         'answer': '1. Abre el ticket cerrado en DeskEli\n2. Click "Reabrir" (disponible hasta 7 días después del cierre)\n3. Explica por qué necesita reapertura\n4. El técnico recibe notificación\n\nSi ya pasaron 7 días:\n- Crea un nuevo ticket\n- En la descripción menciona el ticket anterior: "Relacionado con TKT-XXX"\n- El equipo TI revisa el caso completo'},
        {'category': 'General', 'priority': 'low', 'keywords': 'horarios laborales,jornada',
         'question': '¿Cuáles son los horarios laborales de la empresa?',
         'answer': '**Horarios generales**:\n- Administrativos: 8:00 AM - 5:00 PM\n- Producción turno 1: 6:00 AM - 2:00 PM\n- Producción turno 2: 2:00 PM - 10:00 PM\n- Producción turno 3: 10:00 PM - 6:00 AM\n\n**TI**:\n- Soporte presencial: Lun-Vie 7 AM - 6 PM\n- Soporte 24/7: solo emergencias críticas\n- Sábados: 8 AM - 12 PM\n\nConsulta RRHH para tu horario específico.'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'soporte sabado,festivo,fin semana',
         'question': '¿Hay soporte los fines de semana?',
         'answer': '**Sábados**: 8 AM - 12 PM (atención limitada)\n**Domingos y festivos**: Solo emergencias 24/7\n\n**Para emergencias críticas** (sistemas caídos):\n- WhatsApp ON-CALL: +57 300 555 0911\n- Email: emergencias-ti@empresa.com\n- DeskEli → categoría "Crítico" → activa alerta automática\n\nPara consultas no urgentes, mejor esperar al lunes.'},
        {'category': 'General', 'priority': 'low', 'keywords': 'directorio empleados,buscar persona',
         'question': '¿Cómo encuentro a un empleado?',
         'answer': '**Outlook** (más rápido):\n1. Nuevo correo\n2. En "Para" empieza a escribir el nombre\n3. Auto-completa con datos del directorio\n4. Click → ver más info (cargo, ubicación, jefe)\n\n**Teams**:\n1. Búsqueda superior → escribir nombre\n2. Ve estado online/ocupado\n3. Chat directo\n\n**Intranet**: portal.empresa.local/directorio'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'recibir notificaciones,desactivar alerts',
         'question': '¿Cómo controlo las notificaciones que recibo?',
         'answer': '**Email DeskEli**:\n1. ingresa a DeskEli → Mi perfil\n2. Notificaciones por correo:\n   - ☑ Nuevo comentario\n   - ☑ Cambio de estado\n   - ☑ Asignado a alguien\n   - ☐ Resumen semanal\n3. Guardar\n\n**Notificaciones del navegador**:\n- Permite o bloquea desde el navegador\n- Útiles para tiempo real\n\n**Push móvil** (si usas app): Settings → Notificaciones'},
        {'category': 'General', 'priority': 'low', 'keywords': 'sugerencia mejora,feedback,buzon',
         'question': '¿Cómo doy sugerencias para mejorar?',
         'answer': '**Canales para feedback**:\n\n1. **DeskEli**: crea ticket en categoría "General → Sugerencia"\n2. **Buzón de sugerencias**: sugerencias@empresa.com\n3. **Encuesta anual** (sale por Outlook)\n4. **Reuniones 1:1** con tu gerente\n5. **Teams - canal Ideas**\n\n**Importante**:\n- Las sugerencias son revisadas mensualmente\n- Recibes respuesta en máx. 30 días\n- Las implementadas son reconocidas formalmente'},

        # === BOT / IA / sistema DeskEli (10) ===
        {'category': 'General', 'priority': 'low', 'keywords': 'bot eli,asistente,inteligencia artificial',
         'question': '¿Cómo uso el bot Eli?',
         'answer': '**Eli** es el asistente IA de DeskEli.\n\n**Cómo usarlo**:\n1. Click en el avatar de Eli (esquina inferior derecha del portal)\n2. Escribe tu pregunta como si fuera a una persona\n3. Eli responde con la solución desde su base de conocimiento (56+ soluciones)\n4. Si no tiene respuesta exacta, usa IA Claude para generar una\n5. Si te resuelve: marcalo "✓ Solucionado"\n6. Si NO: "📋 Crear ticket" y asigna a especialista\n\n**Eli funciona 24/7** y aprende de tus preguntas.'},
        {'category': 'General', 'priority': 'low', 'keywords': 'asignacion automatica,ia tickets',
         'question': '¿Cómo asigna DeskEli los tickets automáticamente?',
         'answer': 'DeskEli usa 4 agentes IA:\n\n1. **Classifier**: clasifica el ticket en una categoría (Hardware, SAP, etc.)\n2. **Assignor**: busca el técnico con:\n   - Habilidades requeridas\n   - Menor carga de trabajo actual\n3. **Responder**: genera respuesta inicial si está en la KB\n4. **Escalator**: monitorea SLA y escala automáticamente\n\nResultado: tus tickets llegan al especialista correcto en segundos, sin intervención manual.'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'sla,acuerdo nivel servicio,tiempo respuesta',
         'question': '¿Qué es el SLA y cuáles son los tiempos?',
         'answer': '**SLA** = Service Level Agreement (Acuerdo de Nivel de Servicio).\n\nEs el compromiso de tiempo del equipo TI para resolver tickets según prioridad:\n\n| Prioridad | Respuesta | Resolución |\n|-----------|-----------|------------|\n| 🔴 Crítica | 15 min | 1 h |\n| 🟠 Alta | 1 h | 4 h |\n| 🟡 Media | 4 h | 1 día |\n| 🟢 Baja | 1 día | 5 días |\n\nSi tu ticket no cumple SLA, el sistema escala automáticamente al supervisor.'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'prioridad ticket,como elegir',
         'question': '¿Qué prioridad debo darle a mi ticket?',
         'answer': '**Guía para elegir prioridad**:\n\n🔴 **CRÍTICA**:\n- Sistema caído afectando producción\n- Múltiples usuarios sin trabajar\n- Pérdida de datos en curso\n\n🟠 **ALTA**:\n- No puedes hacer tu trabajo\n- Bloqueo total para ti\n- Sin alternativa\n\n🟡 **MEDIA**:\n- Funciona pero con problemas\n- Tienes workaround temporal\n- Inconveniente pero no urgente\n\n🟢 **BAJA**:\n- Solicitud no urgente\n- Mejora o nueva funcionalidad\n- Información\n\n**Tip**: Marcar todo como CRÍTICA reduce su efectividad. Usa la prioridad real.'},
        {'category': 'General', 'priority': 'low', 'keywords': 'historial ticket,quien hizo',
         'question': '¿Puedo ver el historial de mi ticket?',
         'answer': 'Sí, cada ticket tiene un **historial completo** auditado:\n\n1. Abre el ticket\n2. Sección "Historial" o "Actividad"\n\nVerás:\n- Quién creó el ticket\n- Asignaciones y reasignaciones\n- Cambios de estado\n- Comentarios\n- Adjuntos agregados\n- Escalaciones\n- Tiempo dedicado por cada acción\n\nTodo queda registrado con usuario, fecha/hora exacta.'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'eligir tecnico,asignar a quien',
         'question': '¿Puedo elegir qué técnico atienda mi ticket?',
         'answer': '**No directamente** - el sistema asigna automáticamente al mejor técnico disponible según:\n- Habilidades requeridas\n- Carga actual de trabajo\n- Disponibilidad horaria\n\n**Excepciones**:\n- Si tienes un técnico asignado para casos especiales: menciónalo en la descripción\n- Si requiere un especialista específico: mencionalo en el ticket\n- Si es una continuación: cita el ticket anterior\n\nEsto garantiza distribución equitativa y atención rápida.'},
        {'category': 'General', 'priority': 'low', 'keywords': 'imprimir ticket,exportar ticket',
         'question': '¿Puedo imprimir o exportar un ticket?',
         'answer': '1. Abre el ticket\n2. Botón "Imprimir" o "Exportar":\n   - **PDF**: descarga el ticket completo con historial\n   - **Imprimir**: usa el diálogo de impresión\n\nÚtil para:\n- Reportes a tu gerente\n- Auditorías\n- Documentación de procesos\n- Compartir con compañeros que no usan DeskEli\n\nEl PDF mantiene todo el formato e información del ticket.'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'donde aprender,manual,documentacion',
         'question': '¿Dónde aprendo a usar DeskEli?',
         'answer': '**Recursos disponibles**:\n\n1. **Bot Eli**: pregúntale cualquier cosa\n2. **Base de conocimiento**: 100+ artículos en DeskEli\n3. **Capacitaciones mensuales**: en Teams\n4. **Manual de usuario**: portal.empresa.local/manuales\n5. **Tutoriales en video**: canal interno de Stream\n6. **FAQ**: DeskEli.empresa.local/faq\n\n**Capacitación 1-on-1**: solicítala vía ticket en categoría "Capacitación".'},
        {'category': 'General', 'priority': 'low', 'keywords': 'puedo ver tickets compañeros,otros tickets',
         'question': '¿Puedo ver los tickets de mis compañeros?',
         'answer': '**No por privacidad**. Cada usuario solo ve:\n- SUS propios tickets creados\n- Tickets donde es asignado (técnicos)\n- Todos los de la empresa (admins)\n\n**Excepciones**:\n- Si trabajan en el mismo ticket: agregar como colaborador\n- Si necesitas información compartida: pídela al usuario o al técnico\n\n**Multi-tenant**: NUNCA verás tickets de otras empresas (Eliot/Pash/Primatela están aislados).'},
        {'category': 'General', 'priority': 'medium', 'keywords': 'sistema lento,DeskEli lento',
         'question': 'DeskEli está lento o no responde',
         'answer': '1. Refresca la página (Ctrl+F5)\n2. Limpia caché del navegador (Ctrl+Shift+Supr)\n3. Prueba en navegador distinto (Chrome/Edge)\n4. Verifica conexión a internet (speedtest.net)\n5. Verifica el estado del servicio:\n   - Bot Eli si está accesible (señal de que el servicio responde)\n   - Otros sistemas como Outlook ¿funcionan?\n6. Si solo es DeskEli: reporta vía WhatsApp ON-CALL\n7. Si todo está lento: problema de red general'},
    ]

    created = 0
    skipped = 0
    for entry in kb_entries:
        existing = BotKnowledge.query.filter_by(question=entry['question']).first()
        if existing:
            skipped += 1
            continue
        kb = BotKnowledge(
            keywords=entry['keywords'],
            question=entry['question'],
            answer=entry['answer'],
            category=entry['category'],
            priority=entry['priority']
        )
        db.session.add(kb)
        created += 1

    db.session.commit()
    log_audit('seed_botkb', session['user_id'], 'botkb', None, f'KB inicial: {created} creadas, {skipped} omitidas')

    return jsonify({
        'success': True,
        'message': f'Base de conocimiento ampliada: {created} soluciones nuevas ({skipped} ya existían)',
        'created': created,
        'skipped': skipped,
        'total': BotKnowledge.query.count()
    })


@app.route('/api/admin/templates', methods=['GET'])
def api_admin_templates_list():
    """Listar plantillas de tickets de la empresa"""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    company = session['company']
    templates = Template.query.filter_by(company=company).order_by(Template.category, Template.name).all()
    result = []
    for t in templates:
        # Parsear form_fields JSON
        form_fields = []
        if t.form_fields:
            try:
                form_fields = json.loads(t.form_fields)
            except Exception:
                form_fields = []
        result.append({
            'id': t.id,
            'name': t.name,
            'description': t.description or '',
            'title_template': t.title_template,
            'description_template': t.description_template or '',
            'category': t.category or 'General',
            'priority': t.priority or 'medium',
            'is_system': bool(t.is_system),
            'form_fields': form_fields
        })
    return jsonify({'success': True, 'templates': result})


@app.route('/api/admin/templates/regenerate-forms', methods=['POST'])
def api_admin_templates_regenerate_forms():
    """Regenera form_fields de todas las plantillas del sistema parseando description_template."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    count = convert_legacy_templates_to_forms(force=True)
    return jsonify({
        'success': True,
        'message': f'{count} plantillas regeneradas con campos específicos extraídos de su descripción.',
        'count': count
    })


@app.route('/api/admin/templates', methods=['POST'])
def api_admin_templates_create():
    """Crear plantilla"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    try:
        data = request.get_json()
        name = (data.get('name') or '').strip()
        title_template = (data.get('title_template') or '').strip()

        if not name or not title_template:
            return jsonify({'success': False, 'error': 'Nombre y título son requeridos'}), 400

        t = Template(
            name=name,
            description=(data.get('description') or '').strip(),
            title_template=title_template,
            description_template=(data.get('description_template') or '').strip(),
            category=(data.get('category') or 'General').strip(),
            priority=(data.get('priority') or 'medium').strip(),
            company=session['company'],
            is_system=False
        )
        db.session.add(t)
        db.session.commit()

        log_audit('create_template', session['user_id'], 'template', t.id, f'Plantilla creada: {name}')
        return jsonify({'success': True, 'id': t.id, 'message': 'Plantilla creada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/templates/<int:template_id>', methods=['PUT'])
def api_admin_templates_update(template_id):
    """Actualizar plantilla"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    t = Template.query.get(template_id)
    if not t or t.company != session['company']:
        return jsonify({'success': False, 'error': 'No encontrada'}), 404

    try:
        data = request.get_json()
        if 'name' in data: t.name = data['name'].strip()
        if 'description' in data: t.description = data['description'].strip()
        if 'title_template' in data: t.title_template = data['title_template'].strip()
        if 'description_template' in data: t.description_template = data['description_template'].strip()
        if 'category' in data: t.category = data['category'].strip()
        if 'priority' in data: t.priority = data['priority'].strip()
        db.session.commit()

        log_audit('update_template', session['user_id'], 'template', t.id, f'Plantilla actualizada: {t.name}')
        return jsonify({'success': True, 'message': 'Plantilla actualizada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/templates/<int:template_id>', methods=['DELETE'])
def api_admin_templates_delete(template_id):
    """Eliminar plantilla"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    t = Template.query.get(template_id)
    if not t or t.company != session['company']:
        return jsonify({'success': False, 'error': 'No encontrada'}), 404

    name = t.name
    db.session.delete(t)
    db.session.commit()
    log_audit('delete_template', session['user_id'], 'template', template_id, f'Plantilla eliminada: {name}')
    return jsonify({'success': True, 'message': 'Plantilla eliminada'})


@app.route('/api/admin/templates/seed', methods=['POST'])
def api_admin_templates_seed():
    """Crear plantillas iniciales predefinidas para la empresa"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']
    seed_templates = [
        # HARDWARE
        {'name': 'Computador no enciende', 'category': 'Hardware', 'priority': 'high',
         'title_template': 'Computador no enciende',
         'description_template': 'Equipo: [Marca/Modelo]\nUbicación: [Oficina/Piso]\nUsuario afectado: [Nombre]\nFecha de incidente: [Fecha]\n\nDescripción del problema:\n- ¿LED encendido al presionar botón?\n- ¿Hace algún sonido?\n- ¿Conectado a la corriente?\n\nAcciones realizadas:\n- '},
        {'name': 'Pantalla con problemas', 'category': 'Hardware', 'priority': 'medium',
         'title_template': 'Monitor / Pantalla presenta fallas',
         'description_template': 'Tipo de problema:\n[ ] Pantalla negra\n[ ] Líneas o píxeles muertos\n[ ] Parpadea constantemente\n[ ] Colores incorrectos\n[ ] No detecta señal\n\nMarca/Modelo del monitor:\nDuración del problema:\nUsuario afectado: '},
        {'name': 'Teclado o mouse no funciona', 'category': 'Hardware', 'priority': 'medium',
         'title_template': 'Periférico (teclado/mouse) no responde',
         'description_template': 'Dispositivo afectado:\n[ ] Teclado\n[ ] Mouse\n[ ] Ambos\n\nTipo de conexión:\n[ ] USB cableado\n[ ] Inalámbrico USB\n[ ] Bluetooth\n\nMarca/Modelo:\nUsuario: '},
        {'name': 'Impresora no imprime', 'category': 'Hardware', 'priority': 'medium',
         'title_template': 'Impresora con falla',
         'description_template': 'Impresora afectada: [Nombre/Modelo]\nUbicación: [Piso/Área]\nUsuarios afectados: [1 / Varios / Todos]\n\nSíntoma:\n[ ] No imprime nada\n[ ] Imprime mal calidad\n[ ] Atasco de papel\n[ ] Error en pantalla\n[ ] No detecta papel/tóner\n\nMensaje de error mostrado: '},
        {'name': 'Disco duro lleno', 'category': 'Hardware', 'priority': 'medium',
         'title_template': 'Espacio insuficiente en disco',
         'description_template': 'Equipo afectado: [Nombre PC]\nUsuario: \nUnidad afectada: [C: / D:]\nEspacio actual disponible: [GB]\n\nArchivos que requiere conservar:\n- \n\n¿Tiene respaldos? [Sí/No]'},

        # SOFTWARE
        {'name': 'Software no abre', 'category': 'Software', 'priority': 'medium',
         'title_template': 'Aplicación no inicia',
         'description_template': 'Aplicación afectada: \nVersión: \nSistema operativo: \n\nMensaje de error (si aparece):\n\nAcciones intentadas:\n[ ] Reiniciar PC\n[ ] Reinstalar app\n[ ] Ejecutar como administrador\n\nFrecuencia del problema: [Siempre/Intermitente]'},
        {'name': 'Instalación de software', 'category': 'Software', 'priority': 'low',
         'title_template': 'Solicitud de instalación de software',
         'description_template': 'Software solicitado: \nVersión específica: \nMotivo de la solicitud: \n\nUsuario que solicita: \nGerente/Aprobador: \n\n¿Es licencia gratuita o comercial?\n[ ] Gratuito\n[ ] Comercial (adjuntar comprobante de compra)\n\nUrgencia: '},
        {'name': 'Actualización de software', 'category': 'Software', 'priority': 'low',
         'title_template': 'Solicitud de actualización de software',
         'description_template': 'Software a actualizar: \nVersión actual: \nVersión deseada: \n\nMotivo:\n[ ] Funcionalidad nueva requerida\n[ ] Vulnerabilidad de seguridad\n[ ] Compatibilidad\n[ ] Otro: '},
        {'name': 'Office no funciona', 'category': 'Software', 'priority': 'high',
         'title_template': 'Microsoft Office presenta fallas',
         'description_template': 'Aplicación afectada:\n[ ] Word\n[ ] Excel\n[ ] PowerPoint\n[ ] Outlook\n[ ] Otra: \n\nVersión Office: \nSíntoma: \nMensaje de error: \nArchivos afectados: '},
        {'name': 'Navegador con problemas', 'category': 'Software', 'priority': 'low',
         'title_template': 'Navegador web no funciona correctamente',
         'description_template': 'Navegador:\n[ ] Chrome\n[ ] Edge\n[ ] Firefox\n\nSíntoma:\n[ ] No carga páginas\n[ ] Cierre inesperado\n[ ] Páginas lentas\n[ ] Errores de certificado\n\nSitios afectados: '},

        # RED Y CONECTIVIDAD
        {'name': 'Sin acceso a internet', 'category': 'Red', 'priority': 'high',
         'title_template': 'No tengo conexión a internet',
         'description_template': 'Equipo afectado: \nUsuario: \nUbicación: [Piso/Oficina]\n\nTipo de conexión:\n[ ] WiFi\n[ ] Cable Ethernet\n\nSíntoma:\n[ ] No conecta a la red\n[ ] Conecta pero sin internet\n[ ] Internet lento\n\n¿Otros usuarios cerca afectados? '},
        {'name': 'WiFi no conecta', 'category': 'Red', 'priority': 'high',
         'title_template': 'Problema de conexión WiFi',
         'description_template': 'Red WiFi: [SSID]\nEquipo: \nUbicación: \n\nProblema:\n[ ] No detecta la red\n[ ] Detecta pero no conecta\n[ ] Conecta y desconecta\n[ ] Conexión muy lenta\n\nDispositivos similares funcionan? [Sí/No]'},
        {'name': 'VPN no conecta', 'category': 'Red', 'priority': 'high',
         'title_template': 'No puedo conectarme a la VPN',
         'description_template': 'Cliente VPN: \nUsuario VPN: \nUbicación: [Casa/Oficina/Externo]\n\nMensaje de error: \n\n¿Cuándo dejó de funcionar?\n¿Funciona en otra red? '},
        {'name': 'Red corporativa lenta', 'category': 'Red', 'priority': 'medium',
         'title_template': 'Conexión a red corporativa muy lenta',
         'description_template': 'Tipo de conexión: [LAN/WiFi/VPN]\nUbicación: \nUsuarios afectados: [1/Varios/Todos]\n\nRecursos afectados:\n[ ] Compartidos de red\n[ ] Aplicaciones internas\n[ ] Internet\n[ ] Email\n\nHorario del problema: '},

        # CUENTAS Y ACCESOS
        {'name': 'Reseteo de contraseña', 'category': 'Accesos', 'priority': 'medium',
         'title_template': 'Solicitud de reseteo de contraseña',
         'description_template': 'Usuario afectado: \nSistema/Aplicación: [Windows/Email/SAP/Otro]\nFecha último acceso exitoso: \n\nMotivo:\n[ ] Olvidé la contraseña\n[ ] Cuenta bloqueada\n[ ] Contraseña expirada\n\nMétodo de verificación de identidad: '},
        {'name': 'Cuenta bloqueada', 'category': 'Accesos', 'priority': 'high',
         'title_template': 'Mi cuenta está bloqueada',
         'description_template': 'Usuario: \nSistema afectado: \nÚltimo acceso exitoso: \n\nMensaje de error: \n¿Cuántos intentos hizo? \n\nNecesidad urgente: '},
        {'name': 'Solicitud de acceso a sistema', 'category': 'Accesos', 'priority': 'medium',
         'title_template': 'Solicitud de acceso a sistema/aplicación',
         'description_template': 'Sistema solicitado: \nMódulo / Permisos requeridos: \nMotivo de la solicitud: \n\nGerente que aprueba: \nUsuario para acceso: \nFecha de necesidad: '},
        {'name': 'Nuevo usuario', 'category': 'Accesos', 'priority': 'medium',
         'title_template': 'Alta de nuevo usuario',
         'description_template': 'Nombre completo: \nCargo: \nÁrea/Departamento: \nFecha de ingreso: \nGerente directo: \n\nAccesos requeridos:\n[ ] Email corporativo\n[ ] Equipo y Windows\n[ ] SAP\n[ ] VPN\n[ ] Carpetas de red\n[ ] Otros: '},
        {'name': 'Baja de usuario', 'category': 'Accesos', 'priority': 'high',
         'title_template': 'Baja de usuario - desactivación de accesos',
         'description_template': 'Usuario a desactivar: \nÚltimo día de trabajo: \nMotivo: [Renuncia/Despido/Vacaciones largas]\n\n¿Conservar archivos? [Sí/No]\nReceptor de archivos: \n\nGerente que solicita: '},

        # EMAIL
        {'name': 'Correo no funciona', 'category': 'Email', 'priority': 'high',
         'title_template': 'Outlook / Correo no funciona',
         'description_template': 'Cuenta de correo: \nCliente: [Outlook/Webmail/Móvil]\n\nSíntoma:\n[ ] No envía\n[ ] No recibe\n[ ] No abre\n[ ] Solicita contraseña\n[ ] Bandeja entrada llena\n\nMensaje de error: '},
        {'name': 'Correos no llegan', 'category': 'Email', 'priority': 'high',
         'title_template': 'No estoy recibiendo correos',
         'description_template': 'Cuenta: \nDesde cuándo no recibe: \n\n¿Remitentes específicos o todos?\nEjemplo de remitente esperado: \n\n¿Revisó spam/correo no deseado?\n¿Bandeja llena? '},
        {'name': 'Lista de distribución', 'category': 'Email', 'priority': 'low',
         'title_template': 'Solicitud crear/modificar lista de distribución',
         'description_template': 'Tipo:\n[ ] Crear nueva lista\n[ ] Agregar miembros\n[ ] Quitar miembros\n[ ] Cambiar nombre\n\nNombre de lista: \nMiembros: \nPropósito: '},

        # SAP
        {'name': 'SAP no abre', 'category': 'SAP', 'priority': 'critical',
         'title_template': 'No puedo acceder a SAP',
         'description_template': 'Usuario SAP: \nAmbiente: [PRD/QAS/DEV]\nMandante: \n\nMensaje de error: \nHora del problema: \n\n¿Otros usuarios afectados? '},
        {'name': 'Error transacción SAP', 'category': 'SAP', 'priority': 'high',
         'title_template': 'Error en transacción SAP',
         'description_template': 'Transacción: \nMódulo: [FI/CO/MM/SD/PP/HCM]\nUsuario: \nMandante: \n\nPasos para reproducir el error:\n1. \n2. \n3. \n\nMensaje de error exacto: \nNúmero de mensaje: '},
        {'name': 'Reporte SAP no genera', 'category': 'SAP', 'priority': 'medium',
         'title_template': 'Reporte SAP no se genera',
         'description_template': 'Reporte/Transacción: \nFiltros aplicados: \nFechas del reporte: \nUsuario que ejecuta: \n\nSíntoma:\n[ ] Se queda cargando\n[ ] Sale en blanco\n[ ] Error de timeout\n[ ] Datos incorrectos\n\nMensaje de error: '},
        {'name': 'Permisos SAP', 'category': 'SAP', 'priority': 'medium',
         'title_template': 'Solicitud de permisos SAP',
         'description_template': 'Usuario SAP: \nMandante: \n\nTransacciones requeridas:\n- \n\nAutorización solicitada:\n[ ] Visualización\n[ ] Crear\n[ ] Modificar\n[ ] Eliminar\n\nMotivo: \nAprobador: '},

        # SEGURIDAD
        {'name': 'Equipo con virus', 'category': 'Seguridad', 'priority': 'critical',
         'title_template': 'Sospecha de virus/malware en equipo',
         'description_template': 'Equipo afectado: \nUsuario: \n\nSíntoma observado:\n[ ] Equipo muy lento\n[ ] Ventanas emergentes\n[ ] Archivos cifrados/desaparecidos\n[ ] Redirige a sitios extraños\n[ ] Antivirus deshabilitado\n\n¿Qué archivo/email recibió antes del problema? '},
        {'name': 'Phishing reportado', 'category': 'Seguridad', 'priority': 'high',
         'title_template': 'Correo sospechoso de phishing',
         'description_template': 'Remitente del correo: \nAsunto: \nFecha recibido: \n\n¿Hizo clic en algún link? [Sí/No]\n¿Descargó archivos adjuntos? [Sí/No]\n¿Ingresó credenciales? [Sí/No]\n\nUsuario que reporta: '},
        {'name': 'Equipo perdido o robado', 'category': 'Seguridad', 'priority': 'critical',
         'title_template': 'Reporte de equipo perdido/robado',
         'description_template': 'Equipo perdido: [Laptop/Móvil/Tablet]\nMarca/Modelo: \nNúmero de serie: \nÚltimo usuario asignado: \n\nFecha del incidente: \nLugar del incidente: \n\n¿Contenía información confidencial? [Sí/No]\nReporte policial: [Sí/No]'},

        # TELEFONIA
        {'name': 'Teléfono no funciona', 'category': 'Telefonía', 'priority': 'medium',
         'title_template': 'Teléfono IP / Extensión no funciona',
         'description_template': 'Extensión afectada: \nUsuario: \nUbicación: \n\nSíntoma:\n[ ] No tiene tono\n[ ] No recibe llamadas\n[ ] No realiza llamadas\n[ ] Calidad de audio mala\n[ ] No registra en la central\n\nMarca/Modelo del teléfono: '},
        {'name': 'Solicitud de extensión', 'category': 'Telefonía', 'priority': 'low',
         'title_template': 'Solicitud nueva extensión telefónica',
         'description_template': 'Usuario solicitante: \nÁrea/Departamento: \nUbicación física: \n\nTipo de extensión:\n[ ] Solo interna\n[ ] Salida local\n[ ] Salida nacional\n[ ] Internacional\n\nAprobador: '},

        # OTROS
        {'name': 'Solicitud reunión virtual', 'category': 'General', 'priority': 'low',
         'title_template': 'Solicitud configuración reunión Teams/Zoom',
         'description_template': 'Tipo de soporte:\n[ ] Configurar cuenta\n[ ] Configurar sala virtual\n[ ] Capacitación de uso\n[ ] Problema técnico durante reunión\n\nFecha del evento: \nParticipantes: \nDuración estimada: '},
        {'name': 'Capacitación', 'category': 'General', 'priority': 'low',
         'title_template': 'Solicitud de capacitación',
         'description_template': 'Tema de capacitación: \nUsuarios a capacitar: \nNivel: [Básico/Intermedio/Avanzado]\n\nModalidad preferida:\n[ ] Presencial\n[ ] Virtual\n[ ] Material escrito\n\nFecha sugerida: '},
        {'name': 'Mantenimiento programado', 'category': 'General', 'priority': 'low',
         'title_template': 'Mantenimiento preventivo de equipo',
         'description_template': 'Equipos a mantener: \nUbicación: \nUsuarios afectados: \nFecha propuesta: \nDuración estimada: \n\nTipo de mantenimiento:\n[ ] Limpieza física\n[ ] Actualización de software\n[ ] Revisión de hardware\n[ ] Backup\n[ ] Otro: '},
    ]

    created = 0
    skipped = 0
    for tpl in seed_templates:
        existing = Template.query.filter_by(company=company, name=tpl['name']).first()
        if existing:
            skipped += 1
            continue
        t = Template(
            name=tpl['name'],
            title_template=tpl['title_template'],
            description_template=tpl['description_template'],
            category=tpl['category'],
            priority=tpl['priority'],
            company=company,
            is_system=True
        )
        db.session.add(t)
        created += 1

    db.session.commit()
    log_audit('seed_templates', session['user_id'], 'template', None, f'Plantillas iniciales: {created} creadas, {skipped} omitidas')

    return jsonify({
        'success': True,
        'message': f'Se crearon {created} plantillas ({skipped} ya existían)',
        'created': created,
        'skipped': skipped
    })


# ===== BUZONES DE CORREO (Email → Ticket) =====

@app.route('/api/admin/mailboxes', methods=['GET'])
def api_admin_mailboxes_list():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    company = session['company']
    boxes = MailboxConfig.query.filter_by(company=company).order_by(MailboxConfig.created_at.desc()).all()
    return jsonify({
        'success': True,
        'mailboxes': [{
            'id': m.id,
            'name': m.name,
            'imap_host': m.imap_host,
            'imap_port': m.imap_port,
            'imap_user': m.imap_user,
            'use_ssl': bool(m.use_ssl),
            'folder': m.folder,
            'default_priority': m.default_priority,
            'default_category': m.default_category,
            'poll_interval_minutes': m.poll_interval_minutes,
            'is_active': bool(m.is_active),
            'last_check_at': m.last_check_at.strftime('%Y-%m-%d %H:%M:%S') if m.last_check_at else None,
            'last_status': m.last_status,
            'last_error': m.last_error,
            'tickets_created': m.tickets_created or 0,
            'has_password': bool(m.imap_password),
            'auth_type': m.auth_type or 'password',
            'oauth_tenant_id': m.oauth_tenant_id or '',
            'oauth_client_id': m.oauth_client_id or '',
            'has_oauth_secret': bool(m.oauth_client_secret),
        } for m in boxes]
    })


@app.route('/api/admin/mailboxes', methods=['POST'])
def api_admin_mailboxes_create():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    company = session['company']
    # Límite: máximo 2 buzones por empresa
    count = MailboxConfig.query.filter_by(company=company).count()
    if count >= 2:
        return jsonify({'success': False, 'error': 'Máximo 2 buzones por empresa. Elimina uno antes de agregar otro.'}), 400

    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    host = (data.get('imap_host') or '').strip()
    user = (data.get('imap_user') or '').strip()
    pwd = (data.get('imap_password') or '').strip()
    if not name or not host or not user:
        return jsonify({'success': False, 'error': 'Nombre, host y usuario son requeridos'}), 400

    auth_type = (data.get('auth_type') or 'password').lower()
    if auth_type not in ('password', 'oauth2'):
        auth_type = 'password'

    mb = MailboxConfig(
        name=name[:100],
        company=company,
        imap_host=host[:200],
        imap_port=int(data.get('imap_port', 993)),
        imap_user=user[:200],
        imap_password=encrypt_secret(pwd or None),
        use_ssl=bool(data.get('use_ssl', True)),
        folder=(data.get('folder') or 'INBOX')[:100],
        default_priority=(data.get('default_priority') or 'medium'),
        default_category=(data.get('default_category') or 'Email')[:100],
        poll_interval_minutes=int(data.get('poll_interval_minutes', 5)),
        is_active=bool(data.get('is_active', True)),
        auth_type=auth_type,
        oauth_tenant_id=(data.get('oauth_tenant_id') or '').strip() or None,
        oauth_client_id=(data.get('oauth_client_id') or '').strip() or None,
        oauth_client_secret=encrypt_secret((data.get('oauth_client_secret') or '').strip() or None),
    )
    db.session.add(mb)
    db.session.commit()
    log_audit('mailbox_create', session['user_id'], 'mailbox', mb.id, f'Buzón "{name}" creado ({auth_type})')
    return jsonify({'success': True, 'id': mb.id, 'message': f'Buzón "{name}" creado'})


@app.route('/api/admin/mailboxes/<int:mb_id>', methods=['PUT'])
def api_admin_mailboxes_update(mb_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    mb = MailboxConfig.query.get_or_404(mb_id)
    if mb.company != session['company']:
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    data = request.get_json() or {}
    if 'name' in data: mb.name = data['name'][:100]
    if 'imap_host' in data: mb.imap_host = data['imap_host'][:200]
    if 'imap_port' in data: mb.imap_port = int(data['imap_port'])
    if 'imap_user' in data: mb.imap_user = data['imap_user'][:200]
    if 'imap_password' in data and data['imap_password']:
        mb.imap_password = encrypt_secret(data['imap_password'])
    if 'use_ssl' in data: mb.use_ssl = bool(data['use_ssl'])
    if 'folder' in data: mb.folder = data['folder'][:100]
    if 'default_priority' in data: mb.default_priority = data['default_priority']
    if 'default_category' in data: mb.default_category = data['default_category'][:100]
    if 'poll_interval_minutes' in data: mb.poll_interval_minutes = int(data['poll_interval_minutes'])
    if 'is_active' in data: mb.is_active = bool(data['is_active'])
    # OAuth fields
    if 'auth_type' in data:
        at = (data['auth_type'] or 'password').lower()
        if at in ('password', 'oauth2'):
            mb.auth_type = at
    if 'oauth_tenant_id' in data:
        mb.oauth_tenant_id = (data.get('oauth_tenant_id') or '').strip() or None
    if 'oauth_client_id' in data:
        mb.oauth_client_id = (data.get('oauth_client_id') or '').strip() or None
    if 'oauth_client_secret' in data and data.get('oauth_client_secret'):
        # Solo actualizar si se manda un valor nuevo (cifrado en BD)
        mb.oauth_client_secret = encrypt_secret(data['oauth_client_secret'])
    db.session.commit()
    log_audit('mailbox_update', session['user_id'], 'mailbox', mb.id,
              f'Buzón "{mb.name}" actualizado (auth: {mb.auth_type})')
    return jsonify({'success': True, 'message': 'Buzón actualizado'})


@app.route('/api/admin/mailboxes/<int:mb_id>', methods=['DELETE'])
def api_admin_mailboxes_delete(mb_id):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    mb = MailboxConfig.query.get_or_404(mb_id)
    if mb.company != session['company']:
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    name = mb.name
    # Eliminar histórico también
    MailboxEmail.query.filter_by(mailbox_id=mb.id).delete()
    db.session.delete(mb)
    db.session.commit()
    log_audit('mailbox_delete', session['user_id'], 'mailbox', mb_id, f'Buzón "{name}" eliminado')
    return jsonify({'success': True, 'message': f'Buzón "{name}" eliminado'})


def _diagnose_mailbox_error(mb, raw_err):
    """Analiza un error de IMAP y devuelve un diagnóstico estructurado.
    Retorna dict: {category, title, explanation, next_steps[], raw}"""
    err = str(raw_err or '').lower()
    auth_type = mb.auth_type or 'password'
    is_office365 = 'outlook.office365.com' in (mb.imap_host or '').lower() or 'office365' in (mb.imap_host or '').lower()
    is_gmail = 'gmail.com' in (mb.imap_host or '').lower()

    # ─── OAuth: errores de token ───
    if 'token request failed' in err or 'aadsts' in err:
        steps = []
        if 'aadsts7000215' in err or 'invalid client secret' in err:
            return {
                'category': 'OAuth — Client Secret inválido',
                'title': '❌ El Client Secret es incorrecto o expiró',
                'explanation': 'Microsoft rechazó el secret. Suele ser por: (a) copiaste el "Secret ID" en vez del "Value", (b) el secret expiró, o (c) hubo un error de copia/pegado.',
                'next_steps': [
                    '1. Ir a portal.azure.com → App registrations → tu app → Certificates & secrets',
                    '2. Si el secret expiró, crear uno nuevo (botón "+ New client secret")',
                    '3. ⚠ Copiá el campo "Value" (la cadena random) — NO el "Secret ID"',
                    '4. Pegalo limpio (sin espacios ni saltos de línea)',
                ],
                'raw': str(raw_err)[:300]
            }
        if 'aadsts70011' in err or 'invalid scope' in err:
            return {
                'category': 'OAuth — Permisos faltantes',
                'title': '❌ La app no tiene permisos para acceder a IMAP',
                'explanation': 'El scope solicitado no está autorizado. Falta asignar IMAP.AccessAsApp con admin consent.',
                'next_steps': [
                    '1. portal.azure.com → tu app → API permissions',
                    '2. Add permission → APIs my organization uses → buscar "Office 365 Exchange Online"',
                    '3. Application permissions → marcar IMAP.AccessAsApp',
                    '4. Click "Grant admin consent for [tenant]" (botón violeta)',
                    '5. Esperar 5-10 minutos para que se propague',
                ],
                'raw': str(raw_err)[:300]
            }
        if 'aadsts700016' in err or 'application with identifier' in err:
            return {
                'category': 'OAuth — Client ID inválido',
                'title': '❌ La aplicación no existe en este tenant',
                'explanation': 'El Client ID no corresponde a ninguna app registrada en el Tenant especificado.',
                'next_steps': [
                    '1. Verificar Client ID en portal.azure.com → App registrations',
                    '2. Confirmar que el Tenant ID corresponde a tu organización',
                    '3. Asegurar que ambos valores son del MISMO tenant',
                ],
                'raw': str(raw_err)[:300]
            }
        return {
            'category': 'OAuth — Error de token',
            'title': '❌ No se pudo obtener el token de Microsoft',
            'explanation': 'Microsoft rechazó la solicitud de token. Revisar credenciales OAuth y permisos.',
            'next_steps': [
                '1. Verificar Tenant ID, Client ID y Client Secret en Azure Portal',
                '2. Confirmar que IMAP.AccessAsApp tiene admin consent',
                '3. Verificar que el secret no expiró',
            ],
            'raw': str(raw_err)[:300]
        }

    # ─── IMAP: AUTHENTICATE failed (Basic Auth bloqueado en M365) ───
    if 'authenticate failed' in err or 'authentication failed' in err:
        if is_office365 and auth_type == 'password':
            return {
                'category': 'Microsoft 365 — Basic Auth bloqueado',
                'title': '❌ Microsoft bloquea Basic Auth para IMAP',
                'explanation': 'Microsoft 365 deshabilita la autenticación básica para IMAP desde 2022. La contraseña (incluso App Password) ya no es suficiente — necesitás OAuth 2.0.',
                'next_steps': [
                    '✓ RECOMENDADO: Cambiar tipo de auth a "🔐 OAuth 2.0 (Microsoft 365)"',
                    '✓ Seguir la guía paso a paso para registrar app en Azure AD',
                    '✓ O pedir a IT que habilite Authentication Policy con AllowBasicAuthImap',
                    '  Comando PowerShell: New-AuthenticationPolicy "AllowBasicImap" -AllowBasicAuthImap',
                    '  Set-User -Identity ' + (mb.imap_user or 'usuario') + ' -AuthenticationPolicy "AllowBasicImap"',
                ],
                'raw': str(raw_err)[:300]
            }
        if is_office365 and auth_type == 'oauth2':
            return {
                'category': 'Microsoft 365 — Sin permisos en el buzón',
                'title': '❌ La app no puede acceder a este buzón',
                'explanation': 'El token OAuth se generó correctamente, pero el Service Principal no tiene permisos sobre el buzón. Microsoft requiere asignar permisos explícitos.',
                'next_steps': [
                    'Ejecutar en PowerShell de Exchange Online:',
                    '  Connect-ExchangeOnline',
                    '  Add-MailboxPermission -Identity "' + (mb.imap_user or 'user@empresa.com') + '" -User "<CLIENT_ID>" -AccessRights FullAccess',
                    'Verificar que IMAP esté habilitado:',
                    '  Set-CASMailbox -Identity ' + (mb.imap_user or 'user@empresa.com') + ' -ImapEnabled $true',
                    'Esperar 5-10 minutos para propagación.',
                ],
                'raw': str(raw_err)[:300]
            }
        if is_gmail:
            return {
                'category': 'Gmail — Credenciales inválidas',
                'title': '❌ Gmail rechazó la autenticación',
                'explanation': 'Para Gmail necesitás un App Password si tenés 2FA activado, o habilitar "Less secure apps" (no recomendado).',
                'next_steps': [
                    '1. Ir a myaccount.google.com/apppasswords',
                    '2. Generar App Password de 16 caracteres',
                    '3. Pegar acá SIN espacios (gmail muestra con espacios, removelos)',
                    '4. Verificar que IMAP esté habilitado en gmail.com → Settings → Forwarding and POP/IMAP',
                ],
                'raw': str(raw_err)[:300]
            }
        return {
            'category': 'IMAP — Autenticación fallida',
            'title': '❌ Usuario o contraseña incorrectos',
            'explanation': 'El servidor rechazó las credenciales. Puede ser usuario equivocado, contraseña incorrecta, o el servidor bloqueó la cuenta.',
            'next_steps': [
                '1. Verificar el usuario (suele ser el email completo)',
                '2. Probar la contraseña en el webmail del proveedor',
                '3. Si la cuenta tiene 2FA, usar App Password',
                '4. Verificar que IMAP esté habilitado en la cuenta',
            ],
            'raw': str(raw_err)[:300]
        }

    # ─── Conexión: red ───
    if 'getaddrinfo' in err or 'name or service not known' in err or 'no address associated' in err:
        return {
            'category': 'Red — DNS',
            'title': '❌ No se pudo resolver el servidor',
            'explanation': f'No existe registro DNS para "{mb.imap_host}". El nombre del servidor está mal escrito o hay un problema de DNS.',
            'next_steps': [
                '1. Verificar el nombre del servidor (común: imap.gmail.com, outlook.office365.com)',
                '2. Probar en otra red por si hay bloqueo de DNS interno',
                '3. Si usás dominio propio (mail.empresa.com), confirmar con IT que el registro exista',
            ],
            'raw': str(raw_err)[:300]
        }
    if 'timed out' in err or 'timeout' in err:
        return {
            'category': 'Red — Timeout',
            'title': f'❌ Timeout al conectar a {mb.imap_host}:{mb.imap_port}',
            'explanation': 'El servidor existe pero no respondió en el tiempo esperado. Puede ser firewall corporativo bloqueando el puerto saliente.',
            'next_steps': [
                f'1. Verificar que el puerto {mb.imap_port} esté abierto de salida (firewall corporativo)',
                '2. Probar puerto 143 (sin SSL) o 993 (con SSL) según corresponda',
                '3. Confirmar con IT que la red permite IMAP saliente',
                '4. Si estás detrás de un proxy, configurar las variables HTTP_PROXY/HTTPS_PROXY',
            ],
            'raw': str(raw_err)[:300]
        }
    if 'connection refused' in err:
        return {
            'category': 'Red — Conexión rechazada',
            'title': f'❌ El servidor rechaza la conexión en {mb.imap_host}:{mb.imap_port}',
            'explanation': 'El servidor está accesible pero el puerto está cerrado o no acepta esa conexión.',
            'next_steps': [
                '1. Verificar el puerto (estándar: 993 con SSL, 143 sin SSL)',
                '2. Si pusiste 993 sin SSL o 143 con SSL, corregir el toggle SSL',
                '3. Verificar con IT/proveedor que IMAP esté habilitado en el servidor',
            ],
            'raw': str(raw_err)[:300]
        }
    if 'ssl' in err and ('handshake' in err or 'certificate' in err or 'wrong version' in err):
        return {
            'category': 'SSL — Error de protocolo',
            'title': '❌ Error de SSL/TLS',
            'explanation': 'El servidor no acepta SSL en ese puerto, o el certificado tiene problemas.',
            'next_steps': [
                f'1. Verificar si el puerto {mb.imap_port} es realmente SSL (estándar: 993 SSL, 143 sin SSL)',
                '2. Probar cambiando el toggle SSL (Sí ↔ No)',
                '3. Si el certificado del servidor es self-signed, contactar a IT para usar uno válido',
            ],
            'raw': str(raw_err)[:300]
        }

    # ─── Folder/buzón ───
    if 'no such folder' in err or 'no such mailbox' in err or 'select failed' in err:
        return {
            'category': 'Carpeta no encontrada',
            'title': f'❌ La carpeta "{mb.folder}" no existe',
            'explanation': 'La cuenta autenticó OK pero esa carpeta no existe en el servidor.',
            'next_steps': [
                '1. El valor estándar es INBOX (mayúsculas)',
                '2. Para subcarpetas usar separador del proveedor (ej: "INBOX.Soporte" en Gmail)',
                '3. Verificar en el webmail el nombre exacto de la carpeta',
            ],
            'raw': str(raw_err)[:300]
        }

    # ─── Genérico ───
    return {
        'category': 'Error desconocido',
        'title': '❌ Error al conectar al buzón',
        'explanation': 'No reconozco este error. El detalle técnico está abajo.',
        'next_steps': [
            '1. Revisar los datos del buzón (servidor, puerto, usuario)',
            '2. Probar la conexión desde el webmail del proveedor',
            '3. Si persiste, copiar el detalle técnico y consultarlo con soporte',
        ],
        'raw': str(raw_err)[:500]
    }


@app.route('/api/admin/mailboxes/<int:mb_id>/test', methods=['POST'])
def api_admin_mailboxes_test(mb_id):
    """Prueba la conexión IMAP y muestra cuántos correos no leídos hay (sin crearlos como tickets).
    Devuelve diagnóstico detallado en caso de error."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    mb = MailboxConfig.query.get_or_404(mb_id)
    if mb.company != session['company']:
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403

    auth_label = 'OAuth 2.0' if mb.auth_type == 'oauth2' else 'Password'

    # Diagnóstico pre-conexión: campos faltantes
    pre_errors = []
    if not mb.imap_host: pre_errors.append('Servidor IMAP vacío')
    if not mb.imap_user: pre_errors.append('Usuario vacío')
    if mb.auth_type == 'oauth2':
        if not mb.oauth_tenant_id: pre_errors.append('Tenant ID OAuth vacío')
        if not mb.oauth_client_id: pre_errors.append('Client ID OAuth vacío')
        if not mb.oauth_client_secret: pre_errors.append('Client Secret OAuth vacío')
    else:
        if not mb.imap_password: pre_errors.append('Contraseña vacía')
    if pre_errors:
        return jsonify({
            'success': False,
            'error': 'Configuración incompleta: ' + ', '.join(pre_errors),
            'diagnosis': {
                'category': 'Configuración incompleta',
                'title': '⚠ Faltan datos obligatorios',
                'explanation': 'Necesitás completar los siguientes campos antes de poder probar la conexión:',
                'next_steps': ['• ' + e for e in pre_errors],
                'raw': ''
            },
            'context': {
                'host': mb.imap_host or '(vacío)',
                'port': mb.imap_port,
                'ssl': bool(mb.use_ssl),
                'user': mb.imap_user or '(vacío)',
                'folder': mb.folder,
                'auth_type': auth_label
            }
        }), 400

    try:
        conn, err = _imap_connect_and_login(mb)
        if err:
            diag = _diagnose_mailbox_error(mb, err)
            return jsonify({
                'success': False,
                'error': diag['title'] + ' — ' + diag['explanation'],
                'diagnosis': diag,
                'context': {
                    'host': mb.imap_host, 'port': mb.imap_port,
                    'ssl': bool(mb.use_ssl), 'user': mb.imap_user,
                    'folder': mb.folder, 'auth_type': auth_label
                }
            }), 400
        try:
            conn.select(mb.folder)
        except Exception as folder_err:
            diag = _diagnose_mailbox_error(mb, folder_err)
            try: conn.logout()
            except Exception: pass
            return jsonify({
                'success': False,
                'error': diag['title'] + ' — ' + diag['explanation'],
                'diagnosis': diag,
                'context': {
                    'host': mb.imap_host, 'port': mb.imap_port,
                    'ssl': bool(mb.use_ssl), 'user': mb.imap_user,
                    'folder': mb.folder, 'auth_type': auth_label
                }
            }), 400

        typ, ids = conn.search(None, 'UNSEEN')
        unread = len(ids[0].split()) if typ == 'OK' and ids and ids[0] else 0
        typ2, all_ids = conn.search(None, 'ALL')
        total = len(all_ids[0].split()) if typ2 == 'OK' and all_ids and all_ids[0] else 0
        try: conn.logout()
        except Exception: pass

        return jsonify({
            'success': True,
            'message': f'✓ Conexión exitosa a {mb.imap_host}:{mb.imap_port} ({auth_label})',
            'unread': unread,
            'total': total,
            'context': {
                'host': mb.imap_host, 'port': mb.imap_port,
                'ssl': bool(mb.use_ssl), 'user': mb.imap_user,
                'folder': mb.folder, 'auth_type': auth_label
            }
        })
    except Exception as e:
        diag = _diagnose_mailbox_error(mb, e)
        return jsonify({
            'success': False,
            'error': diag['title'] + ' — ' + diag['explanation'],
            'diagnosis': diag,
            'context': {
                'host': mb.imap_host, 'port': mb.imap_port,
                'ssl': bool(mb.use_ssl), 'user': mb.imap_user,
                'folder': mb.folder, 'auth_type': auth_label
            }
        }), 400


@app.route('/api/admin/mailboxes/<int:mb_id>/sync', methods=['POST'])
def api_admin_mailboxes_sync(mb_id):
    """Sincroniza manualmente el buzón (no espera al scheduler)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    mb = MailboxConfig.query.get_or_404(mb_id)
    if mb.company != session['company']:
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    count, err = fetch_emails_from_mailbox(mb.id)
    if err:
        return jsonify({'success': False, 'error': err, 'tickets_created': count}), 500
    return jsonify({
        'success': True,
        'tickets_created': count,
        'message': f'✓ Sincronización completada: {count} tickets creados'
    })


# ===== SUBROLES (catálogo de especializaciones técnicas) =====

@app.route('/api/admin/subroles', methods=['GET'])
def api_admin_subroles_list():
    """Listar subroles disponibles (sistema global + propios de la empresa)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    company = session.get('company')
    subroles = Subrole.query.filter(
        (Subrole.company == None) | (Subrole.company == company)
    ).order_by(Subrole.is_system.desc(), Subrole.name).all()

    return jsonify({
        'success': True,
        'subroles': [{
            'id': s.id,
            'name': s.name,
            'description': s.description or '',
            'icon': s.icon or '🔧',
            'company': s.company,
            'is_system': bool(s.is_system),
            'is_active': bool(s.is_active),
            'is_global': s.company is None
        } for s in subroles]
    })


@app.route('/api/admin/subroles', methods=['POST'])
def api_admin_subroles_create():
    """Crear subrol personalizado para la empresa actual."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name or len(name) < 2:
        return jsonify({'success': False, 'error': 'El nombre es requerido (mínimo 2 caracteres)'}), 400
    company = session.get('company')
    # Validar duplicado dentro de la empresa o global con mismo nombre
    existing = Subrole.query.filter(
        Subrole.name == name,
        (Subrole.company == None) | (Subrole.company == company)
    ).first()
    if existing:
        return jsonify({'success': False, 'error': f'Ya existe un subrol "{name}".'}), 400
    s = Subrole(
        name=name[:100],
        description=(data.get('description') or '').strip()[:500] or None,
        icon=(data.get('icon') or '🔧').strip()[:10],
        company=company,
        is_system=False,
        is_active=True
    )
    db.session.add(s)
    db.session.commit()
    log_audit('create_subrole', session['user_id'], 'subrole', s.id, f'Subrol "{name}" creado')
    return jsonify({'success': True, 'id': s.id, 'message': f'Subrol "{name}" creado'})


@app.route('/api/admin/subroles/<int:subrole_id>', methods=['PUT'])
def api_admin_subroles_update(subrole_id):
    """Actualizar subrol (sistema solo permite cambiar icon/description; custom permite todo)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    s = Subrole.query.get_or_404(subrole_id)
    data = request.get_json() or {}
    if s.is_system:
        # Solo permite editar icon, description, is_active
        if 'icon' in data: s.icon = data['icon'][:10]
        if 'description' in data: s.description = (data['description'] or '').strip()[:500] or None
        if 'is_active' in data: s.is_active = bool(data['is_active'])
    else:
        if 'name' in data:
            new_name = (data['name'] or '').strip()
            if new_name and new_name != s.name:
                existing = Subrole.query.filter(Subrole.name == new_name, Subrole.id != s.id).first()
                if existing:
                    return jsonify({'success': False, 'error': 'Ya existe un subrol con ese nombre'}), 400
                s.name = new_name[:100]
        if 'icon' in data: s.icon = data['icon'][:10]
        if 'description' in data: s.description = (data['description'] or '').strip()[:500] or None
        if 'is_active' in data: s.is_active = bool(data['is_active'])
    db.session.commit()
    log_audit('update_subrole', session['user_id'], 'subrole', s.id, f'Subrol "{s.name}" actualizado')
    return jsonify({'success': True, 'message': 'Subrol actualizado'})


@app.route('/api/admin/subroles/<int:subrole_id>', methods=['DELETE'])
def api_admin_subroles_delete(subrole_id):
    """Eliminar subrol. Los is_system no se pueden borrar, solo desactivar."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    s = Subrole.query.get_or_404(subrole_id)
    if s.is_system:
        return jsonify({'success': False, 'error': 'No se puede eliminar un subrol del sistema. Desactívalo en su lugar.'}), 400
    # Verificar si está asignado a usuarios
    count = UserSubrole.query.filter_by(subrole_id=s.id).count()
    if count > 0:
        return jsonify({'success': False, 'error': f'Hay {count} usuario(s) con este subrol. Quita las asignaciones antes de eliminar.'}), 400
    name = s.name
    db.session.delete(s)
    db.session.commit()
    log_audit('delete_subrole', session['user_id'], 'subrole', subrole_id, f'Subrol "{name}" eliminado')
    return jsonify({'success': True, 'message': f'Subrol "{name}" eliminado'})


@app.route('/api/admin/subroles/export', methods=['GET'])
def api_admin_subroles_export():
    """Exporta todos los subroles (globales + de la empresa) a JSON descargable."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    company = session.get('company')
    subroles = Subrole.query.filter(
        (Subrole.company == None) | (Subrole.company == company)
    ).order_by(Subrole.is_system.desc(), Subrole.name).all()

    payload = {
        '_meta': {
            'exported_at': datetime.now().isoformat(timespec='seconds'),
            'source_company': company,
            'format_version': '1',
            'total': len(subroles),
        },
        'subroles': [{
            'name': s.name,
            'description': s.description or '',
            'icon': s.icon or '🔧',
            'is_global': s.company is None,
            'is_system': bool(s.is_system),
            'is_active': bool(s.is_active),
        } for s in subroles]
    }

    log_audit('subroles_export', session['user_id'], 'subrole', None,
              f'Exportó {len(subroles)} subroles para empresa {company}')

    from flask import Response
    import json as _json
    filename = f'subroles_{company}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    return Response(
        _json.dumps(payload, ensure_ascii=False, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/api/admin/subroles/import', methods=['POST'])
def api_admin_subroles_import():
    """Importa subroles desde un JSON (compatible con el exportado por /export).
    Modo merge: salta los que ya existen por nombre en la empresa."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    company = session.get('company')

    # Acepta JSON en body o archivo subido en form-data
    data = None
    if request.is_json:
        data = request.get_json(silent=True)
    elif 'file' in request.files:
        try:
            import json as _json
            data = _json.loads(request.files['file'].read().decode('utf-8'))
        except Exception as e:
            return jsonify({'success': False, 'error': f'JSON inválido: {e}'}), 400

    if not data or 'subroles' not in data:
        return jsonify({'success': False, 'error': 'Formato inválido: falta la clave "subroles"'}), 400

    incoming = data['subroles']
    if not isinstance(incoming, list):
        return jsonify({'success': False, 'error': '"subroles" debe ser una lista'}), 400

    created = 0
    skipped = 0
    errors = []

    # Cache de nombres existentes en la empresa
    existing_names = {
        s.name.lower() for s in Subrole.query.filter(
            (Subrole.company == None) | (Subrole.company == company)
        ).all()
    }

    for item in incoming[:200]:  # Límite de 200 por request
        try:
            name = (item.get('name') or '').strip()[:100]
            if not name or len(name) < 2:
                errors.append(f'Nombre inválido: {item}')
                continue
            if name.lower() in existing_names:
                skipped += 1
                continue
            s = Subrole(
                name=name,
                description=(item.get('description') or '').strip()[:500] or None,
                icon=(item.get('icon') or '🔧').strip()[:10],
                company=company,  # Siempre se importan como propios de la empresa
                is_system=False,  # Nunca importar como sistema
                is_active=bool(item.get('is_active', True)),
            )
            db.session.add(s)
            existing_names.add(name.lower())
            created += 1
        except Exception as e:
            errors.append(f'Error en {item.get("name","?")}: {e}')

    db.session.commit()
    log_audit('subroles_import', session['user_id'], 'subrole', None,
              f'Import subroles: {created} creados, {skipped} omitidos (ya existían), {len(errors)} errores')

    return jsonify({
        'success': True,
        'created': created,
        'skipped': skipped,
        'errors': errors[:10],  # Solo primeros 10 errores para no saturar
        'message': f'✓ {created} subroles importados. {skipped} omitidos (ya existían).'
    })


@app.route('/api/admin/users/<int:user_id>/subroles', methods=['GET'])
def api_user_subroles_get(user_id):
    """Subroles asignados a un usuario."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    user = User.query.get_or_404(user_id)
    if user.company != session.get('company') and not is_master_admin():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    assignments = UserSubrole.query.filter_by(user_id=user_id).all()
    return jsonify({
        'success': True,
        'subroles': [{
            'id': a.subrole.id,
            'name': a.subrole.name,
            'icon': a.subrole.icon
        } for a in assignments if a.subrole]
    })


@app.route('/api/admin/users/<int:user_id>/subroles', methods=['POST'])
def api_user_subroles_set(user_id):
    """Reemplaza completamente la lista de subroles asignados al usuario.
    Body: {subrole_ids: [1, 2, 3]}"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    user = User.query.get_or_404(user_id)
    if user.company != session.get('company') and not is_master_admin():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    data = request.get_json() or {}
    new_ids = data.get('subrole_ids') or []
    try:
        new_ids = [int(x) for x in new_ids]
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'subrole_ids debe ser una lista de IDs'}), 400

    # Borrar todas las asignaciones actuales y reemplazar
    UserSubrole.query.filter_by(user_id=user_id).delete()
    added = 0
    for sid in new_ids:
        s = Subrole.query.get(sid)
        if not s or not s.is_active:
            continue
        # Verificar que el subrol sea accesible (global o de la empresa del user)
        if s.company is not None and s.company != user.company:
            continue
        db.session.add(UserSubrole(user_id=user_id, subrole_id=sid))
        added += 1
    db.session.commit()
    log_audit('set_user_subroles', session['user_id'], 'user', user_id, f'{added} subroles asignados a {user.name}')
    return jsonify({'success': True, 'message': f'{added} subroles asignados', 'count': added})


@app.route('/api/admin/specialists', methods=['GET'])
def api_admin_specialists_list():
    """Listar especialistas (perfiles de IA) de la empresa"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']
    profiles = TechnicianProfile.query.filter_by(company=company).all()

    result = []
    for p in profiles:
        user = User.query.get(p.user_id)
        if not user:
            continue

        # Contar tickets activos
        active_count = Ticket.query.filter(
            Ticket.assignee_id == p.user_id,
            Ticket.status.in_(['open', 'in_progress'])
        ).count()

        skills_list = [s.strip() for s in (p.skills or '').split(',') if s.strip()]

        result.append({
            'id': p.id,
            'user_id': user.id,
            'user_name': user.name,
            'username': user.username,
            'skills': skills_list,
            'max_tickets': p.max_tickets,
            'active_tickets': active_count,
            'is_available': bool(p.is_available),
            'tickets_resolved_total': p.tickets_resolved_total or 0
        })

    result.sort(key=lambda x: x['user_name'])
    return jsonify({'success': True, 'specialists': result})


@app.route('/api/admin/specialists', methods=['POST'])
def api_admin_specialists_create():
    """Crear o actualizar especialista"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    try:
        data = request.get_json()
        user_id = data.get('user_id')
        skills_list = data.get('skills', [])
        max_tickets = int(data.get('max_tickets', 10))

        if not user_id:
            return jsonify({'success': False, 'error': 'user_id requerido'}), 400

        user = User.query.get(user_id)
        if not user or user.company != session['company']:
            return jsonify({'success': False, 'error': 'Usuario no encontrado en tu empresa'}), 404

        if not skills_list:
            return jsonify({'success': False, 'error': 'Agrega al menos una habilidad'}), 400

        skills_csv = ','.join([s.strip() for s in skills_list if s.strip()])

        # Buscar perfil existente (uno por usuario)
        profile = TechnicianProfile.query.filter_by(user_id=user_id).first()
        if profile:
            profile.skills = skills_csv
            profile.max_tickets = max_tickets
            profile.company = session['company']
            profile.updated_at = datetime.now()
            action = 'update_specialist'
            msg = f'Especialista {user.name} actualizado'
        else:
            profile = TechnicianProfile(
                user_id=user_id,
                company=session['company'],
                skills=skills_csv,
                max_tickets=max_tickets,
                is_available=True
            )
            db.session.add(profile)
            action = 'create_specialist'
            msg = f'Especialista {user.name} creado'

        db.session.commit()
        log_audit(action, session['user_id'], 'specialist', profile.id,
                  f'{msg} - Habilidades: {skills_csv}')

        return jsonify({
            'success': True,
            'message': msg,
            'id': profile.id
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/specialists/<int:specialist_id>', methods=['DELETE'])
def api_admin_specialists_delete(specialist_id):
    """Eliminar perfil de especialista"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    profile = TechnicianProfile.query.get(specialist_id)
    if not profile or profile.company != session['company']:
        return jsonify({'success': False, 'error': 'No encontrado'}), 404

    user = User.query.get(profile.user_id)
    name = user.name if user else 'Desconocido'
    db.session.delete(profile)
    db.session.commit()

    log_audit('delete_specialist', session['user_id'], 'specialist', specialist_id,
              f'Especialista eliminado: {name}')

    return jsonify({'success': True, 'message': 'Especialista eliminado'})


@app.route('/api/technicians', methods=['GET'])
def api_technicians_list():
    """Listar tecnicos y admins de la empresa (para reasignacion).
    Si se pasa ?ticket_id=X, devuelve técnicos de la empresa del ticket
    (útil cuando el admin master ve tickets de otras empresas)."""
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    target_company = session['company']
    ticket_id = request.args.get('ticket_id', type=int)
    if ticket_id:
        t = Ticket.query.get(ticket_id)
        if t and t.company in admin_companies_scope():
            target_company = t.company

    users = User.query.filter(
        User.company == target_company,
        User.role.in_(['technician', 'admin']),
        User.is_active == True
    ).order_by(User.name).all()

    return jsonify({
        'success': True,
        'company': target_company,
        'technicians': [{
            'id': u.id,
            'username': u.username,
            'name': u.name,
            'role': u.role,
            'email': u.email or ''
        } for u in users]
    })


@app.route('/api/tickets/list', methods=['GET'])
def api_tickets_list():
    """Lista tickets de la empresa en JSON con paginación opcional.
    Params: ?page=N&per_page=M (default page=1, per_page=500). Sin params, devuelve los 500 más recientes."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401

    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = max(1, min(int(request.args.get('per_page', 500)), 2000))  # tope 2000
    except (ValueError, TypeError):
        page, per_page = 1, 500

    scope = admin_companies_scope()
    # Excluir tickets internos del sistema (DMs y chats grupales)
    base_query = Ticket.query.filter(
        Ticket.company.in_(scope),
        ~Ticket.ticket_number.like('DM-%'),
        ~Ticket.ticket_number.like('CHAT-%'),
    ).order_by(Ticket.created_at.desc())
    total = base_query.count()
    tickets = base_query.offset((page - 1) * per_page).limit(per_page).all()

    result = []
    for t in tickets:
        # URL según rol
        href = f'/admin/ticket/{t.id}' if session.get('role') == 'admin' else f'/technician/ticket/{t.id}'
        result.append({
            'id': t.id,
            'ticket_number': t.ticket_number,
            'title': t.title,
            'status': t.status or '',
            'priority': t.priority or '',
            'assignee': t.assignee.name if t.assignee else 'Sin asignar',
            'creator': t.creator.name if hasattr(t, 'creator') and t.creator else '',
            'category': t.category or '',
            'company': t.company,
            'created_at': t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            'sla_deadline': t.sla_deadline.strftime('%Y-%m-%d %H:%M') if t.sla_deadline else '',
            'sla_expired': (t.sla_deadline and t.sla_deadline < datetime.now() and t.status not in ['resolved', 'closed']) if t.sla_deadline else False,
            'href': href
        })

    return jsonify({
        'success': True,
        'tickets': result,
        'total': total,                              # total de tickets accesibles
        'returned': len(result),                     # cuántos vienen en esta página
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })


@app.route('/api/reports/metrics', methods=['GET'])
def api_reports_metrics():
    """Generar reporte gerencial de métricas en Excel o PDF, opcionalmente filtrado por rango de fechas."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    format_type = request.args.get('format', 'excel').lower()
    company = session['company']
    company_obj = Company.query.filter_by(code=company).first()
    company_name = company_obj.name if company_obj else company

    # === FILTROS DE FECHA ===
    # Formato: YYYY-MM-DD; opcional ?month=YYYY-MM
    date_from_raw = request.args.get('date_from', '').strip()
    date_to_raw = request.args.get('date_to', '').strip()
    month_raw = request.args.get('month', '').strip()  # ej "2026-06"

    date_from = None
    date_to = None
    period_label = 'Histórico completo'

    try:
        if month_raw:
            yr, mo = month_raw.split('-')
            yr, mo = int(yr), int(mo)
            date_from = datetime(yr, mo, 1)
            if mo == 12:
                date_to = datetime(yr + 1, 1, 1)
            else:
                date_to = datetime(yr, mo + 1, 1)
            period_label = f"Mes: {date_from.strftime('%B %Y')}"
        else:
            if date_from_raw:
                date_from = datetime.strptime(date_from_raw, '%Y-%m-%d')
            if date_to_raw:
                # Incluir el día completo
                date_to = datetime.strptime(date_to_raw, '%Y-%m-%d') + timedelta(days=1)
            if date_from or date_to:
                df = date_from.strftime('%Y-%m-%d') if date_from else '—'
                dt = (date_to - timedelta(days=1)).strftime('%Y-%m-%d') if date_to else '—'
                period_label = f"Del {df} al {dt}"
    except (ValueError, TypeError) as e:
        return jsonify({'success': False, 'error': f'Fechas inválidas: {e}'}), 400

    # === RECOPILAR DATOS ===
    query = Ticket.query.filter_by(company=company)
    if date_from:
        query = query.filter(Ticket.created_at >= date_from)
    if date_to:
        query = query.filter(Ticket.created_at < date_to)
    all_tickets = query.all()
    total = len(all_tickets)
    by_status = {
        'open': len([t for t in all_tickets if t.status == 'open']),
        'in_progress': len([t for t in all_tickets if t.status == 'in_progress']),
        'resolved': len([t for t in all_tickets if t.status == 'resolved']),
        'closed': len([t for t in all_tickets if t.status == 'closed'])
    }
    by_priority = {
        'critical': len([t for t in all_tickets if t.priority == 'critical']),
        'high': len([t for t in all_tickets if t.priority == 'high']),
        'medium': len([t for t in all_tickets if t.priority == 'medium']),
        'low': len([t for t in all_tickets if t.priority == 'low'])
    }

    resolved_today = len([t for t in all_tickets
                          if t.status == 'resolved' and t.resolved_at
                          and t.resolved_at.date() == datetime.now().date()])

    sla_ok = len([t for t in all_tickets if not t.sla_deadline or t.sla_deadline > datetime.now()])
    sla_pct = round((sla_ok / total * 100), 1) if total > 0 else 100

    worked = [t for t in all_tickets if t.time_worked_seconds]
    avg_resolution_h = round(sum([t.time_worked_seconds or 0 for t in worked]) / len(worked) / 3600, 2) if worked else 0

    rated = [t for t in all_tickets if t.rating]
    avg_rating = round(sum([t.rating for t in rated]) / len(rated), 1) if rated else 0

    # Top técnicos
    technicians = User.query.filter_by(company=company, role='technician').all()
    tech_stats = []
    for tech in technicians:
        my_tickets = [t for t in all_tickets if t.assignee_id == tech.id]
        resolved = [t for t in my_tickets if t.status == 'resolved']
        tech_stats.append({
            'name': tech.name,
            'username': tech.username,
            'assigned': len(my_tickets),
            'resolved': len(resolved),
            'rate': round(len(resolved) / len(my_tickets) * 100, 1) if my_tickets else 0
        })
    tech_stats.sort(key=lambda x: x['resolved'], reverse=True)

    # === GENERAR EXCEL ===
    if format_type == 'excel':
        from io import BytesIO
        wb = openpyxl.Workbook()

        # Estilos
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        subheader_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
        thin_border = Border(left=Side(style='thin', color='D1D5DB'),
                            right=Side(style='thin', color='D1D5DB'),
                            top=Side(style='thin', color='D1D5DB'),
                            bottom=Side(style='thin', color='D1D5DB'))

        # ==== Hoja 1: RESUMEN EJECUTIVO ====
        ws = wb.active
        ws.title = 'Resumen Ejecutivo'

        ws['A1'] = f'📊 REPORTE GERENCIAL - DeskEli'
        ws['A1'].font = Font(bold=True, size=18, color='1E40AF')
        ws.merge_cells('A1:D1')

        ws['A2'] = f'Empresa: {company_name}'
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:D2')

        ws['A3'] = f'Período: {period_label}'
        ws['A3'].font = Font(bold=True, size=11, color='065F46')
        ws.merge_cells('A3:D3')

        ws['A4'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        ws['A4'].font = Font(italic=True, color='6B7280')
        ws.merge_cells('A4:D4')

        # KPIs principales
        ws['A6'] = 'INDICADOR'
        ws['B6'] = 'VALOR'
        ws['A6'].font = header_font
        ws['B6'].font = header_font
        ws['A6'].fill = header_fill
        ws['B6'].fill = header_fill

        kpis = [
            ('Tickets Totales', total),
            ('Resueltos Hoy', resolved_today),
            ('En Progreso', by_status['in_progress']),
            ('Abiertos', by_status['open']),
            ('Cumplimiento SLA (%)', f'{sla_pct}%'),
            ('Tiempo Promedio Resolución (h)', avg_resolution_h),
            ('Calificación Promedio Cliente', f'{avg_rating}/5'),
            ('Tickets Críticos Activos', by_priority['critical']),
        ]
        for i, (label, value) in enumerate(kpis, start=7):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].border = thin_border
            ws[f'B{i}'].border = thin_border
            if i % 2 == 0:
                ws[f'A{i}'].fill = subheader_fill
                ws[f'B{i}'].fill = subheader_fill

        # Distribución por estado
        ws[f'A{len(kpis)+8}'] = 'DISTRIBUCIÓN POR ESTADO'
        ws[f'A{len(kpis)+8}'].font = Font(bold=True, size=14, color='1E40AF')
        ws.merge_cells(f'A{len(kpis)+8}:B{len(kpis)+8}')

        row = len(kpis) + 9
        ws[f'A{row}'] = 'Estado'
        ws[f'B{row}'] = 'Cantidad'
        ws[f'A{row}'].font = header_font; ws[f'B{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill; ws[f'B{row}'].fill = header_fill

        status_labels = {'open': 'Abierto', 'in_progress': 'En Progreso', 'resolved': 'Resuelto', 'closed': 'Cerrado'}
        for k, v in by_status.items():
            row += 1
            ws[f'A{row}'] = status_labels[k]
            ws[f'B{row}'] = v
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border

        # Distribución por prioridad
        row += 3
        ws[f'A{row}'] = 'DISTRIBUCIÓN POR PRIORIDAD'
        ws[f'A{row}'].font = Font(bold=True, size=14, color='1E40AF')
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = 'Prioridad'
        ws[f'B{row}'] = 'Cantidad'
        ws[f'A{row}'].font = header_font; ws[f'B{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill; ws[f'B{row}'].fill = header_fill

        priority_labels = {'critical': '🔴 Crítica', 'high': '🟠 Alta', 'medium': '🟡 Media', 'low': '🟢 Baja'}
        for k, v in by_priority.items():
            row += 1
            ws[f'A{row}'] = priority_labels[k]
            ws[f'B{row}'] = v
            ws[f'A{row}'].border = thin_border
            ws[f'B{row}'].border = thin_border

        # Ancho de columnas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

        # ==== Hoja 2: RANKING DE TÉCNICOS ====
        ws2 = wb.create_sheet('Top Técnicos')
        ws2['A1'] = '🏆 RANKING DE TÉCNICOS'
        ws2['A1'].font = Font(bold=True, size=18, color='1E40AF')
        ws2.merge_cells('A1:E1')

        headers = ['#', 'Técnico', 'Usuario', 'Tickets Asignados', 'Resueltos', 'Tasa Resolución']
        for col, h in enumerate(headers, start=1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for i, tech in enumerate(tech_stats[:10], start=4):
            ws2.cell(row=i, column=1, value=i-3).border = thin_border
            ws2.cell(row=i, column=2, value=tech['name']).border = thin_border
            ws2.cell(row=i, column=3, value=tech['username']).border = thin_border
            ws2.cell(row=i, column=4, value=tech['assigned']).border = thin_border
            ws2.cell(row=i, column=5, value=tech['resolved']).border = thin_border
            ws2.cell(row=i, column=6, value=f"{tech['rate']}%").border = thin_border

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws2.column_dimensions[col_letter].width = 22

        # ==== Hoja 3: TICKETS DETALLADOS ====
        ws3 = wb.create_sheet('Tickets Detallados')
        ws3['A1'] = '📋 LISTADO COMPLETO DE TICKETS'
        ws3['A1'].font = Font(bold=True, size=16, color='1E40AF')
        ws3.merge_cells('A1:G1')

        headers = ['Número', 'Título', 'Estado', 'Prioridad', 'Asignado', 'Creado', 'Resuelto']
        for col, h in enumerate(headers, start=1):
            cell = ws3.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for i, t in enumerate(all_tickets[:200], start=4):
            ws3.cell(row=i, column=1, value=t.ticket_number)
            ws3.cell(row=i, column=2, value=t.title)
            ws3.cell(row=i, column=3, value=status_labels.get(t.status, t.status))
            ws3.cell(row=i, column=4, value=priority_labels.get(t.priority, t.priority))
            ws3.cell(row=i, column=5, value=t.assignee.name if t.assignee else 'Sin asignar')
            ws3.cell(row=i, column=6, value=t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '')
            ws3.cell(row=i, column=7, value=t.resolved_at.strftime('%Y-%m-%d %H:%M') if t.resolved_at else '')

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws3.column_dimensions[col_letter].width = 18

        # Guardar en buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        log_audit('download_report', session['user_id'], 'report', None, f'Reporte Excel descargado')

        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_metricas_{company}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    # === GENERAR PDF (con gráficas) ===
    elif format_type == 'pdf':
        from io import BytesIO
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # ====== GENERAR GRÁFICAS CON MATPLOTLIB ======
        import matplotlib
        matplotlib.use('Agg')  # backend sin GUI
        import matplotlib.pyplot as plt
        from matplotlib.patches import Wedge

        def fig_to_buffer(fig):
            buf = BytesIO()
            fig.savefig(buf, format='PNG', bbox_inches='tight', dpi=120)
            buf.seek(0)
            plt.close(fig)
            return buf

        chart_buffers = {}

        # 1. Gráfica DONUT - Distribución por Estado
        if total > 0:
            fig, ax = plt.subplots(figsize=(7, 4))
            labels = ['Abierto', 'En Progreso', 'Resuelto', 'Cerrado']
            sizes = [by_status['open'], by_status['in_progress'], by_status['resolved'], by_status['closed']]
            colors_s = ['#3b82f6', '#ea580c', '#16a34a', '#6b7280']
            # Filtrar zeros
            filtered = [(l,s,c) for l,s,c in zip(labels, sizes, colors_s) if s > 0]
            if filtered:
                labels_f, sizes_f, colors_f = zip(*filtered)
                wedges, texts, autotexts = ax.pie(sizes_f, labels=labels_f, colors=colors_f,
                                                  autopct='%1.1f%%', startangle=90,
                                                  wedgeprops=dict(width=0.4, edgecolor='white', linewidth=2))
                for t in texts: t.set_fontsize(10)
                for t in autotexts: t.set_color('white'); t.set_fontweight('bold')
                ax.set_title('Distribución por Estado', fontsize=14, fontweight='bold', pad=15)
                chart_buffers['status'] = fig_to_buffer(fig)
            else:
                plt.close(fig)

        # 2. Gráfica BARRAS - Por Prioridad
        fig, ax = plt.subplots(figsize=(8, 4))
        labels = ['Crítica', 'Alta', 'Media', 'Baja']
        values = [by_priority['critical'], by_priority['high'], by_priority['medium'], by_priority['low']]
        colors_p = ['#dc2626', '#ea580c', '#f59e0b', '#16a34a']
        bars = ax.bar(labels, values, color=colors_p, edgecolor='white', linewidth=2)
        ax.set_title('Tickets por Prioridad', fontsize=14, fontweight='bold', pad=15)
        ax.set_ylabel('Cantidad', fontsize=11)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(axis='y', alpha=0.3)
        for bar, val in zip(bars, values):
            if val > 0:
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                       str(val), ha='center', va='bottom', fontweight='bold')
        chart_buffers['priority'] = fig_to_buffer(fig)

        # 3. Gráfica BARRAS HORIZONTALES - Top Técnicos
        if tech_stats:
            fig, ax = plt.subplots(figsize=(8, 4))
            top5 = tech_stats[:5]
            names = [t['name'][:25] for t in top5]
            resolved_vals = [t['resolved'] for t in top5]
            colors_t = ['#2563eb', '#7c3aed', '#16a34a', '#ea580c', '#dc2626'][:len(top5)]
            bars = ax.barh(names, resolved_vals, color=colors_t, edgecolor='white', linewidth=2)
            ax.set_title('Top 5 Técnicos (Tickets Resueltos)', fontsize=14, fontweight='bold', pad=15)
            ax.set_xlabel('Tickets Resueltos', fontsize=11)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.grid(axis='x', alpha=0.3)
            for bar, val in zip(bars, resolved_vals):
                if val > 0:
                    ax.text(val + 0.3, bar.get_y() + bar.get_height()/2,
                           str(val), va='center', fontweight='bold')
            ax.invert_yaxis()
            chart_buffers['tech'] = fig_to_buffer(fig)

        # 4. Gráfica GAUGE - SLA Compliance
        fig, ax = plt.subplots(figsize=(7, 3.5), subplot_kw=dict(aspect='equal'))
        # Semicírculo de fondo (gris)
        bg_wedge = Wedge((0.5, 0), 0.4, 0, 180, width=0.15, facecolor='#e5e7eb')
        ax.add_patch(bg_wedge)
        # Semicírculo coloreado según SLA
        sla_color = '#16a34a' if sla_pct >= 80 else ('#f59e0b' if sla_pct >= 50 else '#dc2626')
        sla_angle = 180 - (sla_pct / 100 * 180)
        fg_wedge = Wedge((0.5, 0), 0.4, sla_angle, 180, width=0.15, facecolor=sla_color)
        ax.add_patch(fg_wedge)
        # Texto central
        ax.text(0.5, 0.15, f'{sla_pct}%', ha='center', va='center',
                fontsize=36, fontweight='bold', color=sla_color)
        ax.text(0.5, 0.02, 'Cumplimiento SLA', ha='center', va='center',
                fontsize=11, color='#6b7280')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 0.6)
        ax.axis('off')
        ax.set_title('Cumplimiento de SLA', fontsize=14, fontweight='bold', pad=15)
        chart_buffers['sla'] = fig_to_buffer(fig)

        # 5. Gráfica LÍNEAS - Tendencia (últimos 7 días simulada)
        fig, ax = plt.subplots(figsize=(8, 4))
        days_data = []
        days_labels = []
        for i in range(6, -1, -1):
            day = (datetime.now() - timedelta(days=i)).date()
            created = len([t for t in all_tickets if t.created_at and t.created_at.date() == day])
            resolved = len([t for t in all_tickets if t.resolved_at and t.resolved_at.date() == day])
            days_data.append((created, resolved))
            days_labels.append(day.strftime('%a %d'))

        creados = [d[0] for d in days_data]
        resueltos = [d[1] for d in days_data]
        ax.plot(days_labels, creados, marker='o', linewidth=2, label='Creados', color='#2563eb')
        ax.plot(days_labels, resueltos, marker='s', linewidth=2, label='Resueltos', color='#16a34a')
        ax.fill_between(days_labels, creados, alpha=0.1, color='#2563eb')
        ax.fill_between(days_labels, resueltos, alpha=0.1, color='#16a34a')
        ax.set_title('Tendencia - Últimos 7 días', fontsize=14, fontweight='bold', pad=15)
        ax.set_ylabel('Cantidad', fontsize=11)
        ax.legend(loc='upper left', framealpha=0.95)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(alpha=0.3)
        chart_buffers['trend'] = fig_to_buffer(fig)

        # ====== CONSTRUIR PDF ======
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()

        # Estilos custom
        title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                     fontSize=22, textColor=colors.HexColor('#1e40af'),
                                     spaceAfter=10, alignment=TA_CENTER, fontName='Helvetica-Bold')
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'],
                                        fontSize=12, textColor=colors.HexColor('#6b7280'),
                                        spaceAfter=20, alignment=TA_CENTER)
        section_style = ParagraphStyle('Section', parent=styles['Heading2'],
                                       fontSize=16, textColor=colors.HexColor('#1e40af'),
                                       spaceAfter=12, spaceBefore=20, fontName='Helvetica-Bold')
        normal = styles['Normal']

        elements = []

        # ===== CALCULAR INSIGHTS Y RECOMENDACIONES =====
        # Periodo: comparar últimos 30 días vs 30 días previos
        now_dt = datetime.now()
        period_30 = now_dt - timedelta(days=30)
        period_60 = now_dt - timedelta(days=60)
        current_period = [t for t in all_tickets if t.created_at and t.created_at >= period_30]
        prev_period = [t for t in all_tickets if t.created_at and period_60 <= t.created_at < period_30]

        # Variación
        var_total = ((len(current_period) - len(prev_period)) / len(prev_period) * 100) if prev_period else 0
        var_resolved_cur = len([t for t in current_period if t.status == 'resolved'])
        var_resolved_prev = len([t for t in prev_period if t.status == 'resolved'])
        var_resolved = ((var_resolved_cur - var_resolved_prev) / var_resolved_prev * 100) if var_resolved_prev else 0

        # SLA breach count
        sla_breached = len([t for t in all_tickets if t.sla_deadline and t.sla_deadline < datetime.now() and t.status != 'resolved'])

        # Health Score (0-100)
        health_score = round((sla_pct * 0.5) + (min(avg_rating * 20, 100) * 0.3) +
                            ((100 - min(by_priority['critical'] / max(total, 1) * 100, 100)) * 0.2), 0)
        health_status = 'EXCELENTE' if health_score >= 85 else ('BUENO' if health_score >= 70 else ('REGULAR' if health_score >= 50 else 'CRÍTICO'))
        health_color = '#16a34a' if health_score >= 85 else ('#84cc16' if health_score >= 70 else ('#f59e0b' if health_score >= 50 else '#dc2626'))

        # Insights automáticos
        insights = []
        if sla_pct >= 90:
            insights.append(('🟢 EXCELENTE', f'Cumplimiento SLA de {sla_pct}% supera el objetivo del 85%.', '#16a34a'))
        elif sla_pct >= 80:
            insights.append(('🟡 ACEPTABLE', f'Cumplimiento SLA de {sla_pct}% está en rango aceptable pero debe mejorar.', '#f59e0b'))
        else:
            insights.append(('🔴 ATENCIÓN', f'Cumplimiento SLA de {sla_pct}% por debajo del objetivo. Requiere acción inmediata.', '#dc2626'))

        if by_priority['critical'] > 5:
            insights.append(('🔴 ALERTA', f'{by_priority["critical"]} tickets críticos activos. Riesgo operacional alto.', '#dc2626'))
        elif by_priority['critical'] > 0:
            insights.append(('🟡 MONITOREAR', f'{by_priority["critical"]} ticket(s) crítico(s) requieren atención.', '#f59e0b'))

        if var_total > 20:
            insights.append(('📈 TENDENCIA', f'Aumento del {round(var_total,1)}% en tickets vs período anterior. Investigar causa raíz.', '#7c3aed'))
        elif var_total < -10:
            insights.append(('📉 MEJORA', f'Reducción del {abs(round(var_total,1))}% en tickets vs período anterior. Buena tendencia.', '#16a34a'))

        if avg_rating >= 4.5:
            insights.append(('⭐ SATISFACCIÓN', f'Calificación de {avg_rating}/5 indica alta satisfacción del cliente.', '#16a34a'))
        elif avg_rating < 3.5 and avg_rating > 0:
            insights.append(('⚠️ SATISFACCIÓN', f'Calificación de {avg_rating}/5 está baja. Revisar calidad del servicio.', '#dc2626'))

        if sla_breached > 0:
            insights.append(('🚨 SLA VENCIDOS', f'{sla_breached} ticket(s) con SLA vencido sin resolver.', '#dc2626'))

        # Recomendaciones
        recommendations = []
        if by_status['open'] > total * 0.3 and total > 10:
            recommendations.append('Asignar más rápido los tickets abiertos. Más del 30% sin asignar indica cuello de botella.')
        if avg_resolution_h > 24:
            recommendations.append(f'Tiempo de resolución promedio de {avg_resolution_h}h es alto. Identificar bottlenecks.')
        if sla_pct < 85:
            recommendations.append('Reforzar capacitación del equipo o revisar umbrales de SLA por prioridad.')
        if by_priority['critical'] > 0:
            recommendations.append('Atender tickets críticos prioritariamente. Considerar refuerzo de personal.')
        if not tech_stats or tech_stats[0]['resolved'] < 5:
            recommendations.append('Productividad de técnicos baja. Revisar carga de trabajo y capacitación.')
        if len(recommendations) == 0:
            recommendations.append('Mantener el ritmo actual. Continuar con mejores prácticas establecidas.')

        # ===== PORTADA EJECUTIVA =====
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph('REPORTE EJECUTIVO', ParagraphStyle('CoverTitle',
            fontSize=14, textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER, fontName='Helvetica', letterSpacing=4)))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph('Gestión de Incidencias TI', ParagraphStyle('CoverSubtitle',
            fontSize=28, textColor=colors.HexColor('#111827'),
            alignment=TA_CENTER, fontName='Helvetica-Bold')))
        elements.append(Spacer(1, 0.2*cm))
        elements.append(Paragraph(company_name, ParagraphStyle('CoverCompany',
            fontSize=20, textColor=colors.HexColor('#2563eb'),
            alignment=TA_CENTER, fontName='Helvetica')))
        elements.append(Spacer(1, 1.5*cm))

        # Health Score Card (gran indicador visual)
        health_table = Table([
            ['HEALTH SCORE GENERAL'],
            [f'{int(health_score)}/100'],
            [health_status]
        ], colWidths=[17*cm])
        health_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor(health_color)),
            ('BACKGROUND', (0,1), (-1,2), colors.HexColor('#f9fafb')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('TEXTCOLOR', (0,1), (-1,1), colors.HexColor(health_color)),
            ('TEXTCOLOR', (0,2), (-1,2), colors.HexColor('#374151')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTNAME', (0,2), (-1,2), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('FONTSIZE', (0,1), (-1,1), 56),
            ('FONTSIZE', (0,2), (-1,2), 18),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 14),
            ('LINEBELOW', (0,0), (-1,0), 1, colors.HexColor('#d1d5db')),
        ]))
        elements.append(health_table)
        elements.append(Spacer(1, 1.5*cm))

        # Metadata del reporte
        meta_data = [
            ['EMPRESA', company_name],
            ['PERÍODO ANALIZADO', period_label],
            ['FECHA DE EMISIÓN', now_dt.strftime("%Y-%m-%d %H:%M")],
            ['GENERADO POR', session.get("name", "Admin")],
            ['CONFIDENCIALIDAD', 'USO INTERNO · Solo Dirección y Gerencia TI'],
        ]
        meta_table = Table(meta_data, colWidths=[5*cm, 12*cm])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0,0), (0,-1), colors.white),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (0,-1), 9),
            ('FONTSIZE', (1,0), (1,-1), 10),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#d1d5db')),
            ('PADDING', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(meta_table)

        # ===== PÁGINA 2: RESUMEN EJECUTIVO Y INSIGHTS =====
        elements.append(PageBreak())
        elements.append(Paragraph('RESUMEN EJECUTIVO', section_style))
        elements.append(Spacer(1, 0.2*cm))

        exec_summary = (f'<b>El sistema DeskEli de {company_name}</b> ha gestionado un total de '
                       f'<b>{total} tickets</b> durante su operación, con un '
                       f'<b style="color:{health_color};">Health Score de {int(health_score)}/100</b> ({health_status}). '
                       f'En los últimos 30 días se crearon <b>{len(current_period)} tickets</b> '
                       f'({"+"if var_total >= 0 else ""}{round(var_total,1)}% vs período anterior). ')

        if sla_pct >= 85:
            exec_summary += f'El <b>cumplimiento de SLA del {sla_pct}%</b> indica un servicio dentro de objetivos. '
        else:
            exec_summary += f'El <b>cumplimiento de SLA del {sla_pct}%</b> está por debajo del objetivo y requiere atención. '

        if avg_rating > 0:
            exec_summary += f'La satisfacción del cliente promedio es <b>{avg_rating}/5</b>.'

        elements.append(Paragraph(exec_summary, ParagraphStyle('Summary', parent=normal,
            fontSize=11, leading=18, textColor=colors.HexColor('#374151'), alignment=4, spaceAfter=20)))

        # INSIGHTS CLAVE
        elements.append(Paragraph('🔍 HALLAZGOS CLAVE', section_style))
        if insights:
            for badge, text, color in insights[:6]:
                insight_para = Paragraph(
                    f'<font color="{color}"><b>{badge}</b></font> &nbsp; {text}',
                    ParagraphStyle('Insight', parent=normal,
                        fontSize=10, leading=15, spaceAfter=8,
                        leftIndent=10, borderColor=colors.HexColor(color), borderWidth=0,
                        backColor=colors.HexColor('#fafafa'))
                )
                elements.append(insight_para)
        else:
            elements.append(Paragraph('No hay alertas relevantes en este período.', normal))
        elements.append(Spacer(1, 0.4*cm))

        # RECOMENDACIONES ESTRATÉGICAS
        elements.append(Paragraph('💡 RECOMENDACIONES ESTRATÉGICAS', section_style))
        for i, rec in enumerate(recommendations[:5], 1):
            elements.append(Paragraph(f'<b>{i}.</b> {rec}',
                ParagraphStyle('Rec', parent=normal, fontSize=10, leading=15, spaceAfter=6, leftIndent=10)))
        elements.append(Spacer(1, 0.5*cm))

        # ===== PÁGINA 3: KPIs DETALLADOS =====
        elements.append(PageBreak())
        elements.append(Paragraph('📌 INDICADORES OPERATIVOS DETALLADOS', section_style))

        def trend_arrow(value, good_up=True):
            if abs(value) < 1:
                return f'<font color="#6b7280">→ {round(value,1)}%</font>'
            if (value > 0 and good_up) or (value < 0 and not good_up):
                return f'<font color="#16a34a">▲ +{round(value,1)}%</font>'
            else:
                return f'<font color="#dc2626">▼ {round(value,1)}%</font>'

        kpi_data = [
            ['INDICADOR', 'VALOR', 'VARIACIÓN', 'EVALUACIÓN'],
            ['Volumen Total Tickets', str(total),
             Paragraph(trend_arrow(var_total, good_up=False), normal),
             'Actividad' if var_total > 0 else 'Estabilidad'],
            ['Tickets Período Actual (30d)', str(len(current_period)),
             Paragraph(trend_arrow(var_total, good_up=False), normal),
             'vs 30d anteriores'],
            ['Resueltos Hoy', str(resolved_today),
             Paragraph('▶ Diario', normal), 'Productividad diaria'],
            ['Tickets en Progreso', str(by_status['in_progress']),
             Paragraph(f'<font color="#ea580c">●</font> Activos', normal), 'Carga actual'],
            ['Tickets Abiertos Sin Asignar', str(by_status['open']),
             Paragraph(f'<font color="#3b82f6">●</font> Por asignar', normal), 'Pendientes'],
            ['Cumplimiento SLA', f'{sla_pct}%',
             Paragraph(f'<font color="{"#16a34a" if sla_pct>=85 else "#dc2626"}"><b>{"✓" if sla_pct>=85 else "✗"}</b></font>', normal),
             'Objetivo: ≥ 85%'],
            ['SLA Vencidos Sin Resolver', str(sla_breached),
             Paragraph(f'<font color="{"#dc2626" if sla_breached>0 else "#16a34a"}">●</font> {"Riesgo" if sla_breached>0 else "OK"}', normal),
             'Objetivo: 0'],
            ['Tiempo Promedio Resolución', f'{avg_resolution_h} h',
             Paragraph(f'<font color="{"#16a34a" if avg_resolution_h<8 else "#f59e0b"}">{"Bueno" if avg_resolution_h<8 else "Mejorable"}</font>', normal),
             'Objetivo: < 8h'],
            ['Calificación Cliente', f'{avg_rating}/5',
             Paragraph(f'<font color="{"#16a34a" if avg_rating>=4 else "#dc2626"}">{"★" * int(avg_rating)}</font>', normal),
             'Objetivo: ≥ 4.0'],
            ['Tickets Críticos Activos', str(by_priority['critical']),
             Paragraph(f'<font color="{"#dc2626" if by_priority["critical"]>0 else "#16a34a"}"><b>{"⚠ Alerta" if by_priority["critical"]>0 else "✓ Limpio"}</b></font>', normal),
             'Objetivo: 0'],
        ]
        kpi_table = Table(kpi_data, colWidths=[7*cm, 3*cm, 3.5*cm, 3.5*cm])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9fafb')]),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 7),
        ]))
        elements.append(kpi_table)

        # DISTRIBUCIÓN POR ESTADO
        elements.append(Paragraph('📊 Distribución por Estado', section_style))
        status_data = [
            ['Estado', 'Cantidad', 'Porcentaje'],
            ['Abierto', str(by_status['open']), f"{round(by_status['open']/total*100,1) if total else 0}%"],
            ['En Progreso', str(by_status['in_progress']), f"{round(by_status['in_progress']/total*100,1) if total else 0}%"],
            ['Resuelto', str(by_status['resolved']), f"{round(by_status['resolved']/total*100,1) if total else 0}%"],
            ['Cerrado', str(by_status['closed']), f"{round(by_status['closed']/total*100,1) if total else 0}%"],
        ]
        status_table = Table(status_data, colWidths=[6*cm, 4*cm, 4*cm])
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(status_table)

        # DISTRIBUCIÓN POR PRIORIDAD
        elements.append(Paragraph('🎯 Distribución por Prioridad', section_style))
        prio_data = [
            ['Prioridad', 'Cantidad', 'Porcentaje'],
            ['Crítica', str(by_priority['critical']), f"{round(by_priority['critical']/total*100,1) if total else 0}%"],
            ['Alta', str(by_priority['high']), f"{round(by_priority['high']/total*100,1) if total else 0}%"],
            ['Media', str(by_priority['medium']), f"{round(by_priority['medium']/total*100,1) if total else 0}%"],
            ['Baja', str(by_priority['low']), f"{round(by_priority['low']/total*100,1) if total else 0}%"],
        ]
        prio_table = Table(prio_data, colWidths=[6*cm, 4*cm, 4*cm])
        prio_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dc2626')),
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#fee2e2')),
            ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#fed7aa')),
            ('BACKGROUND', (0,3), (-1,3), colors.HexColor('#fef3c7')),
            ('BACKGROUND', (0,4), (-1,4), colors.HexColor('#d1fae5')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (1,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('PADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(prio_table)

        # ============== NUEVA PÁGINA: GRÁFICAS ==============
        elements.append(PageBreak())
        elements.append(Paragraph('📈 Visualizaciones Gráficas', section_style))
        elements.append(Spacer(1, 0.3*cm))

        # Gauge SLA + Donut Estado lado a lado
        if 'sla' in chart_buffers:
            elements.append(Image(chart_buffers['sla'], width=14*cm, height=7*cm))
            elements.append(Spacer(1, 0.5*cm))

        if 'status' in chart_buffers:
            elements.append(Image(chart_buffers['status'], width=14*cm, height=8*cm))
            elements.append(Spacer(1, 0.5*cm))

        # Nueva página para más gráficas
        elements.append(PageBreak())
        elements.append(Paragraph('📊 Análisis de Tickets', section_style))
        elements.append(Spacer(1, 0.3*cm))

        if 'priority' in chart_buffers:
            elements.append(Image(chart_buffers['priority'], width=14*cm, height=7*cm))
            elements.append(Spacer(1, 0.5*cm))

        if 'trend' in chart_buffers:
            elements.append(Image(chart_buffers['trend'], width=14*cm, height=7*cm))
            elements.append(Spacer(1, 0.5*cm))

        # NUEVA PÁGINA - TOP TÉCNICOS con gráfica
        elements.append(PageBreak())
        elements.append(Paragraph('🏆 Ranking de Técnicos', section_style))

        if 'tech' in chart_buffers:
            elements.append(Image(chart_buffers['tech'], width=14*cm, height=7*cm))
            elements.append(Spacer(1, 0.5*cm))

        elements.append(Paragraph('Detalle Top 10:', ParagraphStyle('h3', parent=styles['Heading3'],
                                                                    fontSize=12, textColor=colors.HexColor('#374151'),
                                                                    spaceAfter=8)))
        tech_data = [['#', 'Técnico', 'Asignados', 'Resueltos', 'Tasa %']]
        for i, t in enumerate(tech_stats[:10], 1):
            tech_data.append([str(i), t['name'], str(t['assigned']), str(t['resolved']), f"{t['rate']}%"])

        tech_table = Table(tech_data, colWidths=[1*cm, 7*cm, 3*cm, 3*cm, 3*cm])
        tech_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (1,0), (1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9fafb')]),
            ('PADDING', (0,0), (-1,-1), 8),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))
        elements.append(tech_table)

        # ===== PÁGINA FINAL: CONCLUSIONES Y FIRMA =====
        elements.append(PageBreak())
        elements.append(Paragraph('📋 CONCLUSIONES Y PRÓXIMOS PASOS', section_style))
        elements.append(Spacer(1, 0.3*cm))

        # Resumen final con bullets
        conclusion_text = f'''
        Este reporte refleja el estado operativo de DeskEli para <b>{company_name}</b>
        al corte del <b>{now_dt.strftime("%Y-%m-%d")}</b>. Los hallazgos principales son:
        '''
        elements.append(Paragraph(conclusion_text, ParagraphStyle('Conclusion', parent=normal,
            fontSize=11, leading=18, alignment=4, spaceAfter=15)))

        # Tabla resumen final
        summary_final = [
            ['ÁREA', 'ESTADO', 'ACCIÓN REQUERIDA'],
            ['Cumplimiento SLA', f'{sla_pct}%',
             'Mantener' if sla_pct >= 85 else 'Reforzar capacitación'],
            ['Tickets Críticos', str(by_priority['critical']),
             'Sin acción' if by_priority['critical'] == 0 else 'Atención inmediata'],
            ['Productividad Equipo', f'{len(tech_stats)} técnicos',
             'OK' if tech_stats else 'Sin datos'],
            ['Satisfacción Cliente', f'{avg_rating}/5',
             'Mantener' if avg_rating >= 4 else 'Revisar procesos'],
            ['Salud General', health_status, 'Monitoreo continuo'],
        ]
        summary_table = Table(summary_final, colWidths=[5*cm, 4*cm, 8*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9fafb')]),
            ('PADDING', (0,0), (-1,-1), 8),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (1,0), (1,-1), 'CENTER'),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 1*cm))

        # Próximos hitos
        elements.append(Paragraph('🎯 PRÓXIMOS HITOS', ParagraphStyle('h3', parent=styles['Heading3'],
            fontSize=13, textColor=colors.HexColor('#374151'), spaceAfter=8)))
        next_steps = [
            f'Revisar este reporte en comité de TI antes del {(now_dt + timedelta(days=7)).strftime("%Y-%m-%d")}.',
            'Implementar las recomendaciones estratégicas en orden de prioridad.',
            'Re-generar reporte en 30 días para medir mejoras.',
            'Comunicar resultados al equipo de operaciones.',
            'Escalar tickets críticos abiertos a comité ejecutivo si persisten.',
        ]
        for i, step in enumerate(next_steps, 1):
            elements.append(Paragraph(f'<font color="#2563eb"><b>□</b></font> &nbsp; {step}',
                ParagraphStyle('Step', parent=normal, fontSize=10, leading=16, spaceAfter=4, leftIndent=12)))

        elements.append(Spacer(1, 1.5*cm))

        # Firma
        sign_data = [
            ['', '', ''],
            ['_______________________', '', '_______________________'],
            [session.get("name", "Admin"), '', 'Gerencia TI'],
            ['Administrador del Sistema', '', 'Aprobación'],
        ]
        sign_table = Table(sign_data, colWidths=[6*cm, 5*cm, 6*cm])
        sign_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,2), (-1,2), 'Helvetica-Bold'),
            ('FONTSIZE', (0,2), (-1,2), 10),
            ('FONTSIZE', (0,3), (-1,3), 9),
            ('TEXTCOLOR', (0,3), (-1,3), colors.HexColor('#6b7280')),
        ]))
        elements.append(sign_table)

        # FOOTER FINAL
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
                                      fontSize=8, textColor=colors.HexColor('#9ca3af'),
                                      alignment=TA_CENTER, leading=12)
        elements.append(Paragraph('━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━', footer_style))
        elements.append(Paragraph(
            f'<b>DeskEli</b> · Sistema de Gestión de Incidencias TI<br/>'
            f'Reporte generado el {now_dt.strftime("%Y-%m-%d %H:%M:%S")} · '
            f'Documento confidencial · Solo para uso interno',
            footer_style))

        doc.build(elements)
        buffer.seek(0)

        log_audit('download_report', session['user_id'], 'report', None, f'Reporte PDF descargado')

        from flask import send_file
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'reporte_metricas_{company}_{datetime.now().strftime("%Y%m%d")}.pdf'
        )

    return jsonify({'success': False, 'error': 'Formato no soportado'}), 400


# ═════════════════════════════════════════════════════════════════════════════
# REPORTES EJECUTIVOS PARA GERENCIA / DIRECCIÓN
# ═════════════════════════════════════════════════════════════════════════════

EXECUTIVE_REPORTS = {
    'sla_compliance': {
        'name': 'Cumplimiento de SLA',
        'icon': '⏱️',
        'description': 'Análisis mensual de cumplimiento de SLA por prioridad y técnico. Identifica brechas y tendencias.'
    },
    'tech_performance': {
        'name': 'Rendimiento del Equipo',
        'icon': '🏆',
        'description': 'Ranking de técnicos: tickets resueltos, tiempo promedio, calificación cliente, productividad.'
    },
    'category_analysis': {
        'name': 'Análisis de Categorías',
        'icon': '📊',
        'description': 'Distribución de tickets por categoría, prioridad y empresa. Identifica problemas recurrentes.'
    },
    'backlog_aging': {
        'name': 'Backlog y Antigüedad',
        'icon': '🗂️',
        'description': 'Tickets pendientes ordenados por antigüedad. Detecta casos olvidados y cuellos de botella.'
    },
    'customer_satisfaction': {
        'name': 'Satisfacción del Cliente',
        'icon': '⭐',
        'description': 'CSAT promedio, distribución de calificaciones, tickets con peor calificación para acción.'
    }
}


def _exec_report_data(report_type, all_tickets, period_label, company_name):
    """Recopila datos específicos para cada tipo de reporte ejecutivo."""
    now_dt = datetime.now()
    data = {
        'type': report_type,
        'meta': EXECUTIVE_REPORTS.get(report_type, {}),
        'period': period_label,
        'company': company_name,
        'generated_at': now_dt.strftime('%Y-%m-%d %H:%M'),
        'total_tickets': len(all_tickets)
    }

    if report_type == 'sla_compliance':
        # Por prioridad
        by_priority = {}
        for prio in ['critical', 'high', 'medium', 'low']:
            t_prio = [t for t in all_tickets if t.priority == prio]
            total = len(t_prio)
            on_time = 0
            for t in t_prio:
                if not t.sla_deadline:
                    on_time += 1
                elif t.resolved_at:
                    if t.resolved_at <= t.sla_deadline:
                        on_time += 1
                elif t.sla_deadline > now_dt:
                    on_time += 1
            pct = round((on_time / total * 100), 1) if total else 100
            by_priority[prio] = {'total': total, 'on_time': on_time, 'pct': pct}
        data['by_priority'] = by_priority

        # Por técnico
        by_tech = {}
        techs_with_tickets = set(t.assignee_id for t in all_tickets if t.assignee_id)
        for tech_id in techs_with_tickets:
            tech = User.query.get(tech_id)
            if not tech: continue
            t_tech = [t for t in all_tickets if t.assignee_id == tech_id]
            on_time = sum(1 for t in t_tech if not t.sla_deadline or
                          (t.resolved_at and t.resolved_at <= t.sla_deadline) or
                          (not t.resolved_at and t.sla_deadline > now_dt))
            total = len(t_tech)
            by_tech[tech.name] = {
                'total': total, 'on_time': on_time,
                'pct': round((on_time / total * 100), 1) if total else 100
            }
        data['by_tech'] = dict(sorted(by_tech.items(), key=lambda x: -x[1]['pct']))

        # Tendencia mensual últimos 6 meses
        by_month = {}
        for i in range(5, -1, -1):
            month_start = (now_dt.replace(day=1) - timedelta(days=i*30)).replace(day=1)
            month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            key = month_start.strftime('%Y-%m')
            m_tickets = [t for t in all_tickets if t.created_at and month_start <= t.created_at < month_end]
            m_total = len(m_tickets)
            m_on_time = sum(1 for t in m_tickets if not t.sla_deadline or
                            (t.resolved_at and t.resolved_at <= t.sla_deadline))
            by_month[key] = {
                'total': m_total,
                'on_time': m_on_time,
                'pct': round((m_on_time / m_total * 100), 1) if m_total else 0
            }
        data['by_month'] = by_month

    elif report_type == 'tech_performance':
        techs_with_tickets = set(t.assignee_id for t in all_tickets if t.assignee_id)
        rows = []
        for tech_id in techs_with_tickets:
            tech = User.query.get(tech_id)
            if not tech: continue
            t_tech = [t for t in all_tickets if t.assignee_id == tech_id]
            resolved = [t for t in t_tech if t.status == 'resolved' and t.resolved_at]
            in_progress = [t for t in t_tech if t.status == 'in_progress']
            ratings = [t.rating for t in t_tech if t.rating]
            avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else 0
            durations = [(t.resolved_at - t.created_at).total_seconds() / 3600
                         for t in resolved if t.created_at]
            avg_hrs = round(sum(durations) / len(durations), 1) if durations else 0
            on_time = sum(1 for t in resolved if t.sla_deadline and t.resolved_at <= t.sla_deadline)
            sla_pct = round((on_time / len(resolved) * 100), 1) if resolved else 0
            rows.append({
                'name': tech.name,
                'role': tech.role,
                'assigned': len(t_tech),
                'resolved': len(resolved),
                'in_progress': len(in_progress),
                'avg_hours': avg_hrs,
                'sla_pct': sla_pct,
                'rating': avg_rating,
                'rating_count': len(ratings)
            })
        rows.sort(key=lambda x: (-x['resolved'], -x['rating']))
        data['rows'] = rows

    elif report_type == 'category_analysis':
        # Por categoría
        by_category = {}
        for t in all_tickets:
            cat = t.category or 'Sin categoría'
            if cat not in by_category:
                by_category[cat] = {'total': 0, 'resolved': 0, 'open': 0, 'critical': 0,
                                    'avg_hrs': []}
            by_category[cat]['total'] += 1
            if t.status == 'resolved': by_category[cat]['resolved'] += 1
            elif t.status in ('open', 'in_progress'): by_category[cat]['open'] += 1
            if t.priority == 'critical': by_category[cat]['critical'] += 1
            if t.resolved_at and t.created_at:
                by_category[cat]['avg_hrs'].append(
                    (t.resolved_at - t.created_at).total_seconds() / 3600
                )
        result = []
        for cat, info in by_category.items():
            durations = info.pop('avg_hrs', [])
            info['avg_hrs'] = round(sum(durations) / len(durations), 1) if durations else 0
            info['category'] = cat
            info['resolution_pct'] = round(info['resolved'] / info['total'] * 100, 1) if info['total'] else 0
            result.append(info)
        result.sort(key=lambda x: -x['total'])
        data['rows'] = result

    elif report_type == 'backlog_aging':
        # Tickets abiertos/en progreso, ordenados por antigüedad
        backlog = [t for t in all_tickets if t.status in ('open', 'in_progress')]
        buckets = {'<24h': 0, '24h-3d': 0, '3-7d': 0, '7-30d': 0, '>30d': 0}
        rows = []
        for t in backlog:
            age_hrs = (now_dt - t.created_at).total_seconds() / 3600 if t.created_at else 0
            if age_hrs < 24: bucket = '<24h'
            elif age_hrs < 72: bucket = '24h-3d'
            elif age_hrs < 168: bucket = '3-7d'
            elif age_hrs < 720: bucket = '7-30d'
            else: bucket = '>30d'
            buckets[bucket] += 1
            sla_status = 'OK'
            if t.sla_deadline and t.sla_deadline < now_dt: sla_status = 'VENCIDO'
            elif t.sla_deadline and (t.sla_deadline - now_dt).total_seconds() / 60 < 30: sla_status = 'POR VENCER'
            rows.append({
                'ticket_number': t.ticket_number,
                'title': t.title,
                'priority': t.priority,
                'status': t.status,
                'assignee': t.assignee.name if t.assignee else 'Sin asignar',
                'created_at': t.created_at.strftime('%Y-%m-%d') if t.created_at else '',
                'age_hours': round(age_hrs, 1),
                'age_bucket': bucket,
                'sla_status': sla_status,
                'company': t.company
            })
        rows.sort(key=lambda x: -x['age_hours'])
        data['buckets'] = buckets
        data['rows'] = rows[:50]  # Top 50 más viejos
        data['backlog_total'] = len(backlog)

    elif report_type == 'customer_satisfaction':
        rated = [t for t in all_tickets if t.rating]
        avg_rating = round(sum(t.rating for t in rated) / len(rated), 2) if rated else 0
        dist = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        for t in rated:
            dist[t.rating] = dist.get(t.rating, 0) + 1
        # Peores ratings (1-2 estrellas) para acción
        worst = sorted([t for t in rated if t.rating <= 2], key=lambda x: x.rating)[:20]
        data['total_rated'] = len(rated)
        data['total_resolved'] = sum(1 for t in all_tickets if t.status == 'resolved')
        data['response_rate'] = round(len(rated) / max(1, data['total_resolved']) * 100, 1)
        data['avg_rating'] = avg_rating
        data['distribution'] = dist
        data['worst_tickets'] = [{
            'ticket_number': t.ticket_number,
            'title': t.title,
            'rating': t.rating,
            'assignee': t.assignee.name if t.assignee else 'N/A',
            'category': t.category or 'General'
        } for t in worst]

    return data


def _exec_report_to_excel(data):
    """Genera Excel multi-hoja para el reporte ejecutivo."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = data['meta'].get('name', 'Reporte')[:31]

    header_fill = PatternFill('solid', fgColor='1E40AF')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_font = Font(bold=True, size=18, color='1E40AF')
    thin = Side(border_style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabecera común
    ws['A1'] = f"{data['meta'].get('icon','')} {data['meta'].get('name','Reporte')}".strip()
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    ws['A2'] = f"Empresa: {data['company']}"
    ws['A2'].font = Font(bold=True, size=11)
    ws.merge_cells('A2:F2')
    ws['A3'] = f"Período: {data['period']}"
    ws['A3'].font = Font(bold=True, color='065F46', size=11)
    ws.merge_cells('A3:F3')
    ws['A4'] = f"Generado: {data['generated_at']} · Total tickets analizados: {data['total_tickets']}"
    ws['A4'].font = Font(italic=True, color='6B7280', size=10)
    ws.merge_cells('A4:F4')

    row = 6

    def write_header(cells_values):
        nonlocal row
        for i, v in enumerate(cells_values):
            c = ws.cell(row=row, column=i+1, value=v)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        row += 1

    def write_row(values, alt=False):
        nonlocal row
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=i+1, value=v)
            c.border = border
            if alt:
                c.fill = PatternFill('solid', fgColor='F3F4F6')
        row += 1

    rtype = data['type']

    if rtype == 'sla_compliance':
        # Por prioridad
        write_header(['Prioridad', 'Total Tickets', 'A Tiempo', '% Cumplimiento'])
        prio_names = {'critical':'🔴 Crítica','high':'🟠 Alta','medium':'🟡 Media','low':'🟢 Baja'}
        for i, (p, info) in enumerate(data['by_priority'].items()):
            write_row([prio_names[p], info['total'], info['on_time'], f"{info['pct']}%"], alt=i%2)
        row += 1

        ws.cell(row=row, column=1, value='📊 Tendencia Mensual (últimos 6 meses)').font = Font(bold=True, size=13, color='1E40AF')
        row += 1
        write_header(['Mes', 'Total', 'A Tiempo', '% Cumplimiento'])
        for i, (m, info) in enumerate(data['by_month'].items()):
            write_row([m, info['total'], info['on_time'], f"{info['pct']}%"], alt=i%2)
        row += 1

        ws.cell(row=row, column=1, value='👥 Cumplimiento SLA por Técnico').font = Font(bold=True, size=13, color='1E40AF')
        row += 1
        write_header(['Técnico', 'Total', 'A Tiempo', '% Cumplimiento'])
        for i, (name, info) in enumerate(data['by_tech'].items()):
            write_row([name, info['total'], info['on_time'], f"{info['pct']}%"], alt=i%2)

    elif rtype == 'tech_performance':
        write_header(['Técnico', 'Rol', 'Asignados', 'Resueltos', 'En Progreso', 'Tiempo Prom (h)', '% SLA', '⭐ Rating', 'Encuestas'])
        for i, r in enumerate(data['rows']):
            write_row([r['name'], r['role'], r['assigned'], r['resolved'], r['in_progress'],
                       r['avg_hours'], f"{r['sla_pct']}%", r['rating'] if r['rating'] else '—', r['rating_count']], alt=i%2)

    elif rtype == 'category_analysis':
        write_header(['Categoría', 'Total', 'Resueltos', 'Abiertos', 'Críticos', 'Tiempo Prom (h)', '% Resolución'])
        for i, r in enumerate(data['rows']):
            write_row([r['category'], r['total'], r['resolved'], r['open'], r['critical'],
                       r['avg_hrs'], f"{r['resolution_pct']}%"], alt=i%2)

    elif rtype == 'backlog_aging':
        ws.cell(row=row, column=1, value='📦 Distribución por Antigüedad').font = Font(bold=True, size=13, color='1E40AF')
        row += 1
        write_header(['Rango', 'Cantidad de Tickets'])
        for i, (b, c) in enumerate(data['buckets'].items()):
            write_row([b, c], alt=i%2)
        row += 1

        ws.cell(row=row, column=1, value=f'🗂️ Top 50 tickets más antiguos (de {data["backlog_total"]} totales en backlog)').font = Font(bold=True, size=13, color='1E40AF')
        row += 1
        write_header(['Ticket', 'Título', 'Prioridad', 'Estado', 'Asignado', 'Creado', 'Días', 'SLA'])
        for i, r in enumerate(data['rows']):
            write_row([r['ticket_number'], r['title'][:50], r['priority'], r['status'],
                       r['assignee'], r['created_at'], round(r['age_hours']/24, 1), r['sla_status']], alt=i%2)

    elif rtype == 'customer_satisfaction':
        ws.cell(row=row, column=1, value='📈 KPIs Generales').font = Font(bold=True, size=13, color='1E40AF')
        row += 1
        write_header(['Métrica', 'Valor'])
        write_row(['Tickets resueltos', data['total_resolved']])
        write_row(['Tickets con calificación', data['total_rated']], alt=True)
        write_row(['Tasa de respuesta', f"{data['response_rate']}%"])
        write_row(['⭐ Rating promedio', f"{data['avg_rating']}/5"], alt=True)
        row += 1

        ws.cell(row=row, column=1, value='⭐ Distribución de Calificaciones').font = Font(bold=True, size=13, color='1E40AF')
        row += 1
        write_header(['Estrellas', 'Cantidad'])
        for i, (s, c) in enumerate(sorted(data['distribution'].items(), reverse=True)):
            label = '⭐'*s + (' ' + '☆'*(5-s) if s<5 else '')
            write_row([label, c], alt=i%2)
        row += 1

        if data['worst_tickets']:
            ws.cell(row=row, column=1, value='⚠️ Tickets con Peor Calificación (Acción Recomendada)').font = Font(bold=True, size=13, color='DC2626')
            row += 1
            write_header(['Ticket', 'Título', '⭐', 'Asignado', 'Categoría'])
            for i, t in enumerate(data['worst_tickets']):
                write_row([t['ticket_number'], t['title'][:50], t['rating'], t['assignee'], t['category']], alt=i%2)

    # Ajustar ancho columnas
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            try:
                col_letter = cell.column_letter
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _exec_report_to_pdf(data):
    """Genera PDF ejecutivo para el reporte."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'],
        fontSize=20, textColor=colors.HexColor('#1E40AF'), spaceAfter=8)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'],
        fontSize=14, textColor=colors.HexColor('#1E40AF'), spaceAfter=10, spaceBefore=14)
    normal = styles['Normal']

    elements = []
    elements.append(Paragraph(f"{data['meta'].get('icon','')} {data['meta'].get('name','Reporte Ejecutivo')}", title_style))
    elements.append(Paragraph(f"<b>Empresa:</b> {data['company']}", normal))
    elements.append(Paragraph(f"<b>Período:</b> {data['period']}", normal))
    elements.append(Paragraph(f"<b>Generado:</b> {data['generated_at']} | <b>Total tickets:</b> {data['total_tickets']}", normal))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(f"<i>{data['meta'].get('description','')}</i>", normal))
    elements.append(Spacer(1, 0.5*cm))

    def make_table(headers, rows, col_widths=None):
        tbl_data = [headers] + rows
        tbl = Table(tbl_data, colWidths=col_widths)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 5),
        ]))
        return tbl

    rtype = data['type']

    if rtype == 'sla_compliance':
        elements.append(Paragraph('Cumplimiento por Prioridad', section_style))
        prio_names = {'critical':'🔴 Crítica','high':'🟠 Alta','medium':'🟡 Media','low':'🟢 Baja'}
        rows = [[prio_names[p], str(info['total']), str(info['on_time']), f"{info['pct']}%"]
                for p, info in data['by_priority'].items()]
        elements.append(make_table(['Prioridad','Total','A Tiempo','% SLA'], rows, [5*cm,3*cm,3*cm,3*cm]))
        elements.append(Spacer(1, 0.4*cm))

        elements.append(Paragraph('Tendencia Mensual (últimos 6 meses)', section_style))
        rows = [[m, str(info['total']), str(info['on_time']), f"{info['pct']}%"]
                for m, info in data['by_month'].items()]
        elements.append(make_table(['Mes','Total','A Tiempo','% SLA'], rows, [4*cm,3*cm,3*cm,3*cm]))
        elements.append(PageBreak())

        elements.append(Paragraph('Cumplimiento SLA por Técnico', section_style))
        rows = [[n, str(i['total']), str(i['on_time']), f"{i['pct']}%"]
                for n, i in data['by_tech'].items()]
        elements.append(make_table(['Técnico','Total','A Tiempo','% SLA'], rows, [7*cm,3*cm,3*cm,3*cm]))

    elif rtype == 'tech_performance':
        elements.append(Paragraph('Ranking de Rendimiento', section_style))
        rows = [[r['name'][:25], r['role'][:10], str(r['assigned']), str(r['resolved']),
                 str(r['avg_hours']), f"{r['sla_pct']}%",
                 str(r['rating']) if r['rating'] else '—']
                for r in data['rows']]
        elements.append(make_table(
            ['Técnico','Rol','Asign.','Resuel.','Prom h','% SLA','⭐'],
            rows, [4*cm,2*cm,1.5*cm,1.7*cm,1.7*cm,1.5*cm,1.5*cm]))

    elif rtype == 'category_analysis':
        elements.append(Paragraph('Tickets por Categoría', section_style))
        rows = [[r['category'][:25], str(r['total']), str(r['resolved']),
                 str(r['open']), str(r['critical']), str(r['avg_hrs']),
                 f"{r['resolution_pct']}%"]
                for r in data['rows']]
        elements.append(make_table(
            ['Categoría','Total','Resuel.','Abier.','Críticos','Prom h','% Resol.'],
            rows, [4.5*cm,1.7*cm,1.7*cm,1.7*cm,1.7*cm,1.7*cm,2*cm]))

    elif rtype == 'backlog_aging':
        elements.append(Paragraph(f'Backlog actual: {data["backlog_total"]} tickets pendientes', section_style))
        rows = [[b, str(c)] for b, c in data['buckets'].items()]
        elements.append(make_table(['Antigüedad','Cantidad'], rows, [6*cm,4*cm]))
        elements.append(PageBreak())

        elements.append(Paragraph('Top 30 tickets más antiguos en backlog', section_style))
        rows = [[r['ticket_number'][:18], r['title'][:30], r['priority'][:10],
                 r['assignee'][:18], str(round(r['age_hours']/24, 1)), r['sla_status']]
                for r in data['rows'][:30]]
        elements.append(make_table(
            ['Ticket','Título','Prio','Asignado','Días','SLA'],
            rows, [3.5*cm,5.5*cm,1.8*cm,3.5*cm,1.2*cm,2*cm]))

    elif rtype == 'customer_satisfaction':
        elements.append(Paragraph('Indicadores Generales', section_style))
        rows = [
            ['Tickets resueltos', str(data['total_resolved'])],
            ['Tickets calificados', str(data['total_rated'])],
            ['Tasa de respuesta', f"{data['response_rate']}%"],
            ['⭐ Rating promedio', f"{data['avg_rating']}/5"]
        ]
        elements.append(make_table(['Métrica','Valor'], rows, [8*cm,4*cm]))
        elements.append(Spacer(1, 0.4*cm))

        elements.append(Paragraph('Distribución de Calificaciones', section_style))
        rows = [[f"{'⭐'*s}{'☆'*(5-s)}", str(c)]
                for s, c in sorted(data['distribution'].items(), reverse=True)]
        elements.append(make_table(['Estrellas','Cantidad'], rows, [6*cm,4*cm]))

        if data['worst_tickets']:
            elements.append(PageBreak())
            elements.append(Paragraph('⚠️ Tickets con Peor Calificación', section_style))
            rows = [[t['ticket_number'][:18], t['title'][:30],
                     '⭐'*t['rating'], t['assignee'][:18], t['category'][:15]]
                    for t in data['worst_tickets']]
            elements.append(make_table(
                ['Ticket','Título','⭐','Asignado','Categoría'],
                rows, [3.5*cm,5.5*cm,2*cm,3.5*cm,2.5*cm]))

    doc.build(elements)
    buffer.seek(0)
    return buffer


@app.route('/api/reports/executive')
def api_reports_executive():
    """Reportes ejecutivos para gerencia.
    Params: ?type=<sla_compliance|tech_performance|category_analysis|backlog_aging|customer_satisfaction>
            &format=excel|pdf
            &month=YYYY-MM | &date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
    """
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    report_type = request.args.get('type', 'sla_compliance')
    if report_type not in EXECUTIVE_REPORTS:
        return jsonify({'success': False, 'error': f'Tipo de reporte inválido. Opciones: {list(EXECUTIVE_REPORTS.keys())}'}), 400

    fmt = request.args.get('format', 'pdf').lower()
    if fmt not in ('excel', 'pdf'):
        return jsonify({'success': False, 'error': 'format debe ser excel o pdf'}), 400

    # Filtros de fecha (mismos que /api/reports/metrics)
    date_from_raw = request.args.get('date_from', '').strip()
    date_to_raw = request.args.get('date_to', '').strip()
    month_raw = request.args.get('month', '').strip()

    date_from = date_to = None
    period_label = 'Histórico completo'
    try:
        if month_raw:
            yr, mo = map(int, month_raw.split('-'))
            date_from = datetime(yr, mo, 1)
            date_to = datetime(yr+1, 1, 1) if mo == 12 else datetime(yr, mo+1, 1)
            period_label = f"Mes: {date_from.strftime('%B %Y')}"
        elif date_from_raw or date_to_raw:
            if date_from_raw:
                date_from = datetime.strptime(date_from_raw, '%Y-%m-%d')
            if date_to_raw:
                date_to = datetime.strptime(date_to_raw, '%Y-%m-%d') + timedelta(days=1)
            df = date_from.strftime('%Y-%m-%d') if date_from else '—'
            dt = (date_to - timedelta(days=1)).strftime('%Y-%m-%d') if date_to else '—'
            period_label = f"Del {df} al {dt}"
    except (ValueError, TypeError) as e:
        return jsonify({'success': False, 'error': f'Fechas inválidas: {e}'}), 400

    company = session['company']
    company_obj = Company.query.filter_by(code=company).first()
    company_name = company_obj.name if company_obj else company

    scope = admin_companies_scope()
    query = Ticket.query.filter(Ticket.company.in_(scope))
    if date_from:
        query = query.filter(Ticket.created_at >= date_from)
    if date_to:
        query = query.filter(Ticket.created_at < date_to)
    all_tickets = query.all()

    data = _exec_report_data(report_type, all_tickets, period_label, company_name)

    if fmt == 'excel':
        buffer = _exec_report_to_excel(data)
        filename = f"reporte_{report_type}_{company}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(buffer,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    else:
        buffer = _exec_report_to_pdf(data)
        filename = f"reporte_{report_type}_{company}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(buffer, mimetype='application/pdf',
                         as_attachment=True, download_name=filename)


def _parse_role_labels(text_value):
    """Convierte el JSON serializado de extra_role_labels en lista de strings."""
    if not text_value:
        return []
    try:
        data = json.loads(text_value)
        if isinstance(data, list):
            return [str(x) for x in data if str(x).strip()]
    except Exception:
        pass
    return []


def _serialize_role_labels(labels):
    """Lista de strings → JSON. Deduplica y filtra vacíos."""
    seen = set()
    out = []
    for x in (labels or []):
        s = str(x).strip()
        if s and s not in seen:
            seen.add(s)
            out.append(s)
    return json.dumps(out, ensure_ascii=False) if out else None


def _all_role_labels(user):
    """Combina role_label (primario) + extra_role_labels en lista única."""
    primary = (user.role_label or '').strip()
    extras = _parse_role_labels(user.extra_role_labels)
    result = []
    if primary:
        result.append(primary)
    for e in extras:
        if e != primary and e not in result:
            result.append(e)
    return result


@app.route('/api/admin/team', methods=['GET'])
def api_admin_team():
    """Listar miembros del equipo de la empresa"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company = session['company']
    users = User.query.filter_by(company=company).order_by(User.role.desc(), User.name).all()
    return jsonify({
        'success': True,
        'team': [{
            'id': u.id,
            'username': u.username,
            'name': u.name,
            'email': u.email or '',
            'role': u.role,
            'role_label': u.role_label or '',
            'all_role_labels': _all_role_labels(u),
            'extra_role_labels': _parse_role_labels(u.extra_role_labels),
            'is_active': bool(u.is_active),
            'last_login': u.last_login.strftime('%d/%m %H:%M') if u.last_login else 'Nunca',
            'created_at': u.created_at.strftime('%d/%m/%Y') if u.created_at else ''
        } for u in users]
    })


@app.route('/api/admin/team/import-template', methods=['GET'])
def api_admin_team_import_template():
    """Descarga una plantilla Excel para importar usuarios masivamente."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Usuarios'

    # Cabecera con instrucciones
    headers = ['username', 'name', 'email', 'role', 'password']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.fill = PatternFill('solid', fgColor='1E40AF')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 22

    # Filas de ejemplo
    examples = [
        ['jperez',   'Juan Pérez',         'jperez@empresa.com',   'employee',   ''],
        ['mlopez',   'María López',        'mlopez@empresa.com',   'technician', ''],
        ['agomez',   'Andrea Gómez',       'agomez@empresa.com',   'admin',      'MiClaveSegura123'],
    ]
    for i, ex in enumerate(examples, start=2):
        for col, val in enumerate(ex, 1):
            ws.cell(row=i, column=col, value=val)

    # Hoja 2 con instrucciones
    ws2 = wb.create_sheet('Instrucciones')
    instructions = [
        ('📋 Instrucciones de importación', ''),
        ('', ''),
        ('Columna', 'Descripción'),
        ('username', 'Identificador único del usuario (sin espacios). Ej: jperez'),
        ('name', 'Nombre completo. Ej: Juan Pérez'),
        ('email', 'Correo corporativo único'),
        ('role', 'Rol: admin, technician, o employee'),
        ('password', '(Opcional) Contraseña inicial. Si está vacía, se usa "demo"'),
        ('', ''),
        ('Reglas', ''),
        ('• Si el username ya existe en tu empresa, esa fila se OMITE.', ''),
        ('• Los roles permitidos son: admin, technician, employee.', ''),
        ('• El sistema crea los usuarios EN TU EMPRESA actual automáticamente.', ''),
        ('• Las contraseñas se hashean con PBKDF2.', ''),
    ]
    for i, (a, b) in enumerate(instructions, 1):
        c1 = ws2.cell(row=i, column=1, value=a)
        c2 = ws2.cell(row=i, column=2, value=b)
        if i == 1:
            c1.font = Font(bold=True, size=16, color='1E40AF')
            ws2.merge_cells('A1:B1')
        elif i == 3:
            c1.font = Font(bold=True, color='FFFFFF')
            c2.font = Font(bold=True, color='FFFFFF')
            c1.fill = PatternFill('solid', fgColor='1E40AF')
            c2.fill = PatternFill('solid', fgColor='1E40AF')
        elif i == 10:
            c1.font = Font(bold=True, size=14, color='1E40AF')
    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 60

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name='plantilla_importacion_usuarios.xlsx')


@app.route('/api/admin/team/import', methods=['POST'])
def api_admin_team_import():
    """Importa usuarios desde un archivo Excel o CSV."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se recibió archivo'}), 400
    f = request.files['file']
    if not f or not f.filename:
        return jsonify({'success': False, 'error': 'Archivo vacío'}), 400

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    company = session['company']

    # Leer filas
    rows = []
    try:
        if ext in ('xlsx', 'xls'):
            import openpyxl
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            header_row = None
            for r in ws.iter_rows(values_only=True):
                if header_row is None:
                    header_row = [str(c).strip().lower() if c else '' for c in r]
                    continue
                rows.append({
                    header_row[i]: (str(r[i]).strip() if r[i] is not None else '')
                    for i in range(len(header_row)) if i < len(r)
                })
        elif ext == 'csv':
            import csv as _csv
            from io import TextIOWrapper
            wrapper = TextIOWrapper(f.stream, encoding='utf-8-sig')
            reader = _csv.DictReader(wrapper)
            for row in reader:
                # Normalizar keys a lower
                normalized = {(k or '').strip().lower(): (str(v).strip() if v else '') for k, v in row.items()}
                rows.append(normalized)
        else:
            return jsonify({'success': False, 'error': 'Formato no soportado. Usa .xlsx, .xls o .csv'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error leyendo archivo: {str(e)[:200]}'}), 400

    if not rows:
        return jsonify({'success': False, 'error': 'El archivo no tiene filas válidas'}), 400

    # Procesar filas
    import re
    created = 0
    skipped = 0
    errors = []
    skipped_details = []
    valid_roles = {'admin', 'technician', 'employee'}

    for idx, row in enumerate(rows, start=2):  # fila 2 (después de cabecera)
        username = (row.get('username') or '').strip()
        name = (row.get('name') or '').strip()
        email = (row.get('email') or '').strip()
        role = (row.get('role') or '').strip().lower()
        password = (row.get('password') or '').strip() or 'demo'

        # Validaciones por fila
        if not username:
            errors.append({'row': idx, 'error': 'username vacío'})
            continue
        if not name:
            errors.append({'row': idx, 'error': f'{username}: name vacío'})
            continue
        if not email or '@' not in email:
            errors.append({'row': idx, 'error': f'{username}: email inválido'})
            continue
        if role not in valid_roles:
            errors.append({'row': idx, 'error': f'{username}: role "{role}" inválido (usa: admin/technician/employee)'})
            continue
        if not re.match(r'^[a-zA-Z0-9._-]+$', username):
            errors.append({'row': idx, 'error': f'{username}: username solo permite letras, números, ".", "_", "-"'})
            continue

        # ¿Ya existe en esta empresa?
        existing = User.query.filter_by(username=username, company=company).first()
        if existing:
            skipped += 1
            skipped_details.append({'username': username, 'reason': 'ya existe en esta empresa'})
            continue

        # Hash password
        try:
            pwd_hash = hashlib.pbkdf2_hmac('sha256', password.encode(), username.encode(), 100000).hex()
        except Exception as e:
            errors.append({'row': idx, 'error': f'{username}: error generando password ({e})'})
            continue

        try:
            user = User(
                username=username,
                name=name[:120],
                email=email[:120],
                role=role,
                company=company,
                password_hash=pwd_hash,
                is_active=True
            )
            db.session.add(user)
            created += 1
        except Exception as e:
            errors.append({'row': idx, 'error': f'{username}: {str(e)[:150]}'})

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': f'Error guardando: {str(e)[:200]}'}), 500

    log_audit('bulk_import_users', session['user_id'], 'user', None,
              f'Importación masiva: {created} creados, {skipped} omitidos, {len(errors)} errores en {len(rows)} filas')

    return jsonify({
        'success': True,
        'total_rows': len(rows),
        'created': created,
        'skipped': skipped,
        'errors': errors[:200],  # tope para no inflar response
        'skipped_details': skipped_details[:200]
    })


# ─── Mirroring de técnicos Eliot → Pash/Primatela ───────────────────────────
# Regla: si un técnico se crea en la empresa master (eliot), automáticamente
# se replica en pash y primatela con el mismo hash de contraseña, para que
# pueda loguear a las 3 empresas con la misma credencial y aparezca en la
# lista de técnicos disponibles de cada una. Los admins de pash/primatela
# siguen pudiendo crear sus propios técnicos locales (mirrored_from_id NULL).

MIRROR_SOURCE_COMPANY = 'eliot'
MIRROR_TARGET_COMPANIES = ('pash', 'primatela')


def mirror_technician_to_other_companies(src_user):
    """Crea o actualiza espejos del técnico Eliot en pash/primatela.
    - Solo aplica si src_user.company == 'eliot' y src_user.role == 'technician'.
    - Idempotente: si el espejo ya existe, sincroniza name/email/password/is_active.
    - Si el usuario destino existía como local (mirrored_from_id=NULL), NO lo pisa
      — respeta ese registro local.
    """
    if src_user.company != MIRROR_SOURCE_COMPANY:
        return 0
    if src_user.role != 'technician':
        return 0

    created_or_updated = 0
    for target_co in MIRROR_TARGET_COMPANIES:
        # Buscar espejo existente por mirrored_from_id
        mirror = User.query.filter_by(
            mirrored_from_id=src_user.id,
            company=target_co
        ).first()

        if mirror:
            # Sincronizar datos
            mirror.name = src_user.name
            mirror.email = src_user.email
            mirror.role = src_user.role
            mirror.password_hash = src_user.password_hash
            mirror.is_active = src_user.is_active
            mirror.must_change_password = src_user.must_change_password
            mirror.area = src_user.area
            mirror.location = src_user.location
            mirror.phone = src_user.phone
            created_or_updated += 1
            continue

        # No existe el espejo. Verificar si hay conflicto con un usuario LOCAL
        # de la empresa destino (mismo username o mismo email → no pisar).
        conflict = User.query.filter(
            User.company == target_co,
            db.or_(
                User.username == src_user.username,
                db.func.lower(User.email) == (src_user.email or '').lower()
            )
        ).first()
        if conflict:
            # Ya hay un usuario local con ese username/email — no lo tocamos.
            # Si querés que Eliot se apropie de ese registro, seteá su mirrored_from_id
            # manualmente desde el script de backfill.
            continue

        # Crear el espejo
        m = User(
            username=src_user.username,
            name=src_user.name,
            email=src_user.email,
            role=src_user.role,
            company=target_co,
            password_hash=src_user.password_hash,
            is_active=src_user.is_active,
            must_change_password=src_user.must_change_password,
            area=src_user.area,
            location=src_user.location,
            phone=src_user.phone,
            mirrored_from_id=src_user.id,
        )
        db.session.add(m)
        created_or_updated += 1
    return created_or_updated


def delete_technician_mirrors(src_user):
    """Elimina los espejos cuando se elimina el técnico origen en Eliot."""
    if src_user.company != MIRROR_SOURCE_COMPANY:
        return 0
    mirrors = User.query.filter_by(mirrored_from_id=src_user.id).all()
    for m in mirrors:
        db.session.delete(m)
    return len(mirrors)


@app.route('/api/admin/team/sync-from-master', methods=['POST'])
def api_admin_team_sync_from_master():
    """Sincroniza los tecnicos de Eliot (empresa master) a la empresa del admin
    logueado. Idempotente: crea los que faltan y actualiza los ya espejados.
    Solo disponible para admins de pash y primatela (Eliot es el origen)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    admin_company = session.get('company')
    if admin_company == MIRROR_SOURCE_COMPANY:
        return jsonify({
            'success': False,
            'error': 'Este boton es solo para empresas target (pash/primatela). Eliot es la fuente.'
        }), 400
    if admin_company not in MIRROR_TARGET_COMPANIES:
        return jsonify({'success': False, 'error': 'Empresa no configurada como target de sincronizacion.'}), 400

    # Iterar sobre todos los tecnicos de Eliot y disparar el mirror
    eliot_techs = User.query.filter_by(
        company=MIRROR_SOURCE_COMPANY, role='technician'
    ).all()

    stats = {'created': 0, 'updated': 0, 'skipped_conflict': 0}
    for t in eliot_techs:
        existing = User.query.filter_by(
            mirrored_from_id=t.id, company=admin_company
        ).first()

        if existing:
            existing.name = t.name
            existing.email = t.email
            existing.role = t.role
            existing.password_hash = t.password_hash
            existing.is_active = t.is_active
            existing.must_change_password = t.must_change_password
            existing.area = t.area
            existing.location = t.location
            existing.phone = t.phone
            stats['updated'] += 1
            continue

        # Chequear conflicto con usuario local
        conflict = User.query.filter(
            User.company == admin_company,
            db.or_(
                User.username == t.username,
                db.func.lower(User.email) == (t.email or '').lower()
            )
        ).first()
        if conflict:
            stats['skipped_conflict'] += 1
            continue

        m = User(
            username=t.username, name=t.name, email=t.email, role=t.role,
            company=admin_company, password_hash=t.password_hash,
            is_active=t.is_active, must_change_password=t.must_change_password,
            area=t.area, location=t.location, phone=t.phone,
            mirrored_from_id=t.id,
        )
        db.session.add(m)
        stats['created'] += 1

    db.session.commit()
    total = stats['created'] + stats['updated']
    log_audit('sync_technicians_from_master', session['user_id'], 'user', 0,
              f'Sincronizados {total} tecnicos desde eliot a {admin_company} '
              f'(nuevos: {stats["created"]}, actualizados: {stats["updated"]}, '
              f'conflicto local: {stats["skipped_conflict"]})')

    parts = []
    if stats['created']:
        parts.append(f'{stats["created"]} nuevo(s)')
    if stats['updated']:
        parts.append(f'{stats["updated"]} actualizado(s)')
    if stats['skipped_conflict']:
        parts.append(f'{stats["skipped_conflict"]} omitido(s) por conflicto local')
    msg = 'Sincronizacion completada. ' + (', '.join(parts) if parts else 'Sin cambios.')

    return jsonify({'success': True, 'message': msg, 'stats': stats, 'source_count': len(eliot_techs)})


@app.route('/api/admin/team', methods=['POST'])
def api_admin_team_create():
    """Crear nuevo usuario en la empresa"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json()
        username = (data.get('username') or '').strip()
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip()
        role = (data.get('role') or 'employee').strip()
        role_label = (data.get('role_label') or '').strip() or None
        password = (data.get('password') or 'demo').strip()
        company = session['company']

        if not username or not name or not email:
            return jsonify({'success': False, 'error': 'Usuario, nombre y email son requeridos'}), 400

        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
            return jsonify({'success': False, 'error': 'Email inválido'}), 400

        if role not in ['admin', 'technician', 'employee']:
            return jsonify({'success': False, 'error': 'Rol inválido'}), 400

        existing = User.query.filter_by(username=username, company=company).first()
        if existing:
            return jsonify({'success': False, 'error': f'El usuario "{username}" ya existe en esta empresa'}), 400

        # Email único por empresa (necesario para login por email)
        existing_email = User.query.filter(
            db.func.lower(User.email) == email.lower(),
            User.company == company
        ).first()
        if existing_email:
            return jsonify({'success': False, 'error': f'El email "{email}" ya está registrado en esta empresa'}), 400

        # Validar política de contraseñas
        ok, err = validate_password(password, username=username)
        if not ok:
            return jsonify({'success': False, 'error': err}), 400

        # Hash de contraseña
        password_hash = hashlib.pbkdf2_hmac('sha256', password.encode(), username.encode(), 100000).hex()

        u = User(
            username=username,
            name=name,
            email=email,
            role=role,
            role_label=role_label,
            company=company,
            password_hash=password_hash,
            is_active=True,
            must_change_password=True,  # Forzar cambio en primer login
        )
        db.session.add(u)
        db.session.commit()

        # Auto-replicar tecnicos de Eliot en pash/primatela
        mirrored = mirror_technician_to_other_companies(u)
        if mirrored:
            db.session.commit()
            log_audit('mirror_technician', session['user_id'], 'user', u.id,
                      f'Tecnico {username} replicado en {mirrored} empresa(s) target')

        log_audit('create_user', session['user_id'], 'user', u.id, f'Usuario creado: {username} ({role})')
        return jsonify({'success': True, 'id': u.id, 'message': f'Usuario "{name}" creado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/team/<int:user_id>', methods=['PUT'])
def api_admin_team_update(user_id):
    """Actualizar usuario (incluyendo reset de contraseña)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    u = User.query.get(user_id)
    if not u or u.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    try:
        data = request.get_json() or {}
        if 'name' in data: u.name = data['name'].strip()
        if 'email' in data:
            new_email = (data['email'] or '').strip()
            if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', new_email):
                return jsonify({'success': False, 'error': 'Email inválido'}), 400
            # Verificar email único por empresa (excluyendo el propio usuario)
            dup = User.query.filter(
                db.func.lower(User.email) == new_email.lower(),
                User.company == u.company,
                User.id != u.id
            ).first()
            if dup:
                return jsonify({'success': False, 'error': f'Ese email ya está usado por "{dup.username}" en esta empresa'}), 400
            u.email = new_email
        if 'role' in data and data['role'] in ['admin', 'technician', 'employee']:
            u.role = data['role']
        if 'role_label' in data:
            lbl = (data.get('role_label') or '').strip()
            u.role_label = lbl or None
        if 'is_active' in data: u.is_active = bool(data['is_active'])
        password_changed = False
        if 'password' in data and data['password']:
            pwd = str(data['password'])
            ok, err = validate_password(pwd, username=u.username)
            if not ok:
                return jsonify({'success': False, 'error': err}), 400
            u.password_hash = hashlib.pbkdf2_hmac('sha256', pwd.encode(), u.username.encode(), 100000).hex()
            u.must_change_password = True  # Forzar cambio en próximo login
            password_changed = True
        db.session.commit()

        # Sincronizar espejos si el usuario origen es un tecnico de Eliot
        if u.company == MIRROR_SOURCE_COMPANY and u.role == 'technician':
            mirror_technician_to_other_companies(u)
            db.session.commit()

        if password_changed:
            log_audit('reset_password', session['user_id'], 'user', u.id, f'Contraseña reseteada: {u.username} ({u.company})')
        else:
            log_audit('update_user', session['user_id'], 'user', u.id, f'Usuario actualizado: {u.username}')
        return jsonify({'success': True, 'message': 'Contraseña actualizada' if password_changed else 'Usuario actualizado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/users/<int:user_id>/role-labels', methods=['GET'])
def api_user_role_labels_get(user_id):
    """Devuelve la lista de roles personalizados del usuario."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    u = User.query.get(user_id)
    if not u or u.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    return jsonify({
        'success': True,
        'user_id': u.id,
        'user_name': u.name,
        'primary': u.role_label or '',
        'extras': _parse_role_labels(u.extra_role_labels),
        'all_labels': _all_role_labels(u),
    })


@app.route('/api/admin/users/<int:user_id>/role-labels', methods=['PUT'])
def api_user_role_labels_set(user_id):
    """Reemplaza la lista de roles personalizados del usuario.
    Body: {primary: str|null, extras: [str, ...]}
    """
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    u = User.query.get(user_id)
    if not u or u.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    try:
        data = request.get_json() or {}
        # Validar contra roles personalizados existentes
        valid_labels = {r['label'] for r in _load_custom_roles()}

        primary = (data.get('primary') or '').strip()
        if primary and primary not in valid_labels:
            return jsonify({'success': False, 'error': f'Rol primario "{primary}" no existe'}), 400

        extras = data.get('extras') or []
        clean_extras = []
        for e in extras:
            es = str(e).strip()
            if not es or es == primary:
                continue
            if es not in valid_labels:
                return jsonify({'success': False, 'error': f'Rol "{es}" no existe'}), 400
            if es not in clean_extras:
                clean_extras.append(es)

        u.role_label = primary or None
        u.extra_role_labels = _serialize_role_labels(clean_extras)
        db.session.commit()
        total = (1 if primary else 0) + len(clean_extras)
        log_audit('update_user_role_labels', session['user_id'], 'user', u.id,
                  f'Roles asignados a {u.username}: {total} (primario: {primary or "—"})')
        return jsonify({
            'success': True,
            'message': f'{total} rol{"es" if total != 1 else ""} asignado{"s" if total != 1 else ""}',
            'primary': primary,
            'extras': clean_extras,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/team/<int:user_id>', methods=['DELETE'])
def api_admin_team_delete(user_id):
    """Eliminar usuario (con protección)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    if user_id == session['user_id']:
        return jsonify({'success': False, 'error': 'No puedes eliminarte a ti mismo'}), 400

    u = User.query.get(user_id)
    if not u or u.company != session['company']:
        return jsonify({'success': False, 'error': 'No encontrado'}), 404

    # Verificar si tiene tickets asignados o creados
    tickets_assigned = Ticket.query.filter_by(assignee_id=u.id).count()
    tickets_created = Ticket.query.filter_by(creator_id=u.id).count()

    if tickets_assigned > 0 or tickets_created > 0:
        return jsonify({
            'success': False,
            'error': f'No se puede eliminar. Tiene {tickets_assigned} tickets asignados y {tickets_created} creados. Desactívalo en su lugar.'
        }), 400

    name = u.name

    # Si es tecnico de Eliot, tambien eliminar sus espejos en pash/primatela.
    # Solo si los espejos no tienen tickets propios asignados.
    mirrors_deleted = 0
    if u.company == MIRROR_SOURCE_COMPANY and u.role == 'technician':
        mirrors = User.query.filter_by(mirrored_from_id=u.id).all()
        for m in mirrors:
            m_assigned = Ticket.query.filter_by(assignee_id=m.id).count()
            m_created = Ticket.query.filter_by(creator_id=m.id).count()
            if m_assigned == 0 and m_created == 0:
                db.session.delete(m)
                mirrors_deleted += 1
            else:
                # No podemos borrar el espejo — desactivarlo en su lugar
                m.is_active = False

    db.session.delete(u)
    db.session.commit()
    audit_msg = f'Usuario eliminado: {name}'
    if mirrors_deleted:
        audit_msg += f' (+ {mirrors_deleted} espejo(s) en pash/primatela)'
    log_audit('delete_user', session['user_id'], 'user', user_id, audit_msg)
    return jsonify({'success': True, 'message': 'Usuario eliminado'})


# ============ ROLES PERSONALIZADOS ============
# Almacenados como JSON en Config(key='custom_roles')
# Estructura: [{key, label, base_role, color, icon, description}, ...]
# base_role determina los permisos (admin/technician/employee).

def _load_custom_roles():
    c = Config.query.filter_by(key='custom_roles').first()
    if not c or not c.value:
        return []
    try:
        return json.loads(c.value)
    except Exception:
        return []


def _save_custom_roles(roles):
    c = Config.query.filter_by(key='custom_roles').first()
    payload = json.dumps(roles, ensure_ascii=False)
    if c:
        c.value = payload
    else:
        db.session.add(Config(key='custom_roles', value=payload))
    db.session.commit()


@app.route('/api/admin/custom-roles', methods=['GET'])
def api_custom_roles_list():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    return jsonify({'success': True, 'roles': _load_custom_roles()})


@app.route('/api/admin/custom-roles', methods=['POST'])
def api_custom_roles_create():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json() or {}
        key = (data.get('key') or '').strip().lower()
        label = (data.get('label') or '').strip()
        base_role = (data.get('base_role') or '').strip().lower()
        color = (data.get('color') or '#6b7280').strip()
        icon = (data.get('icon') or '👤').strip()
        description = (data.get('description') or '').strip()

        if not key or not label:
            return jsonify({'success': False, 'error': 'Clave y etiqueta son requeridas'}), 400
        if not re.match(r'^[a-z][a-z0-9_]{1,29}$', key):
            return jsonify({'success': False, 'error': 'Clave inválida: solo minúsculas/números/_, debe iniciar con letra'}), 400
        if base_role not in ['admin', 'technician', 'employee']:
            return jsonify({'success': False, 'error': 'Rol base inválido'}), 400
        if key in ('admin', 'technician', 'employee'):
            return jsonify({'success': False, 'error': 'Clave reservada'}), 400

        roles = _load_custom_roles()
        if any(r.get('key') == key for r in roles):
            return jsonify({'success': False, 'error': f'Ya existe un rol con la clave "{key}"'}), 400

        roles.append({
            'key': key, 'label': label, 'base_role': base_role,
            'color': color, 'icon': icon, 'description': description
        })
        _save_custom_roles(roles)
        log_audit('create_custom_role', session['user_id'], 'role', None, f'Rol personalizado creado: {label} ({key} → {base_role})')
        return jsonify({'success': True, 'message': f'Rol "{label}" creado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/custom-roles/<key>', methods=['PUT'])
def api_custom_roles_update(key):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json() or {}
        roles = _load_custom_roles()
        target = next((r for r in roles if r.get('key') == key), None)
        if not target:
            return jsonify({'success': False, 'error': 'Rol no encontrado'}), 404

        if 'label' in data:
            target['label'] = (data.get('label') or target['label']).strip()
        if 'base_role' in data and data['base_role'] in ['admin', 'technician', 'employee']:
            target['base_role'] = data['base_role']
        if 'color' in data:
            target['color'] = (data.get('color') or target.get('color', '#6b7280')).strip()
        if 'icon' in data:
            target['icon'] = (data.get('icon') or target.get('icon', '👤')).strip()
        if 'description' in data:
            target['description'] = (data.get('description') or '').strip()
        _save_custom_roles(roles)
        log_audit('update_custom_role', session['user_id'], 'role', None, f'Rol personalizado actualizado: {key}')
        return jsonify({'success': True, 'message': 'Rol actualizado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/custom-roles/<key>', methods=['DELETE'])
def api_custom_roles_delete(key):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        roles = _load_custom_roles()
        target = next((r for r in roles if r.get('key') == key), None)
        if not target:
            return jsonify({'success': False, 'error': 'Rol no encontrado'}), 404

        # Verificar si hay usuarios usando este rol_label
        users_using = User.query.filter_by(role_label=target['label']).count()
        if users_using > 0:
            return jsonify({
                'success': False,
                'error': f'No se puede eliminar: {users_using} usuario(s) tienen este rol asignado. Cambia su rol primero.'
            }), 400

        roles = [r for r in roles if r.get('key') != key]
        _save_custom_roles(roles)
        log_audit('delete_custom_role', session['user_id'], 'role', None, f'Rol personalizado eliminado: {key}')
        return jsonify({'success': True, 'message': 'Rol eliminado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/custom-roles/export', methods=['GET'])
def api_custom_roles_export():
    """Exporta todos los roles personalizados a JSON descargable."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    roles = _load_custom_roles()
    company = session.get('company')
    payload = {
        '_meta': {
            'exported_at': datetime.now().isoformat(timespec='seconds'),
            'source_company': company,
            'format_version': '1',
            'total': len(roles),
        },
        'custom_roles': roles,
    }

    log_audit('custom_roles_export', session['user_id'], 'role', None,
              f'Exportó {len(roles)} roles personalizados')

    from flask import Response
    import json as _json
    filename = f'roles_personalizados_{company}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    return Response(
        _json.dumps(payload, ensure_ascii=False, indent=2),
        mimetype='application/json',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/api/admin/custom-roles/import', methods=['POST'])
def api_custom_roles_import():
    """Importa roles personalizados desde JSON. Modo merge: salta duplicados por 'key'."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401

    # Acepta JSON en body o archivo subido en form-data
    data = None
    if request.is_json:
        data = request.get_json(silent=True)
    elif 'file' in request.files:
        try:
            import json as _json
            data = _json.loads(request.files['file'].read().decode('utf-8'))
        except Exception as e:
            return jsonify({'success': False, 'error': f'JSON inválido: {e}'}), 400

    if not data or 'custom_roles' not in data:
        return jsonify({'success': False, 'error': 'Formato inválido: falta la clave "custom_roles"'}), 400

    incoming = data['custom_roles']
    if not isinstance(incoming, list):
        return jsonify({'success': False, 'error': '"custom_roles" debe ser una lista'}), 400

    existing_roles = _load_custom_roles()
    existing_keys = {r.get('key', '').lower() for r in existing_roles}

    created = 0
    skipped = 0
    errors = []

    for item in incoming[:200]:
        try:
            key = (item.get('key') or '').strip().lower()
            label = (item.get('label') or '').strip()
            base_role = (item.get('base_role') or '').strip().lower()

            if not key or not label:
                errors.append(f'Falta key o label en: {item}')
                continue
            if not re.match(r'^[a-z][a-z0-9_]{1,29}$', key):
                errors.append(f'Clave inválida "{key}": debe empezar con minúscula y solo tener letras/números/_')
                continue
            if base_role not in ['admin', 'technician', 'employee']:
                errors.append(f'"{key}": base_role inválido ({base_role}). Debe ser admin/technician/employee.')
                continue
            if key in ('admin', 'technician', 'employee'):
                errors.append(f'"{key}": clave reservada del sistema')
                continue
            if key in existing_keys:
                skipped += 1
                continue

            existing_roles.append({
                'key': key,
                'label': label[:80],
                'base_role': base_role,
                'color': (item.get('color') or '#7c3aed').strip()[:20],
                'icon': (item.get('icon') or '👔').strip()[:4],
                'description': (item.get('description') or '').strip()[:255],
            })
            existing_keys.add(key)
            created += 1
        except Exception as e:
            errors.append(f'Error en {item.get("key", "?")}: {e}')

    _save_custom_roles(existing_roles)
    log_audit('custom_roles_import', session['user_id'], 'role', None,
              f'Import roles: {created} creados, {skipped} omitidos, {len(errors)} errores')

    return jsonify({
        'success': True,
        'created': created,
        'skipped': skipped,
        'errors': errors[:10],
        'message': f'✓ {created} roles importados. {skipped} omitidos (ya existían).'
    })


# ============ REPORTES AUTOMÁTICOS ============

@app.route('/api/admin/report-recipients', methods=['GET'])
def api_report_recipients_list():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    scope = admin_companies_scope()
    recipients = ReportRecipient.query.filter(ReportRecipient.company.in_(scope)).order_by(ReportRecipient.company, ReportRecipient.name).all()
    return jsonify({
        'success': True,
        'recipients': [{
            'id': r.id,
            'name': r.name,
            'email': r.email,
            'company': r.company,
            'title': r.title or '',
            'team_user_ids': r.get_team_ids(),
            'team_size': len(r.get_team_ids()),
            'cc_user_ids': r.get_cc_ids(),
            'cc_size': len(r.get_cc_ids()),
            'send_quincenal': bool(r.send_quincenal),
            'send_monthly': bool(r.send_monthly),
            'send_annual': bool(r.send_annual),
            'send_monday_stuck': bool(r.send_monday_stuck),
            'is_active': bool(r.is_active),
            'last_sent_at': r.last_sent_at.strftime('%d/%m/%Y %H:%M') if r.last_sent_at else None,
        } for r in recipients]
    })


@app.route('/api/admin/report-recipients', methods=['POST'])
def api_report_recipients_create():
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip()
        company = (data.get('company') or '').strip().lower()
        if not name or not email or not company:
            return jsonify({'success': False, 'error': 'Nombre, email y empresa son requeridos'}), 400
        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
            return jsonify({'success': False, 'error': 'Email inválido'}), 400
        if company not in admin_companies_scope():
            return jsonify({'success': False, 'error': 'No tenés acceso a esa empresa'}), 403

        r = ReportRecipient(
            name=name, email=email, company=company,
            title=(data.get('title') or '').strip() or None,
            send_quincenal=bool(data.get('send_quincenal', True)),
            send_monthly=bool(data.get('send_monthly', True)),
            send_annual=bool(data.get('send_annual', True)),
            send_monday_stuck=bool(data.get('send_monday_stuck', False)),
            is_active=True,
        )
        db.session.add(r)
        db.session.commit()
        log_audit('create_report_recipient', session['user_id'], 'recipient', r.id, f'{name} ({email}) → {company}')
        return jsonify({'success': True, 'id': r.id, 'message': f'Destinatario "{name}" agregado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/report-recipients/<int:rid>', methods=['PUT'])
def api_report_recipients_update(rid):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    r = ReportRecipient.query.get(rid)
    if not r or r.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    try:
        data = request.get_json() or {}
        if 'name' in data: r.name = (data.get('name') or r.name).strip()
        if 'email' in data:
            new_email = (data.get('email') or '').strip()
            if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', new_email):
                return jsonify({'success': False, 'error': 'Email inválido'}), 400
            r.email = new_email
        if 'company' in data:
            new_comp = (data.get('company') or '').strip().lower()
            if new_comp and new_comp in admin_companies_scope():
                r.company = new_comp
        if 'title' in data: r.title = (data.get('title') or '').strip() or None
        if 'team_user_ids' in data:
            r.set_team_ids(data.get('team_user_ids') or [])
        if 'send_quincenal' in data: r.send_quincenal = bool(data['send_quincenal'])
        if 'send_monthly' in data: r.send_monthly = bool(data['send_monthly'])
        if 'send_annual' in data: r.send_annual = bool(data['send_annual'])
        if 'send_monday_stuck' in data: r.send_monday_stuck = bool(data['send_monday_stuck'])
        if 'is_active' in data: r.is_active = bool(data['is_active'])
        db.session.commit()
        log_audit('update_report_recipient', session['user_id'], 'recipient', r.id, f'Destinatario actualizado: {r.name}')
        return jsonify({'success': True, 'message': 'Destinatario actualizado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/report-recipients/<int:rid>/team', methods=['GET'])
def api_report_recipients_team_get(rid):
    """Devuelve la lista de usuarios disponibles (técnicos/admins) de la empresa
    del destinatario, con flag si están en el equipo."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    r = ReportRecipient.query.get(rid)
    if not r or r.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    team_ids = set(r.get_team_ids())
    users = User.query.filter(
        User.company == r.company,
        User.role.in_(['technician', 'admin'])
    ).order_by(User.role.desc(), User.name).all()
    return jsonify({
        'success': True,
        'recipient': {'id': r.id, 'name': r.name, 'company': r.company},
        'team_user_ids': list(team_ids),
        'users': [{
            'id': u.id,
            'name': u.name,
            'username': u.username,
            'role': u.role,
            'role_label': u.role_label or '',
            'email': u.email,
            'is_active': bool(u.is_active),
            'in_team': u.id in team_ids,
        } for u in users]
    })


@app.route('/api/admin/report-recipients/<int:rid>/team', methods=['PUT'])
def api_report_recipients_team_set(rid):
    """Reemplaza el equipo del destinatario.
    Body: {team_user_ids: [1,2,3]}"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    r = ReportRecipient.query.get(rid)
    if not r or r.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    try:
        data = request.get_json() or {}
        ids = data.get('team_user_ids') or []
        # Validar que los IDs pertenezcan a la misma empresa
        valid_ids = []
        if ids:
            users = User.query.filter(User.id.in_([int(i) for i in ids])).all()
            for u in users:
                if u.company == r.company:
                    valid_ids.append(u.id)
        r.set_team_ids(valid_ids)
        db.session.commit()
        log_audit('update_report_recipient_team', session['user_id'], 'recipient', r.id,
                  f'Equipo de {r.name}: {len(valid_ids)} miembros')
        return jsonify({'success': True, 'team_size': len(valid_ids), 'message': f'{len(valid_ids)} miembros asignados'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/report-recipients/<int:rid>/cc', methods=['GET'])
def api_report_recipients_cc_get(rid):
    """Devuelve la lista de tecnicos/admins de la empresa del destinatario
    marcando quienes estan configurados como CC del reporte."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    r = ReportRecipient.query.get(rid)
    if not r or r.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    cc_ids = set(r.get_cc_ids())
    users = User.query.filter(
        User.company == r.company,
        User.role.in_(['technician', 'admin'])
    ).order_by(User.role.desc(), User.name).all()
    return jsonify({
        'success': True,
        'recipient': {'id': r.id, 'name': r.name, 'company': r.company, 'email': r.email},
        'cc_user_ids': list(cc_ids),
        'users': [{
            'id': u.id,
            'name': u.name,
            'username': u.username,
            'role': u.role,
            'role_label': u.role_label or '',
            'email': u.email,
            'is_active': bool(u.is_active),
            'in_cc': u.id in cc_ids,
        } for u in users]
    })


@app.route('/api/admin/report-recipients/<int:rid>/cc', methods=['PUT'])
def api_report_recipients_cc_set(rid):
    """Reemplaza la lista de especialistas CC del destinatario.
    Body: {cc_user_ids: [1,2,3]}"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    r = ReportRecipient.query.get(rid)
    if not r or r.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    try:
        data = request.get_json() or {}
        ids = data.get('cc_user_ids') or []
        valid_ids = []
        if ids:
            users = User.query.filter(User.id.in_([int(i) for i in ids])).all()
            for u in users:
                if u.company == r.company and u.email:
                    valid_ids.append(u.id)
        r.set_cc_ids(valid_ids)
        db.session.commit()
        log_audit('update_report_recipient_cc', session['user_id'], 'recipient', r.id,
                  f'CC de {r.name}: {len(valid_ids)} especialistas')
        return jsonify({'success': True, 'cc_size': len(valid_ids), 'message': f'{len(valid_ids)} especialistas en CC'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/report-recipients/<int:rid>', methods=['DELETE'])
def api_report_recipients_delete(rid):
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    r = ReportRecipient.query.get(rid)
    if not r or r.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrado'}), 404
    try:
        name = r.name
        db.session.delete(r)
        db.session.commit()
        log_audit('delete_report_recipient', session['user_id'], 'recipient', rid, f'Destinatario eliminado: {name}')
        return jsonify({'success': True, 'message': 'Destinatario eliminado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


def _company_display_name(company_code):
    c = Company.query.filter_by(code=company_code).first()
    return c.name if c else (company_code or '').title()


def _send_report_to_recipient(recipient, period_type, period_start, period_end):
    """Genera Excel + PDF y los envía por email al destinatario.
    Retorna (ok, message).
    period_type: 'quincenal' | 'mensual' | 'anual'.
    """
    try:
        from reports_gen import (_collect_metrics, generate_excel_report,
                                 generate_pdf_report, build_email_body)
    except Exception as e:
        return False, f'Módulo reports_gen no disponible: {e}'

    period_label = {
        'quincenal': 'Quincenal',
        'mensual': 'Mensual',
        'anual': 'Anual',
    }.get(period_type, 'del período')

    company_display = _company_display_name(recipient.company)
    team_ids = recipient.get_team_ids()

    metrics = _collect_metrics(db, Ticket, User, TechnicianProfile,
                               recipient.company, period_start, period_end,
                               team_user_ids=team_ids)
    if team_ids:
        # Incluir nombre del equipo en el display de la empresa
        team_size = len(team_ids)
        company_display = f"{company_display} — Equipo de {recipient.name} ({team_size} miembro{'s' if team_size != 1 else ''})"
    if metrics['total'] == 0:
        scope = f'equipo de {recipient.name}' if team_ids else 'empresa'
        print(f'[report] {recipient.email}: sin tickets del {scope} en el período, no se envía.')
        return False, f'Sin tickets en el período ({scope})'

    try:
        xlsx_buf = generate_excel_report(metrics, period_label, company_display)
        pdf_buf = generate_pdf_report(metrics, period_label, company_display)
    except Exception as e:
        print(f'[report] Error generando reportes: {e}')
        return False, f'Error generando reportes: {e}'

    date_tag = period_start.strftime('%Y%m%d') + '_' + period_end.strftime('%Y%m%d')
    suffix = f'{recipient.company}_{date_tag}'
    attachments = [
        (f'Reporte_{period_label}_{suffix}.xlsx', xlsx_buf,
         'vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
        (f'Reporte_{period_label}_{suffix}.pdf', pdf_buf, 'pdf'),
    ]

    subject = f'[DeskEli] Reporte {period_label} {company_display} — {period_start.strftime("%d/%m")} a {period_end.strftime("%d/%m/%Y")}'
    body = build_email_body(metrics, period_label, company_display, recipient.name)

    # Recolectar emails de los especialistas en CC
    cc_ids = recipient.get_cc_ids()
    cc_emails = []
    if cc_ids:
        cc_users = User.query.filter(
            User.id.in_(cc_ids),
            User.company == recipient.company,
            User.is_active == True,
            User.email.isnot(None)
        ).all()
        cc_emails = [u.email for u in cc_users if u.email and u.email.lower() != (recipient.email or '').lower()]

    ok = send_email(recipient.email, subject, body, attachments=attachments,
                    company=recipient.company, cc_emails=cc_emails or None)
    if ok:
        recipient.last_sent_at = datetime.now()
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        log_audit('send_report', None, 'recipient', recipient.id,
                  f'Reporte {period_type} enviado a {recipient.email} ({recipient.company})')
    return ok, ('Enviado' if ok else 'Falló el envío SMTP')


def _send_monday_stuck_report(recipient):
    """Envia por email un listado de casos del grupo del destinatario que
    esten VENCIDOS (SLA expirado) o lleven mas de 6 dias abiertos.
    Se dispara los lunes automaticamente.
    Retorna (ok, message)."""
    now = datetime.now()
    six_days_ago = now - timedelta(days=6)
    team_ids = recipient.get_team_ids()

    # Base query: tickets abiertos/en progreso de la empresa
    q = Ticket.query.filter(
        Ticket.company == recipient.company,
        Ticket.status.in_(['open', 'in_progress'])
    )
    # Filtrar por grupo del destinatario (si tiene). Si no, todos los tickets
    # abiertos/en progreso de la empresa.
    if team_ids:
        q = q.filter(Ticket.assignee_id.in_(team_ids))

    # Vencidos o antiguos (>6 dias desde creado)
    q = q.filter(
        db.or_(
            db.and_(Ticket.sla_deadline.isnot(None), Ticket.sla_deadline < now),
            Ticket.created_at < six_days_ago
        )
    )
    stuck_tickets = q.order_by(Ticket.sla_deadline.asc().nullslast()).all()

    # Contar tickets activos del grupo (para contexto en el email "al día")
    active_q = Ticket.query.filter(
        Ticket.company == recipient.company,
        Ticket.status.in_(['open', 'in_progress'])
    )
    if team_ids:
        active_q = active_q.filter(Ticket.assignee_id.in_(team_ids))
    active_total = active_q.count()

    # Emails de CC (comun a los dos flujos)
    cc_ids = recipient.get_cc_ids()
    cc_emails = []
    if cc_ids:
        cc_users = User.query.filter(
            User.id.in_(cc_ids),
            User.company == recipient.company,
            User.is_active == True,
            User.email.isnot(None)
        ).all()
        cc_emails = [u.email for u in cc_users if u.email and u.email.lower() != (recipient.email or '').lower()]

    company_display = _company_display_name(recipient.company)
    scope_line = f'Grupo: <strong>{recipient.name}</strong> ({len(team_ids)} especialistas)' if team_ids else f'Alcance: <strong>Toda la empresa {company_display}</strong>'

    # ── Flujo "AL DIA": no hay casos vencidos ni antiguos ──────────
    if not stuck_tickets:
        body_al_dia = f'''
        <div style="font-family:'Segoe UI',Tahoma,sans-serif;max-width:900px;margin:auto;color:#1f2937;">
            <div style="background:linear-gradient(135deg,#16a34a,#15803d);padding:24px;border-radius:10px 10px 0 0;color:white;">
                <h1 style="margin:0;font-size:22px;">✅ Equipo al día</h1>
                <p style="margin:6px 0 0 0;font-size:13px;opacity:0.9;">{company_display} · Lunes {now.strftime('%d/%m/%Y')}</p>
            </div>
            <div style="background:white;padding:20px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 10px 10px;">
                <p style="font-size:14px;">Hola <strong>{recipient.name}</strong>,</p>
                <p style="font-size:14px;">Buenas noticias: <strong>no hay casos vencidos ni con más de 6 días abiertos</strong> en tu área. {scope_line}.</p>

                <div style="background:#d1fae5;border-left:4px solid #16a34a;padding:14px 18px;border-radius:6px;margin:16px 0;">
                    <div style="font-size:15px;color:#065f46;font-weight:700;">🎉 El equipo está al día con los casos</div>
                    <div style="font-size:13px;color:#047857;margin-top:6px;">
                        {active_total} caso(s) activo(s), todos dentro del SLA y con menos de 6 días de antigüedad.
                    </div>
                </div>

                <p style="font-size:13px;color:#374151;">Este es el resumen semanal de lunes:</p>
                <table style="border-collapse:collapse;font-size:13px;margin:10px 0;">
                    <tr><td style="padding:6px 12px;color:#6b7280;">Casos activos del grupo:</td><td style="padding:6px 12px;font-weight:700;color:#111827;">{active_total}</td></tr>
                    <tr><td style="padding:6px 12px;color:#6b7280;">Casos con SLA vencido:</td><td style="padding:6px 12px;font-weight:700;color:#16a34a;">0</td></tr>
                    <tr><td style="padding:6px 12px;color:#6b7280;">Casos con &gt;6 días abiertos:</td><td style="padding:6px 12px;font-weight:700;color:#16a34a;">0</td></tr>
                </table>

                <p style="font-size:11px;color:#9ca3af;margin-top:20px;text-align:center;">
                    Este reporte se envía automáticamente los lunes.<br>
                    DeskEli — Sistema de Gestión de Incidencias
                </p>
            </div>
        </div>
        '''
        subject_al_dia = f'[DeskEli] ✅ Equipo al día — {company_display} · Lunes {now.strftime("%d/%m")}'
        ok = send_email(recipient.email, subject_al_dia, body_al_dia,
                        company=recipient.company, cc_emails=cc_emails or None)
        if ok:
            log_audit('send_monday_stuck', None, 'recipient', recipient.id,
                      f'Alerta de lunes (equipo AL DIA — {active_total} activos) enviada a {recipient.email}')
        return ok, (f'Enviado (equipo al día, {active_total} activos)' if ok else 'Falló envío SMTP')

    # Cachear assignees para el listado
    assignee_ids = {t.assignee_id for t in stuck_tickets if t.assignee_id}
    assignees = {u.id: u for u in User.query.filter(User.id.in_(assignee_ids)).all()} if assignee_ids else {}

    # Build HTML rows
    rows = []
    for t in stuck_tickets:
        age_days = (now - t.created_at).days if t.created_at else 0
        if t.sla_deadline and t.sla_deadline < now:
            sla_overdue_h = int((now - t.sla_deadline).total_seconds() / 3600)
            sla_status = f'<span style="color:#dc2626;font-weight:700;">⏰ SLA VENCIDO hace {sla_overdue_h}h</span>'
        elif age_days > 6:
            sla_status = f'<span style="color:#d97706;font-weight:700;">📅 {age_days} días abierto</span>'
        else:
            sla_status = f'<span style="color:#6b7280;">—</span>'

        assignee = assignees.get(t.assignee_id) if t.assignee_id else None
        assignee_text = f'{assignee.name}' if assignee else '<em style="color:#9ca3af;">Sin asignar</em>'

        priority_colors = {
            'critical': '#dc2626', 'high': '#ea580c', 'medium': '#ca8a04', 'low': '#16a34a'
        }
        prio_color = priority_colors.get(t.priority or 'medium', '#6b7280')
        prio_label = {'critical': 'Crítica', 'high': 'Alta', 'medium': 'Media', 'low': 'Baja'}.get(t.priority or 'medium', t.priority)

        rows.append(f'''
            <tr>
                <td style="padding:8px 10px;border-bottom:1px solid #e5e7eb;font-family:monospace;color:#7c3aed;font-weight:700;">{t.ticket_number}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #e5e7eb;color:#1f2937;">{(t.title or '')[:80]}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #e5e7eb;color:{prio_color};font-weight:600;">{prio_label}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #e5e7eb;color:#374151;">{assignee_text}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #e5e7eb;">{sla_status}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #e5e7eb;color:#6b7280;font-size:12px;">{t.created_at.strftime('%d/%m/%Y') if t.created_at else '—'}</td>
            </tr>
        ''')

    # company_display, scope_line y cc_emails ya se calcularon arriba
    base_url = get_public_base_url() if 'get_public_base_url' in globals() else ''
    link_hint = f'<p style="font-size:13px;color:#6b7280;">Podés abrir el sistema en <a href="{base_url}" style="color:#7c3aed;">{base_url or "DeskEli"}</a>.</p>' if base_url else ''

    body = f'''
    <div style="font-family:'Segoe UI',Tahoma,sans-serif;max-width:900px;margin:auto;color:#1f2937;">
        <div style="background:linear-gradient(135deg,#7c3aed,#5b21b6);padding:24px;border-radius:10px 10px 0 0;color:white;">
            <h1 style="margin:0;font-size:22px;">🚨 Casos vencidos o &gt;6 días abiertos</h1>
            <p style="margin:6px 0 0 0;font-size:13px;opacity:0.9;">{company_display} · Lunes {now.strftime('%d/%m/%Y')}</p>
        </div>
        <div style="background:white;padding:20px;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 10px 10px;">
            <p style="font-size:14px;">Hola <strong>{recipient.name}</strong>,</p>
            <p style="font-size:14px;">Este es tu resumen semanal de casos que requieren atención. {scope_line}.</p>
            <p style="font-size:14px;">Se listan casos <strong>abiertos o en progreso</strong> que están <strong>vencidos por SLA</strong> o <strong>llevan más de 6 días</strong>:</p>
            <div style="background:#fef3c7;border-left:4px solid #d97706;padding:10px 14px;border-radius:5px;margin:14px 0;font-size:14px;">
                <strong>Total: {len(stuck_tickets)} caso(s)</strong>
            </div>
            <table style="width:100%;border-collapse:collapse;font-size:13px;background:white;border:1px solid #e5e7eb;border-radius:6px;overflow:hidden;">
                <thead style="background:#f9fafb;">
                    <tr>
                        <th style="padding:10px;text-align:left;font-size:11px;text-transform:uppercase;color:#6b7280;">N° Caso</th>
                        <th style="padding:10px;text-align:left;font-size:11px;text-transform:uppercase;color:#6b7280;">Asunto</th>
                        <th style="padding:10px;text-align:left;font-size:11px;text-transform:uppercase;color:#6b7280;">Prioridad</th>
                        <th style="padding:10px;text-align:left;font-size:11px;text-transform:uppercase;color:#6b7280;">Asignado</th>
                        <th style="padding:10px;text-align:left;font-size:11px;text-transform:uppercase;color:#6b7280;">SLA / Antigüedad</th>
                        <th style="padding:10px;text-align:left;font-size:11px;text-transform:uppercase;color:#6b7280;">Creado</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
            {link_hint}
            <p style="font-size:11px;color:#9ca3af;margin-top:20px;text-align:center;">
                Este reporte se envía automáticamente los lunes.<br>
                DeskEli — Sistema de Gestión de Incidencias
            </p>
        </div>
    </div>
    '''

    subject = f'[DeskEli] 🚨 {len(stuck_tickets)} caso(s) vencido(s) o >6 días — {company_display}'

    ok = send_email(recipient.email, subject, body, company=recipient.company,
                    cc_emails=cc_emails or None)
    if ok:
        log_audit('send_monday_stuck', None, 'recipient', recipient.id,
                  f'Alerta de lunes ({len(stuck_tickets)} casos) enviada a {recipient.email}')
    return ok, (f'Enviado ({len(stuck_tickets)} casos)' if ok else 'Falló envío SMTP')


def _calc_periods_for_today(today=None):
    """Calcula qué períodos toca enviar HOY. Reglas:
      - QUINCENAL: días 1 y 15 (si cae sábado/domingo, se envía el lunes siguiente).
        - Día 1 → cubre del 16 al fin del mes anterior.
        - Día 15 → cubre del 1 al 14 del mes en curso.
      - MENSUAL: día 1 (mismo ajuste de fin de semana) → cubre el mes anterior completo.
      - ANUAL: 1 de enero (con ajuste fin de semana) → cubre el año anterior.
    Retorna lista de tuplas (period_type, period_start, period_end).
    """
    if today is None:
        today = datetime.now()
    today = today.replace(hour=0, minute=0, second=0, microsecond=0)
    weekday = today.weekday()  # 0=Lun, 6=Dom

    # Helper: dada una fecha "objetivo", ¿hoy es el día apropiado de envío?
    def is_send_day(target_day):
        """target_day = día del mes (1 o 15). True si hoy es ese día,
        o si hoy es lunes y el target cayó sábado o domingo de este mes."""
        if today.day == target_day:
            # No es fin de semana → enviar hoy
            t_weekday = today.weekday()
            return t_weekday < 5
        if weekday == 0:  # Lunes
            # Verificar si el target del mismo mes cae viernes/sábado/domingo y hoy es el lunes después
            try:
                d_target = today.replace(day=target_day)
            except ValueError:
                return False
            # El lunes siguiente al sábado/domingo del target
            if d_target.weekday() in (5, 6) and (today - d_target).days in (2, 3):
                return True
        return False

    periods = []

    # QUINCENAL día 15: cubre 1→14 del mes actual
    if is_send_day(15):
        start = today.replace(day=1)
        end = today.replace(day=15)
        periods.append(('quincenal', start, end))

    # QUINCENAL día 1: cubre 16→fin del mes anterior
    if is_send_day(1):
        last_day_prev = today.replace(day=1) - timedelta(days=1)
        start = last_day_prev.replace(day=16)
        end = today.replace(day=1)
        periods.append(('quincenal', start, end))

        # MENSUAL día 1: cubre todo el mes anterior
        start_m = last_day_prev.replace(day=1)
        end_m = today.replace(day=1)
        periods.append(('mensual', start_m, end_m))

        # ANUAL: 1 de enero (con ajuste finde) cubre año anterior
        if today.month == 1:
            start_y = today.replace(year=today.year - 1, month=1, day=1)
            end_y = today.replace(month=1, day=1)
            periods.append(('anual', start_y, end_y))

    return periods


def report_dispatch_today():
    """Función principal del scheduler: revisa qué reportes toca enviar HOY y los manda."""
    periods = _calc_periods_for_today()
    sent_total = 0

    with app.app_context():
        for period_type, period_start, period_end in periods:
            flag_attr = {'quincenal': 'send_quincenal', 'mensual': 'send_monthly', 'anual': 'send_annual'}[period_type]
            recipients = ReportRecipient.query.filter(
                ReportRecipient.is_active == True,
                getattr(ReportRecipient, flag_attr) == True
            ).all()

            print(f'[report-scheduler] {period_type.upper()}: {len(recipients)} destinatario(s) candidato(s) ({period_start.date()} → {period_end.date()})')

            for r in recipients:
                # Evitar reenvío en el mismo día
                if r.last_sent_at and r.last_sent_at.date() == datetime.now().date():
                    print(f'  → {r.email}: ya recibió hoy, saltando.')
                    continue
                ok, msg = _send_report_to_recipient(r, period_type, period_start, period_end)
                print(f'  → {r.email} [{period_type}]: {msg}')
                if ok:
                    sent_total += 1

        # ── Alerta LUNES: casos vencidos o >6 dias del grupo ──────────
        if datetime.now().weekday() == 0:  # 0 = Lunes
            monday_recipients = ReportRecipient.query.filter(
                ReportRecipient.is_active == True,
                ReportRecipient.send_monday_stuck == True
            ).all()
            print(f'[report-scheduler] LUNES-STUCK: {len(monday_recipients)} destinatario(s) candidato(s)')
            for r in monday_recipients:
                ok, msg = _send_monday_stuck_report(r)
                print(f'  → {r.email} [lunes-stuck]: {msg}')
                if ok:
                    sent_total += 1

    return sent_total


def start_report_scheduler():
    """Scheduler diario para reportes automáticos.
    Corre cada hora pero solo dispara entre las 7:00 y 7:59 del día."""
    import threading
    import time as _t

    def _loop():
        # Esperar 60s al inicio para no chocar con el arranque
        _t.sleep(60)
        last_run_date = None
        while True:
            try:
                now = datetime.now()
                # Disparar una vez al día entre las 7:00 y 7:59
                if now.hour == 7 and (last_run_date != now.date()):
                    try:
                        n = report_dispatch_today()
                        if n > 0:
                            print(f'[report-scheduler] {n} reporte(s) enviado(s) exitosamente')
                        last_run_date = now.date()
                    except Exception as e:
                        print(f'[report-scheduler] Error: {e}')
            except Exception as e:
                print(f'[report-scheduler] Loop error: {e}')
            _t.sleep(60 * 15)  # verificar cada 15 min

    t = threading.Thread(target=_loop, name='report-scheduler', daemon=True)
    t.start()
    print('[report-scheduler] Iniciado (envío diario a las 7:00, días 1 y 15)')


@app.route('/api/admin/reports/send-now', methods=['POST'])
def api_admin_reports_send_now():
    """Endpoint manual para disparar el envío AHORA (testing).
    Body: {recipient_id?, period_type='quincenal'|'mensual'|'anual', period_start?, period_end?}
    Si no se pasan fechas, usa el período actual (mes en curso para mensual, etc.)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json() or {}
        period_type = data.get('period_type', 'quincenal')
        if period_type not in ('quincenal', 'mensual', 'anual'):
            return jsonify({'success': False, 'error': 'period_type inválido'}), 400

        # Período por defecto: últimos 15/30/365 días
        now = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        if period_type == 'quincenal':
            period_start = now - timedelta(days=15)
            period_end = now
        elif period_type == 'mensual':
            period_start = now - timedelta(days=30)
            period_end = now
        else:
            period_start = now - timedelta(days=365)
            period_end = now

        # Si pasaron fechas, usar esas
        if data.get('period_start'):
            try:
                period_start = datetime.strptime(data['period_start'], '%Y-%m-%d')
            except Exception:
                pass
        if data.get('period_end'):
            try:
                period_end = datetime.strptime(data['period_end'], '%Y-%m-%d')
            except Exception:
                pass

        recipient_id = data.get('recipient_id')
        if recipient_id:
            r = ReportRecipient.query.get(int(recipient_id))
            if not r or r.company not in admin_companies_scope():
                return jsonify({'success': False, 'error': 'Destinatario no encontrado'}), 404
            recipients = [r]
        else:
            scope = admin_companies_scope()
            recipients = ReportRecipient.query.filter(
                ReportRecipient.is_active == True,
                ReportRecipient.company.in_(scope)
            ).all()

        if not recipients:
            return jsonify({'success': False, 'error': 'No hay destinatarios activos'}), 400

        results = []
        for r in recipients:
            ok, msg = _send_report_to_recipient(r, period_type, period_start, period_end)
            results.append({'email': r.email, 'company': r.company, 'success': ok, 'message': msg})

        return jsonify({
            'success': True,
            'period_type': period_type,
            'period_start': period_start.strftime('%Y-%m-%d'),
            'period_end': period_end.strftime('%Y-%m-%d'),
            'results': results,
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/reports/send-monday-stuck-now', methods=['POST'])
def api_admin_reports_send_monday_stuck_now():
    """Dispara MANUALMENTE la alerta de lunes (casos vencidos/+6 dias) para
    un destinatario específico. Útil para probar sin esperar al lunes.
    Body: {recipient_id: int}"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    try:
        data = request.get_json() or {}
        recipient_id = data.get('recipient_id')
        if not recipient_id:
            return jsonify({'success': False, 'error': 'recipient_id requerido'}), 400
        r = ReportRecipient.query.get(int(recipient_id))
        if not r or r.company not in admin_companies_scope():
            return jsonify({'success': False, 'error': 'Destinatario no encontrado'}), 404
        ok, msg = _send_monday_stuck_report(r)
        return jsonify({'success': ok, 'message': msg, 'email': r.email})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/general', methods=['GET', 'POST'])
def api_config_general():
    """Configuración general (sistema, idioma, zona horaria)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    if request.method == 'GET':
        cfg = {}
        for key in ['system_name', 'language', 'timezone']:
            c = Config.query.filter_by(key=f'general_{key}').first()
            cfg[key] = c.value if c else None
        return jsonify({'success': True, 'config': cfg})

    try:
        data = request.get_json()
        for key in ['system_name', 'language', 'timezone']:
            val = data.get(key)
            if val is None:
                continue
            c = Config.query.filter_by(key=f'general_{key}').first()
            if c:
                c.value = str(val)
            else:
                db.session.add(Config(key=f'general_{key}', value=str(val)))
        db.session.commit()
        log_audit('update_config_general', session['user_id'], 'config', None, 'Config general actualizada')
        return jsonify({'success': True, 'message': 'Configuración guardada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/email/company/<company_code>', methods=['GET', 'POST'])
def api_config_email_per_company(company_code):
    """Configuración SMTP independiente POR EMPRESA.
    GET: devuelve config actual de esa empresa.
    POST: guarda la config en Company.smtp_*.
    """
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    company_code = (company_code or '').lower().strip()
    if company_code not in admin_companies_scope():
        return jsonify({'success': False, 'error': f'No tienes acceso a la empresa "{company_code}"'}), 403

    c = Company.query.filter_by(code=company_code).first()
    if not c:
        return jsonify({'success': False, 'error': 'Empresa no encontrada'}), 404

    if request.method == 'GET':
        return jsonify({
            'success': True,
            'company_code': c.code,
            'company_name': c.name,
            'config': {
                'smtp_host': c.smtp_host or '',
                'smtp_port': c.smtp_port or 587,
                'smtp_user': c.smtp_user or '',
                'smtp_from': c.smtp_from or '',
                'smtp_security': c.smtp_security or 'tls',
                'has_password': bool(c.smtp_password),
            }
        })

    try:
        data = request.get_json() or {}
        if 'smtp_host' in data: c.smtp_host = (data.get('smtp_host') or '').strip() or None
        if 'smtp_port' in data:
            try: c.smtp_port = int(data.get('smtp_port') or 587)
            except (ValueError, TypeError): c.smtp_port = 587
        if 'smtp_user' in data: c.smtp_user = (data.get('smtp_user') or '').strip() or None
        if 'smtp_from' in data: c.smtp_from = (data.get('smtp_from') or '').strip() or None
        if 'smtp_security' in data:
            sec = (data.get('smtp_security') or '').strip().lower()
            c.smtp_security = sec if sec in ('tls', 'ssl') else None
        # Contraseña solo si se envía nueva (cifrada en BD)
        if data.get('smtp_password'):
            c.smtp_password = encrypt_secret(data['smtp_password'])
        if data.get('clear_password'):
            c.smtp_password = None

        db.session.commit()
        log_audit('update_email_config_company', session['user_id'], 'config', None,
                  f'SMTP por empresa actualizado: {company_code}')
        return jsonify({'success': True, 'message': f'SMTP de {c.name} guardado'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/email/company/<company_code>/test', methods=['POST'])
def api_config_email_per_company_test(company_code):
    """Prueba la conexión SMTP de una empresa específica."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401
    company_code = (company_code or '').lower().strip()
    if company_code not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso a esa empresa'}), 403

    cfg = _get_smtp_config(company=company_code)
    if cfg['source'] != f'company:{company_code}':
        return jsonify({'success': False, 'error': f'La empresa "{company_code}" no tiene SMTP propio configurado (usa "{cfg["source"]}").'}), 400
    if not cfg['user'] or not cfg['password']:
        return jsonify({'success': False, 'error': 'Falta usuario o contraseña SMTP'}), 400

    import smtplib, ssl, socket
    try:
        if cfg['security'] == 'ssl' or cfg['port'] == 465:
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(cfg['server'], cfg['port'], timeout=15, context=ctx) as s:
                s.ehlo(); s.login(cfg['user'], cfg['password'])
        else:
            with smtplib.SMTP(cfg['server'], cfg['port'], timeout=15) as s:
                s.ehlo(); s.starttls(); s.ehlo(); s.login(cfg['user'], cfg['password'])
        return jsonify({'success': True, 'message': f'✓ Conectado a {cfg["server"]}:{cfg["port"]} como {cfg["user"]}'})
    except smtplib.SMTPAuthenticationError as e:
        return jsonify({'success': False, 'error': f'Credenciales inválidas: {str(e)[:200]}'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'No se pudo conectar: {str(e)[:200]}'}), 400


@app.route('/api/config/email', methods=['GET', 'POST'])
def api_config_email():
    """Configuración SMTP de Office 365 + eventos de notificación (GLOBAL — para todas las empresas si no tienen una propia)"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    if request.method == 'GET':
        cfg = {}
        # Config SMTP
        for key in ['smtp_host', 'smtp_port', 'smtp_user', 'smtp_from', 'smtp_security', 'smtp_state']:
            c = Config.query.filter_by(key=f'email_{key}').first()
            cfg[key] = c.value if c else None
        # Defaults Office 365
        cfg['smtp_host'] = cfg.get('smtp_host') or 'smtp.office365.com'
        cfg['smtp_port'] = cfg.get('smtp_port') or '587'
        cfg['smtp_security'] = cfg.get('smtp_security') or 'tls'
        # Eventos
        for evt in ['ticket_created', 'ticket_assigned',
                    'sla_30', 'sla_60', 'sla_100', 'sla_overdue',
                    'ticket_comment', 'ticket_resolved', 'ticket_escalated', 'server_down']:
            c = Config.query.filter_by(key=f'email_evt_{evt}').first()
            cfg[f'evt_{evt}'] = (c.value == '1') if c else True  # default activos
        return jsonify({'success': True, 'config': cfg})

    try:
        data = request.get_json()

        # Guardar campos SMTP
        smtp_fields = ['smtp_host', 'smtp_port', 'smtp_user', 'smtp_from', 'smtp_security']
        for key in smtp_fields:
            if key in data:
                val = str(data[key]).strip()
                c = Config.query.filter_by(key=f'email_{key}').first()
                if c:
                    c.value = val
                else:
                    db.session.add(Config(key=f'email_{key}', value=val))

        # Contraseña SMTP (solo si se envía nueva)
        if 'smtp_pass' in data and data['smtp_pass']:
            c = Config.query.filter_by(key='email_smtp_pass').first()
            if c:
                c.value = data['smtp_pass']
            else:
                db.session.add(Config(key='email_smtp_pass', value=data['smtp_pass']))

        # Eventos (claves que empiezan con evt_)
        for key, val in data.items():
            if key.startswith('evt_'):
                evt_name = key[4:]
                c = Config.query.filter_by(key=f'email_evt_{evt_name}').first()
                value_str = '1' if val else '0'
                if c:
                    c.value = value_str
                else:
                    db.session.add(Config(key=f'email_evt_{evt_name}', value=value_str))

        db.session.commit()
        log_audit('update_email_config', session['user_id'], 'config', None, 'Config email actualizada')
        return jsonify({'success': True, 'message': 'Configuración guardada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/email/test-connection', methods=['POST'])
def api_email_test_connection():
    """Probar conexión SMTP sin enviar correo"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    try:
        # Leer config de BD
        host = Config.query.filter_by(key='email_smtp_host').first()
        port = Config.query.filter_by(key='email_smtp_port').first()
        user = Config.query.filter_by(key='email_smtp_user').first()
        pwd = Config.query.filter_by(key='email_smtp_pass').first()
        security = Config.query.filter_by(key='email_smtp_security').first()

        if not host or not user or not pwd:
            return jsonify({'success': False, 'error': 'Configuración incompleta. Guarda primero usuario y contraseña.'}), 400

        host_val = host.value
        port_val = int(port.value) if port else 587
        user_val = user.value
        pwd_val = pwd.value
        sec_val = security.value if security else 'tls'

        # Probar conexión
        import smtplib
        import socket
        try:
            if sec_val == 'ssl':
                server = smtplib.SMTP_SSL(host_val, port_val, timeout=10)
            else:
                server = smtplib.SMTP(host_val, port_val, timeout=10)
                if sec_val == 'tls':
                    server.starttls()
            server.login(user_val, pwd_val)
            server.quit()

            # Marcar estado como verificado
            state = Config.query.filter_by(key='email_smtp_state').first()
            if state:
                state.value = 'Verificado'
            else:
                db.session.add(Config(key='email_smtp_state', value='Verificado'))
            db.session.commit()

            return jsonify({
                'success': True,
                'message': f'Conectado a {host_val}:{port_val} como {user_val}'
            })
        except smtplib.SMTPAuthenticationError as e:
            err_text = str(e)
            # Diagnóstico específico según el código de error de Microsoft 365
            if '5.7.139' in err_text and 'did not meet the criteria' in err_text:
                hint = ('Microsoft 365 bloqueó la autenticación. Causa más probable: SMTP AUTH '
                        'deshabilitado en el buzón o tenant. Pídele a IT que ejecute en Exchange Online: '
                        'Set-CASMailbox -Identity ' + user_val + ' -SmtpClientAuthenticationDisabled $false')
            elif '5.7.139' in err_text and 'credentials were incorrect' in err_text:
                hint = ('Contraseña incorrecta. Si la cuenta tiene MFA, necesitás una App Password (16 chars). '
                        'Si NO tiene MFA, verificá que la contraseña ingresada sea exactamente la del buzón.')
            else:
                hint = 'Verificá usuario y contraseña. Si la cuenta tiene MFA, generá una App Password.'
            return jsonify({'success': False, 'error': f'{hint}\n\nDetalle SMTP: {err_text[:200]}'}), 400
        except (smtplib.SMTPException, socket.error, ConnectionError) as e:
            return jsonify({'success': False, 'error': f'No se pudo conectar a {host_val}:{port_val}: {str(e)[:120]}'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/email/test-send', methods=['POST'])
def api_email_test_send():
    """Enviar email de prueba"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    try:
        data = request.get_json()
        to_email = (data.get('to') or '').strip()
        if not to_email or '@' not in to_email:
            return jsonify({'success': False, 'error': 'Email destinatario inválido'}), 400

        # Leer config
        host = Config.query.filter_by(key='email_smtp_host').first()
        port = Config.query.filter_by(key='email_smtp_port').first()
        user = Config.query.filter_by(key='email_smtp_user').first()
        pwd = Config.query.filter_by(key='email_smtp_pass').first()
        sender = Config.query.filter_by(key='email_smtp_from').first()
        security = Config.query.filter_by(key='email_smtp_security').first()

        if not host or not user or not pwd:
            return jsonify({'success': False, 'error': 'Configuración SMTP incompleta'}), 400

        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        msg = MIMEMultipart('alternative')
        msg['Subject'] = '✅ DeskEli - Email de Prueba'
        msg['From'] = sender.value if sender else user.value
        msg['To'] = to_email

        html = f'''
        <html><body style="font-family: Arial, sans-serif; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px;">
                <h1 style="color: #2563eb;">🎫 DeskEli</h1>
                <p>Hola,</p>
                <p>Este es un <strong>email de prueba</strong> desde tu sistema DeskEli.</p>
                <p>Si recibes este mensaje, significa que la configuración SMTP con Office 365 está funcionando correctamente.</p>
                <hr>
                <p style="font-size: 12px; color: #666;">
                    Enviado por: {session.get('name', 'Admin')}<br>
                    Servidor: {host.value}:{port.value if port else 587}<br>
                    Empresa: {session.get('company', '')}<br>
                    Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}
                </p>
            </div>
        </body></html>
        '''
        msg.attach(MIMEText(html, 'html'))

        port_val = int(port.value) if port else 587
        sec_val = security.value if security else 'tls'

        if sec_val == 'ssl':
            server = smtplib.SMTP_SSL(host.value, port_val, timeout=15)
        else:
            server = smtplib.SMTP(host.value, port_val, timeout=15)
            if sec_val == 'tls':
                server.starttls()
        server.login(user.value, pwd.value)
        server.send_message(msg)
        server.quit()

        log_audit('email_test_sent', session['user_id'], 'email', None, f'Email de prueba enviado a {to_email}')
        return jsonify({'success': True, 'message': f'Email enviado a {to_email}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)[:200]}), 500


@app.route('/api/config/teams', methods=['GET', 'POST'])
def api_config_teams():
    """Configuración de webhooks de Microsoft Teams"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    if request.method == 'GET':
        cfg = {}
        for key in ['webhook_general', 'webhook_critical', 'webhook_servers']:
            c = Config.query.filter_by(key=f'teams_{key}').first()
            cfg[key] = c.value if c else ''
        for evt in ['ticket_created', 'ticket_critical', 'sla_warning', 'sla_expired', 'server_down', 'ticket_resolved']:
            c = Config.query.filter_by(key=f'teams_evt_{evt}').first()
            cfg[f'evt_{evt}'] = c.value == '1' if c else False
        return jsonify({'success': True, 'config': cfg})

    try:
        data = request.get_json()
        for key in ['webhook_general', 'webhook_critical', 'webhook_servers']:
            val = data.get(key, '').strip()
            c = Config.query.filter_by(key=f'teams_{key}').first()
            if c:
                c.value = val
            else:
                db.session.add(Config(key=f'teams_{key}', value=val))

        events = data.get('events', {})
        for evt, enabled in events.items():
            c = Config.query.filter_by(key=f'teams_evt_{evt}').first()
            val = '1' if enabled else '0'
            if c:
                c.value = val
            else:
                db.session.add(Config(key=f'teams_evt_{evt}', value=val))

        db.session.commit()
        log_audit('update_config_teams', session['user_id'], 'config', None, 'Config Teams webhooks actualizada')
        return jsonify({'success': True, 'message': 'Configuración Teams guardada'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/config/teams/test', methods=['POST'])
def api_config_teams_test():
    """Enviar mensaje de prueba a Teams"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False}), 401

    try:
        import urllib.request
        import json as _json

        data = request.get_json()
        webhook_url = data.get('webhook_url', '').strip()
        if not webhook_url or not webhook_url.startswith('https://'):
            return jsonify({'success': False, 'error': 'URL de webhook inválida'}), 400

        payload = {
            '@type': 'MessageCard',
            '@context': 'http://schema.org/extensions',
            'themeColor': '0078D7',
            'summary': 'Prueba DeskEli',
            'sections': [{
                'activityTitle': 'DeskEli - Prueba',
                'activitySubtitle': f'Empresa: {session["company"]}',
                'text': f'Mensaje de prueba enviado por {session.get("name", "Admin")} a las {datetime.now().strftime("%H:%M:%S")}'
            }]
        }
        req = urllib.request.Request(
            webhook_url,
            data=_json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'}
        )
        try:
            urllib.request.urlopen(req, timeout=5)
            return jsonify({'success': True, 'message': 'Mensaje enviado a Teams'})
        except Exception as e:
            return jsonify({'success': False, 'error': f'No se pudo conectar: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/monitor/escalations', methods=['GET'])
def api_monitor_escalations():
    """Monitor de escalaciones con filtro por período.
    Acepta ?period=day|week|month|year|all (default: all)
    Devuelve:
      - escalations: tickets con SLA crítico (>80%) AHORA, filtrados por created_at en el período
      - history: histórico de escalaciones (AgentAction escalator) del período
      - trend: serie temporal para gráfica (por día/semana según el rango)
    """
    if 'user_id' not in session or session['role'] not in ['technician', 'admin']:
        return jsonify({'success': False}), 401

    company = session['company']
    now = datetime.now()
    period = (request.args.get('period') or 'all').lower()

    # Definir rango temporal
    if period == 'day':
        period_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Hoy'
        bucket_format = '%H:00'   # bucket por hora
        bucket_delta_hours = 1
    elif period == 'week':
        period_start = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Últimos 7 días'
        bucket_format = '%d/%m'   # bucket por día
        bucket_delta_hours = 24
    elif period == 'month':
        period_start = (now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Últimos 30 días'
        bucket_format = '%d/%m'
        bucket_delta_hours = 24
    elif period == 'year':
        period_start = (now - timedelta(days=365)).replace(hour=0, minute=0, second=0, microsecond=0)
        period_label = 'Últimos 12 meses'
        bucket_format = '%Y-%m'   # bucket por mes
        bucket_delta_hours = 24 * 30
    else:  # all
        period_start = None
        period_label = 'Histórico completo'
        bucket_format = '%Y-%m'
        bucket_delta_hours = 24 * 30

    # 1. Escalaciones ACTIVAS (tickets activos con SLA crítico)
    q_active = Ticket.query.filter(
        Ticket.company == company,
        Ticket.status.in_(['open', 'in_progress'])
    )
    if period_start:
        q_active = q_active.filter(Ticket.created_at >= period_start)
    active_tickets = q_active.all()

    escalated = []
    for t in active_tickets:
        if not t.sla_deadline or not t.sla_minutes:
            continue
        total_minutes = t.sla_minutes
        elapsed = (now - t.created_at).total_seconds() / 60
        pct = (elapsed / total_minutes) * 100
        if pct >= 80:
            remaining_min = int((t.sla_deadline - now).total_seconds() / 60)
            escalated.append({
                'id': t.id,
                'ticket_number': t.ticket_number,
                'title': t.title,
                'priority': t.priority,
                'status': t.status,
                'assignee': t.assignee.name if t.assignee else 'Sin asignar',
                'sla_pct': round(pct, 1),
                'remaining_min': remaining_min,
                'is_expired': pct >= 100,
                'severity': 'critical' if pct >= 100 else ('warning' if pct >= 90 else 'attention'),
                'created_at': t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else ''
            })
    escalated.sort(key=lambda x: -x['sla_pct'])

    # 2. Histórico de escalaciones (AgentAction del agent 'escalator')
    q_hist = AgentAction.query.filter(
        AgentAction.company == company,
        AgentAction.agent_name == 'escalator'
    )
    if period_start:
        q_hist = q_hist.filter(AgentAction.created_at >= period_start)
    historical_actions = q_hist.order_by(AgentAction.created_at.desc()).all()

    history = []
    for a in historical_actions[:50]:  # limitar a 50 más recientes
        ticket = Ticket.query.get(a.ticket_id)
        reason = ''
        try:
            if a.output_data:
                import json as _json
                d = _json.loads(a.output_data)
                reason = d.get('reason', '')
        except Exception:
            pass
        history.append({
            'id': a.id,
            'ticket_id': a.ticket_id,
            'ticket_number': ticket.ticket_number if ticket else '?',
            'ticket_title': ticket.title if ticket else '(eliminado)',
            'ticket_status': ticket.status if ticket else '?',
            'reason': reason,
            'created_at': a.created_at.strftime('%Y-%m-%d %H:%M') if a.created_at else '',
            'confidence': a.confidence or 0,
            'success': bool(a.success),
        })

    # 3. Tendencia: contar escalaciones por bucket de tiempo
    trend_buckets = {}
    for a in historical_actions:
        if a.created_at:
            key = a.created_at.strftime(bucket_format)
            trend_buckets[key] = trend_buckets.get(key, 0) + 1
    trend = [{'label': k, 'count': v} for k, v in sorted(trend_buckets.items())]

    # 4. Stats por prioridad (de tickets que generaron escalaciones)
    by_priority = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    by_status = {'open': 0, 'in_progress': 0, 'resolved': 0, 'closed': 0}
    ticket_ids_seen = set()
    for a in historical_actions:
        if a.ticket_id in ticket_ids_seen:
            continue
        ticket_ids_seen.add(a.ticket_id)
        t = Ticket.query.get(a.ticket_id)
        if t:
            by_priority[t.priority or 'medium'] = by_priority.get(t.priority or 'medium', 0) + 1
            by_status[t.status or 'open'] = by_status.get(t.status or 'open', 0) + 1

    return jsonify({
        'success': True,
        'period': period,
        'period_label': period_label,
        'period_start': period_start.isoformat() if period_start else None,
        'escalations': escalated,
        'summary': {
            'total': len(escalated),
            'expired': len([e for e in escalated if e['is_expired']]),
            'warning': len([e for e in escalated if e['severity'] == 'warning']),
            'attention': len([e for e in escalated if e['severity'] == 'attention']),
            'historical_total': len(historical_actions),
            'by_priority': by_priority,
            'by_status': by_status,
        },
        'history': history,
        'trend': trend,
    })

# ═════════════════════════════════════════════════════════════════════════════
# API PÚBLICA v1 - Para integraciones externas (proveedores, sistemas legacy)
# Autenticación por Bearer token (API Key generada en el panel admin)
# ═════════════════════════════════════════════════════════════════════════════

def _validate_api_key():
    """Valida el header 'X-Authorization: Bearer <token>' (o Authorization).
    Devuelve (api_key, None) si es válido, o (None, error_response) si no."""
    auth = request.headers.get('X-Authorization') or request.headers.get('Authorization') or ''
    if not auth.lower().startswith('bearer '):
        return None, (jsonify({'success': False, 'error': 'Falta header X-Authorization: Bearer <token>'}), 401)
    token = auth[7:].strip()
    if not token:
        return None, (jsonify({'success': False, 'error': 'Token vacío'}), 401)

    token_hash = hashlib.sha256(token.encode()).hexdigest()
    api_key = ApiKey.query.filter_by(token_hash=token_hash, is_active=True).first()
    if not api_key:
        return None, (jsonify({'success': False, 'error': 'Token inválido o revocado'}), 401)

    if api_key.expires_at and api_key.expires_at < datetime.now():
        return None, (jsonify({'success': False, 'error': 'Token expirado'}), 401)

    # Registrar uso
    api_key.last_used_at = datetime.now()
    api_key.last_used_ip = request.remote_addr
    api_key.usage_count = (api_key.usage_count or 0) + 1
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    return api_key, None


def _api_key_has_scope(api_key, scope):
    """True si la API key tiene el scope solicitado."""
    scopes = [s.strip() for s in (api_key.scopes or '').split(',') if s.strip()]
    return scope in scopes or '*' in scopes


@app.route('/api/v1/external/tickets', methods=['POST'])
def api_v1_external_create_ticket():
    """Crear un ticket desde una integración externa, con subtareas (controls) y adjuntos.

    Auth: header 'X-Authorization: Bearer <TOKEN>' (o 'Authorization').

    Body JSON esperado:
    {
        "subject": "Título del ticket",           (obligatorio)
        "description": "Descripción del caso",     (obligatorio)
        "category": "Software",                    (opcional, default: "General")
        "priority": "high",                        (opcional: low/medium/high/critical)

        "applicantEmail": "juan@empresa.com",      (obligatorio — o applicantId)
        "applicantId": 15192,
        "authorId": 10376,
        "assigneeEmail": "tecnico@empresa.com",
        "assigneeId": 12,
        "userArea": "Contabilidad",
        "userLocation": "Piso 3",
        "userPhone": "555-1234",
        "externalRef": "SR-98765",
        "listAdditionalField": [...],

        // Subtareas (aka "controls" para compatibilidad con la plataforma origen).
        // Cada una se crea como una Subtask del ticket padre.
        "subtasks": [
            {
                "title": "Control 1: Validar backups",  (obligatorio)
                "description": "Detalle del control",    (opcional)
                "priority": "medium",                    (opcional, hereda del padre)
                "category": "SAP",                       (opcional)
                "assigneeEmail": "tec@empresa.com"       (opcional, técnico asignado)
            }
        ],
        // Alias aceptado: "controls" (para compat con Aranda/otras plataformas)

        // Adjuntos codificados en base64. Van al ticket, subtareas, o ambos.
        "attachments": [
            {
                "filename": "acta_aprobada.pdf",         (obligatorio)
                "content_base64": "JVBERi0xLjMK...",     (obligatorio, bytes en base64)
                "mime": "application/pdf",               (opcional, se infiere)
                "attach_to": "both"                      (opcional: "ticket" | "subtasks" | "both", default "both")
            }
        ]
    }

    Respuesta 201:
    {
        "success": true,
        "id": 42,
        "ticket_number": "TKT-ELIOT-00042",
        "url": "https://.../technician/ticket/42",
        "subtasks": [
            {"id": 12, "subtask_number": "TKT-ELIOT-00042-S01", "title": "Control 1"},
            ...
        ],
        "attachments": {
            "ticket": 1,       // cantidad guardada en el ticket padre
            "subtasks": 3      // cantidad de registros en subtareas
        }
    }
    """
    api_key, err = _validate_api_key()
    if err:
        return err
    if not _api_key_has_scope(api_key, 'tickets:create'):
        return jsonify({'success': False, 'error': 'Token sin scope tickets:create'}), 403

    data = request.get_json(silent=True) or {}
    subject = (data.get('subject') or data.get('title') or '').strip()[:200]
    description = (data.get('description') or '').strip()
    if not subject or not description:
        return jsonify({'success': False, 'error': 'Faltan campos requeridos: subject y description'}), 400

    # Sanitizar (los caller externos son menos confiables)
    description = sanitize_html(description) if 'sanitize_html' in globals() else description

    # Resolver solicitante
    applicant = None
    applicant_email = (data.get('applicantEmail') or data.get('applicant_email') or '').strip().lower()
    applicant_id_in = data.get('applicantId') or data.get('applicant_id')
    if applicant_email:
        applicant = User.query.filter(
            db.func.lower(User.email) == applicant_email,
            User.company == api_key.company
        ).first()
    if not applicant and applicant_id_in:
        applicant = User.query.get(applicant_id_in)
    if not applicant:
        return jsonify({'success': False, 'error': 'Solicitante no encontrado. Envía applicantEmail o applicantId válido.'}), 400
    if applicant.company != api_key.company:
        return jsonify({'success': False, 'error': f'Solicitante no pertenece a la empresa {api_key.company}'}), 403

    # Resolver autor (quién lo crea)
    author = None
    author_id_in = data.get('authorId') or data.get('author_id')
    if author_id_in:
        author = User.query.get(author_id_in)
    if not author or author.company != api_key.company:
        author = applicant  # fallback

    # Resolver asignatario opcional
    assignee = None
    assignee_email = (data.get('assigneeEmail') or data.get('assignee_email') or '').strip().lower()
    assignee_id_in = data.get('assigneeId') or data.get('assignee_id')
    if assignee_email:
        assignee = User.query.filter(
            db.func.lower(User.email) == assignee_email,
            User.company == api_key.company,
            User.role.in_(['technician', 'admin'])
        ).first()
    if not assignee and assignee_id_in:
        assignee = User.query.get(assignee_id_in)
        if assignee and assignee.company != api_key.company:
            assignee = None

    # Categoría y prioridad
    category = (data.get('category') or 'General').strip()[:100]
    priority = (data.get('priority') or 'medium').lower().strip()
    if priority not in ('low', 'medium', 'high', 'critical'):
        priority = 'medium'

    # Contacto del solicitante
    user_area = (data.get('userArea') or data.get('user_area') or applicant.area or '').strip()[:120] or None
    user_location = (data.get('userLocation') or data.get('user_location') or applicant.location or '').strip()[:120] or None
    user_phone = (data.get('userPhone') or data.get('user_phone') or applicant.phone or '').strip()[:40] or None

    # Referencia externa (Aranda style) y campos adicionales
    external_ref = (data.get('externalRef') or data.get('external_ref') or '').strip()[:100]
    additional_fields = data.get('listAdditionalField') or data.get('additional_fields') or []
    if external_ref or additional_fields:
        meta_lines = []
        if external_ref:
            meta_lines.append(f'📌 Ref. externa: {external_ref}')
        if additional_fields and isinstance(additional_fields, list):
            meta_lines.append('📎 Campos adicionales:')
            for f in additional_fields[:20]:
                if not isinstance(f, dict):
                    continue
                name = f.get('name') or f.get('fieldName') or f'field_{f.get("fieldId")}'
                value = f.get('stringValue') or f.get('intValue') or f.get('boolValue')
                meta_lines.append(f'   - {name}: {value}')
        description = description + '\n\n---\n' + '\n'.join(meta_lines)

    # SLA en minutos según prioridad
    sla_map = {'critical': 120, 'high': 240, 'medium': 480, 'low': 1440}
    sla_minutes = sla_map.get(priority, 480)

    # Generar número de ticket
    ticket_number = generate_ticket_number(api_key.company) if 'generate_ticket_number' in globals() else \
                    f'API-{api_key.company.upper()}-{int(time.time())}'

    try:
        ticket = Ticket(
            ticket_number=ticket_number,
            title=subject,
            description=description,
            category=category,
            priority=priority,
            status='open',
            company=api_key.company,
            creator_id=author.id,
            assignee_id=assignee.id if assignee else None,
            sla_minutes=sla_minutes,
            sla_deadline=datetime.now() + timedelta(minutes=sla_minutes),
            user_area=user_area,
            user_location=user_location,
            user_phone=user_phone,
        )
        db.session.add(ticket)
        db.session.commit()

        log_audit('api_ticket_created', author.id if author else None, 'ticket', ticket.id,
                  f'Ticket #{ticket.ticket_number} creado por API key "{api_key.name}" desde {request.remote_addr}. '
                  f'Solicitante: {applicant.email}. Ref externa: {external_ref or "N/A"}.')

        # ─── SUBTAREAS ─────────────────────────────────────────────────
        # Prioridad: primero se resuelve por guion (guion_id / guion_code),
        # y si no viene, se usan los subtasks/controls del payload.
        subtasks_created = []
        source_guion = None

        # Opción 1: usar un guion preconfigurado
        guion_id_in = data.get('guion_id') or data.get('guionId')
        guion_code_in = (data.get('guion_code') or data.get('guionCode') or '').strip().lower()
        if guion_id_in or guion_code_in:
            gq = Guion.query
            if guion_id_in:
                gq = gq.filter_by(id=guion_id_in)
            else:
                gq = gq.filter_by(code=guion_code_in)
            source_guion = gq.filter_by(company=api_key.company, is_active=True).first()
            if not source_guion:
                # Si el proveedor mandó guion pero no existe/está inactivo → hard error para evitar
                # que se creen tickets huerfanos sin las subtareas esperadas
                db.session.delete(ticket)
                db.session.commit()
                return jsonify({
                    'success': False,
                    'error': f'Guión no encontrado o inactivo: {guion_code_in or guion_id_in} (empresa {api_key.company})'
                }), 400

        if source_guion:
            gs_list = GuionSubtask.query.filter_by(guion_id=source_guion.id).order_by(GuionSubtask.order_idx).all()

            # Fallback pool: especialistas asignados a este guion desde Gestión de Usuarios.
            # Se usan (round-robin por carga) cuando una subtarea no tiene assignee_id fijo.
            pool_ids = [ug.user_id for ug in UserGuion.query.filter_by(guion_id=source_guion.id).all()]
            pool_users = User.query.filter(
                User.id.in_(pool_ids),
                User.company == source_guion.company,
                User.is_active == True,
                User.role.in_(['technician', 'admin'])
            ).all() if pool_ids else []
            pool_load = {}
            for u in pool_users:
                pool_load[u.id] = Subtask.query.filter(
                    Subtask.assignee_id == u.id,
                    Subtask.status.in_(['open', 'in_progress'])
                ).count()

            def pick_from_pool():
                if not pool_load:
                    return None
                least_id = min(pool_load, key=pool_load.get)
                pool_load[least_id] += 1  # simular la carga que agrega esta subtarea
                return least_id

            for idx, gs in enumerate(gs_list):
                st_priority = (gs.priority or priority).lower()
                if st_priority not in ('low', 'medium', 'high', 'critical'):
                    st_priority = 'medium'
                sla_min = sla_map.get(st_priority, 480)
                # 1º prioridad: técnico fijo de la subtarea del guion. 2º: pool de especialistas del guion.
                resolved_assignee = gs.assignee_id or pick_from_pool()
                st = Subtask(
                    ticket_id=ticket.id,
                    subtask_number=f'{ticket.ticket_number}-S{(idx+1):02d}',
                    title=gs.title[:255],
                    description=gs.description,
                    category=gs.category or category,
                    status='open',
                    priority=st_priority,
                    sla_minutes=sla_min,
                    sla_deadline=datetime.now() + timedelta(minutes=sla_min),
                    assignee_id=resolved_assignee,
                    created_by_id=author.id,
                    order_idx=idx,
                )
                db.session.add(st)
                subtasks_created.append(st)
            if subtasks_created:
                db.session.commit()
                log_audit('api_subtasks_from_guion', author.id if author else None, 'ticket', ticket.id,
                          f'{len(subtasks_created)} subtareas creadas desde guion "{source_guion.code}" ({source_guion.name}) '
                          f'para ticket {ticket.ticket_number}')
        else:
            # Opción 2 (fallback / uso ad-hoc): subtareas del payload
            subtasks_raw = data.get('subtasks') or data.get('controls') or []
            if isinstance(subtasks_raw, list):
                for idx, st_data in enumerate(subtasks_raw[:50]):
                    if not isinstance(st_data, dict):
                        continue
                    st_title = (st_data.get('title') or st_data.get('name') or '').strip()
                    if not st_title:
                        continue
                    st_desc = sanitize_html((st_data.get('description') or '').strip()) if 'sanitize_html' in globals() else (st_data.get('description') or '').strip()
                    st_priority = (st_data.get('priority') or priority).lower().strip()
                    if st_priority not in ('low', 'medium', 'high', 'critical'):
                        st_priority = priority
                    st_category = (st_data.get('category') or category).strip()[:100]
                    st_assignee = None
                    st_ass_email = (st_data.get('assigneeEmail') or st_data.get('assignee_email') or '').strip().lower()
                    if st_ass_email:
                        st_assignee = User.query.filter(
                            db.func.lower(User.email) == st_ass_email,
                            User.company == api_key.company,
                            User.role.in_(['technician', 'admin'])
                        ).first()
                    if not st_assignee:
                        st_assignee = assignee
                    st_num = f'{ticket.ticket_number}-S{(idx+1):02d}'
                    sla_min = sla_map.get(st_priority, 480)
                    st = Subtask(
                        ticket_id=ticket.id,
                        subtask_number=st_num,
                        title=st_title[:255],
                        description=st_desc or None,
                        category=st_category,
                        status='open',
                        priority=st_priority,
                        sla_minutes=sla_min,
                        sla_deadline=datetime.now() + timedelta(minutes=sla_min),
                        assignee_id=st_assignee.id if st_assignee else None,
                        created_by_id=author.id,
                        order_idx=idx,
                    )
                    db.session.add(st)
                    subtasks_created.append(st)
                if subtasks_created:
                    db.session.commit()
                    log_audit('api_subtasks_created', author.id if author else None, 'ticket', ticket.id,
                              f'{len(subtasks_created)} subtareas ad-hoc creadas via API para ticket {ticket.ticket_number}')

        # ─── ADJUNTOS (base64) ────────────────────────────────────────
        import base64 as _b64
        from werkzeug.utils import secure_filename as _sec
        attachments_raw = data.get('attachments') or []
        att_ticket_count = 0
        att_subtask_count = 0
        # Directorio destino
        ticket_upload_dir = app.config.get('TICKET_UPLOAD_FOLDER', 'uploads/tickets')
        subtask_upload_dir = app.config.get('UPLOAD_FOLDER', 'uploads/subtasks')
        os.makedirs(ticket_upload_dir, exist_ok=True)
        os.makedirs(subtask_upload_dir, exist_ok=True)

        if isinstance(attachments_raw, list):
            for att_data in attachments_raw[:20]:  # límite 20 por request
                if not isinstance(att_data, dict):
                    continue
                filename = (att_data.get('filename') or 'archivo').strip()[:255]
                b64_content = att_data.get('content_base64') or att_data.get('content') or ''
                mime = (att_data.get('mime') or att_data.get('mime_type') or '').strip()[:120] or 'application/octet-stream'
                attach_to = (att_data.get('attach_to') or 'both').lower().strip()
                if attach_to not in ('ticket', 'subtasks', 'both'):
                    attach_to = 'both'
                if not b64_content:
                    continue
                # Chequeo permitido
                if '_allowed_attachment' in globals() and not _allowed_attachment(filename):
                    continue
                # Decodificar
                try:
                    raw_bytes = _b64.b64decode(b64_content, validate=False)
                except Exception:
                    continue
                if len(raw_bytes) > 50 * 1024 * 1024:  # 50 MB cap
                    continue

                # Guardar 1 sola vez en disco (para eficiencia)
                safe = _sec(filename) or 'archivo'
                ext = safe.rsplit('.', 1)[1].lower() if '.' in safe else ''
                stored = f"{uuid.uuid4().hex}.{ext}" if ext else uuid.uuid4().hex

                # Guardar en la carpeta de ticket attachments; ambas rutas de descarga apuntan al mismo archivo
                target_path = os.path.join(ticket_upload_dir, stored)
                with open(target_path, 'wb') as fh:
                    fh.write(raw_bytes)

                size = len(raw_bytes)

                # Registro en TicketAttachment
                if attach_to in ('ticket', 'both'):
                    db.session.add(TicketAttachment(
                        ticket_id=ticket.id,
                        original_name=filename,
                        stored_name=stored,
                        mime_type=mime,
                        size_bytes=size,
                        uploaded_by_id=author.id,
                    ))
                    att_ticket_count += 1

                # Registro en cada Subtask (SubtaskAttachment)
                if attach_to in ('subtasks', 'both') and subtasks_created:
                    # Duplicar archivo físico en la carpeta de subtasks para que api_subtask_attachment_download lo encuentre
                    subtask_stored = stored  # reuso el mismo nombre
                    subtask_path = os.path.join(subtask_upload_dir, subtask_stored)
                    if not os.path.exists(subtask_path):
                        try:
                            import shutil as _sh
                            _sh.copyfile(target_path, subtask_path)
                        except Exception:
                            pass
                    for st in subtasks_created:
                        db.session.add(SubtaskAttachment(
                            subtask_id=st.id,
                            original_name=filename,
                            stored_name=subtask_stored,
                            mime_type=mime,
                            size_bytes=size,
                            uploaded_by_id=author.id,
                        ))
                        att_subtask_count += 1

            if att_ticket_count or att_subtask_count:
                db.session.commit()
                log_audit('api_attachments_saved', author.id if author else None, 'ticket', ticket.id,
                          f'API adjuntos: {att_ticket_count} en ticket, {att_subtask_count} en subtareas ({ticket.ticket_number})')

        # Emitir websocket a los técnicos de la empresa
        try:
            socketio.emit('ticket_created', {
                'id': ticket.id,
                'ticket_number': ticket.ticket_number,
                'title': ticket.title,
                'priority': ticket.priority,
                'company': ticket.company,
                'source': 'api_external',
                'subtasks_count': len(subtasks_created),
            }, room=f'company_{api_key.company}')
        except Exception:
            pass

        base_url = get_public_base_url()
        return jsonify({
            'success': True,
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'url': f'{base_url}/technician/ticket/{ticket.id}',
            'status': ticket.status,
            'created_at': ticket.created_at.isoformat(timespec='seconds'),
            'sla_deadline': ticket.sla_deadline.isoformat(timespec='seconds') if ticket.sla_deadline else None,
            'subtasks': [{
                'id': st.id,
                'subtask_number': st.subtask_number,
                'title': st.title,
                'priority': st.priority,
                'assignee_id': st.assignee_id,
            } for st in subtasks_created],
            'attachments': {
                'ticket': att_ticket_count,
                'subtasks': att_subtask_count,
            },
        }), 201
    except Exception as e:
        db.session.rollback()
        log_audit('api_ticket_failed', None, 'ticket', None,
                  f'API key "{api_key.name}" falló al crear ticket: {e}')
        return jsonify({'success': False, 'error': f'Error al crear ticket: {str(e)}'}), 500


@app.route('/api/v1/external/tickets/<int:ticket_id>', methods=['GET'])
def api_v1_external_get_ticket(ticket_id):
    """Consultar estado de un ticket. Auth por Bearer token, scope 'tickets:read'."""
    api_key, err = _validate_api_key()
    if err:
        return err
    if not _api_key_has_scope(api_key, 'tickets:read'):
        return jsonify({'success': False, 'error': 'Token sin scope tickets:read'}), 403

    ticket = Ticket.query.get(ticket_id)
    if not ticket or ticket.company != api_key.company:
        return jsonify({'success': False, 'error': 'Ticket no encontrado'}), 404

    assignee = User.query.get(ticket.assignee_id) if ticket.assignee_id else None
    creator = User.query.get(ticket.creator_id) if ticket.creator_id else None

    return jsonify({
        'success': True,
        'id': ticket.id,
        'ticket_number': ticket.ticket_number,
        'title': ticket.title,
        'description': ticket.description,
        'category': ticket.category,
        'priority': ticket.priority,
        'status': ticket.status,
        'company': ticket.company,
        'creator': {'id': creator.id, 'name': creator.name, 'email': creator.email} if creator else None,
        'assignee': {'id': assignee.id, 'name': assignee.name, 'email': assignee.email} if assignee else None,
        'created_at': ticket.created_at.isoformat(timespec='seconds') if ticket.created_at else None,
        'updated_at': ticket.updated_at.isoformat(timespec='seconds') if ticket.updated_at else None,
        'resolved_at': ticket.resolved_at.isoformat(timespec='seconds') if ticket.resolved_at else None,
        'sla_deadline': ticket.sla_deadline.isoformat(timespec='seconds') if ticket.sla_deadline else None,
    })


# ═════════════════════════════════════════════════════════════════════════════
# GUIONES — Plantillas de subtareas invocables desde la API externa
# ═════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/guiones', methods=['GET'])
def api_admin_guiones_list():
    """Lista todos los guiones de la empresa del admin."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    scope = admin_companies_scope()
    guiones = Guion.query.filter(Guion.company.in_(scope)).order_by(Guion.company, Guion.name).all()
    result = []
    for g in guiones:
        count = GuionSubtask.query.filter_by(guion_id=g.id).count()
        result.append({
            'id': g.id,
            'code': g.code,
            'name': g.name,
            'description': g.description or '',
            'company': g.company,
            'default_priority': g.default_priority,
            'default_category': g.default_category,
            'is_active': bool(g.is_active),
            'subtasks_count': count,
            'created_at': g.created_at.isoformat(timespec='seconds') if g.created_at else None,
        })
    return jsonify({'success': True, 'guiones': result})


@app.route('/api/admin/guiones', methods=['POST'])
def api_admin_guiones_create():
    """Crea un guión nuevo."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    data = request.get_json() or {}
    code = (data.get('code') or '').strip().lower()
    name = (data.get('name') or '').strip()
    company = (data.get('company') or session.get('company') or '').strip()

    if not code or not re.match(r'^[a-z][a-z0-9_-]{1,49}$', code):
        return jsonify({'success': False, 'error': 'Código inválido: minúsculas/números/-/_, empieza con letra, 2-50 chars'}), 400
    if not name:
        return jsonify({'success': False, 'error': 'Nombre requerido'}), 400
    if company not in admin_companies_scope():
        return jsonify({'success': False, 'error': f'Sin permiso sobre empresa {company}'}), 403
    if Guion.query.filter_by(code=code).first():
        return jsonify({'success': False, 'error': f'Ya existe un guión con código "{code}"'}), 400

    g = Guion(
        code=code,
        name=name[:200],
        description=(data.get('description') or '').strip() or None,
        company=company,
        default_priority=(data.get('default_priority') or 'medium').lower(),
        default_category=(data.get('default_category') or 'General').strip()[:100],
        is_active=bool(data.get('is_active', True)),
        created_by_id=session['user_id'],
    )
    db.session.add(g)
    db.session.commit()
    log_audit('guion_create', session['user_id'], 'guion', g.id, f'Guion "{g.name}" (code={g.code}) creado para {g.company}')
    return jsonify({'success': True, 'id': g.id, 'code': g.code}), 201


@app.route('/api/admin/guiones/<int:guion_id>', methods=['GET'])
def api_admin_guion_get(guion_id):
    """Devuelve un guión con sus subtareas."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    g = Guion.query.get_or_404(guion_id)
    if g.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    subs = GuionSubtask.query.filter_by(guion_id=g.id).order_by(GuionSubtask.order_idx).all()
    return jsonify({
        'success': True,
        'guion': {
            'id': g.id, 'code': g.code, 'name': g.name, 'description': g.description or '',
            'company': g.company, 'default_priority': g.default_priority,
            'default_category': g.default_category, 'is_active': bool(g.is_active),
            'subtasks': [{
                'id': s.id,
                'order_idx': s.order_idx,
                'title': s.title,
                'description': s.description or '',
                'category': s.category or '',
                'priority': s.priority or 'medium',
                'assignee_id': s.assignee_id,
                'assignee_name': (s.assignee.name if s.assignee else None),
                'assignee_email': (s.assignee.email if s.assignee else None),
            } for s in subs]
        }
    })


@app.route('/api/admin/guiones/<int:guion_id>', methods=['PUT'])
def api_admin_guion_update(guion_id):
    """Actualiza metadatos del guión (no las subtareas)."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    g = Guion.query.get_or_404(guion_id)
    if g.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    data = request.get_json() or {}
    if 'name' in data: g.name = (data['name'] or '').strip()[:200] or g.name
    if 'description' in data: g.description = (data['description'] or '').strip() or None
    if 'default_priority' in data:
        p = (data['default_priority'] or 'medium').lower()
        if p in ('low', 'medium', 'high', 'critical'):
            g.default_priority = p
    if 'default_category' in data: g.default_category = (data['default_category'] or 'General').strip()[:100]
    if 'is_active' in data: g.is_active = bool(data['is_active'])
    db.session.commit()
    log_audit('guion_update', session['user_id'], 'guion', g.id, f'Guion "{g.name}" actualizado')
    return jsonify({'success': True})


@app.route('/api/admin/guiones/<int:guion_id>', methods=['DELETE'])
def api_admin_guion_delete(guion_id):
    """Elimina un guión y todas sus subtareas asociadas."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    g = Guion.query.get_or_404(guion_id)
    if g.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    name, code = g.name, g.code
    db.session.delete(g)
    db.session.commit()
    log_audit('guion_delete', session['user_id'], 'guion', guion_id, f'Guion "{name}" (code={code}) eliminado')
    return jsonify({'success': True})


@app.route('/api/admin/guiones/<int:guion_id>/subtasks', methods=['POST'])
def api_admin_guion_subtask_create(guion_id):
    """Agrega una subtarea preconfigurada al guión."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    g = Guion.query.get_or_404(guion_id)
    if g.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({'success': False, 'error': 'Título requerido'}), 400
    priority = (data.get('priority') or g.default_priority or 'medium').lower()
    if priority not in ('low', 'medium', 'high', 'critical'):
        priority = 'medium'
    # Resolver assignee por ID o email
    assignee_id = data.get('assignee_id')
    if not assignee_id:
        ass_email = (data.get('assignee_email') or '').strip().lower()
        if ass_email:
            u = User.query.filter(
                db.func.lower(User.email) == ass_email,
                User.company == g.company,
                User.role.in_(['technician', 'admin'])
            ).first()
            assignee_id = u.id if u else None

    # Order: siguiente disponible
    last_order = db.session.query(db.func.coalesce(db.func.max(GuionSubtask.order_idx), -1)).filter_by(guion_id=g.id).scalar()

    gs = GuionSubtask(
        guion_id=g.id,
        order_idx=(last_order or 0) + 1,
        title=title[:255],
        description=(data.get('description') or '').strip() or None,
        category=(data.get('category') or g.default_category or '').strip()[:100] or None,
        priority=priority,
        assignee_id=assignee_id,
    )
    db.session.add(gs)
    db.session.commit()
    log_audit('guion_subtask_create', session['user_id'], 'guion', g.id, f'Subtarea "{title}" agregada al guion {g.code}')
    return jsonify({'success': True, 'id': gs.id}), 201


@app.route('/api/admin/guion-subtasks/<int:subtask_id>', methods=['PUT'])
def api_admin_guion_subtask_update(subtask_id):
    """Actualiza una subtarea del guión."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    gs = GuionSubtask.query.get_or_404(subtask_id)
    g = Guion.query.get(gs.guion_id)
    if not g or g.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    data = request.get_json() or {}
    if 'title' in data: gs.title = (data['title'] or '').strip()[:255] or gs.title
    if 'description' in data: gs.description = (data['description'] or '').strip() or None
    if 'category' in data: gs.category = (data['category'] or '').strip()[:100] or None
    if 'priority' in data:
        p = (data['priority'] or 'medium').lower()
        if p in ('low', 'medium', 'high', 'critical'):
            gs.priority = p
    if 'order_idx' in data:
        try: gs.order_idx = int(data['order_idx'])
        except: pass
    if 'assignee_id' in data:
        gs.assignee_id = data['assignee_id'] or None
    elif 'assignee_email' in data:
        ass_email = (data.get('assignee_email') or '').strip().lower()
        if ass_email:
            u = User.query.filter(
                db.func.lower(User.email) == ass_email,
                User.company == g.company,
                User.role.in_(['technician', 'admin'])
            ).first()
            gs.assignee_id = u.id if u else None
        else:
            gs.assignee_id = None
    db.session.commit()
    log_audit('guion_subtask_update', session['user_id'], 'guion', g.id, f'Subtarea "{gs.title}" del guion {g.code} actualizada')
    return jsonify({'success': True})


@app.route('/api/admin/guion-subtasks/<int:subtask_id>', methods=['DELETE'])
def api_admin_guion_subtask_delete(subtask_id):
    """Elimina una subtarea del guión."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    gs = GuionSubtask.query.get_or_404(subtask_id)
    g = Guion.query.get(gs.guion_id)
    if not g or g.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    title = gs.title
    db.session.delete(gs)
    db.session.commit()
    log_audit('guion_subtask_delete', session['user_id'], 'guion', g.id, f'Subtarea "{title}" del guion {g.code} eliminada')
    return jsonify({'success': True})


# ─── Reporte de apertura de DevTools desde el portal empleado (audit log) ───

@app.route('/api/security/devtools-detected', methods=['POST'])
def api_security_devtools_detected():
    """Registra en audit log cuando el detector de DevTools del portal
    empleado se dispara. Barrera cosmética — util para detectar patrones."""
    if 'user_id' not in session:
        return jsonify({'success': False}), 401
    try:
        data = request.get_json(silent=True) or {}
        url = (data.get('url') or '')[:200]
        ua = (data.get('ua') or '')[:255]
        log_audit(
            'devtools_detected',
            session['user_id'],
            'session',
            session['user_id'],
            f'DevTools abierto en portal empleado. URL={url}. UA={ua}'
        )
    except Exception:
        pass
    return jsonify({'success': True})


# ─── Asignación de guiones a especialistas (M:N desde Gestión de Usuarios) ───

@app.route('/api/admin/users/<int:user_id>/guiones', methods=['GET'])
def api_user_guiones_get(user_id):
    """Guiones asignados a un especialista + catálogo de guiones disponibles de su empresa."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    user = User.query.get_or_404(user_id)
    if user.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403

    assigned = UserGuion.query.filter_by(user_id=user_id).all()
    assigned_ids = {a.guion_id for a in assigned}
    catalog = Guion.query.filter_by(company=user.company, is_active=True).order_by(Guion.name).all()

    return jsonify({
        'success': True,
        'user': {'id': user.id, 'name': user.name, 'role': user.role, 'company': user.company},
        'assigned_guion_ids': list(assigned_ids),
        'catalog': [{
            'id': g.id,
            'code': g.code,
            'name': g.name,
            'description': g.description or '',
            'default_priority': g.default_priority,
            'default_category': g.default_category,
            'subtask_count': GuionSubtask.query.filter_by(guion_id=g.id).count(),
        } for g in catalog]
    })


@app.route('/api/admin/users/<int:user_id>/guiones', methods=['POST'])
def api_user_guiones_set(user_id):
    """Reemplaza la lista de guiones asignados al usuario.
    Body: {guion_ids: [1, 2, 3]}"""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    user = User.query.get_or_404(user_id)
    if user.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'Sin acceso'}), 403
    if user.role not in ('technician', 'admin'):
        return jsonify({'success': False, 'error': 'Solo se pueden asignar guiones a técnicos o administradores'}), 400

    data = request.get_json() or {}
    try:
        new_ids = [int(x) for x in (data.get('guion_ids') or [])]
    except (ValueError, TypeError):
        return jsonify({'success': False, 'error': 'guion_ids debe ser una lista de IDs enteros'}), 400

    UserGuion.query.filter_by(user_id=user_id).delete()
    added = 0
    for gid in new_ids:
        g = Guion.query.get(gid)
        if not g or not g.is_active:
            continue
        if g.company != user.company:
            continue
        db.session.add(UserGuion(user_id=user_id, guion_id=gid))
        added += 1
    db.session.commit()
    log_audit('set_user_guiones', session['user_id'], 'user', user_id, f'{added} guiones asignados a {user.name}')
    return jsonify({'success': True, 'count': added, 'message': f'{added} guiones asignados'})


# ─── Admin endpoints para gestionar API Keys ─────────────────────────────────

@app.route('/api/admin/api-keys', methods=['GET'])
def api_admin_api_keys_list():
    """Listar API Keys de la empresa del admin."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    scope = admin_companies_scope()
    keys = ApiKey.query.filter(ApiKey.company.in_(scope)).order_by(ApiKey.created_at.desc()).all()
    return jsonify({
        'success': True,
        'api_keys': [{
            'id': k.id,
            'name': k.name,
            'token_prefix': k.token_prefix,
            'company': k.company,
            'scopes': (k.scopes or '').split(','),
            'is_active': bool(k.is_active),
            'created_at': k.created_at.isoformat(timespec='seconds') if k.created_at else None,
            'last_used_at': k.last_used_at.isoformat(timespec='seconds') if k.last_used_at else None,
            'last_used_ip': k.last_used_ip,
            'expires_at': k.expires_at.isoformat(timespec='seconds') if k.expires_at else None,
            'usage_count': k.usage_count or 0,
        } for k in keys]
    })


@app.route('/api/admin/api-keys', methods=['POST'])
def api_admin_api_keys_create():
    """Crea una nueva API Key. Devuelve el token EN CLARO SOLO ESTA VEZ."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()[:100]
    company = (data.get('company') or session.get('company') or '').strip()
    scopes_list = data.get('scopes') or ['tickets:create', 'tickets:read']
    if isinstance(scopes_list, list):
        scopes = ','.join(s.strip() for s in scopes_list if s.strip())
    else:
        scopes = str(scopes_list)
    expires_at = None
    if data.get('expires_at'):
        try:
            expires_at = datetime.fromisoformat(data['expires_at'])
        except Exception:
            pass

    if not name:
        return jsonify({'success': False, 'error': 'name es obligatorio'}), 400
    if company not in admin_companies_scope():
        return jsonify({'success': False, 'error': f'Sin permiso sobre la empresa {company}'}), 403

    # Generar token: prefijo reconocible + 40 chars aleatorios seguros
    prefix = 'dsk_' + ('t' if 'tickets:create' in scopes else 'r')
    random_part = secrets.token_urlsafe(32)
    token = f'{prefix}_{random_part}'
    token_hash = hashlib.sha256(token.encode()).hexdigest()
    token_prefix_visible = token[:12]  # Mostrar los primeros 12 en la UI

    key = ApiKey(
        name=name,
        token_prefix=token_prefix_visible,
        token_hash=token_hash,
        company=company,
        scopes=scopes,
        is_active=True,
        created_by=session['user_id'],
        expires_at=expires_at,
    )
    db.session.add(key)
    db.session.commit()
    log_audit('api_key_created', session['user_id'], 'api_key', key.id,
              f'API Key "{name}" creada para empresa {company} con scopes: {scopes}')

    return jsonify({
        'success': True,
        'id': key.id,
        'name': key.name,
        'token': token,  # SOLO se muestra en la respuesta de creación — el admin debe guardarla
        'token_prefix': token_prefix_visible,
        'company': key.company,
        'scopes': scopes.split(','),
        'message': '⚠️ Guardá este token AHORA. No se puede volver a ver.',
    }), 201


@app.route('/api/admin/api-keys/<int:key_id>/toggle', methods=['POST'])
def api_admin_api_keys_toggle(key_id):
    """Activar/desactivar una API Key."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    key = ApiKey.query.get(key_id)
    if not key or key.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrada'}), 404
    key.is_active = not key.is_active
    db.session.commit()
    log_audit('api_key_toggle', session['user_id'], 'api_key', key.id,
              f'API Key "{key.name}" {"activada" if key.is_active else "desactivada"}')
    return jsonify({'success': True, 'is_active': key.is_active})


@app.route('/api/admin/api-keys/<int:key_id>', methods=['DELETE'])
def api_admin_api_keys_delete(key_id):
    """Eliminar (revocar permanentemente) una API Key."""
    if 'user_id' not in session or session['role'] != 'admin':
        return jsonify({'success': False, 'error': 'No autorizado'}), 401
    key = ApiKey.query.get(key_id)
    if not key or key.company not in admin_companies_scope():
        return jsonify({'success': False, 'error': 'No encontrada'}), 404
    name = key.name
    db.session.delete(key)
    db.session.commit()
    log_audit('api_key_deleted', session['user_id'], 'api_key', key_id,
              f'API Key "{name}" eliminada permanentemente')
    return jsonify({'success': True, 'message': 'API Key eliminada'})


# ═════════════════════════════════════════════════════════════════════════════

def bootstrap_app():
    """Inicializa BD + arranca todos los schedulers en background.
    Llamado desde __main__ (dev server) o desde wsgi.py (Gunicorn)."""
    if app.config.get('_bootstrapped'):
        return  # No re-inicializar (evita duplicación en multi-worker)
    app.config['_bootstrapped'] = True

    init_db()
    start_server_monitoring()
    start_backup_scheduler()
    start_watchdog()
    start_token_cleanup_scheduler()
    start_mailbox_poller()
    start_sla_alert_scheduler()
    # NUEVO: Purga automática de AuditLog (12 meses retención)
    try:
        start_audit_log_purge_scheduler()
    except NameError:
        pass  # función definida más abajo
    # NUEVO: Reportes automáticos (quincenal/mensual/anual)
    try:
        start_report_scheduler()
    except NameError:
        pass

    # Agent Orchestrator
    try:
        from agents import AgentOrchestrator
        with app.app_context():
            orchestrator_instance = AgentOrchestrator(app, db)
            orchestrator_instance.start_background_agents()
            app.config['orchestrator'] = orchestrator_instance
        print("  [OK] Agent Orchestrator inicializado")
    except Exception as e:
        print(f"  [WARN] No se pudo inicializar orchestrator: {e}")


if __name__ == '__main__':
    bootstrap_app()

    print("\n" + "="*70)
    print("DeskEli - Sistema Completo")
    print("="*70)
    print("\nServidor en: http://localhost:5050/")
    print("\nCaracterísticas incluidas:")
    print("  [OK] Selector de empresa (Eliot, Pash, Primatela)")
    print("  [OK] LDAP/Active Directory integrado")
    print("  [OK] JWT + Token Blacklist")
    print("  [OK] Formato TKT-EMPRESA-NUMERO")
    print("  [OK] SLA configurable por prioridad")
    print("  [OK] Búsqueda global (Ctrl+K)")
    print("  [OK] Exportación Excel con colores")
    print("  [OK] 12 Temas visuales")
    print("  [OK] Audit trail completo")
    print("  [OK] Tiempo trabajado con cronómetro")
    print("\nSecurity Fixes aplicados:")
    print("  [OK] Password validation en login")
    print("  [OK] Security headers (HSTS, CSP, X-Frame-Options)")
    print("  [OK] XSS prevention con bleach")
    print("  [OK] CSRF protection con Flask-WTF")
    print("  [OK] CORS limitado a dominios específicos")
    print("  [OK] JWT expiration enforcement")
    print("  [OK] Token blacklist auto-cleanup")
    print("  [OK] Rate limiting mejorado")
    print("  [OK] Company_id validation")
    print("  [OK] Input sanitization")
    print("  [OK] Secure session cookies")
    print("  [OK] Audit log sanitization")
    print("\nPara credenciales de prueba, revisar la página de login")
    print("\n" + "="*70 + "\n")

    # Leer HOST y PORT del entorno
    host = os.getenv('HOST', '0.0.0.0')
    port = int(os.getenv('PORT', 5050))
    socketio.run(app, debug=True, host=host, port=port, allow_unsafe_werkzeug=True)
