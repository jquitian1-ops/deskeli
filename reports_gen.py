"""Generación de reportes ejecutivos automáticos (Excel + PDF con gráficas).

Se invoca desde el scheduler quincenal/mensual/anual o desde un endpoint manual.
Recibe un app context y los modelos. Devuelve BytesIO listos para adjuntar a email.
"""
import io
from datetime import datetime, timedelta
from collections import defaultdict


def _matplotlib_safe():
    """Configura matplotlib en modo headless. Llamar UNA vez antes de cualquier figura."""
    import matplotlib
    matplotlib.use('Agg')


def _fmt_pct(num, total):
    if not total:
        return '0%'
    return f"{(num / total) * 100:.1f}%"


def _collect_metrics(db, Ticket, User, TechnicianProfile, company, period_start, period_end, team_user_ids=None):
    """Reúne métricas del período para una empresa.
    Si team_user_ids está definido (lista no vacía), filtra a tickets cuyo assignee_id esté en esa lista.
    Devuelve dict con KPIs + series para gráficas."""
    q = Ticket.query.filter(
        Ticket.company == company,
        Ticket.created_at >= period_start,
        Ticket.created_at < period_end
    )
    if team_user_ids:
        q = q.filter(Ticket.assignee_id.in_(team_user_ids))
    tickets = q.all()

    total = len(tickets)
    resolved = sum(1 for t in tickets if t.status in ('resolved', 'closed'))
    open_ = sum(1 for t in tickets if t.status == 'open')
    in_progress = sum(1 for t in tickets if t.status == 'in_progress')

    # Por prioridad
    by_priority = defaultdict(int)
    for t in tickets:
        by_priority[t.priority or 'medium'] += 1

    # Por categoría (top 8)
    by_category = defaultdict(int)
    for t in tickets:
        by_category[(t.category or 'General').strip()] += 1
    top_categories = sorted(by_category.items(), key=lambda x: -x[1])[:8]

    # Por técnico
    by_technician = defaultdict(lambda: {'total': 0, 'resolved': 0, 'open': 0, 'critical': 0, 'high': 0})
    for t in tickets:
        if not t.assignee_id:
            continue
        user = User.query.get(t.assignee_id)
        if not user:
            continue
        s = by_technician[user.name]
        s['total'] += 1
        if t.status in ('resolved', 'closed'):
            s['resolved'] += 1
        else:
            s['open'] += 1
        if t.priority == 'critical':
            s['critical'] += 1
        elif t.priority == 'high':
            s['high'] += 1
    tech_ranking = sorted(by_technician.items(), key=lambda x: -x[1]['total'])[:10]

    # Tendencia diaria (creación vs resolución)
    days = (period_end - period_start).days
    daily_created = defaultdict(int)
    daily_resolved = defaultdict(int)
    for t in tickets:
        d = t.created_at.date()
        daily_created[d] += 1
        if t.status in ('resolved', 'closed') and t.updated_at:
            d2 = t.updated_at.date()
            if period_start.date() <= d2 < period_end.date():
                daily_resolved[d2] += 1

    daily_series = []
    cursor = period_start.date()
    while cursor < period_end.date():
        daily_series.append({
            'date': cursor,
            'created': daily_created.get(cursor, 0),
            'resolved': daily_resolved.get(cursor, 0)
        })
        cursor += timedelta(days=1)

    # SLA compliance (simplificado: % resueltos dentro de su SLA)
    sla_on_time = 0
    sla_late = 0
    for t in tickets:
        if t.status not in ('resolved', 'closed'):
            continue
        if not t.sla_deadline:
            continue
        if t.updated_at and t.updated_at <= t.sla_deadline:
            sla_on_time += 1
        else:
            sla_late += 1
    sla_total = sla_on_time + sla_late
    sla_compliance = (sla_on_time / sla_total * 100) if sla_total else 0

    return {
        'period_start': period_start,
        'period_end': period_end,
        'company': company,
        'total': total,
        'resolved': resolved,
        'open': open_,
        'in_progress': in_progress,
        'by_priority': dict(by_priority),
        'top_categories': top_categories,
        'tech_ranking': tech_ranking,
        'daily_series': daily_series,
        'sla_on_time': sla_on_time,
        'sla_late': sla_late,
        'sla_compliance': sla_compliance,
    }


def _generate_chart_pngs(metrics):
    """Genera 4 imágenes PNG en memoria con las gráficas clave.
    Retorna dict {name: BytesIO}."""
    _matplotlib_safe()
    import matplotlib.pyplot as plt

    charts = {}

    # 1. Tendencia diaria (líneas)
    fig, ax = plt.subplots(figsize=(9, 4.2))
    dates = [d['date'] for d in metrics['daily_series']]
    created = [d['created'] for d in metrics['daily_series']]
    resolved = [d['resolved'] for d in metrics['daily_series']]
    ax.plot(dates, created, marker='o', color='#2563eb', label='Creados', linewidth=2)
    ax.plot(dates, resolved, marker='s', color='#16a34a', label='Resueltos', linewidth=2)
    ax.set_title('Tendencia: Creación vs Resolución', fontweight='bold')
    ax.set_xlabel('Fecha')
    ax.set_ylabel('Tickets')
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.autofmt_xdate()
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    charts['trend'] = buf

    # 2. Por prioridad (barras)
    fig, ax = plt.subplots(figsize=(7, 3.8))
    prio_order = ['critical', 'high', 'medium', 'low']
    prio_labels = {'critical': '🔴 Crítica', 'high': '🟠 Alta', 'medium': '🟡 Media', 'low': '🟢 Baja'}
    prio_colors = {'critical': '#dc2626', 'high': '#ea580c', 'medium': '#f59e0b', 'low': '#16a34a'}
    counts = [metrics['by_priority'].get(p, 0) for p in prio_order]
    labels = [prio_labels[p] for p in prio_order]
    colors = [prio_colors[p] for p in prio_order]
    bars = ax.bar(labels, counts, color=colors)
    for b, v in zip(bars, counts):
        if v > 0:
            ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.3, str(v), ha='center', fontweight='bold')
    ax.set_title('Distribución por Prioridad', fontweight='bold')
    ax.set_ylabel('Tickets')
    ax.grid(True, axis='y', alpha=0.3)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    charts['priority'] = buf

    # 3. Top categorías (barras horizontales)
    fig, ax = plt.subplots(figsize=(8, 4.2))
    if metrics['top_categories']:
        cats = [c[0][:22] for c in metrics['top_categories']]
        vals = [c[1] for c in metrics['top_categories']]
        bars = ax.barh(cats[::-1], vals[::-1], color='#3b82f6')
        for b, v in zip(bars, vals[::-1]):
            ax.text(b.get_width() + 0.2, b.get_y() + b.get_height() / 2, str(v), va='center', fontweight='bold')
    ax.set_title('Top Categorías Más Reportadas', fontweight='bold')
    ax.set_xlabel('Cantidad')
    ax.grid(True, axis='x', alpha=0.3)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    charts['categories'] = buf

    # 4. Top técnicos (barras horizontales)
    fig, ax = plt.subplots(figsize=(8, 4.2))
    if metrics['tech_ranking']:
        names = [t[0][:20] for t in metrics['tech_ranking']]
        vals = [t[1]['total'] for t in metrics['tech_ranking']]
        resolved_vals = [t[1]['resolved'] for t in metrics['tech_ranking']]
        y_pos = range(len(names))
        ax.barh(y_pos, vals, color='#94a3b8', label='Total asignados')
        ax.barh(y_pos, resolved_vals, color='#16a34a', label='Resueltos')
        ax.set_yticks(y_pos)
        ax.set_yticklabels(names)
        ax.invert_yaxis()
        ax.legend()
    ax.set_title('Ranking de Técnicos por Productividad', fontweight='bold')
    ax.set_xlabel('Tickets')
    ax.grid(True, axis='x', alpha=0.3)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    charts['technicians'] = buf

    # 5. SLA Compliance (donut)
    fig, ax = plt.subplots(figsize=(5, 4.2))
    if metrics['sla_on_time'] + metrics['sla_late'] > 0:
        sizes = [metrics['sla_on_time'], metrics['sla_late']]
        labels = ['En plazo', 'Vencidos']
        colors = ['#16a34a', '#dc2626']
        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90,
               wedgeprops={'width': 0.4, 'edgecolor': 'white'})
        ax.text(0, 0, f"{metrics['sla_compliance']:.0f}%", ha='center', va='center',
                fontsize=22, fontweight='bold', color='#16a34a')
    else:
        ax.text(0.5, 0.5, 'Sin tickets cerrados\nen el período', ha='center', va='center', transform=ax.transAxes)
        ax.axis('off')
    ax.set_title('Cumplimiento SLA', fontweight='bold')
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    charts['sla'] = buf

    return charts


def generate_excel_report(metrics, period_label, company_display):
    """Genera archivo Excel con varias hojas y gráficas embebidas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Estilos comunes
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1f2937', end_color='1f2937', fill_type='solid')
    title_font = Font(bold=True, size=18, color='1f2937')
    subtitle_font = Font(bold=True, size=14, color='374151')
    thin_border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )

    # === HOJA 1: RESUMEN EJECUTIVO ===
    ws = wb.active
    ws.title = 'Resumen Ejecutivo'
    ws['A1'] = f'📊 Reporte {period_label} — {company_display}'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A2'] = f"Período: {metrics['period_start'].strftime('%d/%m/%Y')} → {metrics['period_end'].strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(italic=True, color='6b7280')
    ws.merge_cells('A2:E2')

    # KPIs
    kpis = [
        ('Total tickets', metrics['total'], '3b82f6'),
        ('Resueltos', metrics['resolved'], '16a34a'),
        ('En curso', metrics['in_progress'], 'f59e0b'),
        ('Abiertos sin asignar', metrics['open'], 'dc2626'),
        ('Cumplimiento SLA', f"{metrics['sla_compliance']:.1f}%", '16a34a' if metrics['sla_compliance'] >= 85 else 'dc2626'),
    ]
    ws['A4'] = 'INDICADORES CLAVE'
    ws['A4'].font = subtitle_font
    for i, (label, val, color) in enumerate(kpis, start=5):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True, size=11)
        ws[f'B{i}'] = val
        ws[f'B{i}'].font = Font(bold=True, size=14, color=color)
        ws[f'B{i}'].alignment = Alignment(horizontal='right')

    # === HOJA 2: PRIORIDADES ===
    ws2 = wb.create_sheet('Por Prioridad')
    ws2['A1'] = 'Distribución por Prioridad'
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:C1')
    headers = ['Prioridad', 'Cantidad', '% del total']
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border
    prio_order = ['critical', 'high', 'medium', 'low']
    prio_labels = {'critical': '🔴 Crítica', 'high': '🟠 Alta', 'medium': '🟡 Media', 'low': '🟢 Baja'}
    row = 4
    for p in prio_order:
        v = metrics['by_priority'].get(p, 0)
        ws2.cell(row=row, column=1, value=prio_labels[p]).border = thin_border
        ws2.cell(row=row, column=2, value=v).border = thin_border
        ws2.cell(row=row, column=3, value=_fmt_pct(v, metrics['total'])).border = thin_border
        row += 1

    # === HOJA 3: CATEGORÍAS ===
    ws3 = wb.create_sheet('Top Categorías')
    ws3['A1'] = 'Top 8 Categorías Más Reportadas'
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:C1')
    for col, h in enumerate(['#', 'Categoría', 'Tickets', '% del total'], start=1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    for i, (cat, count) in enumerate(metrics['top_categories'], start=4):
        ws3.cell(row=i, column=1, value=i - 3).border = thin_border
        ws3.cell(row=i, column=2, value=cat).border = thin_border
        ws3.cell(row=i, column=3, value=count).border = thin_border
        ws3.cell(row=i, column=4, value=_fmt_pct(count, metrics['total'])).border = thin_border

    # === HOJA 4: TÉCNICOS ===
    ws4 = wb.create_sheet('Ranking Técnicos')
    ws4['A1'] = 'Productividad por Técnico'
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:F1')
    headers4 = ['Posición', 'Técnico', 'Total', 'Resueltos', 'En curso', '% Resolución', 'Críticos', 'Altas']
    for col, h in enumerate(headers4, start=1):
        c = ws4.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
    for i, (name, s) in enumerate(metrics['tech_ranking'], start=4):
        resolution_rate = (s['resolved'] / s['total'] * 100) if s['total'] else 0
        ws4.cell(row=i, column=1, value=i - 3).border = thin_border
        ws4.cell(row=i, column=2, value=name).border = thin_border
        ws4.cell(row=i, column=3, value=s['total']).border = thin_border
        ws4.cell(row=i, column=4, value=s['resolved']).border = thin_border
        ws4.cell(row=i, column=5, value=s['open']).border = thin_border
        ws4.cell(row=i, column=6, value=f"{resolution_rate:.0f}%").border = thin_border
        ws4.cell(row=i, column=7, value=s['critical']).border = thin_border
        ws4.cell(row=i, column=8, value=s['high']).border = thin_border

    # === HOJA 5: GRÁFICAS ===
    ws5 = wb.create_sheet('Gráficas')
    ws5['A1'] = '📊 Gráficas del período'
    ws5['A1'].font = title_font
    charts = _generate_chart_pngs(metrics)
    row_anchor = 3
    for name, label in [
        ('trend', 'Tendencia diaria'),
        ('priority', 'Distribución por prioridad'),
        ('categories', 'Top categorías'),
        ('technicians', 'Ranking técnicos'),
        ('sla', 'Cumplimiento SLA')
    ]:
        ws5.cell(row=row_anchor, column=1, value=label).font = subtitle_font
        try:
            img = XLImage(charts[name])
            img.anchor = f'A{row_anchor + 1}'
            ws5.add_image(img)
        except Exception as e:
            ws5.cell(row=row_anchor + 1, column=1, value=f'[Error gráfica: {e}]')
        row_anchor += 25  # espacio entre gráficas

    # Auto-anchos
    for sheet in wb.worksheets:
        for col in range(1, 9):
            sheet.column_dimensions[get_column_letter(col)].width = 22

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def generate_pdf_report(metrics, period_label, company_display):
    """Genera PDF con KPIs, tablas y gráficas embebidas."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image,
                                    Table, TableStyle, PageBreak)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2 * cm, rightMargin=2 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                 fontSize=22, textColor=colors.HexColor('#1f2937'),
                                 alignment=1, spaceAfter=10)
    subtitle = ParagraphStyle('Subtitle', parent=styles['Normal'],
                              fontSize=12, textColor=colors.HexColor('#6b7280'),
                              alignment=1, spaceAfter=18)
    h2 = ParagraphStyle('H2', parent=styles['Heading2'],
                        fontSize=15, textColor=colors.HexColor('#1f2937'), spaceBefore=12, spaceAfter=8)

    story = []
    story.append(Paragraph(f'📊 Reporte {period_label}', title_style))
    story.append(Paragraph(f'<b>{company_display}</b>', subtitle))
    story.append(Paragraph(
        f"Período: {metrics['period_start'].strftime('%d/%m/%Y')} → {metrics['period_end'].strftime('%d/%m/%Y')}",
        subtitle))

    # KPIs en tabla 2 columnas
    story.append(Paragraph('Indicadores Clave (KPIs)', h2))
    kpi_data = [
        ['Total tickets', str(metrics['total'])],
        ['Resueltos', f"{metrics['resolved']} ({_fmt_pct(metrics['resolved'], metrics['total'])})"],
        ['En curso', f"{metrics['in_progress']} ({_fmt_pct(metrics['in_progress'], metrics['total'])})"],
        ['Abiertos sin asignar', f"{metrics['open']} ({_fmt_pct(metrics['open'], metrics['total'])})"],
        ['Cumplimiento SLA', f"{metrics['sla_compliance']:.1f}%"],
    ]
    kpi_table = Table(kpi_data, colWidths=[8 * cm, 6 * cm])
    kpi_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1f2937')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 14))

    # Generar gráficas
    charts = _generate_chart_pngs(metrics)

    # Tendencia
    story.append(Paragraph('Tendencia: Creación vs Resolución', h2))
    story.append(Image(charts['trend'], width=17 * cm, height=8 * cm))
    story.append(Spacer(1, 8))

    # Prioridad
    story.append(Paragraph('Distribución por Prioridad', h2))
    story.append(Image(charts['priority'], width=14 * cm, height=7 * cm))
    story.append(PageBreak())

    # Categorías
    story.append(Paragraph('Top Categorías Más Reportadas', h2))
    story.append(Image(charts['categories'], width=17 * cm, height=8 * cm))
    if metrics['top_categories']:
        cat_data = [['#', 'Categoría', 'Tickets', '% Total']]
        for i, (cat, count) in enumerate(metrics['top_categories'], start=1):
            cat_data.append([str(i), cat, str(count), _fmt_pct(count, metrics['total'])])
        cat_table = Table(cat_data, colWidths=[1.5 * cm, 9 * cm, 3 * cm, 3 * cm])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
        ]))
        story.append(Spacer(1, 8))
        story.append(cat_table)
    story.append(PageBreak())

    # Técnicos
    story.append(Paragraph('Ranking de Técnicos por Productividad', h2))
    story.append(Image(charts['technicians'], width=17 * cm, height=8 * cm))
    if metrics['tech_ranking']:
        tech_data = [['#', 'Técnico', 'Total', 'Resueltos', '% Res.', 'Crit.', 'Altas']]
        for i, (name, s) in enumerate(metrics['tech_ranking'], start=1):
            rate = (s['resolved'] / s['total'] * 100) if s['total'] else 0
            tech_data.append([str(i), name, str(s['total']), str(s['resolved']),
                              f"{rate:.0f}%", str(s['critical']), str(s['high'])])
        tech_table = Table(tech_data,
                           colWidths=[1.2 * cm, 6 * cm, 2 * cm, 2.2 * cm, 2 * cm, 1.8 * cm, 1.8 * cm])
        tech_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ]))
        story.append(Spacer(1, 8))
        story.append(tech_table)
    story.append(PageBreak())

    # SLA
    story.append(Paragraph('Cumplimiento de SLA', h2))
    story.append(Image(charts['sla'], width=10 * cm, height=8 * cm))
    story.append(Paragraph(
        f"<b>Meta:</b> ≥ 95%. <b>Resultado del período:</b> {metrics['sla_compliance']:.1f}%",
        styles['Normal']
    ))

    doc.build(story)
    buf.seek(0)
    return buf


def build_email_body(metrics, period_label, company_display, recipient_name):
    """HTML del cuerpo del email."""
    period_str = f"{metrics['period_start'].strftime('%d/%m/%Y')} → {metrics['period_end'].strftime('%d/%m/%Y')}"
    sla_color = '#16a34a' if metrics['sla_compliance'] >= 85 else ('#f59e0b' if metrics['sla_compliance'] >= 70 else '#dc2626')
    top_cat_html = ''
    for i, (cat, count) in enumerate(metrics['top_categories'][:5], start=1):
        top_cat_html += f'<li><strong>{cat}</strong>: {count} tickets ({_fmt_pct(count, metrics["total"])})</li>'

    return f"""
    <html><body style="font-family: Segoe UI, Arial, sans-serif; color: #1f2937; max-width: 720px; margin: 0 auto;">
        <div style="background: linear-gradient(135deg, #1f2937, #7c3aed); color: white; padding: 28px; border-radius: 10px 10px 0 0;">
            <h1 style="margin: 0; font-size: 24px;">📊 DeskEli — Reporte {period_label}</h1>
            <p style="margin: 5px 0 0; opacity: 0.95;">{company_display} · {period_str}</p>
        </div>
        <div style="background: white; padding: 24px; border: 1px solid #e5e7eb; border-radius: 0 0 10px 10px;">
            <p>Hola <strong>{recipient_name}</strong>,</p>
            <p>Adjuntamos el <strong>reporte {period_label.lower()}</strong> de tu área. Resumen rápido:</p>

            <table style="width: 100%; border-collapse: collapse; margin: 16px 0;">
                <tr>
                    <td style="padding: 12px; background: #eff6ff; border-radius: 6px; text-align: center; width: 25%;">
                        <div style="font-size: 28px; font-weight: 800; color: #2563eb;">{metrics['total']}</div>
                        <div style="font-size: 12px; color: #1e40af;">Tickets totales</div>
                    </td>
                    <td style="width: 2%;"></td>
                    <td style="padding: 12px; background: #ecfdf5; border-radius: 6px; text-align: center; width: 25%;">
                        <div style="font-size: 28px; font-weight: 800; color: #16a34a;">{metrics['resolved']}</div>
                        <div style="font-size: 12px; color: #065f46;">Resueltos</div>
                    </td>
                    <td style="width: 2%;"></td>
                    <td style="padding: 12px; background: #fef3c7; border-radius: 6px; text-align: center; width: 25%;">
                        <div style="font-size: 28px; font-weight: 800; color: #d97706;">{metrics['in_progress']}</div>
                        <div style="font-size: 12px; color: #92400e;">En curso</div>
                    </td>
                </tr>
                <tr>
                    <td colspan="5" style="padding: 14px; background: #f9fafb; border-radius: 6px; text-align: center; margin-top: 10px;">
                        <div style="font-size: 14px; color: #4b5563;">Cumplimiento SLA</div>
                        <div style="font-size: 32px; font-weight: 800; color: {sla_color};">{metrics['sla_compliance']:.1f}%</div>
                    </td>
                </tr>
            </table>

            <h3 style="color: #1f2937;">Top 5 categorías del período</h3>
            <ol>{top_cat_html or '<li>Sin datos en el período.</li>'}</ol>

            <p style="margin-top: 18px; padding: 12px; background: #f9fafb; border-left: 3px solid #7c3aed; border-radius: 4px;">
                📎 El <strong>análisis completo con gráficas</strong> está en los archivos adjuntos:
                <br>• <strong>Excel</strong>: hojas detalladas + gráficas embebidas
                <br>• <strong>PDF</strong>: reporte ejecutivo formato impresión
            </p>

            <p style="font-size: 12px; color: #6b7280; margin-top: 20px;">
                Reporte generado automáticamente por DeskEli. Si tenés dudas, contactá al área de TI.
            </p>
        </div>
    </body></html>
    """
